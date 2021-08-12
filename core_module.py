import openpyxl
import time
import random
import math
import pandas as pd
import re


class coreRAPID:
    """Initialises all variables and functions used during run"""

    def __init__(self):

        # Generate output workbook
        self.workbook = openpyxl.load_workbook(v['filename'], data_only=True)
        self.generateOutputWorkbook(v['debugTab'])

        # --------------------------- Import Reference Data -------------------------- #

        # AROT/DROT Lookup Table
        self.rot_lookup = pd.read_csv('utility/AROTDROT_distributions.csv').drop(columns=['Unnamed: 0'])

        # Actual Speed Profiles
        self.speed_profile = pd.read_csv('utility/actual_speed_profiles.csv').drop(columns=['Unnamed: 0'])

        # Aircraft wake lookup (contains UK, RECAT-EU and RECAT20)
        self.wake_lookup = pd.read_csv('utility/wake.csv')

        # Wake separation tables
        self.wake_separations = {
            'RECAT-EU': pd.read_csv('utility/RECAT_EU_separation.csv').set_index("LEAD"),
            'UK': pd.read_csv('utility/UK_wake_separation.csv').set_index("LEAD"),
            'RECAT_PWS': pd.read_csv('utility/RECAT_PWS.csv').fillna(0).set_index('FOLLOW'),
            'RECAT20': pd.read_csv('utility/RECAT20_separation.csv').fillna(0).set_index('LEAD')
        }

        # ------------------------------- Run Constants ------------------------------ #

        # Miscellaneous wake rules
        self.WAKE_RULES = {
            # Departure Wake Rules - SOME UNUSED
            'H_H_d': 90,
            'H_M_d': 139, # Used for H_ UM/M/S/L
            'J_H_d': 139, #120
            'J_M_d': 204, #180 #2016 data
            'J_S_d': 204, #180 #2016 data
            'J_L_d': 204, #180 #2016 data
            'M_L_d': 139, #Used for UM_L & M_L
            'S_L_d': 139, #120

            # (DBS) ICAO wake rules. obs: L=S and M=UM - ALL UNUSED
            'j_h': 6,
            'j_m': 7,
            'j_l': 8,
            'h_h': 4,
            'h_m': 5,
            'h_l': 6,
            'm_l': 5,

            # (DBS) RECAT separation (RECAT-EU edition 1.1: 15/07/2015) - ALL UNUSED
            'A_A': 3, #same for : B_B , C_C , C_D , F_F
            'A_B': 4, #same for : B_C , B_D , C_E , E_F
            'A_C': 5, #same for : A_D , B_E , D_F
            'A_D': 6, #same for : A_E , C_F
            'A_F': 8,
            'B_F': 7,

            # (TBS) RECAT separation (RECAT-EU edition 1.1: 15/07/2015) - ALL UNUSED
            'a_b': 100, #same for : B_C , B_D , C_E , E_F
            'a_c': 120, #same for : B_E , C_F , D_F
            'a_d': 140, #same for : B_F
            'a_e': 160,
            'a_f': 180,
            'c_d': 80 #same for : F_F
        }

        # SID group separation
        self.SID_GROUPS = re.findall(r'\d+', v['SIDgroup_separation']) # Input example: (2,4)(3,4)
        self.SID_GROUPS = [[int(self.SID_GROUPS[x]), int(self.SID_GROUPS[y])] for (x, y) in [(0, 1), (2, 3), (3, 2), (1, 0)]]

        # SID queue
        self.SID_QUEUES = re.findall(r'\d+', v['SID_queue_assign']) # Input example: 1 2 | 3 4
        self.SID_QUEUES = [[int(self.SID_QUEUES[x]), int(self.SID_QUEUES[y])] for (x, y) in [(0, 1), (2, 3)]]

        # Get initial number of arrivals and departures
        self.TOTAL_ARRIVALS = self.workbook['Arrivals'].max_row
        self.TOTAL_DEPARTURES = self.workbook['Departures'].max_row

        self.C_DME = 4
        self.D_DME = 3

        self.MIN_RADAR_SEP_DIST = 3 # NM

        # Standard Taxi Time - used for arrivals.
        # Actually, it is from the landing point to the stand
        self.STT = 600

        self.X_BUFFER = 15

        # ------------------------------- Run Variables ------------------------------ #

        self.RWY_status = "E"

        self.ARRIVALqueue = {} # Dict for storing Taxiing-in Arrivals

        self.ArrHOLDqueue = {} # Dict for the Arrivals hold queu

        self.APPqueue = {}

        self.DepSTANDqueue = {} # for holding Departures on Stands (Push/Start Delay)

        self.TAXIqueue = {}
        self.TAXIhold = {}

        # Place Dep A/C with SID group 1 into self.RWYqueue1 etc
        # UNLESS there's no A/C of this type available
        # NOTE: Queues 3 and 4 unused, instigated for 4x2 (and future 8x1) arrangements
        self.RWYqueue1 = {}
        self.RWYqueue2 = {}
        self.RWYqueue3 = {}
        self.RWYqueue4 = {}

        self.currentGap = 86400

        # ARRIVALS
        self.ARRkey = 2
        self.ArrOutput = 2

        # DEPARTURES
        self.SOBTrow = 2
        self.DepOutput = 2
        self.seqRow = 2

        # --------------------- Execute pre-processing functions --------------------- #

        self.Arrival_Input_pre_process()
        self.Departure_Input_pre_process()

        # Start and end times
        if self.readCell('Arrivals', 'A', 2) is None and self.readCell('Departures', 'A', 2) is None:
            raise ValueError('No Arrival or Departure rows found.')
        elif self.readCell('Arrivals', 'A', 2) is None: # no arrivals
            self.start_time = self.readCell('Departures', 'C', 2) - 3000
            self.end_time = self.readCell('Departures', 'C', self.TOTAL_DEPARTURES) + 10000
        elif self.readCell('Departures', 'A', 2) is None: # no departures
            self.start_time = self.readCell('Arrivals', 'C', 2) - 3000
            self.end_time = self.readCell('Arrivals', 'C', self.TOTAL_ARRIVALS) + 10000
        else:
            self.start_time = min(self.readCell('Arrivals', 'C', 2), self.readCell('Departures', 'C', 2)) - 3000
            self.end_time = min(self.readCell('Arrivals', 'C', self.TOTAL_ARRIVALS), self.readCell('Departures', 'C', self.TOTAL_DEPARTURES)) + 10000

        # ------------------ Throughput/Delay Calculation Variables ------------------ #

        self.throughput = []

        self.GoAroundCount = {} # Referenced during run

        self.number_of_goArounds_queued = 0 # Arrival 'go-around' case


    # -------------------- Workbook and Data Import Functions -------------------- #

    def columnLetter(self, col):
        """Checks and converts column index to letter"""
        if isinstance(col, str):
            return col
        else:
            return openpyxl.utils.get_column_letter(int(col))


    def columnIndex(self, col):
        """Checks and converts column letter to index"""
        if isinstance(col, str):
            return openpyxl.utils.column_index_from_string(col)
        else:
            return col


    def readCell(self, ws, col, row):
        """Handy and compact wrapper for getting values from workbook cells"""
        return self.workbook[ws][self.columnLetter(col) + str(row)].value


    def writeCell(self, ws, col, row, new_value):
        """Handy and compact wrapper for writing values to workbook cells"""
        self.workbook[ws][self.columnLetter(col) + str(row)].value = new_value


    def copyCell(self, ws1, col1, row1, ws2, col2, row2):
        """Copies one cell to another location"""
        self.workbook[ws1][self.columnLetter(col1) + str(row1)].value = self.workbook[ws2][self.columnLetter(col2) + str(row2)].value


    def readRow(self, ws, row, start_col, col_range):
        """Read a worksheet row to a list"""
        output_list = []
        for i in range(1, col_range):
            output_list.append(self.workbook[ws][self.columnLetter(self.columnIndex(start_col) + i) + str(row)].value)
        return output_list


    def writeRow(self, ws, row, start_col, value_list):
        """
        Writes a list to a worksheet row
        Creates worksheet if not exist
        """
        if ws not in self.workbook.sheetnames:
            self.workbook.create_sheet(ws)
        for i in range(len(value_list)):
            if value_list[i] is not None:
                self.workbook[ws][self.columnLetter(self.columnIndex(start_col) + i) + str(row)].value = value_list[i]


    def generateOutputWorkbook(self, createDebugTab):
        """Creates new sheets and columns in Workbook (new header columns are defined here)"""

        self.writeRow('Arrivals', 1, 30, [
            'GS_0_1dme', 'GS_1_2dme', 'GS_2_3dme', 'GS_3_4dme', 'GS_4_5dme',
            'GS_5_6dme', 'GS_6_7dme', 'GS_7_8dme', 'GS_8_9dme', 'GS_9_10dme',
            'IAS_0_1dme', 'IAS_1_2dme', 'IAS_2_3dme', 'IAS_3_4dme', 'IAS_4_5dme',
            'IAS_5_6dme', 'IAS_6_7dme', 'IAS_7_8dme', 'IAS_8_9dme', 'IAS_9_10dme'
        ])

        self.writeRow('Runway_calcs', 1, 1, [
            'Arrival ID', 'TAXI-IN', 'AROT', 'ADA', 'ADDA',
            'ATCO variability', 'WIND1', 'SPEED1', 'WIND2', 'SPEED2',
            'VTGT', 'SAE', 'PREDICTED Landing Time', 'MAX Constraint', 'MAX Constraint Label',
            '', '', 'WAKE SEPARATION', 'MRS', 'Departure ID',
            'TAXI-OUT', 'DROT', 'ARRIVAL actual WAKE'
        ])

        self.writeRow('Arrival_Output', 1, 1, [
            'Arrival ID', 'Arrival HOUR', 'ACTUAL Landing Time', 'Arrival RWY_EXIT', 'WAKE',
            'In Blocks Time', 'AROT', 'TAXI-IN Duration', 'MAX Constraint', 'MAX Constraint Label',
            'len(ArrHOLDqueue)', 'Arrival DELAY'
        ])

        self.writeRow('Departure_Output', 1, 1, [
            'Departure ID', 'Departure HOUR', 'Departure_RWY_ENTRY', 'Departure_RWY_EXIT', 'WAKE',
            'SID GROUP', 'DROT', 'TAXI-OUT', 'Dep MIN Separation', 'Dep MIN Separation Label',
            'currentGap', 'len(DepSTANDqueue)', 'len(TAXIhold)', 'len(RWYqueue1)', 'len(RWYqueue2)',
            'len(RWYqueue3)', 'len(RWYqueue4)', 'DELAY DepSTANDqueue', 'DELAY TAXIhold', 'DELAY RWYqueue',
            'RWY queue USED'
        ])

        self.writeRow('Throughput', 1, 1, [
            'Hour', 'Departure Throughput', 'Arrival Throughput', 'Total Throughput', 'Cum. No. of Go-Arounds'
        ])

        self.writeRow('Delay', 1, 1, [
            'Departure ID', 'HOUR', 'RWY HOLD Delay', 'Push/Start Delay', '',
            '', '', '', 'Arrival ID', 'HOUR', 'Arrival Delay'
        ])

        self.writeRow('Sequence', 1, 1, [
            'Type', 'ID', 'RWY ENTRY', 'RWY EXIT', 'ROT',
            'Arr ID start ADA pair', 'ADA Buffer'
        ])

        if createDebugTab:
            self.writeRow('Debug', 1, 1, [
                'Time', 'Runway status', 'Current Gap - D', 'Current Gap - A', 'Current Gap - E',
                '', '', '', '', '', '',  'Arrival Hold Delay'
            ])


    # ----------------------- Arrival Separation Functions ----------------------- #

    def distance_to_time_assumed_speed_profile_IAS(self, i, distance): #DELIVERED at THR
        #fixed self.D_DME at 3dme, variable self.C_DME because max deceleration speed is 20kts/NM

        ##### JI - THESE THREE LINES LOOK FISHY!
        deceleration_difference = self.readCell('Runway_calcs', 'H', i) - self.readCell('Runway_calcs', 'K', i)
        if deceleration_difference > 20 :
            self.C_DME = deceleration_difference / 20
        #####

        TBS_assumed_speed_profile_value = 0
        #time between self.D_DME - THR:
        t1 = (self.D_DME *3600)/(self.readCell('Runway_calcs', 'K', i))
        #time between self.C_DME - self.D_DME:
        t2 = (2*3600*(self.C_DME-self.D_DME))/(self.readCell('Runway_calcs', 'H', i)+self.readCell('Runway_calcs', 'K', i))

        if distance >= self.C_DME:
            TBS_assumed_speed_profile_value = int(t1+t2+((distance-self.C_DME)*3600/self.readCell('Runway_calcs', 'H', i)))
        elif (distance < self.C_DME) and (distance>self.D_DME):
            d1 = distance-self.D_DME
            speed_at_d1 = (d1*( self.readCell('Runway_calcs', 'H', i) - self.readCell('Runway_calcs', 'K', i) ) /(self.C_DME-self.D_DME)) + self.readCell('Runway_calcs', 'K', i)
            TBS_assumed_speed_profile_value = int(d1*3600*2/(speed_at_d1 + self.readCell('Runway_calcs', 'K', i)) + t1)
        elif distance <= self.D_DME:
            TBS_assumed_speed_profile_value = int(distance*3600/self.readCell('Runway_calcs', 'K', i))
        #print('TBS - on' )

        return TBS_assumed_speed_profile_value


    def DBS_assumed_speed_profile(self, i, distance): # UNUSED!!! (DELIVERED at THR)

        deceleration_difference= (self.readCell('Runway_calcs', 'H', i) - self.readCell('Runway_calcs', 'G', i)) - (self.readCell('Runway_calcs', 'K', i) - self.readCell('Runway_calcs', 'I', i))
        if deceleration_difference > 20 :
            self.C_DME = deceleration_difference / 20

        DBS_assumed_speed_profile_value = 0
        #time between self.D_DME - THR:
        t1 = (self.D_DME *3600)/(self.readCell('Runway_calcs', 'K', i)-self.readCell('Runway_calcs', 'I', i))
        #time between self.C_DME - self.D_DME:
        t2 = (2*3600*(self.C_DME-self.D_DME))/((self.readCell('Runway_calcs', 'H', i)-self.readCell('Runway_calcs', 'G', i))+(self.readCell('Runway_calcs', 'K', i)-self.readCell('Runway_calcs', 'I', i)))

        if distance >= self.C_DME:
            DBS_assumed_speed_profile_value = int(t1+t2+((distance-self.C_DME)*3600/(self.readCell('Runway_calcs', 'H', i)-self.readCell('Runway_calcs', 'G', i))))
        elif (distance < self.C_DME) and (distance>self.D_DME):
            d1 = distance-self.D_DME
            speed_at_d1 = (d1*( (self.readCell('Runway_calcs', 'H', i)-self.readCell('Runway_calcs', 'G', i)) - (self.readCell('Runway_calcs', 'K', i)-self.readCell('Runway_calcs', 'I', i)) ) /(self.C_DME-self.D_DME)) + (self.readCell('Runway_calcs', 'K', i)-self.readCell('Runway_calcs', 'I', i))
            DBS_assumed_speed_profile_value = int(d1*3600*2/(speed_at_d1 + self.readCell('Runway_calcs', 'K', i) - self.readCell('Runway_calcs', 'I', i)) + t1)
        elif distance <= self.D_DME:
            DBS_assumed_speed_profile_value = int(distance*3600/(self.readCell('Runway_calcs', 'K', i)-self.readCell('Runway_calcs', 'I', i)))

        return DBS_assumed_speed_profile_value


    def DBS_actual_speed_profile(self, distance, row): #DELIVERED at THR # use GS

        T=0

        def full_segments(n,row):
            T= 0
            if n >= 1:
                T = 2*3600/(self.readCell('Arrivals', 'AD', row)+self.readCell('Arrivals', 'AE', row))
                if n >=2:
                    T += 2*3600/(self.readCell('Arrivals', 'AE', row) +self.readCell('Arrivals', 'AF', row))
                    if n>=3:
                        T += 2*3600/(self.readCell('Arrivals', 'AF', row) + self.readCell('Arrivals', 'AG', row))
                        if n>=4:
                            T += 2*3600/(self.readCell('Arrivals', 'AG', row) + self.readCell('Arrivals', 'AH', row))
                            if n>=5:
                                T += 2*3600/(self.readCell('Arrivals', 'AH', row)+self.readCell('Arrivals', 'AI', row))
                                if n>=6:
                                    T += 2*3600/(self.readCell('Arrivals', 'AI', row)+self.readCell('Arrivals', 'AJ', row))
                                    if n>=7:
                                        T += 2*3600/(self.readCell('Arrivals', 'AJ', row)+self.readCell('Arrivals', 'AK', row))
                                        if n>=8:
                                            T += 2*3600/(self.readCell('Arrivals', 'AK', row)+self.readCell('Arrivals', 'AL', row))
                                            if n==9:
                                                T += 2*3600/(self.readCell('Arrivals', 'AL', row) + self.readCell('Arrivals', 'AM', row))
                                            elif n>9:
                                                T += (n-9)*3600/self.readCell('Arrivals', 'AM', row)
            return T


        def fraction_of_segments(n,f,row):
            T = 0
            if n== 1:
                S = f*(self.readCell('Arrivals', 'AF', row) - self.readCell('Arrivals', 'AE', row)) + self.readCell('Arrivals', 'AE', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AE', row))
            elif n==2:
                S = f*(self.readCell('Arrivals', 'AG', row) - self.readCell('Arrivals', 'AF', row)) + self.readCell('Arrivals', 'AF', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AF', row))
            elif n==3:
                S = f*(self.readCell('Arrivals', 'AH', row) - self.readCell('Arrivals', 'AG', row)) + self.readCell('Arrivals', 'AG', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AG', row))
            elif n==4:
                S = f*(self.readCell('Arrivals', 'AI', row) - self.readCell('Arrivals', 'AH', row)) + self.readCell('Arrivals', 'AH', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AH', row))
            elif n==5:
                S = f*(self.readCell('Arrivals', 'AJ', row) - self.readCell('Arrivals', 'AI', row)) + self.readCell('Arrivals', 'AI', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AI', row))
            elif n==6:
                S = f*(self.readCell('Arrivals', 'AK', row) - self.readCell('Arrivals', 'AJ', row)) + self.readCell('Arrivals', 'AJ', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AJ', row))
            elif n==7:
                S = f*(self.readCell('Arrivals', 'AL', row) - self.readCell('Arrivals', 'AK', row)) + self.readCell('Arrivals', 'AK', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AK', row))
            elif n==8:
                S = f*(self.readCell('Arrivals', 'AM', row) - self.readCell('Arrivals', 'AL', row)) + self.readCell('Arrivals', 'AL', row)
                T = (f*3600)/(self.readCell('Arrivals', 'AM', row))
            return T


        # if distance > 0:
        X = distance + self.readCell('Runway_calcs', 'F', row) # Actual distance + ATCO var
        D = X - 0.5
        if D <0:
            T = (X*3600)/self.readCell('Arrivals', 'AD', row)
        elif D > 0:
            v['n'] = math.floor(D)
            f = D - v['n']
            T1 = full_segments(v['n'],row)
            if (f != 0) and (v['n']<=8):
                T2 = fraction_of_segments(v['n'],f,row)
                T = T1 + T2 + (0.5*3600)/self.readCell('Arrivals', 'AD', row)
            else:
                T = T1 + (0.5*3600)/self.readCell('Arrivals', 'AD', row)
        return T


    def TBS_actual_speed_profile(self, distance, row): #DELIVERED at THR # use IAS
        def full_segments(n,row):
            if n >= 1:
                T = 2*3600/(self.readCell('Arrivals', 'AN', row)+self.readCell('Arrivals', 'AO', row))
                if n >=2:
                    T += 2*3600/(self.readCell('Arrivals', 'AO', row)+self.readCell('Arrivals', 'AP', row))
                    if n>=3:
                        T += 2*3600/(self.readCell('Arrivals', 'AP', row) + self.readCell('Arrivals', 'AQ', row))
                        if n>=4:
                            T += 2*3600/(self.readCell('Arrivals', 'AQ', row)+self.readCell('Arrivals', 'AR', row))
                            if n>=5:
                                T += 2*3600/(self.readCell('Arrivals', 'AR', row) + self.readCell('Arrivals', 'AS', row))
                                if n>=6:
                                    T += 2*3600/(self.readCell('Arrivals', 'AS', row)+self.readCell('Arrivals', 'AT', row))
                                    if n>=7:
                                        T += 2*3600/(self.readCell('Arrivals', 'AT', row)+self.readCell('Arrivals', 'AU', row))
                                        if n>=8:
                                            T += 2*3600/(self.readCell('Arrivals', 'AU', row)+self.readCell('Arrivals', 'AV', row))
                                            if n==9:
                                                T += 2*3600/(self.readCell('Arrivals', 'AV', row) + self.readCell('Arrivals', 'AW', row))
                                            elif n>9:
                                                T += (n-9)*3600/self.readCell('Arrivals', 'AW', row)
            return T


        def fraction_of_segments(n,f,row):
            if n== 1:
                S = f*(self.readCell('Arrivals', 'AP', row) - self.readCell('Arrivals', 'AO', row)) + self.readCell('Arrivals', 'AO', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AO', row))
            elif n==2:
                S = f*(self.readCell('Arrivals', 'AQ', row) - self.readCell('Arrivals', 'AP', row)) + self.readCell('Arrivals', 'AP', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AP', row))
            elif n==3:
                S = f*(self.readCell('Arrivals', 'AR', row) - self.readCell('Arrivals', 'AQ', row)) + self.readCell('Arrivals', 'AQ', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AQ', row))
            elif n==4:
                S = f*(self.readCell('Arrivals', 'AS', row) - self.readCell('Arrivals', 'AR', row)) + self.readCell('Arrivals', 'AR', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AR', row))
            elif n==5:
                S = f*(self.readCell('Arrivals', 'AT', row) - self.readCell('Arrivals', 'AS', row)) + self.readCell('Arrivals', 'AS', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AS', row))
            elif n==6:
                S = f*(self.readCell('Arrivals', 'AU', row) - self.readCell('Arrivals', 'AT', row)) + self.readCell('Arrivals', 'AT', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AT', row))
            elif n==7:
                S = f*(self.readCell('Arrivals', 'AV', row) - self.readCell('Arrivals', 'AU', row)) + self.readCell('Arrivals', 'AU', row)
                T = (f*2*3600)/(S+self.readCell('Arrivals', 'AU', row))
            elif n==8:
                S = f*(self.readCell('Arrivals', 'AW', row) - self.readCell('Arrivals', 'AV', row)) + self.readCell('Arrivals', 'AV', row)
                T = (f*3600)/self.readCell('Arrivals', 'AW', row)
            return T


        # if distance > 0:
        X = distance + self.readCell('Runway_calcs', 'F', row) # Actual distance + ATCO var
        D = X - 0.5
        if D <0:
            T = (X*3600)/self.readCell('Arrivals', 'AN', row)
        elif D > 0:
            v['n'] = math.floor(D)
            f = D - v['n']
            T1 = full_segments(v['n'],row)
            if (f != 0) and (v['n']<=8):
                T2 = fraction_of_segments(v['n'],f,row)
                T = T1 + T2 + (0.5*3600)/self.readCell('Arrivals', 'AN', row)
            else:
                T = T1
        return T


    def time_to_distance_assumed_speed_profile_IAS(self, row, T): # UNUSED!!!

        deceleration_difference= (self.readCell('Runway_calcs', 'H', row) - self.readCell('Runway_calcs', 'K', row))
        if deceleration_difference > 20 :
            self.C_DME = deceleration_difference / 20
        t1 = self.D_DME*3600/self.readCell('Runway_calcs', 'K', row)
        t2 = (self.C_DME - self.D_DME)*3600*2/(self.readCell('Runway_calcs', 'K', row) + self.readCell('Runway_calcs', 'H', row)) + t1

        if T <= t1 :
            D = (self.readCell('Runway_calcs', 'K', row)*T)/3600
        elif (T > t1) and (T < t2):
            t = T- t1
            S = (t*(self.readCell('Runway_calcs', 'H', row)-self.readCell('Runway_calcs', 'K', row)))/t2 + self.readCell('Runway_calcs', 'K', row)
            D = (t*(self.readCell('Runway_calcs', 'K', row)+S))/(2*3600) + self.D_DME
        elif T >= t2:
            D = self.C_DME + (T-t2)*self.readCell('Runway_calcs', 'H', row)/3600
        return D


    def time_to_distance_assumed_speed_profile_GS(self, row, T):

        deceleration_difference= (self.readCell('Runway_calcs', 'H', row) - self.readCell('Runway_calcs', 'G', row)) - (self.readCell('Runway_calcs', 'K', row) - self.readCell('Runway_calcs', 'I', row))
        if deceleration_difference > 20 :
            self.C_DME = deceleration_difference / 20
        t1 = self.D_DME*3600/(self.readCell('Runway_calcs', 'K', row)- self.readCell('Runway_calcs', 'I', row))
        t2 = (self.C_DME - self.D_DME)*3600*2/((self.readCell('Runway_calcs', 'K', row)- self.readCell('Runway_calcs', 'I', row)) + (self.readCell('Runway_calcs', 'H', row)- self.readCell('Runway_calcs', 'G', row))) + t1

        if T <= t1 :
            D = ((self.readCell('Runway_calcs', 'K', row)- self.readCell('Runway_calcs', 'I', row))*T)/3600
        elif (T > t1) and (T < t2):
            t = T- t1
            S = (t*((self.readCell('Runway_calcs', 'H', row)- self.readCell('Runway_calcs', 'G', row))-(self.readCell('Runway_calcs', 'K', row)- self.readCell('Runway_calcs', 'I', row))))/t2 + (self.readCell('Runway_calcs', 'K', row)- self.readCell('Runway_calcs', 'I', row))
            D = (t*((self.readCell('Runway_calcs', 'K', row)- self.readCell('Runway_calcs', 'I', row))+S))/(2*3600) + self.D_DME
            D = (T*((self.readCell('Runway_calcs', 'K', row)- self.readCell('Runway_calcs', 'I', row))+(self.readCell('Runway_calcs', 'H', row)- self.readCell('Runway_calcs', 'G', row)))/(2*3600)) + self.D_DME
        elif T >= t2:
            D = self.C_DME + (T-t2)*(self.readCell('Runway_calcs', 'H', row)- self.readCell('Runway_calcs', 'G', row))/3600
        return D


    def Arrival_Input_pre_process(self):
        """Function to pre-process the Arrival input file and make initial calculations"""
        def write_actual_speed_profile_to_output(row, AC_type):
            sp_type = self.speed_profile[self.speed_profile['Aircraft_Type'] == AC_type]
            rand_row = random.randint(0, len(sp_type.index) - 1)
            sp_start_col = sp_type.columns.get_loc('GSPD_0_1DME')
            sp_end_col = sp_start_col + 21
            self.writeRow('Arrivals', row, 30, sp_type.iloc[rand_row, sp_start_col:sp_end_col].tolist())

        # Read in Arrival data from an excel workbook
        for row in range(2, self.TOTAL_ARRIVALS + 1):

            if self.readCell('Arrivals', 'A', row) == None: ## Blank space in input
                print("Blank space detected in input file, terminating 'Read Input' here")
                break

            #If SIBTs are in time format convert them into seconds.
            if self.readCell('Arrivals', 'C', row) == None: ## SIBT seconds are not printed:
                SIBT = self.readCell('Arrivals', 'B', row)
                SIBT_sec = (SIBT.hour * 3600) + (SIBT.minute * 60) + SIBT.second
                self.writeCell('Arrivals', 'C', row, SIBT_sec) # Used from initial schedule
            # Write WTC in arrival Input | it will be used for AROT

            AC_type = self.readCell('Arrivals', 'D', row)
            self.writeCell('Arrivals', 'E', row, self.wake_lookup[self.wake_lookup['ICAO']==AC_type]['WTC'].item())

            # Write wake categories in runway calcs | used for wake separation:
            if v['RECAT']:
                AC_type = self.readCell('Arrivals', 'D', row)
                self.writeCell('Runway_calcs', 'U', row, self.wake_lookup[self.wake_lookup['ICAO']==AC_type]['RECAT-EU'].item()) #RECT-EU cat
            elif v['RECAT_PWS']:
                AC_type = self.readCell('Arrivals', 'D', row)
                self.writeCell('Runway_calcs', 'U', row, self.wake_separations['RECAT_PWS'].at[AC_type,'RECAT20'])
            else:
                self.writeCell('Runway_calcs', 'U', row, self.readCell('Arrivals', 'E', row)) #WTC cat

            ################# ACTUAL _ SPEED _ PROFILE ####################

            AC_type = self.readCell('Arrivals', 'D', row)
            if AC_type in self.speed_profile['Aircraft_Type'].unique():
                write_actual_speed_profile_to_output(row, AC_type)
            else:
                # Find other AC types in wake category with speed profiles
                AC_type_other = list(
                    set(self.wake_lookup[self.wake_lookup['RECAT-EU'] == self.wake_lookup[self.wake_lookup['ICAO'] == AC_type]['RECAT-EU'].item()]['ICAO']) &
                    set(self.speed_profile['Aircraft_Type'].unique())
                )
                # Select a random AC type to use
                write_actual_speed_profile_to_output(row, random.choice(AC_type_other))

            ######################## INTERMEDIATE CALCULATIONS ###########################

            # Arrival ID
            self.writeCell('Runway_calcs', 'A', row, self.readCell('Arrivals', 'A', row))
            ##################### TAXI-IN - normal distribution ###############

            Arrival_Taxiin_mean = self.readCell('Arrivals', 'I', row)
            Arrival_Taxiin_SD = self.readCell('Arrivals', 'J', row)
            # taxi_outliers = True
            # Taxiinlookup = self.readCell('Arrivals', 'M', row)
            # if not taxi_outliers:
            tempTaxiIn = random.normalvariate(Arrival_Taxiin_mean, Arrival_Taxiin_SD)
            # else:
                # tempTaxiIn = Taxiinlookup
            self.writeCell('Runway_calcs', 'B', row, round(tempTaxiIn, 0))

            ####################### AROT - from lookup ########################

            if self.readCell('Arrivals', 'E', row) == "H":
                self.writeCell('Runway_calcs', 'C', row, random.choice(self.rot_lookup['AROT_H'].dropna()))
            elif self.readCell('Arrivals', 'E', row) == "M":
                self.writeCell('Runway_calcs', 'C', row, random.choice(self.rot_lookup['AROT_M'].dropna()))
            elif self.readCell('Arrivals', 'E', row) == "L":
                self.writeCell('Runway_calcs', 'C', row, random.choice(self.rot_lookup['AROT_L'].dropna()))
            elif self.readCell('Arrivals', 'E', row) == "UM":
                self.writeCell('Runway_calcs', 'C', row, random.choice(self.rot_lookup['AROT_UM'].dropna()))
            elif self.readCell('Arrivals', 'E', row) == "J":
                self.writeCell('Runway_calcs', 'C', row, random.choice(self.rot_lookup['AROT_J'].dropna()))
            elif self.readCell('Arrivals', 'E', row) == "S":
                self.writeCell('Runway_calcs', 'C', row, random.choice(self.rot_lookup['AROT_S'].dropna()))

            ##################### ADA - normal distribution ###############
            self.writeCell('Runway_calcs', 'D', row, int(random.normalvariate(self.readCell('Arrivals', 'O', row), self.readCell('Arrivals', 'P', row))))

            ##################### ADDA - normal distribution ###############
            self.writeCell('Runway_calcs', 'E', row, int(random.normalvariate(self.readCell('Arrivals', 'Q', row), self.readCell('Arrivals', 'R', row))))

            ################ ATCO variability - normal distribution ###########
            self.writeCell('Runway_calcs', 'F', row, int(random.normalvariate(self.readCell('Arrivals', 'S', row), self.readCell('Arrivals', 'T', row))))

            ################## ASSUMED_SPEED_PROFILE ######################

            # --- WIND 1  ---#
            self.writeCell('Runway_calcs', 'G', row, random.normalvariate(self.readCell('Arrivals', 'V', row), self.readCell('Arrivals', 'W', row)))
            # --- SPEED 1  ---#
            self.writeCell('Runway_calcs', 'H', row, random.normalvariate(self.readCell('Arrivals', 'X', row), self.readCell('Arrivals', 'Y', row)))
            # --- WIND 2  ---#
            actualWIND2 = random.normalvariate(self.readCell('Arrivals', 'Z', row), self.readCell('Arrivals', 'AA', row))
            self.writeCell('Runway_calcs', 'I', row, actualWIND2)
            # --- SPEED 2  ---#
            actualSPEED2 = random.normalvariate(self.readCell('Arrivals', 'AB', row), self.readCell('Arrivals', 'AC', row))
            self.writeCell('Runway_calcs', 'J', row, actualSPEED2)
            # --- VTGT  ---#
            if (actualWIND2 < 5) or (actualWIND2 > 20):
                wind_adjustment = 5
            else:
                wind_adjustment = actualWIND2*0.5
            V_TGT = actualSPEED2 + wind_adjustment
            self.writeCell('Runway_calcs', 'K', row, V_TGT)

            #------ SAE -------#
            self.writeCell('Runway_calcs', 'L', row, self.readCell('Arrivals', 'C', row) - self.STT - 200) # SAE = SIBT - Standard Taxi Time - App length*
            #---- Predicted Landing Time --------#
            self.writeCell('Runway_calcs', 'M', row, self.readCell('Runway_calcs', 'L', row) + 60) # PLT = SAE + MRS*

            ############################ MAX CONSTRAINT CALCS ##################################

            def min_wake_separation_arrs(key_of_nextArrival): # delievered at THR ACTUAL SPEED PROFILE
                minWakeSepArr = 0 # Initialise local variable (reset on each iteration)

                if v['RECAT_PWS']: # analyse by ac type
                    previousArrival = self.readCell('Arrivals', 'D', key_of_nextArrival - 1)
                    currentArrival = self.readCell('Arrivals', 'D', key_of_nextArrival)
                    previousArrivalWake = self.readCell('Runway_calcs', 'U', key_of_nextArrival - 1) #20cat classification
                    currentArrivalWake = self.readCell('Runway_calcs', 'U', key_of_nextArrival) #20cat classification

                    if key_of_nextArrival == 2: #FirstArrival
                        minWakeSepArr = 0
                    else:
                        if (currentArrival in self.wake_separations['RECAT_PWS']) and (previousArrival in self.wake_separations['RECAT_PWS']):
                            wakeDistance = self.wake_separations['RECAT_PWS'].at[currentArrival,previousArrival]
                            if wakeDistance==0:
                                wakeDistance = self.wake_separations['RECAT20'].at[previousArrivalWake,currentArrivalWake]
                        else: # if the pair is not in the 96x96 table, search in the 20cat
                            wakeDistance = self.wake_separations['RECAT20'].at[previousArrivalWake,currentArrivalWake]

                        if wakeDistance == 0:
                            minWakeSepArr =0
                        else:
                            if v['distanceBased']:
                                if v['WAKE_4DME']:
                                    Total_time_follow = int(self.DBS_actual_speed_profile((wakeDistance+4),key_of_nextArrival))
                                    Time_lead_4dme_to_thr = int(self.DBS_actual_speed_profile(4,(key_of_nextArrival-1)))
                                    minWakeSepArr = Total_time_follow - Time_lead_4dme_to_thr
                                elif v['WAKE_THR']:
                                    minWakeSepArr = int(self.DBS_actual_speed_profile(wakeDistance,key_of_nextArrival))  #time
                                else: # the same as the previous one but it's the default condition
                                    minWakeSepArr = int(self.DBS_actual_speed_profile(wakeDistance,key_of_nextArrival))  #time

                            elif v['timeBased']:
                                time1 = self.distance_to_time_assumed_speed_profile_IAS(key_of_nextArrival, wakeDistance) #time
                                distance = self.time_to_distance_assumed_speed_profile_GS(key_of_nextArrival, int(time1))#distance
                                if v['WAKE_4DME']:
                                    Total_time_follow = int(self.DBS_actual_speed_profile((distance+4),key_of_nextArrival))
                                    Time_lead_4dme_to_thr = int(self.DBS_actual_speed_profile(4,(key_of_nextArrival-1)))
                                    minWakeSepArr = Total_time_follow - Time_lead_4dme_to_thr

                                elif v['WAKE_THR']:
                                    minWakeSepArr = int(self.DBS_actual_speed_profile(distance,key_of_nextArrival)) #time
                                else:
                                    minWakeSepArr = int(self.DBS_actual_speed_profile(distance,key_of_nextArrival)) #time

                else: #analyze by wake
                    previousArrivalWake = self.readCell('Runway_calcs', 'U', key_of_nextArrival - 1)
                    currentArrivalWake = self.readCell('Runway_calcs', 'U', key_of_nextArrival)
                    if key_of_nextArrival == 2: #FirstArrival
                        minWakeSepArr = 0
                    else: #next arrivals

                        if v['RECAT']: # delievered to THR
                            wakeDistance = self.wake_separations['RECAT-EU'].at[previousArrivalWake,currentArrivalWake]
                        else: #UK cat *********** should be delievered to 4dme
                            wakeDistance = self.wake_separations['UK'].at[previousArrivalWake,currentArrivalWake] #distance

                        if wakeDistance == 0:
                            minWakeSepArr =0
                        else:
                            if v['distanceBased']:
                                if v['WAKE_4DME']:
                                    Total_time_follow = int(self.DBS_actual_speed_profile((wakeDistance+4),key_of_nextArrival))
                                    Time_lead_4dme_to_thr = int(self.DBS_actual_speed_profile(4,(key_of_nextArrival-1)))
                                    minWakeSepArr = Total_time_follow - Time_lead_4dme_to_thr
                                elif v['WAKE_THR']:
                                    minWakeSepArr = int(self.DBS_actual_speed_profile(wakeDistance,key_of_nextArrival))  #time
                                else: # the same as the previous one but it's the default condition
                                    minWakeSepArr = int(self.DBS_actual_speed_profile(wakeDistance,key_of_nextArrival))  #time

                            elif v['timeBased']:
                                time1 = self.distance_to_time_assumed_speed_profile_IAS(key_of_nextArrival, wakeDistance) #time
                                distance = self.time_to_distance_assumed_speed_profile_GS(row, int(time1))#distance
                                if v['WAKE_4DME']:
                                    Total_time_follow = int(self.DBS_actual_speed_profile((distance+4),key_of_nextArrival))
                                    Time_lead_4dme_to_thr = int(self.DBS_actual_speed_profile(4,(key_of_nextArrival-1)))
                                    minWakeSepArr = Total_time_follow - Time_lead_4dme_to_thr

                                elif v['WAKE_THR']:
                                    minWakeSepArr = int(self.DBS_actual_speed_profile(distance,key_of_nextArrival)) #time
                                else:
                                    minWakeSepArr = int(self.DBS_actual_speed_profile(distance,key_of_nextArrival)) #time

                return(minWakeSepArr)


            self.writeCell('Runway_calcs', 'P', 1, 'WAKE SEPARATION')
            self.writeCell('Runway_calcs', 'P', row, int(min_wake_separation_arrs(row))) #always Distance-based
            self.writeCell('Runway_calcs', 'Q', 1, 'MRS')
            MRSArr = 0

            if (v['MRS_4DME']) and (row>2):
                Total_time_follow = int(self.DBS_actual_speed_profile((self.MIN_RADAR_SEP_DIST+4),row))
                Time_lead_4dme_to_thr = int(self.DBS_actual_speed_profile(4,(row-1)))
                MRSArr = Total_time_follow - Time_lead_4dme_to_thr
            elif v['MRS_THR']:
                MRSArr = int(self.DBS_actual_speed_profile(self.MIN_RADAR_SEP_DIST,row))  #time
            else: # the same as the previous one but it's the default condition
                MRSArr = int(self.DBS_actual_speed_profile(self.MIN_RADAR_SEP_DIST,row))  #time

            self.writeCell('Runway_calcs', 'Q', row, MRSArr)

            def max_constraint_generator(row):
                wake_constraint = self.readCell('Runway_calcs', 'P', row)
                MRS_constraint = self.readCell('Runway_calcs', 'Q', row)
                spFLAG = "None"
                max_constraint = 0

                if row == 2 :
                    max_constraint = max(wake_constraint,MRS_constraint)
                    spFLAG = "First Arrival"
                else: #not he first arrival
                    AROT_constraint = self.readCell('Runway_calcs', 'C', row - 1) + 5

                    if self.TOTAL_DEPARTURES == 0: #no departures
                        max_constraint = int(max(wake_constraint, MRS_constraint ,AROT_constraint))
                        if max_constraint == wake_constraint:
                            spFLAG = "WAKE"
                        elif max_constraint == MRS_constraint:
                            spFLAG = "MRS"
                        else:
                            spFLAG = "AROT"
                    elif self.TOTAL_DEPARTURES > 0 and self.TOTAL_ARRIVALS > 0: #there are both arrivals and departures scheduled
                        if v['timeBased']:

                            max_constraint = int(max(wake_constraint, MRS_constraint, AROT_constraint))
                            if max_constraint == wake_constraint:
                                spFLAG = "WAKE"
                            elif max_constraint == MRS_constraint:
                                spFLAG = "MRS"
                            else:
                                spFLAG = "AROT"
                        elif v['distanceBased']:
                            if (self.readCell('Arrivals', 'U', row)) == "ADDA" :
                                ADDA_distance = self.readCell('Runway_calcs', 'E', row)
                                if (v['ADDA_4DME']) and (row>2):
                                    Total_time_follow = int(self.DBS_actual_speed_profile((ADDA_distance+4),row))
                                    Time_lead_4dme_to_thr = int(self.DBS_actual_speed_profile(4,(row-1)))
                                    ADDA_separation = Total_time_follow - Time_lead_4dme_to_thr
                                elif v['ADDA_THR']:
                                    ADDA_separation = int(self.DBS_actual_speed_profile(ADDA_distance,row))  #time
                                else: # the same as the previous one but it's the default condition
                                    ADDA_separation = int(self.DBS_actual_speed_profile(ADDA_distance,row))  #time
                                # ADDA_separation = int(self.DBS_actual_speed_profile(ADDA_distance,row))

                                max_constraint = int(max(wake_constraint, ADDA_separation, MRS_constraint,AROT_constraint))
                                if max_constraint ==wake_constraint:
                                    spFLAG = "WAKE"
                                elif max_constraint == ADDA_separation:
                                    spFLAG = "ADDA"
                                elif max_constraint == MRS_constraint:
                                    spFLAG = "MRS"
                                else:
                                    spFLAG = "AROT"
                            elif (self.readCell('Arrivals', 'U', row)) == "ADA" :

                                ADA_distance = self.readCell('Runway_calcs', 'D', row)

                                if (v['ADA_4DME']) and (row>2):
                                    Total_time_follow = int(self.DBS_actual_speed_profile((ADA_distance+4),row))
                                    Time_lead_4dme_to_thr = int(self.DBS_actual_speed_profile(4,(row-1)))
                                    ADA_separation = Total_time_follow - Time_lead_4dme_to_thr
                                elif v['ADA_THR']:
                                    ADA_separation = int(self.DBS_actual_speed_profile(ADA_distance,row))  #time
                                else: # the same as the previous one but it's the default condition
                                    ADA_separation = int(self.DBS_actual_speed_profile(ADA_distance,row))  #time

                                max_constraint = int(max(wake_constraint, ADA_separation, MRS_constraint,AROT_constraint))
                                if max_constraint == wake_constraint:
                                    spFLAG = "WAKE"
                                elif max_constraint == ADA_separation:
                                    spFLAG = "ADA"
                                elif max_constraint == MRS_constraint:
                                    spFLAG = "MRS"
                                else:
                                    spFLAG = "AROT"
                            else:
                                max_constraint = int(max(wake_constraint, MRS_constraint,AROT_constraint))
                                if max_constraint == wake_constraint:
                                    spFLAG = "WAKE"
                                elif max_constraint == MRS_constraint:
                                    spFLAG = "MRS"
                                else:
                                    spFLAG = "AROT"

                return{'a' : max_constraint, 'b' : spFLAG}


            ################################### MAX CONSTRAINT PRINT ###################################

            self.writeCell('Runway_calcs', 'N', row, max_constraint_generator(row)['a'])
            self.writeCell('Runway_calcs', 'O', row, max_constraint_generator(row)['b'])


    def Departure_Input_pre_process(self):
        """Function to pre-process the Departure input file and make initial calculations"""
        # Initialise Departure input
        # Read in Departure data from an excel workbook
        for row in range(2, self.TOTAL_DEPARTURES + 1):
            if self.readCell('Departures', 'A', row) == None:  # NO (more) DEPARTURES
                print("Blank space detected in input file, terminating 'Read Input' here")
                break
            if self.readCell('Departures', 'C', row) == None: #SOBT are in time-format
                SOBT = self.readCell('Departures', 'B', row)
                SOBT_sec = (SOBT.hour * 3600) + (SOBT.minute * 60) + SOBT.second
                self.writeCell('Departures', 'C', row, SOBT_sec) # Used from initial schedule

            ################### INTERMEDIATE CALCULATIONS #####################

            #-----Departure WAKE category-----#
            AC_type = self.readCell('Departures', 'F', row)
            self.writeCell('Departures', 'H', row, self.wake_lookup[self.wake_lookup['ICAO']==AC_type]['WTC'].item())

            #----Departure ID -----#
            self.writeCell('Runway_calcs', 'R', row, self.readCell('Departures', 'A', row))

            #------TAXI-OUT------#

            Departure_Taxiout_mean = self.readCell('Departures', 'K', row)
            Departure_Taxiout_SD = self.readCell('Departures', 'L', row)
            actualTAXIOUT = random.normalvariate(Departure_Taxiout_mean, Departure_Taxiout_SD)
            self.writeCell('Runway_calcs', 'S', row, round(actualTAXIOUT,0))

            #------ DROT-------#
            if self.readCell('Departures', 'H', row) == "H":
                self.writeCell('Runway_calcs', 'T', row, random.choice(self.rot_lookup['DROT_H'].dropna()))
            elif self.readCell('Departures', 'H', row) == "M":
                self.writeCell('Runway_calcs', 'T', row, random.choice(self.rot_lookup['DROT_M'].dropna()))
            elif self.readCell('Departures', 'H', row) == "L":
                self.writeCell('Runway_calcs', 'T', row, random.choice(self.rot_lookup['DROT_L'].dropna()))
            elif self.readCell('Departures', 'H', row) == "UM":
                self.writeCell('Runway_calcs', 'T', row, random.choice(self.rot_lookup['DROT_UM'].dropna()))
            elif self.readCell('Departures', 'H', row) == "J":
                self.writeCell('Runway_calcs', 'T', row, random.choice(self.rot_lookup['DROT_J'].dropna()))
            elif self.readCell('Departures', 'H', row) == "S":
                self.writeCell('Runway_calcs', 'T', row, random.choice(self.rot_lookup['DROT_S'].dropna()))


    # ------------------------ Arrival Movement Functions ------------------------ #

    def update_ArrHOLDqueue_Delay(self, Current_time):
        for AC in list(self.ArrHOLDqueue.keys()):
            ArrHOLDqueue_Delay = Current_time - self.ArrHOLDqueue[AC][1] # Delay = Current_time - SAE
            self.ArrHOLDqueue[AC][4] = ArrHOLDqueue_Delay


    def SAE_lookup(self, Current_time):
        if self.ARRkey != (self.TOTAL_ARRIVALS - 1):
            if Current_time >= self.readCell('Runway_calcs', 'L', self.ARRkey): # Current_time = SAE
                self.ArrHOLDqueue[self.ARRkey] = [
                    self.readCell('Arrivals', 'A', self.ARRkey),
                    self.readCell('Runway_calcs', 'L', self.ARRkey),
                    self.readCell('Runway_calcs', 'C', self.ARRkey),
                    self.readCell('Runway_calcs', 'M', self.ARRkey),
                    0
                ]
                self.ARRkey += 1
        self.update_ArrHOLDqueue_Delay(Current_time)
        return (self.ARRkey)


    def update_APPqueue(self, Current_time): # add to self.APPqueue
        #print(Current_time, ' app queue called')
        if (len(self.ArrHOLDqueue)>0) and (len(self.APPqueue)==0): # There is something in the hold but nothing on approach
            first_in_line_ArrHOLDqueue = min(list(self.ArrHOLDqueue.keys()))

            max_constraint = 0
            self.copyCell('Arrival_Output', 'I', self.ArrOutput, 'Runway_calcs', 'N', first_in_line_ArrHOLDqueue)
            self.copyCell('Arrival_Output', 'J', self.ArrOutput, 'Runway_calcs', 'O', first_in_line_ArrHOLDqueue)
            #target time, optimised gaps
            if v['timeBased']:
                if (len(self.RWYqueue1) + len(self.RWYqueue2))>0: #ther is a departure ready to go
                    if (self.readCell('Arrivals', 'U', first_in_line_ArrHOLDqueue)) == "ADDA" :    #*********to be changed
                        AROT = self.ArrHOLDqueue[first_in_line_ArrHOLDqueue][2]
                        firstDeparture, currentRWYqueue = self.first_in_line_RWYqueue_funct()
                        if currentRWYqueue == 1:
                            DROT1 = self.RWYqueue1[firstDeparture][2]
                        else:
                            DROT1 = self.RWYqueue2[firstDeparture][2]
                        # secondDeparture, nextRWYqueue = self.second_in_line_RWYqueues(currentRWYqueue)
                        # if nextRWYqueue == 1:
                        #     DROT2 = self.RWYqueue1[secondDeparture][2]
                        # else:
                        #     DROT2 = self.RWYqueue2[secondDeparture][2]
                        ADDA_target_time = AROT + DROT1 + DROT1 + self.X_BUFFER# AROT + NextDep DROT + NextDep2 DROT
                        ADDA_target_distance = self.time_to_distance_assumed_speed_profile_GS(first_in_line_ArrHOLDqueue, int(ADDA_target_time))#distance
                        if (v['ADDA_4DME']) and (self.ArrOutput>2):
                            Total_time_follow = int(self.DBS_actual_speed_profile((ADDA_target_distance+4),first_in_line_ArrHOLDqueue))
                            Time_lead_4dme_to_thr = int(self.DBS_actual_speed_profile(4,self.ArrOutput))
                            ADDA_separation = Total_time_follow - Time_lead_4dme_to_thr

                        elif v['ADDA_THR']:
                            ADDA_separation = int(self.DBS_actual_speed_profile(ADDA_target_distance,first_in_line_ArrHOLDqueue)) #time
                        else:
                            ADDA_separation = int(self.DBS_actual_speed_profile(ADDA_target_distance,first_in_line_ArrHOLDqueue)) #time - default

                        if ADDA_separation > self.readCell('Arrival_Output', 'I', self.ArrOutput):
                            self.writeCell('Arrival_Output', 'J', self.ArrOutput, 'ADDA')
                            self.writeCell('Arrival_Output', 'I', self.ArrOutput, ADDA_separation)

                    elif (self.readCell('Arrivals', 'U', first_in_line_ArrHOLDqueue)) == "ADA" :
                        AROT = self.ArrHOLDqueue[first_in_line_ArrHOLDqueue][2]
                        firstDeparture, currentRWYqueue = self.first_in_line_RWYqueue_funct()
                        if currentRWYqueue == 1:
                            DROT1 = self.RWYqueue1[firstDeparture][2]
                        else:
                            DROT1 = self.RWYqueue2[firstDeparture][2]
                        ADA_target_time = AROT + DROT1 + self.X_BUFFER# AROT + NextDep DROT + NextDep2 DROT
                        ADA_target_distance = self.time_to_distance_assumed_speed_profile_GS(first_in_line_ArrHOLDqueue, int(ADA_target_time))#distance
                        if (v['ADA_4DME']) and (self.ArrOutput>2):
                            Total_time_follow = int(self.DBS_actual_speed_profile((ADA_target_distance+4),first_in_line_ArrHOLDqueue))
                            Time_lead_4dme_to_thr = int(self.DBS_actual_speed_profile(4,self.ArrOutput-1))
                            ADA_separation = Total_time_follow - Time_lead_4dme_to_thr

                        elif v['ADA_THR']:
                            ADA_separation = int(self.DBS_actual_speed_profile(ADA_target_distance,first_in_line_ArrHOLDqueue)) #time
                        else:
                            ADA_separation = int(self.DBS_actual_speed_profile(ADA_target_distance,first_in_line_ArrHOLDqueue)) #time

                        if ADA_separation > self.readCell('Arrival_Output', 'I', self.ArrOutput):
                            self.writeCell('Arrival_Output', 'J', self.ArrOutput, 'ADA')
                            self.writeCell('Arrival_Output', 'I', self.ArrOutput, ADA_separation)
                # else: # no departure ready to go
                    # max_constraint = self.readCell('Arrival_Output', 'I', self.ArrOutput)
            # elif v['distanceBased']:
            max_constraint = self.readCell('Arrival_Output', 'I', self.ArrOutput)

            # print(Current_time, self.ArrOutput, ' | max_constraint = ', max_constraint)

            # if max_constraint != 0:

            self.ArrHOLDqueue[first_in_line_ArrHOLDqueue].append(Current_time)#self.APPqueue entry time
            ALT = self.ArrHOLDqueue[first_in_line_ArrHOLDqueue][5]+ int(max_constraint) #(ALT = APPqueue_entry_time + max_constraint)
            self.ArrHOLDqueue[first_in_line_ArrHOLDqueue].append(ALT)
            RWY_EXIT = self.ArrHOLDqueue[first_in_line_ArrHOLDqueue][6] + self.ArrHOLDqueue[first_in_line_ArrHOLDqueue][2] # ALT + AROT
            self.ArrHOLDqueue[first_in_line_ArrHOLDqueue].append(RWY_EXIT)

            self.APPqueue[first_in_line_ArrHOLDqueue]=self.ArrHOLDqueue[first_in_line_ArrHOLDqueue]

            del self.ArrHOLDqueue[first_in_line_ArrHOLDqueue]
            #print to sequence tab


    def Arr_LANDING(self, Current_time, first_in_line_APPqueue):
        # if len(self.APPqueue)!=0:
        #     first_in_line_APPqueue = min(list(self.APPqueue.keys()))#there is only one AC in the self.APPqueue
        #     #print('There is something in the self.APPqueue')

        #     #print('NEXT ARRIVAL = ', AC)
        #     if Current_time == self.APPqueue[first_in_line_APPqueue][6]: #it's time to land
        if self.RWY_status == "D":
            print('*** GO AROUND ***', self.APPqueue[first_in_line_APPqueue])
            del self.APPqueue[first_in_line_APPqueue]
            goAroundHour = int(Current_time/3600)
            if goAroundHour in list(self.GoAroundCount.keys()): #if there was already a goAround at that hour:
                self.GoAroundCount[goAroundHour].append(1)
            else:
                self.GoAroundCount[goAroundHour]=[1]

        elif self.RWY_status == "E":
            self.writeRow('Arrival_Output', self.ArrOutput, 1, [
                self.readCell('Arrivals', 'A', first_in_line_APPqueue), # ARR ID
                int(Current_time/3600), # LANDING HOUR
                Current_time, # ACTUAL LANDING TIME
                self.APPqueue[first_in_line_APPqueue][7], # RWY EXIT
                self.readCell('Runway_calcs', 'U', first_in_line_APPqueue), # WAKE
                self.APPqueue[first_in_line_APPqueue][7] + self.readCell('Runway_calcs', 'B', first_in_line_APPqueue), # In blocks time
                self.readCell('Runway_calcs', 'C', first_in_line_APPqueue), # AROT
                self.readCell('Runway_calcs', 'B', first_in_line_APPqueue) # Taxi-in duration
            ])

            self.writeRow('Arrival_Output', self.ArrOutput, 11, [
                len(self.ArrHOLDqueue), # length self.ArrHOLDqueue
                self.APPqueue[first_in_line_APPqueue][4], # self.ArrHOLDqueue delay
            ])

            AIBT = self.readCell('Arrival_Output', 'F', self.ArrOutput)

            #Add Arrival to self.ARRIVALqueue
            self.ARRIVALqueue[first_in_line_APPqueue] = [self.readCell('Arrival_Output', 'A', self.ArrOutput), AIBT, self.ArrOutput]
            #print('self.ARRIVALqueue = ', list(self.ARRIVALqueue.keys()))
            del self.APPqueue[first_in_line_APPqueue]
            self.writeRow('Sequence', self.seqRow, 1, [
                'A',
                self.readCell('Arrival_Output', 'A', self.ArrOutput),
                self.readCell('Arrival_Output', 'C', self.ArrOutput),
                self.readCell('Arrival_Output', 'D', self.ArrOutput),
                self.readCell('Arrival_Output', 'G', self.ArrOutput)
            ])
            self.ArrOutput+=1
            self.seqRow += 1

        return (self.ArrOutput,self.seqRow)


    def first_in_line_ARRIVALqueue_func(self):
        min_IBT = self.end_time
        first_in_line_ARRIVALqueue = 0
        for AC in list(self.ARRIVALqueue.keys()):
            if self.ARRIVALqueue[AC][1]<min_IBT:
                min_IBT=self.ARRIVALqueue[AC][1]
                first_in_line_ARRIVALqueue = AC
        return(first_in_line_ARRIVALqueue)


    def update_ARRIVALqueue(self, Current_time):
        if len(self.ARRIVALqueue)>0:

            #Check first in line in arrival queue
            first_in_line_ARRIVALqueue = self.first_in_line_ARRIVALqueue_func()

            if Current_time > self.ARRIVALqueue[first_in_line_ARRIVALqueue][1]:
                #print(Current_time, 'ARR { ',AC,' } deleted from self.ARRIVALqueue ')
                del self.ARRIVALqueue[first_in_line_ARRIVALqueue]


    def update_currentGap(self, Current_time):
        if self.RWY_status == "E" or self.RWY_status == "D":
            if len(self.APPqueue)==0: #Nothing in the queue
                self.currentGap = self.end_time # Huuuuge self.currentGap
            else:
                next_Arrival = min(list(self.APPqueue.keys())) # should be only one key in the list
                self.currentGap = self.APPqueue[next_Arrival][6] - Current_time # ALT - Current_time
        elif self.RWY_status == "A":
            self.currentGap = 0
        return (self.currentGap)


    # ----------------------- Departure Movement Functions ----------------------- #

    def update_Departure_Delays(self, Current_time):
        if len(self.DepSTANDqueue)>0:
            for AC in list(self.DepSTANDqueue.keys()):
                DepSTANDqueue_Delay = Current_time - self.DepSTANDqueue[AC][1]
                self.DepSTANDqueue[AC][4] = DepSTANDqueue_Delay
        if len(self.TAXIhold)>0:
            for AC in list(self.TAXIhold.keys()):
                TAXIhold_Delay = Current_time - self.TAXIhold[AC][7]
                self.TAXIhold[AC][7] = TAXIhold_Delay
        if len(self.RWYqueue1)>0:
            for AC in list(self.RWYqueue1.keys()):
                RWYqueue1_delay = Current_time - self.RWYqueue1[AC][9] #Current_time - RWYqueue entry_time
                self.RWYqueue1[AC][10] = RWYqueue1_delay
        if len(self.RWYqueue2)>0:
            for AC in list(self.RWYqueue2.keys()):
                RWYqueue2_delay = Current_time - self.RWYqueue2[AC][9] #Current_time - RWYqueue entry_time
                self.RWYqueue2[AC][10] = RWYqueue2_delay


    def SOBTlookup(self, Current_time):
        if self.SOBTrow < self.TOTAL_DEPARTURES - 1:
            if Current_time >= self.readCell('Departures', 'C', self.SOBTrow) :# Current time = SOBT
                self.DepSTANDqueue[self.SOBTrow] = [
                    self.readCell('Departures', 'A', self.SOBTrow),
                    self.readCell('Departures', 'C', self.SOBTrow),
                    self.readCell('Runway_calcs', 'T', self.SOBTrow),
                    self.readCell('Departures', 'I', self.SOBTrow),
                    0
                ]
                self.SOBTrow += 1
        return self.SOBTrow


    def TAXIqueue_update(self, Current_time):
        #   if ((len(self.TAXIqueue) + len(self.ARRIVALqueue)+ len(self.TAXIhold))<15) and len(self.DepSTANDqueue)> 0: # if there are less than 15 AC moving on the TAXIway and there's something in self.DepSTANDqueue
        # check who has to go first
        first_in_line_DepSTANDqueue = min(list(self.DepSTANDqueue.keys()))

        self.DepSTANDqueue[first_in_line_DepSTANDqueue].append(Current_time) #self.TAXIqueue entry time
        self.DepSTANDqueue[first_in_line_DepSTANDqueue].append(self.readCell('Runway_calcs', 'S', first_in_line_DepSTANDqueue)) #TAXI-out


        #ADD first_in_line_DepSTANDqueue to self.TAXIqueue:
        self.TAXIqueue[first_in_line_DepSTANDqueue]=self.DepSTANDqueue[first_in_line_DepSTANDqueue]
        del self.DepSTANDqueue[first_in_line_DepSTANDqueue]


    def first_in_line_TAXIqueue_func(self):
        first_in_line_TAXIqueue = 0
        min_TAXIqueue_out = self.end_time
        for AC in list(self.TAXIqueue.keys()):
            TAXIqueue_out = self.TAXIqueue[AC][5]+self.TAXIqueue[AC][6]
            if TAXIqueue_out<min_TAXIqueue_out:
                min_TAXIqueue_out = TAXIqueue_out
                first_in_line_TAXIqueue = AC
        return(first_in_line_TAXIqueue)


    def TAXIhold_update(self, Current_time):
        if len(self.TAXIqueue)>0:
            first_in_line_TAXIqueue = self.first_in_line_TAXIqueue_func()
            if Current_time >= (self.TAXIqueue[first_in_line_TAXIqueue][5] + self.TAXIqueue[first_in_line_TAXIqueue][6]): # current_time = TAXIqueue_entry_time + Taxi-out
                self.TAXIqueue[first_in_line_TAXIqueue].append(Current_time) #self.TAXIhold entry time
                self.TAXIqueue[first_in_line_TAXIqueue].append(0) #self.TAXIhold delay

                #ADD first_in_line_TAXIqueue to self.TAXIhold
                self.TAXIhold[first_in_line_TAXIqueue] = self.TAXIqueue[first_in_line_TAXIqueue]
                del self.TAXIqueue[first_in_line_TAXIqueue]


    def transfer_to_2x4_RWYqueues(self, first_in_line_TAXIhold, Current_time):

        # Queue selection
        if v['queueType'] == '1x8':
            maxRWYqueue1Length = 8
            previousRWYqueue = 1
        elif v['queueType'] == '2x4':
            maxRWYqueue1Length = 4
            maxRWYqueue2Length = 4
            previousRWYqueue = 2 # Forces self.RWYqueue1 to go first
        elif v['queueType'] == '4x2':
            maxRWYqueue1Length = 2
            maxRWYqueue2Length = 2
            maxRWYqueue3Length = 2
            maxRWYqueue4Length = 2
            previousRWYqueue = 3 # Forces self.RWYqueue1/2 to go first
        elif v['queueType'] == '8x1':
            maxRWYqueue1Length = 4 # Will use 2x4 methods (accounts for 4 Queues x 1 in length)
            maxRWYqueue2Length = 4 # Will use 2x4 methods (accounts for 4 Queues x 1 in length)
            previousRWYqueue = 2 # Forces self.RWYqueue1 to go first

        if self.TAXIhold[first_in_line_TAXIhold][3] in self.SID_QUEUES[0]: # First check if SID group belongs to self.RWYqueue1
            if len(self.RWYqueue1) < maxRWYqueue1Length: # if there is space in self.RWYqueue1 add A/C to the queue
                self.TAXIhold[first_in_line_TAXIhold].append(Current_time) #self.RWYqueue1 entry time
                self.TAXIhold[first_in_line_TAXIhold].append(0) # self.RWYqueue1 Delay
                self.TAXIhold[first_in_line_TAXIhold].append(1) #RWYqueue used

                self.RWYqueue1[first_in_line_TAXIhold] = self.TAXIhold[first_in_line_TAXIhold]
                del self.TAXIhold[first_in_line_TAXIhold]

        elif self.TAXIhold[first_in_line_TAXIhold][3] in self.SID_QUEUES[1]: # First check if SID group belongs to self.RWYqueue2
            if len(self.RWYqueue2) < maxRWYqueue2Length: # if there is space in self.RWYqueue1 add A/C to the queue
                self.TAXIhold[first_in_line_TAXIhold].append(Current_time) #self.RWYqueue2 entry time
                self.TAXIhold[first_in_line_TAXIhold].append(0) # self.RWYqueue2 Delay
                self.TAXIhold[first_in_line_TAXIhold].append(2) #RWYqueue used

                self.RWYqueue2[first_in_line_TAXIhold] = self.TAXIhold[first_in_line_TAXIhold]
                del self.TAXIhold[first_in_line_TAXIhold]


    def RWYqueues_update(self, Current_time):
        previous_in_line = 0
        first_in_line_TAXIhold = 0
        while (len(self.TAXIhold)>0 and (len(self.RWYqueue1)+len(self.RWYqueue2))<8): # while there is something in self.TAXIhold and there's space in RWY queues
            previous_in_line = first_in_line_TAXIhold
            first_in_line_TAXIhold = min(list(self.TAXIhold.keys()))

            if (first_in_line_TAXIhold!=0) and (previous_in_line!= first_in_line_TAXIhold):
                self.transfer_to_2x4_RWYqueues(first_in_line_TAXIhold,Current_time)
            else:
                break


    def first_in_line_RWYqueue_funct(self):
        first_in_line_RWYqueue = 0
        currentRWYqueue = 0

        #ONLY FOR 2x4
        if self.DepOutput==2: #for the first departure check in which queue is the first departure:
            min_entry_time1 = self.end_time
            min_entry_time2 = self.end_time
            first_in_line_RWYqueue1=0
            first_in_line_RWYqueue2=0
            if len(self.RWYqueue1)>0:
                for AC in list(self.RWYqueue1.keys()):
                    if self.RWYqueue1[AC][9]<min_entry_time1:
                        min_entry_time1 = self.RWYqueue1[AC][9]
                        first_in_line_RWYqueue1 = AC
            if len(self.RWYqueue2)>0:
                for AC in list(self.RWYqueue2.keys()):
                    if self.RWYqueue2[AC][9]<min_entry_time2:
                        min_entry_time2 = self.RWYqueue2[AC][9]
                        first_in_line_RWYqueue2 = AC
            if min_entry_time1<min_entry_time2:
                first_in_line_RWYqueue = first_in_line_RWYqueue1
                currentRWYqueue = 1
            else:
                first_in_line_RWYqueue = first_in_line_RWYqueue2
                currentRWYqueue = 2

        elif self.DepOutput!=2 :
            if self.readCell('Departure_Output', 'U', self.DepOutput - 1) == 1: #If previous departure started from queue 1

                # use self.RWYqueue2
                if len(self.RWYqueue2)>0: #There is smth in the queue

                    min_entry_time = self.end_time
                    for AC in list(self.RWYqueue2.keys()):
                        if self.RWYqueue2[AC][9]<min_entry_time:
                            min_entry_time = self.RWYqueue2[AC][9]
                            first_in_line_RWYqueue = AC

                    currentRWYqueue = 2
                else: #there is nobody in self.RWYqueue2
                    #use self.RWYqueue1 again
                    min_entry_time = self.end_time
                    for AC in list(self.RWYqueue1.keys()):
                        if self.RWYqueue1[AC][9]<min_entry_time:
                            min_entry_time = self.RWYqueue1[AC][9]
                            first_in_line_RWYqueue = AC
                    currentRWYqueue = 1
            elif self.readCell('Departure_Output', 'U', self.DepOutput - 1) == 2: #If previous departure started from queue 2

                # use self.RWYqueue1
                if len(self.RWYqueue1)>0: #There is smth in the queue
                    min_entry_time = self.end_time
                    for AC in list(self.RWYqueue1.keys()):
                        if self.RWYqueue1[AC][9]<min_entry_time:
                            min_entry_time = self.RWYqueue1[AC][9]
                            first_in_line_RWYqueue = AC
                    currentRWYqueue = 1

                else: #there is nobody in self.RWYqueue2
                    #use self.RWYqueue2 again
                    min_entry_time = self.end_time
                    for AC in list(self.RWYqueue2.keys()):
                        if self.RWYqueue2[AC][9]<min_entry_time:
                            min_entry_time = self.RWYqueue2[AC][9]
                            first_in_line_RWYqueue = AC
                    currentRWYqueue = 2
        return(first_in_line_RWYqueue, currentRWYqueue)


    def second_in_line_RWYqueues(self, previousRWYqueue): #used for target ADDA time
        min_entry_time = self.end_time
        second_in_line_RWYqueue = 0
        currentRWYqueue = 0
        if previousRWYqueue == 1: #now use queue2
            if len(self.RWYqueue2)>0:
                currentRWYqueue = 2
                for AC in list(self.RWYqueue2.keys()):
                    if self.RWYqueue2[AC][9]<min_entry_time:
                        min_entry_time = self.RWYqueue2[AC][9]
                        second_in_line_RWYqueue = AC
            else:
                currentRWYqueue = 1
                for AC in list(self.RWYqueue1.keys()):
                    if self.RWYqueue1[AC][9]<min_entry_time:
                        min_entry_time = self.RWYqueue1[AC][9]
                        second_in_line_RWYqueue = AC

        elif previousRWYqueue == 2: #now use 1
            if len(self.RWYqueue1)>0:
                currentRWYqueue = 1
                for AC in list(self.RWYqueue1.keys()):
                    if self.RWYqueue1[AC][9]<min_entry_time:
                        min_entry_time = self.RWYqueue1[AC][9]
                        second_in_line_RWYqueue = AC

            else:
                currentRWYqueue = 2
                for AC in list(self.RWYqueue2.keys()):
                    if self.RWYqueue2[AC][9]<min_entry_time:
                        min_entry_time = self.RWYqueue2[AC][9]
                        second_in_line_RWYqueue = AC
        return(second_in_line_RWYqueue,currentRWYqueue)


    def dep_Wake_separation(self, first_in_line_RWYqueue):
        minWakeSep = 0 # Initialise local variable (reset on each iteration)
        if self.DepOutput == 2: #first departure:
            minWakeSep = 0
        else:
            previousDepartureWake = self.readCell('Departure_Output', 'E', self.DepOutput - 1)
            currentDepartureWake = self.readCell('Departures', 'H', first_in_line_RWYqueue)

            if previousDepartureWake == "J":
                if currentDepartureWake == "J":
                    minWakeSep = 0
                elif currentDepartureWake == "H":
                    minWakeSep = self.WAKE_RULES['J_H_d']
                elif (currentDepartureWake == "UM") or (currentDepartureWake == "M"):
                    minWakeSep = self.WAKE_RULES['J_H_d']
                elif (currentDepartureWake == "S") or (currentDepartureWake == "L"):
                    minWakeSep = self.WAKE_RULES['J_L_d']
                else:
                    print("[J-] Wake Category other than normal detected - check Input file")

            elif previousDepartureWake == "H":
                if currentDepartureWake == "J":
                    minWakeSep = 0
                elif currentDepartureWake == "H":
                    minWakeSep = self.WAKE_RULES['H_H_d']
                elif (currentDepartureWake == "UM") or (currentDepartureWake == "M"):
                    minWakeSep = self.WAKE_RULES['H_M_d']
                elif (currentDepartureWake == "S") or (currentDepartureWake == "L"):
                    minWakeSep = self.WAKE_RULES['H_M_d']
                else:
                    print("[H-] Wake Category other than normal detected - check Input file")

            elif (previousDepartureWake == "UM") or (previousDepartureWake == "M"):
                if currentDepartureWake == "L":
                    minWakeSep = self.WAKE_RULES['M_L_d']
                else:
                    minWakeSep = 0

            elif (previousDepartureWake == "S") or (previousDepartureWake == "S"):
                if currentDepartureWake == "L":
                    minWakeSep = 0

            else:
                minWakeSep = 0

        return(minWakeSep)


    def dep_SID_separation(self, first_in_line_RWYqueue):
        minSIDsep = 0 # Initialise local variable (reset on each iteration)
        if self.DepOutput == 2: #first departure:
            minSIDsep = 0
        else:
            # Compares SID groups between the previous and current A/C - then sets 'minSIDsep' variable as either altSID or sameSID
            previousDepartureSID = self.readCell('Departure_Output', 'F', self.DepOutput - 1)
            nextDepartureSID = self.readCell('Departures', 'I', first_in_line_RWYqueue)

            if nextDepartureSID == previousDepartureSID: #IF the next departure SID is tha same as the previous departure SID => maximum separation
                minSIDsep = v['minDep_sameSID']
            # If they are not equal, check if the SID group has some more separation rules
            elif nextDepartureSID != previousDepartureSID:
                minSIDsep = v['minDep_altSID']
                for item in self.SID_GROUPS:
                    if nextDepartureSID == item[0] and previousDepartureSID == item[1]:
                    #if previousDepartureSID == item[1]: # IF the previous departure SID matches the partner, apply maximum separation
                        minSIDsep = v['minDep_sameSID']
        return (minSIDsep)


    def departure_separation(self, first_in_line_RWYqueue):
        minDeptime = 0
        minDepLabel = ""
        #WAKE
        minWakeSep = self.dep_Wake_separation(first_in_line_RWYqueue)
        #SID
        minSIDsep = self.dep_SID_separation(first_in_line_RWYqueue)
        #compare the two and take the largest constraint
        if minSIDsep>minWakeSep:
            minDeptime = minSIDsep
            minDepLabel = "SID"
        else:
            minDeptime = minWakeSep
            minDepLabel = "WAKE"
        return(minDeptime,minDepLabel)


    def Dep_TAKE_OFF(self, Current_time):
        #if (len(self.RWYqueue1) != 0) or (len(self.RWYqueue2)!=0): #there is something in the queues:
        #print('Something in the RWYqueues')

        first_in_line_RWYqueue, currentRWYqueue = self.first_in_line_RWYqueue_funct()
        if first_in_line_RWYqueue !=0: # there's someone in line
            minDepTime,minDepLabel = self.departure_separation(first_in_line_RWYqueue)

            if self.DepOutput == 2: # First departure, no wake/sid constraints
                if (self.currentGap > v['n']):
                    #TAKE-OFF
                    self.writeRow('Departure_Output', self.DepOutput, 2, [
                        int(Current_time/3600), # Dep HOUR
                        Current_time # Departure RWY Entry
                    ])

                    if currentRWYqueue == 1:
                        self.writeRow('Departure_Output', self.DepOutput, 1, [
                            self.RWYqueue1[first_in_line_RWYqueue][0], # AC ID
                            None, # Skip column B
                            None, # Skip column C
                            self.readCell('Departure_Output', 'C', self.DepOutput) + self.RWYqueue1[first_in_line_RWYqueue][2], # Dep RWY EXIT = Dep RWY ENTRY + DROT
                            self.readCell('Departures', 'H', first_in_line_RWYqueue), #WAKE
                            self.RWYqueue1[first_in_line_RWYqueue][3], # SID
                            self.RWYqueue1[first_in_line_RWYqueue][2], # DROT
                            self.RWYqueue1[first_in_line_RWYqueue][6], # TAXIOUT
                            minDepTime, # DEP MIN SEPARATION
                            minDepLabel, # DEP MIN SEPARATION LABEL
                            self.currentGap, # self.currentGap
                            len(self.DepSTANDqueue),
                            len(self.TAXIhold),
                            len(self.RWYqueue1),
                            len(self.RWYqueue2),
                            None, # len(self.RWYqueue3)
                            None, # len(self.RWYqueue4)
                            self.RWYqueue1[first_in_line_RWYqueue][4], # DELAY self.DepSTANDqueue
                            self.RWYqueue1[first_in_line_RWYqueue][8], # DELAY self.TAXIhold
                            self.RWYqueue1[first_in_line_RWYqueue][10], # DELAY RWYqueue
                            self.RWYqueue1[first_in_line_RWYqueue][11], # RWYqueue USED
                        ])
                        del self.RWYqueue1[first_in_line_RWYqueue]

                    elif currentRWYqueue == 2:
                        self.writeRow('Departure_Output', self.DepOutput, 1, [
                            self.RWYqueue2[first_in_line_RWYqueue][0], # AC ID
                            None, # Skip column B
                            None, # Skip column C
                            self.readCell('Departure_Output', 'C', self.DepOutput) + self.RWYqueue2[first_in_line_RWYqueue][2], # Dep RWY EXIT = Dep RWY ENTRY + DROT
                            self.readCell('Departures', 'H', first_in_line_RWYqueue), #WAKE
                            self.RWYqueue2[first_in_line_RWYqueue][3], # SID
                            self.RWYqueue2[first_in_line_RWYqueue][2], # DROT
                            self.RWYqueue2[first_in_line_RWYqueue][6], # TAXIOUT
                            minDepTime, # DEP MIN SEPARATION
                            minDepLabel, # DEP MIN SEPARATION LABEL
                            self.currentGap, # self.currentGap
                            len(self.DepSTANDqueue),
                            len(self.TAXIhold),
                            len(self.RWYqueue1),
                            len(self.RWYqueue2),
                            None, # len(self.RWYqueue3)
                            None, # len(self.RWYqueue4)
                            self.RWYqueue2[first_in_line_RWYqueue][4], # DELAY self.DepSTANDqueue
                            self.RWYqueue2[first_in_line_RWYqueue][8], # DELAY self.TAXIhold
                            self.RWYqueue2[first_in_line_RWYqueue][10], # DELAY RWYqueue
                            self.RWYqueue2[first_in_line_RWYqueue][11], # RWYqueue USED
                        ])
                        del self.RWYqueue2[first_in_line_RWYqueue]

                    self.DepOutput += 1
            elif self.DepOutput != 2:
                if (self.currentGap > v['n']) and (Current_time > self.readCell('Departure_Output', 'C', self.DepOutput - 1) + minDepTime):
                    #print(first_in_line_RWYqueue,' condition met', self.DepOutput)
                    #TAKE-OFF
                    self.writeRow('Departure_Output', self.DepOutput, 2, [
                        int(Current_time/3600), # Dep HOUR
                        Current_time # Departure RWY Entry
                    ])

                    if currentRWYqueue == 1:
                        self.writeRow('Departure_Output', self.DepOutput, 1, [
                            self.RWYqueue1[first_in_line_RWYqueue][0], # AC ID
                            None, # Skip column B
                            None, # Skip column C
                            self.readCell('Departure_Output', 'C', self.DepOutput) + self.RWYqueue1[first_in_line_RWYqueue][2], # Dep RWY EXIT = Dep RWY ENTRY + DROT
                            self.readCell('Departures', 'H', first_in_line_RWYqueue), #WAKE
                            self.RWYqueue1[first_in_line_RWYqueue][3], # SID
                            self.RWYqueue1[first_in_line_RWYqueue][2], # DROT
                            self.RWYqueue1[first_in_line_RWYqueue][6], # TAXIOUT
                            minDepTime, # DEP MIN SEPARATION
                            minDepLabel, # DEP MIN SEPARATION LABEL
                            self.currentGap, # self.currentGap
                            len(self.DepSTANDqueue),
                            len(self.TAXIhold),
                            len(self.RWYqueue1),
                            len(self.RWYqueue2),
                            None, # len(self.RWYqueue3)
                            None, # len(self.RWYqueue4)
                            self.RWYqueue1[first_in_line_RWYqueue][4], # DELAY self.DepSTANDqueue
                            self.RWYqueue1[first_in_line_RWYqueue][8], # DELAY self.TAXIhold
                            self.RWYqueue1[first_in_line_RWYqueue][10], # DELAY RWYqueue
                            self.RWYqueue1[first_in_line_RWYqueue][11], # RWYqueue USED
                        ])
                        del self.RWYqueue1[first_in_line_RWYqueue]

                    elif currentRWYqueue == 2:
                        self.writeRow('Departure_Output', self.DepOutput, 1, [
                            self.RWYqueue2[first_in_line_RWYqueue][0], # AC ID
                            None, # Skip column B
                            None, # Skip column C
                            self.readCell('Departure_Output', 'C', self.DepOutput) + self.RWYqueue2[first_in_line_RWYqueue][2], # Dep RWY EXIT = Dep RWY ENTRY + DROT
                            self.readCell('Departures', 'H', first_in_line_RWYqueue), #WAKE
                            self.RWYqueue2[first_in_line_RWYqueue][3], # SID
                            self.RWYqueue2[first_in_line_RWYqueue][2], # DROT
                            self.RWYqueue2[first_in_line_RWYqueue][6], # TAXIOUT
                            minDepTime, # DEP MIN SEPARATION
                            minDepLabel, # DEP MIN SEPARATION LABEL
                            self.currentGap, # self.currentGap
                            len(self.DepSTANDqueue),
                            len(self.TAXIhold),
                            len(self.RWYqueue1),
                            len(self.RWYqueue2),
                            None, # len(self.RWYqueue3)
                            None, # len(self.RWYqueue4)
                            self.RWYqueue2[first_in_line_RWYqueue][4], # DELAY self.DepSTANDqueue
                            self.RWYqueue2[first_in_line_RWYqueue][8], # DELAY self.TAXIhold
                            self.RWYqueue2[first_in_line_RWYqueue][10], # DELAY RWYqueue
                            self.RWYqueue2[first_in_line_RWYqueue][11], # RWYqueue USED
                        ])
                        del self.RWYqueue2[first_in_line_RWYqueue]

                    self.writeRow('Sequence', self.seqRow, 1, [
                        'D',
                        self.readCell('Departure_Output', 'A', self.ArrOutput),
                        self.readCell('Departure_Output', 'C', self.ArrOutput),
                        self.readCell('Departure_Output', 'D', self.ArrOutput),
                        self.readCell('Departure_Output', 'G', self.ArrOutput)
                    ])
                    self.seqRow += 1
                    self.DepOutput += 1

        return(self.DepOutput,self.seqRow)


    # ---------------------------- Model Run Function ---------------------------- #

    def runRAPID(self):
        Current_time = self.start_time
        while Current_time < self.end_time:

            if self.RWY_status == "E":
                if self.TOTAL_DEPARTURES > 0: #there are departures
                    self.SOBTrow = self.SOBTlookup(Current_time)
                    if ((len(self.TAXIqueue) + len(self.ARRIVALqueue)+ len(self.TAXIhold))<15) and len(self.DepSTANDqueue)> 0: # if there are less than 15 AC moving on the TAXIway and there's something in self.DepSTANDqueue
                        self.TAXIqueue_update(Current_time)
                    self.TAXIhold_update(Current_time)
                    self.RWYqueues_update(Current_time)
                    self.update_Departure_Delays(Current_time)

                if self.TOTAL_ARRIVALS > 0: # there are arrivals
                    self.ARRkey = self.SAE_lookup(Current_time)
                    if len(self.APPqueue) == 0:
                        self.update_APPqueue(Current_time)
                    self.update_ARRIVALqueue(Current_time)
                    self.currentGap = self.update_currentGap(Current_time)
                else:#if there aren't any arrivals
                    self.currentGap = self.end_time #huuuuuge gap

                #DEPARTURES TAKE OFF
                if self.TOTAL_DEPARTURES > 0:
                    if (len(self.RWYqueue1)+len(self.RWYqueue2))>0:#there is something waiting to takeoff
                        #print('TAKE OFF called')
                        self.DepOutput, self.seqRow = self.Dep_TAKE_OFF(Current_time)
                        #print('dep took off')
                        # Note : DepOurputROW was already increased so (DepOutputROW-1) will reffer to the current departure
                        #if type(self.workbook['Departure_Output']['C' + str(self.DepOutput-1)].value) == int:
                            #print(self.workbook['Departure_Output']['C' + str(self.DepOutput-1)].value)
                        if Current_time < self.workbook['Departure_Output']['D' + str(self.DepOutput-1)].value : # while the Departure is still on the runway
                            #print(Current_time,' Departure {',(self.DepOutput-1),'} is about to take-off')
                            self.RWY_status = "D"

                #ARRIVALS LANDING
                if self.TOTAL_ARRIVALS > 0:
                    if len(self.APPqueue)!=0:
                        first_in_line_APPqueue = min(list(self.APPqueue.keys()))#there is only one AC in the APPqueue
                        if Current_time == self.APPqueue[first_in_line_APPqueue][6]: #it's time to land
                            #print('Current_time = ', Current_time, '| ALT = ',APPqueue[first_in_line_APPqueue][6])
                            self.ArrOutput, self.seqRow = self.Arr_LANDING(Current_time, first_in_line_APPqueue)
                            if Current_time < self.workbook['Arrival_Output']['D' + str(self.ArrOutput-1)].value : #while Arrival is still on the runway
                                #print(Current_time,' Arrival {',self.ArrOutput-1,'} is about to land ')
                                self.RWY_status = "A"

            elif self.RWY_status == "D":
                #print(Current_time,' | ', self.RWY_status)
                self.SOBTrow = self.SOBTlookup(Current_time)
                if ((len(self.TAXIqueue) + len(self.ARRIVALqueue)+ len(self.TAXIhold))<15) and len(self.DepSTANDqueue)> 0: # if there are less than 15 AC moving on the TAXIway and there's something in DepSTANDqueue
                    self.TAXIqueue_update(Current_time)
                self.TAXIhold_update(Current_time)
                self.RWYqueues_update(Current_time)
                self.update_Departure_Delays(Current_time)

                if self.TOTAL_ARRIVALS > 0: #there are arrivals
                    self.ARRkey = self.SAE_lookup(Current_time)
                    if len(self.APPqueue) == 0:
                        self.update_APPqueue(Current_time)
                    self.update_ARRIVALqueue(Current_time)
                    self.currentGap = self.update_currentGap(Current_time)
                else:#if there aren't any arrivals
                    self.currentGap = self.end_time #huuuuuge gap

                if Current_time == self.workbook['Departure_Output']['D' + str(self.DepOutput-1)].value : # when current_time > departure RWY_EXIT the rwy is empty again
                    self.RWY_status = "E"

                #ARRIVALS LANDING (GO-AROUND case)
                if self.TOTAL_ARRIVALS > 0:
                    if len(self.APPqueue)!=0:
                        first_in_line_APPqueue = min(list(self.APPqueue.keys()))#there is only one AC in the APPqueue
                        if Current_time == self.APPqueue[first_in_line_APPqueue][6]: #it's time to land
                            #print('It is time to land but GOaround')
                            self.ArrOutput, self.seqRow = self.Arr_LANDING(Current_time, first_in_line_APPqueue)

                            #print(self.ArrOutput,'******GO AROUND************')

            elif self.RWY_status == "A":
                #print(Current_time,' | ', self.RWY_status)
                if self.TOTAL_DEPARTURES > 0: #there are departures
                    self.SOBTrow = self.SOBTlookup(Current_time)
                    if ((len(self.TAXIqueue) + len(self.ARRIVALqueue)+ len(self.TAXIhold))<15) and len(self.DepSTANDqueue)> 0: # if there are less than 15 AC moving on the TAXIway and there's something in DepSTANDqueue
                        self.TAXIqueue_update(Current_time)
                    self.TAXIhold_update(Current_time)
                    self.RWYqueues_update(Current_time)
                    self.update_Departure_Delays(Current_time)

                self.ARRkey = self.SAE_lookup(Current_time)
                if len(self.APPqueue) == 0:
                    self.update_APPqueue(Current_time)
                self.update_ARRIVALqueue(Current_time)
                self.currentGap = self.update_currentGap(Current_time)

                if Current_time == self.workbook['Arrival_Output']['D' + str(self.ArrOutput-1)].value : #while Arrival is still on the runway
                    self.RWY_status = "E"

            Current_time += 1


    # ------------------------------ Extra Functions ----------------------------- #

    def bufferCalculation(self):

        bufferRow = 2
        for row in range (2, self.workbook['Sequence'].max_row - 2):
            if (self.readCell('Sequence', 'A', row) == "A") and (self.readCell('Sequence', 'A', row + 1) == "D") and  (self.readCell('Sequence', 'A', row + 2)  == "A") :#ADA sequence:
                self.writeCell('Sequence', 'G', bufferRow, self.readCell('Sequence', 'C', row + 1) - self.readCell('Sequence', 'D', row) + self.readCell('Sequence', 'C', row + 2) - self.readCell('Sequence', 'D', row + 1))
                self.writeCell('Sequence', 'F', bufferRow, self.readCell('Sequence', 'B', row))
                bufferRow += 1


    def throughputCalculation(self):

        min_thr_HOUR = min(self.readCell('Arrival_Output', 'B', 2), self.readCell('Departure_Output', 'B', 2))
        max_thr_HOUR = min(self.readCell('Arrival_Output', 'B', self.ArrOutput - 1), self.readCell('Departure_Output', 'B', self.DepOutput - 1))
        diff_thr_HOUR = max_thr_HOUR-min_thr_HOUR

        for row in range(2,(diff_thr_HOUR + 3)):
            dep_thr_count = 0
            arr_thr_count = 0
            self.writeCell('Throughput', 'A', row, min_thr_HOUR)

            for i in range(2, (self.DepOutput)):
                if self.readCell('Departure_Output', 'B', i) == None: # no departures
                    break
                if self.readCell('Departure_Output', 'B', i) == min_thr_HOUR:
                    dep_thr_count +=1
            self.writeCell('Throughput', 'B', row, dep_thr_count)
            for i in range(2, (self.ArrOutput)):
                if self.readCell('Arrival_Output', 'B', i) == None: # no arrivals
                    break
                if self.readCell('Arrival_Output', 'B', i) == min_thr_HOUR:
                    arr_thr_count +=1

            self.writeCell('Throughput', 'C', row, arr_thr_count)
            self.writeCell('Throughput', 'D', row, self.readCell('Throughput', 'B', row) + self.readCell('Throughput', 'C', row))
            total_thr = self.readCell('Throughput', 'D', row)
            self.throughput.append(total_thr)

            if min_thr_HOUR in list(self.GoAroundCount.keys()): # there was at least a goAround at that hour
                self.writeCell('Throughput', 'E', row, sum(self.GoAroundCount[min_thr_HOUR]))
            else:
                self.writeCell('Throughput', 'E', row, 0)

            min_thr_HOUR +=1


    def delayCalculation(self):

        for row in range(2, self.DepOutput):
            self.writeRow('Delay', row, 1, [
                self.readCell('Departure_Output', 'A', row),
                self.readCell('Departure_Output', 'B', row),
                self.readCell('Departure_Output', 'T', row) + self.readCell('Departure_Output', 'S', row),
                self.readCell('Departure_Output', 'R', row)
            ])

        for row in range(2, self.ArrOutput):
            self.writeRow('Delay', row, 9, [
                self.readCell('Departure_Output', 'A', row),
                self.readCell('Departure_Output', 'B', row),
                self.readCell('Departure_Output', 'L', row)
            ])

        self.number_of_goArounds_queued = 0
        for i in list(self.GoAroundCount.keys()):
            self.number_of_goArounds_queued+=sum(self.GoAroundCount[i])


    def printDiagnosticMsgs(self):
        for queue in ['RWYqueue1', 'RWYqueue2', 'DepSTANDqueue', 'TAXIhold', 'ARRIVALqueue', 'APPqueue', 'ArrHOLDqueue']:
            eval_str = f'str(len(self.{queue}))'
            remaining_aircraft = eval(eval_str)
            if int(remaining_aircraft) > 0:
                print(f'WARNING: there are [{remaining_aircraft}] aircraft remaining in {queue}')
        print(f'Final number of queued arrival go-around cases: {self.number_of_goArounds_queued}')


def runModel(parentFrame):

    # Variables from GUI
    global v
    v = {
        'filename': parentFrame.name_input_file,
        'RECAT': bool(int(parentFrame.opt['var6'].get())), # Switch for modelling 'Radar Tower Separation' concept
        'RECAT_PWS': bool(int(parentFrame.opt['var17'].get())),
        'avgThr': bool(int(parentFrame.run['var7'].get())),
        'distanceBased': not bool(int(parentFrame.opt['separation_type'].get())),
        'timeBased': bool(int(parentFrame.opt['separation_type'].get())),
        'MRS_4DME': bool(int(parentFrame.opt['MRS_4dme'].get())),
        'WAKE_4DME': bool(int(parentFrame.opt['WAKE_4dme'].get())),
        'ADA_4DME': bool(int(parentFrame.opt['ADA_4dme'].get())),
        'ADDA_4DME': bool(int(parentFrame.opt['ADDA_4dme'].get())),
        'MRS_THR': bool(int(parentFrame.opt['MRS_thr'].get())),
        'WAKE_THR': bool(int(parentFrame.opt['WAKE_thr'].get())),
        'ADA_THR': bool(int(parentFrame.opt['ADA_thr'].get())),
        'ADDA_THR': bool(int(parentFrame.opt['ADDA_thr'].get())),
        'debugTab': bool(int(parentFrame.run['var14'].get())),
        'queueType': str(parentFrame.req['queue_type'].get()),
        'maxRuns': int(parentFrame.run['n_times_input'].get()),
        'n': int(parentFrame.req['n_input'].get()),
        'ADA_x': int(parentFrame.opt['ADA_x_input'].get()), # UNUSED
        'minDep_altSID': int(parentFrame.req['minDep_altSID_input'].get()),
        'minDep_sameSID': int(parentFrame.req['minDep_sameSID_input'].get()),
        'SIDmax': int(parentFrame.req['SIDmax_input'].get()), #UNUSED
        'SIDgroup_separation': str(parentFrame.req['SIDgroup_separation_input'].get()),
        'SID_queue_assign': str(parentFrame.req['SID_queue_assign_input'].get())
    }

    big_list = []
    averages = []
    difference = []
    iter1 = 0
    maxIter = 10 if v['avgThr'] else v['maxRuns']

    while (iter1 < maxIter):

        timer_start = time.perf_counter()

        # Initialisation and pre-processing
        model = coreRAPID()

        # Model run
        model.runRAPID()

        # Additional calculations
        model.bufferCalculation()
        model.throughputCalculation()
        model.delayCalculation()

        # Print diagnostic messages
        model.printDiagnosticMsgs()


        def saveOutput():
            # Append column F to Throughput tab
            model.writeCell('Throughput', 'F', 1, 'Difference in thr averages')
            difference.append([0] * (model.workbook['Throughput'].max_row - 1))
            model.writeCell('Throughput', 'F', 2, str(difference))

            # Save workbook
            if v['avgThr']:
                parentFrame.name_output_file = f'OUTPUT_RAPID_v3.0_{time.strftime("%Y%m%d_%H%M%S", time.localtime())}_iter_{iter1 + 1}.xlsx'
            else:
                parentFrame.name_output_file = f'OUTPUT_RAPID_v3.0_{time.strftime("%Y%m%d_%H%M%S", time.localtime())}.xlsx'
            model.workbook.save(parentFrame.name_output_file)


        # Calculate multi-run average throughput and save output
        if v['avgThr']:
            big_list.append(model.throughput)
            average_run = []
            diff2 = []
            diff = 0
            average_hour = 0
            summ = 0
            for j in range(0, len(model.throughput)):
                for i in range (0, len(big_list)):
                    print(f'element sum = {big_list[i][j]}')
                    summ += big_list[i][j]
                average_hour = summ / len(big_list)
                average_run.append(average_hour)
                print(f'average_run = {average_run}')
            averages.append(average_run)
            for j in range(0, len(averages[0])):
                print(f'______________________ {j}')
                compare = averages[len(averages)-1][j]
                print(f'last run = {compare}')
                diff = compare - averages[len(averages)-2][j]
                print(f'diff = {diff}')
                diff2.append(diff)
                print(f'diff2 = {diff2}')
            difference.append(diff2)
            if maxIter >= 10:
                for i in range(0, len(difference[0])):
                    if difference[len(difference)-1][i] <= 0.1 and difference[len(difference)-1][i] >= -0.1 :
                        print(f'difference in averages = {difference[len(difference)-1][i]}')
                    else:
                        print('condition false')
                        maxIter += 1
                        print(f'maxRuns 2 = {maxIter}')
                        break
                else:
                    print(f'maxRuns 1 = {maxIter}')
                    saveOutput()
            else:
                maxIter += 1
        else:
            saveOutput()

        print(f'Time elapsed: {round(time.perf_counter() - timer_start, 4)} seconds\n\n')
        iter1 += 1
