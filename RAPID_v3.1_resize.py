"""
 Compatibility : N/A
 Inputs*:1.OPERATIONAL DATA :  Operational_data.csv
         2.INPUT FILE : Input_File_RAPID_v3.0
         3.SCHEDULE : Schedule_File_RAPID_v3.0
 *All the inputs can be found in the folder called 'inputs'
 ! Make sure the utility folder contains: 
         - actual_speed_profile.csv (for the version when the actual speed profiles filtering is not available)
         - RECAT_EU_separation.csv
         - RECAT_PWS.csv
         - RECAT20_separation.csv
         - UK_wake_separation.csv
         - wake.csv
 Outputs: 1.utility/AROTDROT_distributions.csv
          2.Input_File_RAPID_v3.0_ + (time) + .xlsx
          3.OUTPUT_RAPID_v3.0_ + str(output_extension) +  '.xlsx'
         
"""            

#*****************************************************************************#
#=============================================================================#
#                                                                             #
#                           IMPORT FUNCTIONS                                  #
#                                                                             #
# ============================================================================#
#*****************************************************************************#

import openpyxl
# Useful guide for openpyxl here - https://automatetheboringstuff.com/chapter12/
import random
import time
import sys
import math
import pandas as pd
import numpy as np
import json
import tkinter as tk      
import scipy.stats as stats

#from tkinter import Tk, IntVar, StringVar, W, E, S, N, ttk, Frame, filedialog
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure

from tkinter import Tk, IntVar, StringVar, W, E, S, N, ttk, Frame, filedialog, Scale, HORIZONTAL, LabelFrame#, DoubleVar, CENTER
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
#import csv

#*****************************************************************************#
#=============================================================================#
#                                                                             #
#                                 GUI                                         #
#                                                                             #
# ============================================================================#
#*****************************************************************************#


#----------------------------GLOBAL VARIABLES---------------------------------#

#====== FOR CORE: ======#

# Default DROT + buffer value ('n') variable for main runway
n = 50
# Default DROT + buffer value ('n') variable for northern runway
northern_n = 20
# Default minimum Departure Separation (Alt SIDs) in secs
minDep_altSID = 60 # minDep_altSID = 60
# Default minimum Departure Separation (Same SIDs) in secs
minDep_sameSID = 109 #minDep_sameSID = 120
SIDmax = 4
n_times = 1
ADA_x = 10
#SIDmax = 4
SIDgroup_separation = "(2,4)(3,4)"
SID_queue_assign = "1 3 | 2 4"


debugFLAG = False
debugFLAG2 = False # TAXI - RWY queue checks
multipleFLAG = False # used for multiple runs
TBS_Flag = False
RECatFLAG = False
averagethrFLAG = False
taxi_outliers = True

distance_based_FLAG = True
time_based_FLAG = False

MRS_4dme_FLAG = False
WAKE_4dme_FLAG = False
ADA_4dme_FLAG = False
ADDA_4dme_FLAG = False
MRS_thr_FLAG = False
WAKE_thr_FLAG = False
ADA_thr_FLAG = False
ADDA_thr_FLAG = False




open_file1 = ""
open_file = ""
input_excel_sheet = ""
name_input_file = ""
#input_excel_sheet = ""
operational_data = ""
new_data2 = ""
new_data4 = ""
new_data5 = ""
new_data6 = ""



#====FOR VISUAL MODULE======#

m = 0

Thr_FLAG = False
Delay_FLAG = False
Seq_FLAG = False
OP_FLAG = False
new_set_FLAG = False
average_FLAG = False
arr_delay_FLAG = False
convergenceFLAG = False
ADA_buffer_FLAG = False


input_excel_sheet = ""
operational_data = ""
new_data2 = ""
new_data4 = ""
new_data5 = ""
new_data6 = ""

#----------------------------GLOBAL FUNCTIONS---------------------------------#

def load_file():
    input_excel_sheet = filedialog.askopenfilename()
    print(input_excel_sheet)
    name_excel_sheet.set(input_excel_sheet)
    return()

def load_op_data():
    operational_data = filedialog.askopenfilename()
    print(operational_data)
    op_data_sheet.set(operational_data)
    return()
        
def load_new_data2():
    new_data2 = filedialog.askopenfilename()
    print(new_data2)
    new_data_sheet2.set(new_data2)
    return()
        
def load_new_data3():
    new_data3 = filedialog.askopenfilename()
    print(new_data3)
    new_data_sheet3.set(new_data3) 
    return()

def load_new_data4():
    new_data4 = filedialog.askopenfilename()
    print(new_data4)
    new_data_sheet4.set(new_data4) 
    return() 

def load_new_data5():
    new_data5 = filedialog.askopenfilename()
    print(new_data5)
    new_data_sheet5.set(new_data5) 
    return()    
    
def load_new_data6():
    new_data6 = filedialog.askopenfilename()
    print(new_data6)
    new_data_sheet6.set(new_data6) 
    return() 

def define_input_parameters():
     # CORE
    n = int(n_input.get())
    n_output.set(n)
    minDep_altSID = int(minDep_altSID_input.get())
    minDep_altSID_output.set(minDep_altSID)
    minDep_sameSID = int(minDep_sameSID_input.get())
    minDep_sameSID_output.set(minDep_sameSID)
    SIDmax = int(SIDmax_input.get())
    SIDmax_output.set(SIDmax)
    SIDgroup_separation = str(SIDgroup_separation_input.get())
    SIDgroup_separation_output.set(SIDgroup_separation)
    SID_queue_assign = str(SID_queue_assign_input.get())
    SID_queue_assign_output.set(SID_queue_assign)
    n_times = int(n_times_input.get())
    n_times_output.set(n_times)  
    ADA_x = int(ADA_x_input.get())
    ADA_x_output.set(ADA_x)
    
    debug = int(var2.get())
    debug_output.set(debug)    
    q1 = int(queue1.get())
    queue1_output.set(q1)
    q2 = int(queue2.get())
    queue2_output.set(q2)
    q3 = int(queue3.get())
    queue3_output.set(q3)
    q4 = int(queue4.get())
    queue4_output.set(q4)
    button_check.set(True)
#    TBS = int(var5.get())
#    TBS_output.set(TBS)
    RECat = int(var6.get())
    RECat_output.set(RECat)
    RECAT_PWS = int(var17.get())
    RECAT_PWS_output.set(RECAT_PWS)
    averagethr = int(var7.get())
    averagethr_output.set(averagethr)
    distance_based = int(var15.get())
    distance_based_output.set(distance_based)
    time_based = int(var16.get())
    time_based_output.set(time_based)

    MRS4dme = int(MRS_4dme.get())
    MRS4dme_output.set(MRS4dme)
    WAKE4dme = int(WAKE_4dme.get())
    WAKE4dme_output.set(WAKE4dme)
    ADA4dme = int(ADA_4dme.get())
    ADA4dme_output.set(ADA4dme)
    ADDA4dme = int(ADDA_4dme.get())
    ADDA4dme_output.set(ADDA4dme)
    
    MRSthr = int(MRS_thr.get())
    MRSthr_output.set(MRSthr)
    WAKEthr = int(WAKE_thr.get())
    WAKEthr_output.set(WAKEthr)
    ADAthr = int(ADA_thr.get())
    ADAthr_output.set(ADAthr)
    ADDAthr = int(ADDA_thr.get())
    ADDAthr_output.set(ADDAthr)   

    #VISUAL
    m = int(m_input.get())
    m_output.set(m)
    convergence = int(var0.get())
    convergence_output.set(convergence)
    Throughput_check = int(var8.get())
    Throughput_check_output.set(Throughput_check)
    Delay_check = int(var9.get())
    Delay_check_output.set(Delay_check)
    arr_delay = int(var13.get())
    arr_delay_output.set(arr_delay)   
    Seq_check = int(var10.get())
    Seq_check_output.set(Seq_check)     
    op_yes = int(var11.get())
    op_yes_output.set(op_yes)
    new_set = int(var12.get())
    new_set_output.set(new_set)
    ADA_buffer = int(var18.get())
    ADA_buffer_output.set(ADA_buffer)
    button_check.set(True)
    
    win.destroy()

def define_input_parameters2():
    
    #CORE:
    n = int(n_input.get())
    n_output.set(n)
    minDep_altSID = int(minDep_altSID_input.get())
    minDep_altSID_output.set(minDep_altSID)
    minDep_sameSID = int(minDep_sameSID_input.get())
    minDep_sameSID_output.set(minDep_sameSID)
    SIDmax = int(SIDmax_input.get())
    SIDmax_output.set(SIDmax)
    SIDgroup_separation = str(SIDgroup_separation_input.get())
    SIDgroup_separation_output.set(SIDgroup_separation)
    SID_queue_assign = str(SID_queue_assign_input.get())
    SID_queue_assign_output.set(SID_queue_assign)
    n_times = int(n_times_input.get())
    n_times_output.set(n_times)  
    ADA_x = int(ADA_x_input.get())
    ADA_x_output.set(ADA_x)
    
    debug = int(var2.get())
    debug_output.set(debug)    
    q1 = int(queue1.get())
    queue1_output.set(q1)
    q2 = int(queue2.get())
    queue2_output.set(q2)
    q3 = int(queue3.get())
    queue3_output.set(q3)
    q4 = int(queue4.get())
    queue4_output.set(q4)
    button_check.set(True)
#    TBS = int(var5.get())
#    TBS_output.set(TBS)
    RECat = int(var6.get())
    RECat_output.set(RECat)
    RECAT_PWS = int(var17.get())
    RECAT_PWS_output.set(RECAT_PWS)
    averagethr = int(var7.get())
    averagethr_output.set(averagethr)
    distance_based = int(var15.get())
    distance_based_output.set(distance_based)
    time_based = int(var16.get())
    time_based_output.set(time_based)
    
    MRS4dme = int(MRS_4dme.get())
    MRS4dme_output.set(MRS4dme)
    WAKE4dme = int(WAKE_4dme.get())
    WAKE4dme_output.set(WAKE4dme)
    ADA4dme = int(ADA_4dme.get())
    ADA4dme_output.set(ADA4dme)
    ADDA4dme = int(ADDA_4dme.get())
    ADDA4dme_output.set(ADDA4dme)
    
    MRSthr = int(MRS_thr.get())
    MRSthr_output.set(MRSthr)
    WAKEthr = int(WAKE_thr.get())
    WAKEthr_output.set(WAKEthr)
    ADAthr = int(ADA_thr.get())
    ADAthr_output.set(ADAthr)
    ADDAthr = int(ADDA_thr.get())
    ADDAthr_output.set(ADDAthr)
    
    #VISUAL
    m = int(m_input.get())
    m_output.set(m)
    convergence = int(var0.get())
    convergence_output.set(convergence)
    Throughput_check = int(var8.get())
    Throughput_check_output.set(Throughput_check)
    Delay_check = int(var9.get())
    Delay_check_output.set(Delay_check)
    arr_delay = int(var13.get())
    arr_delay_output.set(arr_delay)   
    Seq_check = int(var10.get())
    Seq_check_output.set(Seq_check)     
    op_yes = int(var11.get())
    op_yes_output.set(op_yes)
    new_set = int(var12.get())
    new_set_output.set(new_set)
    ADA_buffer = int(var18.get())
    ADA_buffer_output.set(ADA_buffer)
    button_check.set(True)  
        
    win.destroy()


#-----------------------------------------------------------------------------#

win = tk.Tk()                           # Create instance      
win.title("RAPID V_2.0")                 # Add a title 

# TABS
mainframe = ttk.Frame(win, padding="10 10 30 40")

mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
mainframe.columnconfigure(0, weight=1)
mainframe.rowconfigure(0, weight=1)
    
input_module = ttk.Frame(mainframe)            # Create INPUT tab 
input_module.grid(column=0, row=0, sticky=(N, W, E, S))

core_module = ttk.Frame(mainframe)            # Create CORE tab
core_module.grid(row=0, sticky=(N, W, E, S))

visual_module = ttk.Frame(mainframe)            # Create VISUAL tab 
visual_module.grid(column=0, row=0, sticky=(N, W, E, S))

# Fake TABS

fakeTab_input = ttk.Frame(input_module)
fakeTab_input.grid(row=0, sticky=(N, W, E, S))

tk.Button(fakeTab_input, text='INPUT MODULE', bg='pink', height = 3, padx=100).grid(row=0,column=0)
tk.Button(fakeTab_input, text='CORE MODULE', height = 3, padx=100 ).grid(row=0,column=1)
tk.Button(fakeTab_input, text='VISUAL MODULE', height = 3, padx=100).grid(row=0,column=2)


fakeTab_core = ttk.Frame(core_module)
fakeTab_core.grid(row=0, sticky=(N, W, E, S))

tk.Button(fakeTab_core, text='INPUT MODULE', height = 3, padx=100).grid(row=0,column=0)
tk.Button(fakeTab_core, text='CORE MODULE', bg='pink', height = 3, padx=100 ).grid(row=0,column=1)
tk.Button(fakeTab_core, text='VISUAL MODULE', height = 3, padx=100).grid(row=0,column=2)

fakeTab_visual = ttk.Frame(visual_module)
fakeTab_visual.grid(row=0, sticky=(N, W, E, S))

tk.Button(fakeTab_visual, text='INPUT MODULE', height = 3, padx=100).grid(row=0,column=0)
tk.Button(fakeTab_visual, text='CORE MODULE', height = 3, padx=100 ).grid(row=0,column=1)
tk.Button(fakeTab_visual, text='VISUAL MODULE', bg='pink', height = 3, padx=100).grid(row=0,column=2)



# Main Contentframes

tab1 = ttk.Frame(input_module)
tab1.grid(row=1, sticky=(N, W, E, S))

tab2 = ttk.Frame(core_module)
tab2.grid(row=1, sticky=(N, W, E, S))

tab3 = ttk.Frame(visual_module)
tab3.grid(row=1, sticky=(N, W, E, S))



#-------------------------------INPUT GUI-------------------------------------#


def raise_main_frame(frame):
    frame.tkraise()

for main_frames in (input_module, core_module, visual_module ):
    main_frames.grid(row=0, column=0, sticky='news')
    main_frames.columnconfigure(0, weight=0)
    main_frames.rowconfigure(0, weight=0)

import_airport_data = StringVar()
input_file = StringVar()
open_file = StringVar()

def raise_frame(frame):
    frame.tkraise()




f0 = ttk.Frame(tab1)
f1 = ttk.Frame(tab1)
f2 = ttk.Frame(tab1)
f3 = ttk.Frame(tab1)
f4 = ttk.Frame(tab1)
f5 = ttk.Frame(tab1)
f6 = ttk.Frame(tab1)
f7 = ttk.Frame(tab1)


for frame in (f0,f1,f2):
    frame.grid(row=0, column=0, sticky='news')
    frame.columnconfigure(0, weight=0)
    frame.rowconfigure(0, weight=0)

# Welcome Frame

#f0_help 
   
f0_help = ttk.LabelFrame(f0, text=" Quick Help ")
f0_help.grid(row=0, column=1, sticky=E, \
             padx=5, pady=5, ipadx=5, ipady=5)
ttk.Label(f0_help, text="- Only select ONE file before hitting the 'NEXT' button!", font="Helvetica 12 bold").grid(column=1, row=1, sticky=W) 
ttk.Label(f0_help, text="[ A ] should be selected for first time setup - starts 'Operational Analysis'", font=12).grid(column=1, row=2, sticky=W)  
ttk.Label(f0_help, text="[ B ] should be selected if an INPUT file already exists - skips to the CORE MODULE", font=12).grid(column=1, row=3, sticky=W) 

f0_content = LabelFrame(f0, text="    WELCOME TO RAPID    ", font="Helvetica 16 bold")
f0_content.grid(row=0, column=0, sticky=N, \
             padx=10, pady=40, ipadx=5, ipady=10)


f0_content_a = ttk.LabelFrame(f0_content, text=" [ A ] - Analyse & Filter Operational Data ")
f0_content_a.grid(row=2, column=0, sticky=N, \
             padx=5, pady=15, ipadx=5, ipady=5)


name_excel_sheet1 = StringVar()



tdf = pd.DataFrame()


def airport_data_load_file():
    airport_data = filedialog.askopenfilename()
    import_airport_data.set(airport_data)
    ttk.Label(f0_content_a, text="File Successfully Imported!").grid(column=1, row=3, sticky=N, pady=10)
    #print("#####", import_excel_schedule, "###", import_schedule)
    return()
    
def define_airport_data_parameters():
    button_check.set(True)    
    airport_data = import_airport_data.get()

def create_dataframe_operational_data():
    airport_data = import_airport_data.get()
    df_operational_data = pd.read_csv(airport_data)
    return df_operational_data

def generate_new_input():
    def raise_subframe(frame):
        frame.tkraise()
           
    for frame in (f1, f2, f3, f4, f5,f6):
        frame.grid(row=0, column=0, sticky='news')
        frame.columnconfigure(0, weight=0)
        frame.rowconfigure(0, weight=0)
        
        
    airport_data_load_file()
    define_airport_data_parameters()
#    create_dataframe_operational_data()
    tdf = create_dataframe_operational_data()
    
    #f0_buttons
    f0_buttons= tk.Frame(f0) 
    f0_buttons.grid(column = 0, row=1, columnspan = 7, sticky=(N, W, E, S))
        
    def raise_subframe_f0(frame):
        frame.tkraise()
        create_f1_to_f6()
    
    

   

    def create_f1_to_f6():

        # AROT FRAME ###############################################################################################
        
        #f1_help 
           
        f1_help = ttk.LabelFrame(f1, text=" Quick Help ")
        f1_help.grid(row=0, column=1, sticky=E, \
                     padx=5, pady=5, ipadx=5, ipady=5)
        ttk.Label(f1_help, text="Section used to filter AROT data for when runway is 'Constrained'", font=12).grid(column=1, row=1, sticky=W) 
        ttk.Label(f1_help, text="Select values for the demand and maximum AROT", font=12).grid(column=1, row=2, sticky=W)
        
        #f1_content
        f1_content = LabelFrame(f1, text="    AROT    ", font="Helvetica 14 bold")
        f1_content.grid(row=0, column=0, sticky=E, \
                     padx=10, pady=20, ipadx=5, ipady=5)
        
        #####################################################################
        #                             AROT DATA                             #
        #####################################################################
        
        Final_demand_output = IntVar()
        AROT_output = IntVar()
        
        
        columns_to_drop_AROT =['B1','DROT_Callsign','DROT_Line up time','DROT_Start to roll','DROT_Take off time','DROT_Runway Entry','DROT_Take off speed [kts]','DROT','DROT_Runway','DROT_Demand','DROT_Final Wake','DROT_Aircraft Type ICAO','DROT_SID (shortened)','B2','TAXI_OUT_S1','TAXI_OUT_S2','TAXI_OUT_S3','TAXI_OUT_S4','TAXI_OUT_S5','TAXI_OUT_S6','TAXI_OUT_S7','TAXI_OUT_S8','TAXI_OUT_S9','TAXI_OUT_S10','TAXI_OUT_S11','TAXI_OUT_S12','TAXI_OUT_S13','TAXI_OUT_S14','TAXI_OUT_S15','B3','TAXI_IN_S1','TAXI_IN_S2','TAXI_IN_S3','TAXI_IN_S4','TAXI_IN_S5','TAXI_IN_S6','TAXI_IN_S7','TAXI_IN_S8','TAXI_IN_S9','TAXI_IN_S10','TAXI_IN_S11','TAXI_IN_S12','TAXI_IN_S13','TAXI_IN_S14','TAXI_IN_S15','B4','ADA_id','ADA_ADA','ADA_Combined ROT','ADA_Buffer','ADA_Uniques','ADA_ADA counts','ADA_C_ROT counts','ADA_Buffer_Unique','ADA_Buffer_counts']
        df = tdf.drop(columns = columns_to_drop_AROT)
            
        #xl = pd.ExcelFile(airport_data)
        #xl = pd.ExcelFile("AROT_example_input.xlsx")

        #Find min max values
        max_demand = df['AROT_Demand'].max()
        min_demand = df['AROT_Demand'].min()
        #max_arot = df['AROT'].max()
        max_arot = 160 # Initialise to remove unrealistic outliers from data
        min_arot = df['AROT'].min()

        demand_output = min_demand
        arot_filter_output = max_arot
        
        def getThrottle(event):
            
            ax0.clear() # Needed otherwise creates a new series but old series remains - GG matplotlib...
            ax1.clear()
            ax2.clear()
            
            # Add titles back in every update...
            ax0.set_title('  Filter data by Runway Demand & max AROT ')
            ax0.set_ylabel('Aircraft Count')
            
            demand_output = Throttle.get()
            arot_filter_output = Throttle_arot.get()
            
            df_demand = df.loc[df['AROT_Demand'] >= demand_output]
            df_filtered = df_demand.loc[df_demand['AROT'] <= arot_filter_output]
        
            ## Update label value ##
            ttk.Label(DemandInputFrame, text=str(round((len(df_filtered.index) / total_df_entries)*100,2))).grid(column=2, row=3, sticky=N)
            
            ########## CHANGE ###
            df_plots = df_filtered.groupby(['AROT_Runway', 'AROT'])['AROT_Callsign'].count()
            df_plots = df_plots.reset_index(level=[0,1])
            df_plots.pivot(index='AROT', columns='AROT_Runway', values='AROT_Callsign').plot(kind='line', ax=ax0)
            #####################
            
            df_plots2 = df_filtered.groupby(['AROT_Runway', 'AROT_Final Wake'])['AROT_Callsign'].count()
            df_plots2 = df_plots2.reset_index(level=[0,1])
            df_plots2 = df_plots2.sort_values(by=['AROT_Runway', 'AROT_Final Wake', 'AROT_Callsign'], ascending=True)
            
            arrival_wakes = df_plots2['AROT_Final Wake'].unique()
            arrival_wakes = arrival_wakes.tolist()
        
            df_plots2.pivot(index='AROT_Final Wake', columns='AROT_Runway', values='AROT_Callsign').plot(kind='pie', ax=ax1, subplots=True, labels=arrival_wakes, autopct='%1.1f%%', shadow=False, startangle=90)
        
            df_plots3 = df_filtered.groupby(['AROT_Runway', 'AROT_RwyExit'])['AROT_Callsign'].count()
            df_plots3 = df_plots3.reset_index(level=[0,1])
            df_plots3 = df_plots3.pivot(index='AROT_RwyExit', columns='AROT_Runway', values='AROT_Callsign').fillna(0).plot(kind='bar', subplots=True, ax=ax2)
            plt.show()
        
            canvas.draw()
            canvas1.draw()
            canvas2.draw()
        
        def define_final_AROT():
            #Final_max_arot = 330
            #Final_demand = 45
            Final_demand = Throttle.get()
            Final_demand_output.set(Final_demand)
            Final_max_arot = Throttle_arot.get() 
            #Final_max_arot = arot_filter.get() # Old method (takes last value entered)
            print("AROT=", Final_max_arot, "Demand=", Final_demand)
            #print(Final_max_arot)
            AROT_output.set(Final_max_arot)
            button_check.set(True)
            
            
            ##AROT filtering + save it it a data frame rady to be exported to file
            
            df_final_AROT = pd.DataFrame()
            df_final_AROT_H = pd.DataFrame()
            df_final_AROT_M = pd.DataFrame()
            df_final_AROT_L = pd.DataFrame()
            df_final_AROT_UM = pd.DataFrame()
            df_final_AROT_J = pd.DataFrame()
            df_final_AROT_S = pd.DataFrame()
            
            
            
            df_final_AROT_H['AROT_H'] = ""
            df_final_AROT_H['AROT_H'] = df.loc[(df['AROT']<=Final_max_arot) & (df['AROT_Demand']>=Final_demand)& (df['AROT_Final Wake']=="H"),'AROT']
            df_final_AROT_H=df_final_AROT_H.reset_index()
            df_final_AROT_H = df_final_AROT_H.drop(columns='index')
            
            df_final_AROT_M['AROT_M'] = ""
            df_final_AROT_M['AROT_M'] = df.loc[(df['AROT']<=Final_max_arot) & (df['AROT_Demand']>=Final_demand)& (df['AROT_Final Wake']=="M"),'AROT']
            df_final_AROT_M=df_final_AROT_M.reset_index()
            df_final_AROT_M = df_final_AROT_M.drop(columns='index')
            
            df_final_AROT_L['AROT_L'] = ""
            df_final_AROT_L['AROT_L'] = df.loc[(df['AROT']<=Final_max_arot) & (df['AROT_Demand']>=Final_demand)& (df['AROT_Final Wake']=="L"),'AROT']
            df_final_AROT_L=df_final_AROT_L.reset_index()
            df_final_AROT_L = df_final_AROT_L.drop(columns='index')
            
            df_final_AROT_UM['AROT_UM'] = ""
            df_final_AROT_UM['AROT_UM'] = df.loc[(df['AROT']<=Final_max_arot) & (df['AROT_Demand']>=Final_demand)& (df['AROT_Final Wake']=="UM"),'AROT']
            df_final_AROT_UM=df_final_AROT_UM.reset_index()
            df_final_AROT_UM = df_final_AROT_UM.drop(columns='index')
            
            df_final_AROT_J['AROT_J'] = ""
            df_final_AROT_J['AROT_J'] = df.loc[(df['AROT']<=Final_max_arot) & (df['AROT_Demand']>=Final_demand)& (df['AROT_Final Wake']=="J"),'AROT']
            df_final_AROT_J=df_final_AROT_J.reset_index()
            df_final_AROT_J = df_final_AROT_J.drop(columns='index')
            
            df_final_AROT_S['AROT_S'] = ""
            df_final_AROT_S['AROT_S'] = df.loc[(df['AROT']<=Final_max_arot) & (df['AROT_Demand']>=Final_demand)& (df['AROT_Final Wake']=="S"),'AROT']
            df_final_AROT_S=df_final_AROT_S.reset_index()
            df_final_AROT_S = df_final_AROT_S.drop(columns='index')
            
            frames_AROT = [df_final_AROT_H, df_final_AROT_M,df_final_AROT_L,df_final_AROT_UM,df_final_AROT_J,df_final_AROT_S]
            
            df_final_AROT = pd.concat(frames_AROT, axis=1)
            return df_final_AROT
            #temp_string
        def define_arot_parameters(): 
            df_final_AROT = define_final_AROT()
            print("AROTs defined | Filters AROT = ", Throttle_arot.get(), "Demand=", Throttle.get())
            #print('inside the NEXT function = ', df_final_AROT)
            ttk.Label(inner, text=" AROT Data Exported!  ").grid(column=1, row=2, sticky=N, pady=10)
            raise_subframe(f2)
            
            return df_final_AROT
            #window.destroy()
            
        
        # Specify GUI Structure -------->
        DemandInputFrame = LabelFrame(f1_content, text="    Filter data based on the Demand    ", font="Helvetica 12")
        DemandInputFrame.grid(row=0, column=1, columnspan=1, sticky='N', \
                  padx=5, pady=10, ipadx=5, ipady=5)
        
        
        AROTInputFrame = LabelFrame(f1_content, text="    Filter data based on max. AROT    ", font="Helvetica 12")
        AROTInputFrame.grid(row=0, column=4, columnspan=7, sticky='N', \
                  padx=5, pady=10, ipadx=5, ipady=5)
        
        ResultsFrame = LabelFrame(f1_content, text="   [  Results :  ]   ", font="Helvetica 12")
        ResultsFrame.grid(row=1, columnspan=14, sticky='N', \
                  padx=5, pady=5, ipadx=5, ipady=5)
        
        # Defines expected inputs (i.e. GUI expects integers) and assigns default values
        demand_input = IntVar(f1_content, value=min_demand)
        arot_filter = IntVar(f1_content, value=max_arot)
        button_check = StringVar(f1_content, value='0')
        
        # OLD Filter for AROT
        #input_entry1 = ttk.Entry(AROTInputFrame, width=7, textvariable=arot_filter)
        #input_entry1.grid(row=1, column=2, sticky=N, padx=5, pady=35)
        
        ################## matplotlib figure ##################
        fig0 = plt.Figure()
        canvas = FigureCanvasTkAgg(fig0, ResultsFrame)
        canvas.get_tk_widget().grid(column=1, row=1, sticky=N, rowspan=2, padx=5, pady=5)
        ax0 = fig0.add_subplot(111)
        ax0.set_title('  Filter data by Runway Demand & max AROT ') 
        ax0.set_ylabel('Aircraft Count')
        
        fig1 = plt.Figure(figsize=(6,2.8))
        canvas1 = FigureCanvasTkAgg(fig1, ResultsFrame)
        canvas1.get_tk_widget().grid(column=2, row=1, sticky=N, padx=5, pady=5)    
        ax1 = fig1.add_subplot(111, aspect=1) #aspect=1 #aspect='equal'
        #ax1.axis('equal') # Equal aspect ratio ensures that pie is drawn as a circle.
        fig2 = plt.Figure(figsize=(6,2.8))
        canvas2 = FigureCanvasTkAgg(fig2, ResultsFrame)
        canvas2.get_tk_widget().grid(column=2, row=2, sticky=N, padx=5, pady=5)  
        ax2 = fig2.add_subplot(111) #aspect=1 #aspect='equal'
        fig2.subplots_adjust(hspace=0.4)
        
        ResultsFrame.columnconfigure(1, weight=1)
        ResultsFrame.columnconfigure(2, weight=1)
        ResultsFrame.rowconfigure(1, weight=1)
        ResultsFrame.rowconfigure(2, weight=1)
        
        # PLOT Initialisation
        
        ############ CHANGE #####
        df_demand = df.loc[df['AROT_Demand'] >= min_demand]
        df_filtered = df_demand.loc[df_demand['AROT'] <= max_arot]
        
        total_df_entries = len(df_filtered.index)
        
        df_plots = df_filtered.groupby(['AROT_Runway', 'AROT'])['AROT_Callsign'].count()
        df_plots = df_plots.reset_index(level=[0,1])
        df_plots.pivot(index='AROT', columns='AROT_Runway', values='AROT_Callsign').plot(kind='line', ax=ax0)
        #########################
        
        df_plots2 = df_demand.groupby(['AROT_Runway', 'AROT_Final Wake'])['AROT_Callsign'].count()
        df_plots2 = df_plots2.reset_index(level=[0,1])
        
        arrival_wakes = df_plots2['AROT_Final Wake'].unique()
        arrival_wakes = arrival_wakes.tolist()
        
        df_plots2.pivot(index='AROT_Final Wake', columns='AROT_Runway', values='AROT_Callsign').plot(kind='pie', subplots=True, ax=ax1, labels=arrival_wakes, autopct='%1.1f%%', startangle=90)
        ax1.axis('equal')
        
        df_plots3 = df_demand.groupby(['AROT_Runway', 'AROT_RwyExit'])['AROT_Callsign'].count()
        df_plots3 = df_plots3.reset_index(level=[0,1])
        df_plots3 = df_plots3.pivot(index='AROT_RwyExit', columns='AROT_Runway', values='AROT_Callsign').fillna(0).plot(kind='bar', subplots=True, ax=ax2)
        
        ttk.Label(DemandInputFrame, text=" Select a 'Demand value' from the Input file ->  ", font="Helvetica 10").grid(row=1, column=1, sticky=W, padx=140)
        
        Throttle = Scale(DemandInputFrame, from_=min_demand, to=max_demand, width=10, orient=HORIZONTAL, tickinterval=5, command=getThrottle)#variable = var)
        Throttle.grid(row=2, column=1, sticky='EW', padx=5)
        Throttle.set(0)
        
        Throttle_arot = Scale(AROTInputFrame, from_=min_arot, to=max_arot, width=10, orient=HORIZONTAL, tickinterval=20, command=getThrottle)#variable = var)
        Throttle_arot.grid(row=2, column=1, sticky='EW', padx=5)
        Throttle_arot.set(max_arot)
        
        ttk.Label(DemandInputFrame, text=" Percentage of original entries = ", font="Helvetica 10").grid(row=3, column=1, sticky=N, padx=5)
        ttk.Label(AROTInputFrame, text="      Select a Max. value for AROT ->             ", font="Helvetica 10").grid(row=1, column=1, sticky=N, padx=240)
        
# =============================================================================
#         DemandInputFrame.columnconfigure(1, weight=1)
#         DemandInputFrame.rowconfigure(0, weight=1)
#         DemandInputFrame.rowconfigure(1, weight=1)
#         DemandInputFrame.rowconfigure(2, weight=1)
#         DemandInputFrame.rowconfigure(3, weight=1)
#         
#         AROTInputFrame.columnconfigure(1, weight=1)
#         AROTInputFrame.rowconfigure(0, weight=1)
#         AROTInputFrame.rowconfigure(1, weight=1)
#         AROTInputFrame.rowconfigure(2, weight=1)
#         AROTInputFrame.rowconfigure(3, weight=1)
# =============================================================================
        
        #f1_content.columnconfigure(0, weight=0)
        #f1_content.rowconfigure(0, weight=0)
        
        #f1_help.columnconfigure(1, weight=0)
        #f1_help.rowconfigure(0, weight=0)
        #f1_help.rowconfigure(1, weight=0)
        
        #f1_content.columnconfigure(0, weight=1)
        #f1_content.columnconfigure(1, weight=1)
        #f1_content.columnconfigure(4, weight=1)
        #f1_content.rowconfigure(0, weight=1)
        #f1_content.rowconfigure(1, weight=1)
        
        #inner = LabelFrame(f1_content, bg='pink')
        inner = Frame(f1_content)
        inner.grid(row=2, column=1, sticky='E', \
                      padx=5, pady=10, ipadx=15, ipady=15)
        
        inner.grid_rowconfigure(0, weight=1)
        inner.grid_rowconfigure(2, weight=1)
        inner.grid_columnconfigure(0, weight=1)
        inner.grid_columnconfigure(2, weight=1)
        
        #ttk.Button(inner, text="Confirm Settings and Save", command=define_arot_parameters).grid(column=1, row=1, sticky=N, ipadx=5, ipady=5)
        tk.Button(inner, text='Confirm Settings and Save | NEXT ->', command=define_arot_parameters, activebackground = "pink", font=16, height = 1, overrelief="raised", width = 30).grid(column=1, row=1, sticky=N, ipadx=5, ipady=5)
        
        f1_content.bind('<Return>', define_arot_parameters)
        
        
        
        
        # DROT FRAME ###############################################################################################
        
        #f2_help 
           
        f2_help = ttk.LabelFrame(f2, text=" Quick Help ")
        f2_help.grid(row=0, column=1, sticky=E, \
                     padx=5, pady=5, ipadx=5, ipady=5)
        ttk.Label(f2_help, text="Section used to filter DROT data for when runway is 'Constrained'", font=12).grid(column=1, row=1, sticky=W) 
        ttk.Label(f2_help, text="Select values for the demand and maximum DROT", font=12).grid(column=1, row=2, sticky=W) 
        
        #f2_content
        f2_content = LabelFrame(f2, text="    DROT    ", font="Helvetica 14 bold")
        f2_content.grid(row=0, column=0, sticky=E, \
                     padx=10, pady=20, ipadx=5, ipady=5)
        #ttk.Label(f2_content, text="DROT Content...", font=16).grid(column=1, row=1, sticky=W) 
        
        #####################################################################        
        #                             DROT DATA                             #
        #####################################################################
           
        columns_to_drop_drot = ['AROT_Callsign','AROT_Threshold','AROT_RWY exit time','AROT','AROT_Runway','AROT_Demand','AROT_Final Wake','AROT_RwyExit','AROT_Aircraft Type ICAO','AROT_Threshold Speed [kts]','AROT_Speed @ TDZ [kts]','AROT_RWY Exit Speed 1','AROT_RWY Exit Speed 2','AROT_RWY Exit Speed 3','B1','B2','TAXI_OUT_S1','TAXI_OUT_S2','TAXI_OUT_S3','TAXI_OUT_S4','TAXI_OUT_S5','TAXI_OUT_S6','TAXI_OUT_S7','TAXI_OUT_S8','TAXI_OUT_S9','TAXI_OUT_S10','TAXI_OUT_S11','TAXI_OUT_S12','TAXI_OUT_S13','TAXI_OUT_S14','TAXI_OUT_S15','B3','TAXI_IN_S1','TAXI_IN_S2','TAXI_IN_S3','TAXI_IN_S4','TAXI_IN_S5','TAXI_IN_S6','TAXI_IN_S7','TAXI_IN_S8','TAXI_IN_S9','TAXI_IN_S10','TAXI_IN_S11','TAXI_IN_S12','TAXI_IN_S13','TAXI_IN_S14','TAXI_IN_S15','B4','ADA_id','ADA_ADA','ADA_Combined ROT','ADA_Buffer','ADA_Uniques','ADA_ADA counts','ADA_C_ROT counts','ADA_Buffer_Unique','ADA_Buffer_counts']
        df_drot = tdf.drop(columns = columns_to_drop_drot)
        
        
        Final_demand_output_d = IntVar()
        DROT_output = IntVar()
        
        #Find min max values
        d_max_demand = df_drot['DROT_Demand'].max()
        d_min_demand = df_drot['DROT_Demand'].min()
        
        max_drot = 160 # Initialise to remove unrealistic outliers from data
        min_drot = df_drot['DROT'].min()
        
        #total_drot_entries = len(df_drot.index)
        
        drot_demand_output = d_min_demand
        drot_filter_output = max_drot
        
        #                             DROT GUI                              #
        #####################################################################
        def getThrottle_d(event):
            
            ax4.clear() #ax
            ax5.clear() #ax1
            ax6.clear() #ax2
            
            # Add titles back in every update...
            ax4.set_title('  Filter data by Runway Demand & max DROT ')
            ax4.set_ylabel('Aircraft Count')
            
            drot_demand_output = Throttle_d.get()
            drot_filter_output = Throttle_drot.get()
            
            df_drot_demand = df_drot.loc[df_drot['DROT_Demand'] >= drot_demand_output]
            df_drot_filtered = df_drot_demand.loc[df_drot_demand['DROT'] <= drot_filter_output]
        
            ## Update label value ##
            ttk.Label(DemandInputFrame_d, text=str(round((len(df_drot_filtered.index) / total_drot_entries)*100,2))).grid(column=2, row=3, sticky=N)
            
            ########## CHANGE ###
            df_plots_d = df_drot_filtered.groupby(['DROT_Runway', 'DROT'])['DROT_Callsign'].count()
            df_plots_d = df_plots_d.reset_index(level=[0,1])
            df_plots_d.pivot(index='DROT', columns='DROT_Runway', values='DROT_Callsign').plot(kind='line', ax=ax4)
            #####################
            
            df_plots2_d = df_drot_filtered.groupby(['DROT_Runway', 'DROT_Final Wake'])['DROT_Callsign'].count()
            df_plots2_d = df_plots2_d.reset_index(level=[0,1])
            df_plots2_d = df_plots2_d.sort_values(by=['DROT_Runway', 'DROT_Final Wake', 'DROT_Callsign'], ascending=True)
            
            dep_wakes = df_plots2_d['DROT_Final Wake'].unique()
            dep_wakes = dep_wakes.tolist()
        
            df_plots2_d.pivot(index='DROT_Final Wake', columns='DROT_Runway', values='DROT_Callsign').plot(kind='pie', ax=ax5, subplots=True, labels=dep_wakes, autopct='%1.1f%%', shadow=False, startangle=90)
            ax5.axis('equal')
            
            df_plots3_d = df_drot_filtered.groupby(['DROT_Runway', 'DROT_SID (shortened)'])['DROT_Callsign'].count()
            df_plots3_d = df_plots3_d.reset_index(level=[0,1])
            ######### OPTION ONE - DOUBLE BAR CHART ##############
            #df_plots3_d = df_plots3_d.pivot(index='SID (shortened)', columns='Runway', values='Callsign').fillna(0).plot(kind='bar', subplots=True, ax=ax6)
            
            ######### OPTION TWO - DOUBLE PIE CHART ##############
            dep_SIDs = df_plots3_d['DROT_SID (shortened)'].unique()
            dep_SIDs = dep_SIDs.tolist()
            df_plots3_d = df_plots3_d.rename(columns = {'DROT_SID (shortened)':'DROT_SID'})
            df_plots3_d.pivot(index='DROT_SID', columns='DROT_Runway', values='DROT_Callsign').plot(kind='pie', subplots=True, ax=ax6, labels=dep_SIDs, autopct='%1.1f%%', legend=False, startangle=90)
            ax6.axis('equal')
        
            plt.show()
        
            canvas4.draw() #canvas
            canvas5.draw() #canvas1
            canvas6.draw() #canvas2
        
        def define_final_DROT():
#            Final_max_DROT = 3330
#            Final_demand_d = 45
            Final_demand_d = Throttle_d.get()
            Final_demand_output_d.set(Final_demand_d)
            Final_max_DROT = Throttle_drot.get()
            print("DROT=", Final_max_DROT, "Demand=", Final_demand_d)
            DROT_output.set(Final_max_DROT)
            button_check_d.set(True)
            
            #DROT filtering + save it it a data frame ready to be exported to file
            df_final_DROT = pd.DataFrame()
            df_final_DROT_H = pd.DataFrame()
            df_final_DROT_M = pd.DataFrame()
            df_final_DROT_L = pd.DataFrame()
            df_final_DROT_UM = pd.DataFrame()
            df_final_DROT_J = pd.DataFrame()
            df_final_DROT_S = pd.DataFrame()
            
            
            #F
            df_final_DROT_H['DROT_H'] = ""
            df_final_DROT_H['DROT_H'] = df_drot.loc[(df_drot['DROT']<=Final_max_DROT) & (df_drot['DROT_Demand']>=Final_demand_d)& (df_drot['DROT_Final Wake']=="H"),'DROT']
            df_final_DROT_H=df_final_DROT_H.reset_index()
            df_final_DROT_H = df_final_DROT_H.drop(columns='index')
            
            df_final_DROT_M['DROT_M'] = ""
            df_final_DROT_M['DROT_M'] = df_drot.loc[(df_drot['DROT']<=Final_max_DROT) & (df_drot['DROT_Demand']>=Final_demand_d)& (df_drot['DROT_Final Wake']=="M"),'DROT']
            df_final_DROT_M=df_final_DROT_M.reset_index()
            df_final_DROT_M = df_final_DROT_M.drop(columns='index')
            
            df_final_DROT_L['DROT_L'] = ""
            df_final_DROT_L['DROT_L'] = df_drot.loc[(df_drot['DROT']<=Final_max_DROT) & (df_drot['DROT_Demand']>=Final_demand_d)& (df_drot['DROT_Final Wake']=="L"),'DROT']
            df_final_DROT_L=df_final_DROT_L.reset_index()
            df_final_DROT_L = df_final_DROT_L.drop(columns='index')
            
            df_final_DROT_UM['DROT_UM'] = ""
            df_final_DROT_UM['DROT_UM'] = df_drot.loc[(df_drot['DROT']<=Final_max_DROT) & (df_drot['DROT_Demand']>=Final_demand_d)& (df_drot['DROT_Final Wake']=="UM"),'DROT']
            df_final_DROT_UM=df_final_DROT_UM.reset_index()
            df_final_DROT_UM = df_final_DROT_UM.drop(columns='index')
            
            df_final_DROT_J['DROT_J'] = ""
            df_final_DROT_J['DROT_J'] = df_drot.loc[(df_drot['DROT']<=Final_max_DROT) & (df_drot['DROT_Demand']>=Final_demand_d)& (df_drot['DROT_Final Wake']=="J"),'DROT']
            df_final_DROT_J=df_final_DROT_J.reset_index()
            df_final_DROT_J = df_final_DROT_J.drop(columns='index')
            
            df_final_DROT_S['DROT_S'] = ""
            df_final_DROT_S['DROT_S'] = df_drot.loc[(df_drot['DROT']<=Final_max_DROT) & (df_drot['DROT_Demand']>=Final_demand_d)& (df_drot['DROT_Final Wake']=="S"),'DROT']
            df_final_DROT_S=df_final_DROT_S.reset_index()
            df_final_DROT_S = df_final_DROT_S.drop(columns='index')
            
            frames_DROT = [df_final_DROT_H, df_final_DROT_M,df_final_DROT_L,df_final_DROT_UM,df_final_DROT_J,df_final_DROT_S]
            
            df_final_DROT = pd.concat(frames_DROT, axis=1)
            return df_final_DROT
            
        def define_drot_parameters():
            df_final_DROT = define_final_DROT()
            #ttk.Label(inner_d, text=" DROT Data Exported!  ").grid(column=1, row=2, sticky=N, pady=10) # Grids are banned FNAR!
            print("DROTs defined | Filters DROT = ", Throttle_drot.get(), "Demand=", Throttle_d.get())
            ttk.Label(inner_d, text=" DROT Data Exported!  ").pack(side="right")
            raise_subframe(f5)
            return df_final_DROT
            
            #window.destroy()
        
        # Specify GUI Structure -------->
        DemandInputFrame_d = LabelFrame(f2_content, text="   Filter data based on the Demand   ", font="Helvetica 12")
        DemandInputFrame_d.grid(row=0, column=1, columnspan=1, sticky='N', \
                  padx=5, pady=10, ipadx=5, ipady=5)
        
        DROTInputFrame = LabelFrame(f2_content, text="Filter data by max. DROT     ", font="Helvetica 12")
        DROTInputFrame.grid(row=0, column=4, columnspan=7, sticky='N', \
                  padx=10, pady=10, ipadx=5, ipady=5)
        
        ResultsFrame_d = LabelFrame(f2_content, text="  [  Results :  ]   ", font="Helvetica 12")
        ResultsFrame_d.grid(row=1, columnspan=14, sticky='N', \
                  padx=5, pady=5, ipadx=5, ipady=5)
        
        # Defines expected inputs (i.e. GUI expects integers) and assigns default values
        demand_input_d = IntVar(f2_content, value=d_min_demand)
        drot_filter = IntVar(f2_content, value=max_drot)
        button_check_d = StringVar(f2_content, value='0')
        
        # OLD METHOD for Filtering DROT
        #in_max_drot = df_drot['DROT'].max()
        #input_entry1_d = ttk.Entry(DROTInputFrame, width=7, textvariable=drot_filter)
        #input_entry1_d.grid(row=1, column=2, sticky=N, padx=10, pady=35)
        
        # matplotlib figures
        fig_d = plt.Figure()
        canvas4 = FigureCanvasTkAgg(fig_d, ResultsFrame_d)
        canvas4.get_tk_widget().grid(column=1, row=1, sticky=N, rowspan=2, padx=5, pady=5)
        ax4 = fig_d.add_subplot(111)
        ax4.set_title('  Filter data by Runway Demand & max DROT ')
        ax4.set_ylabel('Aircraft Count')
        
        fig1_d = plt.Figure(figsize=(6,2.8))
        canvas5 = FigureCanvasTkAgg(fig1_d, ResultsFrame_d)
        canvas5.get_tk_widget().grid(column=2, row=1, sticky=N, padx=5, pady=5)    
        ax5 = fig1_d.add_subplot(111, aspect=1)
        
        fig2_d = plt.Figure(figsize=(6,2.8))
        canvas6 = FigureCanvasTkAgg(fig2_d, ResultsFrame_d)
        canvas6.get_tk_widget().grid(column=2, row=2, sticky=N, padx=5, pady=5)  
        ax6 = fig2_d.add_subplot(111)
        fig2_d.subplots_adjust(hspace=0.4)
        
        # PLOT Initialisation

        
        ########## CHANGE ###
        df_drot_demand = df_drot.loc[df_drot['DROT_Demand'] >= d_min_demand]
        df_drot_filtered = df_drot_demand.loc[df_drot_demand['DROT'] <= max_drot]
        
        total_drot_entries = len(df_drot_filtered.index)
        
        df_plots_d = df_drot_filtered.groupby(['DROT_Runway', 'DROT'])['DROT_Callsign'].count()
        df_plots_d = df_plots_d.reset_index(level=[0,1])
        df_plots_d.pivot(index='DROT', columns='DROT_Runway', values='DROT_Callsign').plot(kind='line', ax=ax4)
        #####################
        
        df_plots2_d = df_drot_filtered.groupby(['DROT_Runway', 'DROT_Final Wake'])['DROT_Callsign'].count()
        df_plots2_d = df_plots2_d.reset_index(level=[0,1])
        
        dep_wakes = df_plots2_d['DROT_Final Wake'].unique()
        dep_wakes = dep_wakes.tolist()
        
        df_plots2_d.pivot(index='DROT_Final Wake', columns='DROT_Runway', values='DROT_Callsign').plot(kind='pie', subplots=True, ax=ax5, labels=dep_wakes, autopct='%1.1f%%', startangle=90)
        ax5.axis('equal')
        
        df_plots3_d = df_drot_filtered.groupby(['DROT_Runway', 'DROT_SID (shortened)'])['DROT_Callsign'].count()
        df_plots3_d = df_plots3_d.reset_index(level=[0,1])
        ######### OPTION ONE - DOUBLE BAR CHART ##############
        
        ######### OPTION TWO - DOUBLE PIE CHART ##############
        dep_SIDs = df_plots3_d['DROT_SID (shortened)'].unique()
        dep_SIDs = dep_SIDs.tolist()
        df_plots3_d = df_plots3_d.rename(columns = {'DROT_SID (shortened)':'DROT_SID'})
        df_plots3_d.pivot(index='DROT_SID', columns='DROT_Runway', values='DROT_Callsign').plot(kind='pie', subplots=True, ax=ax6, labels=dep_SIDs, autopct='%1.1f%%', legend=False, startangle=90)
        ax6.axis('equal')
        
        ######################################################
        
        ttk.Label(DemandInputFrame_d, text="Select a 'Demand value' from the Input file ->", font="Helvetica 10").grid(row=1, column=1, sticky=W, padx=140)
        
        Throttle_d = Scale(DemandInputFrame_d, from_=d_min_demand, to=d_max_demand, width=10, orient=HORIZONTAL, tickinterval=5, command=getThrottle_d)#variable = var)
        Throttle_d.grid(row=2, column=1, sticky='EW', padx=5)
        Throttle_d.set(0)
        
        Throttle_drot = Scale(DROTInputFrame, from_=min_drot, to=max_drot, width=10, orient=HORIZONTAL, tickinterval=20, command=getThrottle_d)#variable = var)
        Throttle_drot.grid(row=2, column=1, sticky='EW', padx=5)
        Throttle_drot.set(max_drot)
        
        ttk.Label(DemandInputFrame_d, text="  Percentage of original entries =  ", font="Helvetica 10").grid(row=3, column=1, sticky=N, padx=5)
        ttk.Label(DROTInputFrame, text="      Select a Max. value for DROT ->             ", font="Helvetica 10").grid(row=1, column=1, sticky=N, padx=240)
        
        inner_d = Frame(f2_content)
        inner_d.grid(row=2, column=1, sticky='E', \
                      padx=5, pady=10, ipadx=15, ipady=15)
        
        inner_d.grid_rowconfigure(0, weight=1)
        inner_d.grid_rowconfigure(2, weight=1)
        inner_d.grid_columnconfigure(0, weight=1)
        inner_d.grid_columnconfigure(2, weight=1)
        
        
        tk.Button(inner_d, text='Confirm Settings and Save | NEXT ->', command=define_drot_parameters, activebackground = "pink", font=16, height = 1, overrelief="raised", width = 30).pack(side="right")
        tk.Button(inner_d, text='<- BACK', command=lambda:raise_subframe(f1), activebackground = "pink", font=16, height = 1, overrelief="raised", width = 15).pack(side="right")
        f2_content.bind('<Return>', define_drot_parameters)
        
        
#        # TAXI-OUT FRAME #############################################################################################
#        
#        #f3_help 
#           
#        f3_help = ttk.LabelFrame(f3, text=" Quick Help ")
#        f3_help.grid(row=0, column=1, sticky=E, \
#                     padx=5, pady=5, ipadx=5, ipady=5)
#        ttk.Label(f3_help, text="Provides example analysis for Taxi-out distribution", font=12).grid(column=1, row=1, sticky=W)
#        ttk.Label(f3_help, text="Note the analysis only considers:", font=12).grid(column=1, row=2, sticky=W) 
#        ttk.Label(f3_help, text="  - 'Medium' wake aircraft", font=12).grid(column=1, row=3, sticky=W) 
#        
#        #f3_buttons
#        f3_buttons= ttk.Frame(f3) 
#        f3_buttons.grid(column = 0, row=1, columnspan = 7, sticky=(N, W, E, S))
#        
#        
#        tk.Button(f3_buttons, text='NEXT ->', command=lambda:raise_subframe(f4), activebackground = "pink", font=16, height = 1, overrelief="raised", width = 15).pack(side="right")
#        tk.Button(f3_buttons, text='<- BACK', command=lambda:raise_subframe(f2), activebackground = "pink", font=16, height = 1, overrelief="raised", width = 15).pack(side="right")
#        
#        #f3_content
#        f3_content = LabelFrame(f3, text=" TAXI-OUT ", font="Helvetica 14 bold")
#        f3_content.grid(row=0, column=0, sticky=E, \
#                     padx=10, pady=40, ipadx=5, ipady=10)
#        
##
#        columns_to_drop_taxi_out = ['AROT_Callsign','AROT_Threshold','AROT_RWY exit time','AROT','AROT_Runway','AROT_Demand','AROT_Final Wake','AROT_RwyExit','AROT_Aircraft Type ICAO','AROT_Threshold Speed [kts]','AROT_Speed @ TDZ [kts]','AROT_RWY Exit Speed 1','AROT_RWY Exit Speed 2','AROT_RWY Exit Speed 3','B1','DROT_Callsign','DROT_Line up time','DROT_Start to roll','DROT_Take off time','DROT_Runway Entry','DROT_Take off speed [kts]','DROT','DROT_Runway','DROT_Demand','DROT_Final Wake','DROT_Aircraft Type ICAO','DROT_SID (shortened)','B2','B3','TAXI_IN_S1','TAXI_IN_S2','TAXI_IN_S3','TAXI_IN_S4','TAXI_IN_S5','TAXI_IN_S6','TAXI_IN_S7','TAXI_IN_S8','TAXI_IN_S9','TAXI_IN_S10','TAXI_IN_S11','TAXI_IN_S12','TAXI_IN_S13','TAXI_IN_S14','TAXI_IN_S15','B4','ADA_id','ADA_ADA','ADA_Combined ROT','ADA_Buffer','ADA_Uniques','ADA_ADA counts','ADA_C_ROT counts','ADA_Buffer_Unique','ADA_Buffer_counts']
#        df_final_TAXIOUT = tdf.drop(columns=columns_to_drop_taxi_out)
#        df_taxi_out = df_final_TAXIOUT.rename(columns = {'TAXI_OUT_S1':'S1','TAXI_OUT_S2':'S2','TAXI_OUT_S3':'S3','TAXI_OUT_S4':'S4','TAXI_OUT_S5':'S5','TAXI_OUT_S6':'S6','TAXI_OUT_S7':'S7','TAXI_OUT_S8':'S8','TAXI_OUT_S9':'S9','TAXI_OUT_S10':'S10','TAXI_OUT_S11':'S11','TAXI_OUT_S12':'S12','TAXI_OUT_S13':'S13','TAXI_OUT_S14':'S14','TAXI_OUT_S15':'S15'})
#        
#        TOUT_output = IntVar()
#        
#        max_tout = df_taxi_out['S1'].max() #160 # Initialise to remove unrealistic outliers from data
#        min_tout = df_taxi_out['S1'].min()
#        
#        TOUT_filter_output = max_tout
#        
#        def getThrottle_to(event):
#            
#            ax92.clear() #ax
#            
#            ax92.set_title('  Taxi-out distribution example - Runway Direction = 26L, All Mediums ')
#            ax92.set_xlabel(' Taxi-out time (secs) ')
#            ax92.set_ylabel(' Aircraft count ')
#            
#            TOUT_filter_output = Throttle_tout.get()
#            
#            df_tout_filtered = df_taxi_out.loc[df_taxi_out['S1'] <= TOUT_filter_output]
#            
#            df_tout_filtered['S1'].plot(kind='hist', bins=100, rwidth=0.7, ax=ax92)
#            
#            plt.show()
#        
#            canvas92.draw() #canvas ax92.clear() #ax
#        
#        ResultsFrame_tout = LabelFrame(f3_content, text="  [  Taxi-out Results :  ]  ", font="Helvetica 12")
#        ResultsFrame_tout.grid(row=1, columnspan=14, sticky='N', \
#                  padx=5, pady=15, ipadx=5, ipady=5)
#        
#        
#        ttk.Label(ResultsFrame_tout, text="      Analysis of 2017 Summer Data from Airport X :  ", font="Helvetica 12").grid(column=1, row=0, sticky=N, pady=10, padx=20)
#        
#        fig_tout = plt.Figure()
#        canvas9 = FigureCanvasTkAgg(fig_tout, ResultsFrame_tout)
#        canvas9.get_tk_widget().grid(column=1, row=1, sticky=N, padx=5, pady=5)
#        ax9 = fig_tout.add_subplot(111)
#        ax9.set_title('  Taxi-out movements shown for Runway Direction = 26L ')
#        ax9.set_xlabel('Wake Category = [ Medium ] ')
#        ax9.set_ylabel('Time (secs)')
#        
#        df_taxi_out.plot(kind='box', ax=ax9)
#        
#        ## EXTRA Taxi-out plots!
#        
#        fig_tout_2 = plt.Figure()
#        canvas92 = FigureCanvasTkAgg(fig_tout_2, ResultsFrame_tout)
#        canvas92.get_tk_widget().grid(column=2, row=1, sticky=N, padx=5, pady=5)
#        ax92 = fig_tout_2.add_subplot(111)
#        ax92.set_title('  Taxi-out distribution example - Runway Direction = 26L, All Mediums ')
#        ax92.set_xlabel(' Taxi-out time (secs) ')
#        ax92.set_ylabel(' Aircraft count ')
#        
#        #df_taxi_out.plot(kind='box', ax=ax92)
#        
#        df_taxi_out['S1'].plot(kind='hist', bins=100, rwidth=0.7, ax=ax92)
#        
#        ttk.Label(ResultsFrame_tout, text="   Filter by Max. Taxi-out value ->     ", font="Helvetica 10").grid(row=2, column=2, sticky=N, padx=5, pady=10)
#        
#        Throttle_tout = Scale(ResultsFrame_tout, from_=min_tout, to=max_tout, width=10, orient=HORIZONTAL, tickinterval=100, command=getThrottle_to)#variable = var)
#        Throttle_tout.grid(row=3, column=2, sticky='EW', padx=5)
#        Throttle_tout.set(max_tout)
#        
#        # TAXI-IN FRAME #############################################################################################
#        
#        
#        #f4_help 
#           
#        f4_help = ttk.LabelFrame(f4, text=" Quick Help ")
#        f4_help.grid(row=0, column=1, sticky=E, \
#                     padx=5, pady=5, ipadx=5, ipady=5)
#        ttk.Label(f4_help, text="Provides example analysis for Taxi-in distribution", font=12).grid(column=1, row=1, sticky=W)
#        ttk.Label(f4_help, text="Note the analysis only considers:", font=12).grid(column=1, row=2, sticky=W) 
#        ttk.Label(f4_help, text="  - 'Medium' wake aircraft", font=12).grid(column=1, row=3, sticky=W) 
#        
#        #f4_buttons
#        f4_buttons= ttk.Frame(f4) 
#        f4_buttons.grid(column = 0, row=1, columnspan = 7, sticky=(N, W, E, S))
#        
#        
#        tk.Button(f4_buttons, text='NEXT ->', command=lambda:raise_subframe(f5), activebackground = "pink", font=16, height = 1, overrelief="raised", width = 15).pack(side="right")
#        tk.Button(f4_buttons, text='<- BACK', command=lambda:raise_subframe(f3), activebackground = "pink", font=16, height = 1, overrelief="raised", width = 15).pack(side="right")
#        
#        #f4_content
#        f4_content = LabelFrame(f4, text=" TAXI-IN ", font="Helvetica 14 bold")
#        f4_content.grid(row=0, column=0, sticky=E, \
#                     padx=10, pady=40, ipadx=5, ipady=10)
#        
#        
#        columns_to_drop_taxi_in = ['AROT_Callsign','AROT_Threshold','AROT_RWY exit time','AROT','AROT_Runway','AROT_Demand','AROT_Final Wake','AROT_RwyExit','AROT_Aircraft Type ICAO','AROT_Threshold Speed [kts]','AROT_Speed @ TDZ [kts]','AROT_RWY Exit Speed 1','AROT_RWY Exit Speed 2','AROT_RWY Exit Speed 3','B1','DROT_Callsign','DROT_Line up time','DROT_Start to roll','DROT_Take off time','DROT_Runway Entry','DROT_Take off speed [kts]','DROT','DROT_Runway','DROT_Demand','DROT_Final Wake','DROT_Aircraft Type ICAO','DROT_SID (shortened)','B2','TAXI_OUT_S1','TAXI_OUT_S2','TAXI_OUT_S3','TAXI_OUT_S4','TAXI_OUT_S5','TAXI_OUT_S6','TAXI_OUT_S7','TAXI_OUT_S8','TAXI_OUT_S9','TAXI_OUT_S10','TAXI_OUT_S11','TAXI_OUT_S12','TAXI_OUT_S13','TAXI_OUT_S14','TAXI_OUT_S15','B3','B4','ADA_id','ADA_ADA','ADA_Combined ROT','ADA_Buffer','ADA_Uniques','ADA_ADA counts','ADA_C_ROT counts','ADA_Buffer_Unique','ADA_Buffer_counts']
#        df_final_TAXIIN= tdf.drop(columns=columns_to_drop_taxi_in)
#        df_taxi_in = df_final_TAXIIN.rename(columns = { 'TAXI_IN_S1':'S1','TAXI_IN_S2':'S2','TAXI_IN_S3':'S3','TAXI_IN_S4':'S4','TAXI_IN_S5':'S5','TAXI_IN_S6':'S6','TAXI_IN_S7':'S7','TAXI_IN_S8':'S8','TAXI_IN_S9':'S9','TAXI_IN_S10':'S10','TAXI_IN_S11':'S11','TAXI_IN_S12':'S12','TAXI_IN_S13':'S13','TAXI_IN_S14':'S14','TAXI_IN_S15':'S15'})
#        
#        TIN_output = IntVar()
#        
#        max_tin = df_taxi_out['S1'].max() #160 # Initialise to remove unrealistic outliers from data
#        min_tin = df_taxi_out['S1'].min()
#        
#        TIN_filter_output = max_tin
#        
#        def getThrottle_ti(event):
#            
#            ax102.clear() #ax
#            
#            ax102.set_title('  Taxi-in distribution example - Runway Direction = 26L, All Mediums ')
#            ax102.set_xlabel(' Taxi-in time (secs) ')
#            ax102.set_ylabel(' Aircraft count ')
#            
#            TIN_filter_output = Throttle_tin.get()
#            
#            df_tin_filtered = df_taxi_in.loc[df_taxi_in['S1'] <= TIN_filter_output]
#            
#            df_tin_filtered['S1'].plot(kind='hist', bins=100, rwidth=0.7, ax=ax102)
#            
#            plt.show()
#        
#            canvas102.draw() #canvas
#        
#        ResultsFrame_tin = LabelFrame(f4_content, text="  [  Taxi-in Results :  ]  ", font="Helvetica 12")
#        ResultsFrame_tin.grid(row=1, columnspan=14, sticky='N', \
#                  padx=5, pady=15, ipadx=5, ipady=5)
#        
#        ttk.Label(ResultsFrame_tin, text="      Analysis of 2017 Summer Data from Airport X :  ", font="Helvetica 12").grid(column=1, row=0, sticky=N, pady=10, padx=20)
#        
#        fig_tin = plt.Figure()
#        canvas10 = FigureCanvasTkAgg(fig_tin, ResultsFrame_tin)
#        canvas10.get_tk_widget().grid(column=1, row=1, sticky=N, padx=5, pady=5)
#        ax10 = fig_tin.add_subplot(111)
#        ax10.set_title('  Taxi-in movements shown for Runway Direction = 26L ')
#        ax10.set_xlabel('Wake Category = [ Medium ] ')
#        ax10.set_ylabel('Time (secs)')
#        
#        df_taxi_in.plot(kind='box', ax=ax10) 
#        
#        ## EXTRA Taxi-out plots!
#        
#        fig_tin_2 = plt.Figure()
#        canvas102 = FigureCanvasTkAgg(fig_tin_2, ResultsFrame_tin)
#        canvas102.get_tk_widget().grid(column=2, row=1, sticky=N, padx=5, pady=5)
#        ax102 = fig_tin_2.add_subplot(111)
#        ax102.set_title('  Taxi-out distribution example - Runway Direction = 26L, All Mediums ')
#        ax102.set_xlabel(' Taxi-out time (secs) ')
#        ax102.set_ylabel(' Aircraft count ')
#        
#        #df_taxi_out.plot(kind='box', ax=ax92)
#        
#        df_taxi_in['S1'].plot(kind='hist', bins=100, rwidth=0.7, ax=ax102)
#        
#        ttk.Label(ResultsFrame_tin, text="   Filter by Max. Taxi-in value ->     ", font="Helvetica 10").grid(row=2, column=2, sticky=N, padx=5, pady=10)
#        
#        Throttle_tin = Scale(ResultsFrame_tin, from_=min_tin, to=max_tin, width=10, orient=HORIZONTAL, tickinterval=100, command=getThrottle_ti)#variable = var)
#        Throttle_tin.grid(row=3, column=2, sticky='EW', padx=5)
#        Throttle_tin.set(max_tin)
        
        # ADA/ADDA #############################################################################################
        
        #f5_help 
           
        f5_help = ttk.LabelFrame(f5, text=" Quick Help ")
        f5_help.grid(row=0, column=1, sticky=E, \
                     padx=5, pady=5, ipadx=5, ipady=5)
        ttk.Label(f5_help, text="Provides example analysis for an 'A-D-A' distribution", font=12).grid(column=1, row=1, sticky=W)
        ttk.Label(f5_help, text="Note the analysis only considers: a single Runway direction (26L)", font=12).grid(column=1, row=2, sticky=W) 
        ttk.Label(f5_help, text="  - A single Runway direction (26L)", font=12).grid(column=1, row=3, sticky=W)
        ttk.Label(f5_help, text="  - A single (on the day) wind condition", font=12).grid(column=1, row=4, sticky=W) 
        
        f5_content = LabelFrame(f5, text="   Actual Speed Profiles   ", font="Helvetica 14 bold")
        f5_content.grid(row=0, column=0, sticky=E, \
                     padx=10, pady=40, ipadx=5, ipady=10)
        ttk.Label(f5_content, text = "Not available yet.", font=12).grid(column=1, row=1, sticky=W)
        
        
        # NOTE ! : this frame used to be for ADA, but now we use it for Speed profile as the ADA comes from a normal distribution
        
#        #f5_content
#        f5_content = LabelFrame(f5, text="   ADA / ADDA   ", font="Helvetica 14 bold")
#        f5_content.grid(row=0, column=0, sticky=E, \
#                     padx=10, pady=40, ipadx=5, ipady=10)
         
#        #####################################################################        
#        #                        ADA/ADDA  DATA                             #
#        #####################################################################
#        
#        columns_to_drop_ada = ['AROT_Callsign','AROT_Threshold','AROT_RWY exit time','AROT','AROT_Runway','AROT_Demand','AROT_Final Wake','AROT_RwyExit','AROT_Aircraft Type ICAO','AROT_Threshold Speed [kts]','AROT_Speed @ TDZ [kts]','AROT_RWY Exit Speed 1','AROT_RWY Exit Speed 2','AROT_RWY Exit Speed 3','B1','DROT_Callsign','DROT_Line up time','DROT_Start to roll','DROT_Take off time','DROT_Runway Entry','DROT_Take off speed [kts]','DROT','DROT_Runway','DROT_Demand','DROT_Final Wake','DROT_Aircraft Type ICAO','DROT_SID (shortened)','B2','TAXI_OUT_S1','TAXI_OUT_S2','TAXI_OUT_S3','TAXI_OUT_S4','TAXI_OUT_S5','TAXI_OUT_S6','TAXI_OUT_S7','TAXI_OUT_S8','TAXI_OUT_S9','TAXI_OUT_S10','TAXI_OUT_S11','TAXI_OUT_S12','TAXI_OUT_S13','TAXI_OUT_S14','TAXI_OUT_S15','B3','TAXI_IN_S1','TAXI_IN_S2','TAXI_IN_S3','TAXI_IN_S4','TAXI_IN_S5','TAXI_IN_S6','TAXI_IN_S7','TAXI_IN_S8','TAXI_IN_S9','TAXI_IN_S10','TAXI_IN_S11','TAXI_IN_S12','TAXI_IN_S13','TAXI_IN_S14','TAXI_IN_S15','B4']
#        df_ada = tdf.drop(columns=columns_to_drop_ada)
#
#        ResultsFrame_ada = LabelFrame(f5_content, text="  [  ADA Results :  ]   ", font="Helvetica 12")
#        ResultsFrame_ada.grid(row=1, columnspan=14, sticky='N', \
#                  padx=5, pady=15, ipadx=5, ipady=5)
#        
#        ttk.Label(ResultsFrame_ada, text="      Analysis of 2017 Summer Data from Airport X :  ", font="Helvetica 12").grid(column=1, row=0, sticky=N, pady=10, padx=20)
#        
#        fig_ada = plt.Figure()
#        canvas7 = FigureCanvasTkAgg(fig_ada, ResultsFrame_ada)
#        canvas7.get_tk_widget().grid(column=1, row=1, sticky=N, padx=5, pady=5)
#        ax7 = fig_ada.add_subplot(111)
#        ax7.set_title('  ADA data shown for Runway Direction [ 26L ] ')
#        ax7.set_xlabel('Time (secs)')
#        ax7.set_ylabel('Frequency')
#        
#        df_ada_main_plot = df_ada
#        
#        df_ada_main_plot = df_ada_main_plot.drop(columns=['ADA_id','ADA_ADA','ADA_Combined ROT', 'ADA_Buffer', 'ADA_Buffer_Unique', 'ADA_Buffer_counts'])
#        df_ada_main_plot = df_ada_main_plot.rename(columns = {'ADA_ADA counts':'ADA_ADA'})
#        df_ada_main_plot = df_ada_main_plot.rename(columns = {'ADA_C_ROT counts':'ADA_Combined ROT'})
#        df_ada_main_plot = df_ada_main_plot.rename(columns = {'ADA_Uniques':'ADA_Time (secs)'})
#        df_ada_main_plot = df_ada_main_plot.set_index('ADA_Time (secs)')
#        
#        df_ada_main_plot.plot(kind='line', ax=ax7)
#        
#        fig_buffer = plt.Figure()
#        canvas8 = FigureCanvasTkAgg(fig_buffer, ResultsFrame_ada)
#        canvas8.get_tk_widget().grid(column=2, row=1, sticky=N, padx=5, pady=5)
#        ax8 = fig_buffer.add_subplot(111)
#        ax8.set_title('  Resulting Buffer before next Arrival aircraft  ')
#        ax8.set_xlabel('Buffer Spacing (secs)')
#        ax8.set_ylabel('Frequency')
#        
#        df_ada_sub_plot = df_ada
#        
#        df_ada_sub_plot = df_ada_sub_plot.drop(columns=['ADA_id','ADA_ADA','ADA_Combined ROT', 'ADA_Buffer', 'ADA_Uniques', 'ADA_ADA counts', 'ADA_C_ROT counts'])
#        df_ada_sub_plot = df_ada_sub_plot.rename(columns = {'ADA_Buffer_counts':'ADA_Buffer'})
#        df_ada_sub_plot = df_ada_sub_plot.rename(columns = {'ADA_Buffer_Unique':'ADA_Time (secs)'})
#        df_ada_sub_plot = df_ada_sub_plot.set_index('ADA_Time (secs)')
#        
#        df_ada_sub_plot.plot(kind='line', ax=ax8)
#        
#        df_buffer_limit =  df_ada_sub_plot.loc[df_ada_sub_plot.index <= 15, 'ADA_Buffer']
#        df_buffer_limit = df_buffer_limit.reset_index(level=[0])
#        
#        ############ ADA data filtering! ############
#        
#        
#        ######### Interesting Buffer stats! #########
#        
#        Buffer_Analysis = LabelFrame(ResultsFrame_ada, text="  [  Buffer Analysis :  ]   ", font="Helvetica 12")
#        Buffer_Analysis.grid(row=2, column=2, sticky='N', \
#                  padx=5, pady=15, ipadx=5, ipady=5)
#        
#        ttk.Label(Buffer_Analysis, text="   Average buffer value from selection :     ", font="Helvetica 10").grid(row=1, column=1, sticky=N, padx=5, pady=10)
#        ## Update label value ##
#        ttk.Label(Buffer_Analysis, text=str(round(df_ada['ADA_Buffer'].mean(),2))).grid(column=1, row=2, sticky=N) # Mean Buffer value
#        
#        
#        #Option 2
#        ttk.Label(Buffer_Analysis, text="   Count of instances with buffer < 15 seconds :     ", font="Helvetica 10").grid(row=1, column=2, sticky=N, padx=5, pady=10)
#        ttk.Label(Buffer_Analysis, text=str(df_buffer_limit['ADA_Buffer'].count())).grid(column=2, row=2, sticky=N) # Equivalent to numpy.percentile
        
        output_extension_empty_file = time.strftime("%H_%M", time.localtime(time.time()))
        input_files_name = 'Input_File_RAPID_v3.0_' + output_extension_empty_file
        
        def define_final_distribution_parameters(): 
            
            ################################################################################################
            df_final_AROT = define_final_AROT()
            print("AROTs defined FINAL | Filters AROT = ", Throttle_arot.get(), "Demand=", Throttle.get())
            
            df_final_DROT = define_final_DROT()
            print("DROTs defined FINAL | Filters DROT = ", Throttle_drot.get(), "Demand=", Throttle_d.get())
            
            
            ################################################################################################
            
            #final_distribution_labels=[df_final_AROT,df_final_DROT,df_final_TAXIOUT, df_final_TAXIIN]
            final_distribution_labels=[df_final_AROT,df_final_DROT]
            df_final_distribution = pd.concat(final_distribution_labels, axis=1)
            
            distribution_file_name = 'utility/AROTDROT_distributions.csv'
            print(' !#!#!#!# CSV FILE NAME= ',distribution_file_name)
            df_final_distribution.to_csv(distribution_file_name)
            raise_subframe(f6)
        #f5_buttons
        f5_buttons= ttk.Frame(f5) 
        f5_buttons.grid(column = 0, row=1, columnspan = 7, sticky=(N, W, E, S))
        
        tk.Button(f5_buttons, text='NEXT ->', command=define_final_distribution_parameters, activebackground = "pink", font=16, height = 1, overrelief="raised", width = 15).pack(side="right")
        tk.Button(f5_buttons, text='<- BACK', command=lambda:raise_subframe(f2), activebackground = "pink", font=16, height = 1, overrelief="raised", width = 15).pack(side="right")
        
        # STRATEGY #############################################################################################
        
        #f6_help 
           
        f6_help = ttk.LabelFrame(f6, text=" Quick Help ")
        f6_help.grid(row=0, column=1, sticky=E, \
                     padx=5, pady=5, ipadx=5, ipady=5)
        ttk.Label(f6_help, text="Strategy tool assigns 'ADA, ADDA, or none' based on % of Scheduled Arrivals", font=12).grid(column=1, row=1, sticky=W)
        ttk.Label(f6_help, text="Note - the Operational Analysis saves the key distributions to the INPUT file", font=12).grid(column=1, row=2, sticky=W)  
        
        #f6_buttons
        f6_buttons= ttk.Frame(f6) 
        f6_buttons.grid(column = 0, row=1, columnspan = 7, sticky=(N, W, E, S))
        
        tk.Button(f6_buttons, text='TO CORE MODULE ->', command=lambda:raise_main_frame(core_module), activebackground = "pink", font=16, height = 1, overrelief="raised", width = 25).pack(side="right")
        tk.Button(f6_buttons, text='<- BACK', command=lambda:raise_subframe(f5), activebackground = "pink", font=16, height = 1, overrelief="raised", width = 15).pack(side="right")
        
        #f6_content
        f6_content = LabelFrame(f6, text="    RAPID - INPUT FILE GENERATION    ", font="Helvetica 14 bold")
        f6_content.grid(row=0, column=0, sticky=E, \
                     padx=10, pady=40, ipadx=5, ipady=10)
        
        
        #####################################################################        
        #                         SCHEDULE  DATA                            #
        #####################################################################
        
        #import_excel_schedule = ""
        import_excel_schedule = StringVar()
        
        ##########################################################################################################################################
        def actual_strat_tool():
            input_excel_sheet = import_excel_schedule.get()
            xl = pd.ExcelFile(input_excel_sheet)
            #xl = pd.ExcelFile('inputs/Input_File_RAPID_2.05_CORE_Gatwick_SINGLE.xlsx')
            df_arr = xl.parse("Arrivals")
            df_dep = xl.parse("Departures")
            
            #                       ARRIVAL Intervals                           #
            # Find min max arrival limits  # Only care about Arrivals - If no arrivals set strategy to 'none'
            
            #replace nan with 0:
            df_arr['SIBT'] = df_arr['SIBT'].fillna(0)
            df_dep['SOBT'] = df_dep['SOBT'].fillna(0)
           
            max_arr_SIBT = df_arr['SIBT'].max()
            min_arr_SIBT = df_arr['SIBT'].min()
            #df_arr['SIBT'] = df_arr['SIBT'].fillna(0)
            if df_arr.at[0,'SIBT']==0: # seconds
                df_arr['SIBT'] = pd.to_datetime(df_arr['SIBT(sec)'],unit='s')
                max_arr_SIBT = df_arr['SIBT'].max()
                min_arr_SIBT = df_arr['SIBT'].min()
            
            if df_dep.at[0,'SOBT']==0: # seconds
                df_dep['SOBT'] = pd.to_datetime(df_dep['SOBTsec)'],unit='s')
          
            
            # Generate list of all schedule times (15min intervals)
            SIBT_intervals = []
            
            min_arr_SIBT = str(min_arr_SIBT)
            h1, m1, s1 = min_arr_SIBT.split(':')
            min_interval = (math.floor((int(m1)/60)*(60/15)))*15
            min_arr_SIBT = datetime(2000, 1, 1, int(h1), int(min_interval), int(0))
            #min_arr_SIBT = min_arr_SIBT.datetime.time
            
            max_arr_SIBT = str(max_arr_SIBT) 
            h2, m2, s2 = max_arr_SIBT.split(':')
            max_interval = (math.floor((int(m2)/60)*(60/15)))*15
            max_arr_SIBT = datetime(2000, 1, 1, int(h2), int(max_interval), int(0))
            #max_arr_SIBT = max_arr_SIBT.datetime.time
            
            temp = min_arr_SIBT
            while temp <= max_arr_SIBT:
                SIBT_intervals.append(temp)
                temp += timedelta(minutes=15)
            
            sch_list = pd.DataFrame(SIBT_intervals, columns = ["SIBT"])
            sch_list['SIBT'] = sch_list['SIBT'].dt.time
            
            #                            ARRIVAL SORT                            #
            
            df_arr['SIBT'] = df_arr['SIBT'].astype(str) # Buggy data - have to convert to str then datetime ?!?
            df_arr['SIBT'] = pd.DatetimeIndex(df_arr['SIBT']) # Converts to dtype <M8[ns]
            df_arr['rounded_SIBT'] = df_arr['SIBT'].dt.round('15min')
            df_arr['rounded_SIBT'] = df_arr['rounded_SIBT'].dt.time
            
            df_arr_count = df_arr.groupby(['rounded_SIBT'])['ID'].count()  
            df_arr_count = df_arr_count.reset_index(level=[0])
            df_arr_count = df_arr_count.rename(columns = {'ID':'Arrival_counts', 'rounded_SIBT':'SIBT'})
            
            df_arr_final = pd.merge(sch_list, df_arr_count, how='left', on=['SIBT'], copy=True).fillna(0)
            
            #                           DEPARTURE SORT                            #
            
            df_dep['SOBT'] = df_dep['SOBT'].astype(str) # Buggy data - have to convert to str then datetime ?!?
            df_dep['SOBT'] = pd.DatetimeIndex(df_dep['SOBT']) # Converts to dtype <M8[ns]
            df_dep['rounded_SOBT'] = df_dep['SOBT'].dt.round('15min')
            df_dep['rounded_SOBT'] = df_dep['rounded_SOBT'].dt.time
            
            df_dep_count = df_dep.groupby(['rounded_SOBT'])['ID'].count()  
            df_dep_count = df_dep_count.reset_index(level=[0])
            df_dep_count = df_dep_count.rename(columns = {'ID':'Departure_counts', 'rounded_SOBT':'SIBT'})
            
            #                           ASSIGN STRATEGY                            #
            
            df_schedule_counts = pd.merge(df_arr_final, df_dep_count, how='left', on=['SIBT'], copy=True).fillna(0)
            
            df_schedule_counts['Perc_Arrs'] = round(df_schedule_counts['Arrival_counts'] / (df_schedule_counts['Departure_counts'] + df_schedule_counts['Arrival_counts']), 2) * 100
            
            df_schedule_counts['Strategy'] = ""
            
            df_schedule_counts.loc[(df_schedule_counts['Perc_Arrs'] > 40),'Strategy'] = 'ADA'
            df_schedule_counts.loc[(df_schedule_counts['Perc_Arrs'] <= 40) & (df_schedule_counts['Perc_Arrs'] >= 30),'Strategy'] = 'ADDA'
            df_schedule_counts.loc[(df_schedule_counts['Perc_Arrs'] < 30),'Strategy'] = 'none'
            
            # No assumption here about when aircraft are expected to be on runway
            # I.e. Arrival will land approx 10 mins before SIBT? Departure uses runway 10mins after SOBT? This does not account for delay!!!!
            
            #####################################################################
            #                      1st Pass Complete                            #
            #####################################################################
            
            df_schedule_counts['Next_Strategy'] = df_schedule_counts['Strategy'].shift(-1)
            df_schedule_counts['Prev_Strategy'] = df_schedule_counts['Strategy'].shift(+1)
            df_schedule_counts['P2_Strategy'] = df_schedule_counts['Strategy'].shift(+2)
            
            #                             METHOD ONE                            #
            
            df_schedule_counts['Strategy_v2'] = df_schedule_counts['Strategy']
            
            df_schedule_counts.loc[(df_schedule_counts['Strategy'].astype(str) == str('none')) & (df_schedule_counts['Prev_Strategy'] != df_schedule_counts['Strategy']) & (df_schedule_counts['Next_Strategy'] != df_schedule_counts['Strategy']),'Strategy_v2'] = 'ADDA'
            
            df_schedule_counts.loc[(df_schedule_counts['Prev_Strategy'] ==  df_schedule_counts['Next_Strategy']) & (df_schedule_counts['Prev_Strategy'] ==  df_schedule_counts['P2_Strategy']),'Strategy_v2'] = df_schedule_counts['Prev_Strategy']
            
            df_schedule_counts['Final_Strategy'] = df_schedule_counts['Strategy_v2']
            
            df_schedule_counts = df_schedule_counts.drop(columns=['Next_Strategy', 'Prev_Strategy', 'P2_Strategy'])
            
            df_schedule_counts['Next_Strategy'] = df_schedule_counts['Strategy_v2'].shift(-1)
            df_schedule_counts['Prev_Strategy'] = df_schedule_counts['Strategy_v2'].shift(+1)
            df_schedule_counts['P2_Strategy'] = df_schedule_counts['Strategy_v2'].shift(+2)
            
            df_schedule_counts.loc[(df_schedule_counts['Prev_Strategy'] == df_schedule_counts['Next_Strategy']) & (df_schedule_counts['Prev_Strategy'] ==  df_schedule_counts['P2_Strategy']),'Final_Strategy'] = df_schedule_counts['Prev_Strategy']
            
            df_schedule_counts = df_schedule_counts.drop(columns=['Next_Strategy', 'Prev_Strategy', 'P2_Strategy'])
            # Extra drops for final result
            df_schedule_counts = df_schedule_counts.drop(columns=['Arrival_counts', 'Departure_counts', 'Perc_Arrs', 'Strategy', 'Strategy_v2'])
            
            df_schedule_counts = df_schedule_counts.rename(columns = {'SIBT':'rounded_SIBT'})
            
            #####################################################################
            #                     2nd Pass Complete                             #
            #####################################################################
        
            #                           FINAL MERGE                             # 
            df_temp = df_arr
        
            df_temp = pd.merge(df_temp, df_schedule_counts, how='left', on=['rounded_SIBT'], copy=True)
            
            #####################################################################
            #                        WRITE OUTPUT                               #
            #####################################################################
            #output_ex = time.strftime("%H_%M", time.localtime(time.time()))       
            master_column = df_temp['Final_Strategy']
            
            output_file_to_edit = openpyxl.load_workbook(input_excel_sheet)
            arrival_sheet = output_file_to_edit.get_sheet_by_name('Arrivals')
            arrival_sheet['U' + str(1)].value = 'Master Column'
            for i in range (0, len(master_column)):
                arrival_sheet['U' + str(i+2)].value = master_column[i]
            
            name_input_file = input_files_name + '.xlsx'
            output_file_to_edit.save(name_input_file)
            print(' !#!#!#!# XLSX FILE NAME= ', name_input_file)
            name_excel_sheet.set(name_input_file)
            
            
            return()
        ##########################################################################################################################################
            
        #input_excel_sheet = 'C:/Users/Think - Joe Irwin/Dropbox/Think Users/JoeIrwin/RAPID/RAPID_DEMO/AROT_example_input.xlsx'
        
        def sch_load_file():
            import_schedule = filedialog.askopenfilename()
            import_excel_schedule.set(import_schedule)
            ttk.Label(SchImportFrame, text="File Successfully Imported!").grid(column=1, row=3, sticky=N, pady=10)
            #print("#####", import_excel_schedule, "###", import_schedule)
            return()
            
        def define_gen_parameters():
            button_check.set(True)
            ttk.Label(SchGenFrame, text="Generation Successful!").grid(column=1, row=4, sticky=N, pady=10)
            import_schedule = import_excel_schedule.get()
        #name_input_file = ""    
        def assign_strat_tool():
            button_check.set(True)
            actual_strat_tool()
            
            ttk.Label(StratInputFrame, text="Strategies Successfully Applied!").grid(column=1, row=3, sticky=N, pady=10)
        
        
        # Specify GUI Structure -------->
            
        # Left Side -------->
        IntroFrameLeft = LabelFrame(f6_content, text="   [ STEP 1  -  SCHEDULE INPUT ]   ", font="Helvetica 12")
        IntroFrameLeft.grid(row=0, column=1, sticky='N', \
                      padx=5, pady=40, ipadx=5, ipady=5)
        
        # Right Side -------->
        IntroFrameRight = LabelFrame(f6_content, text="   [ STEP 2  -  ASSIGN STRATEGY ]   ", font="Helvetica 12")
        IntroFrameRight.grid(row=0, column=2, sticky='N', \
                      padx=5, pady=40, ipadx=5, ipady=5)
        
        SchImportFrame = LabelFrame(IntroFrameLeft, text="   [ A ] - Import a Flight Schedule   ", font="Helvetica 12 bold")
        SchImportFrame.grid(row=1, column=1, columnspan=7, sticky='N', \
                  padx=40, pady=0, ipadx=5, ipady=5)
        
        ttk.Label(IntroFrameLeft, text="Select one of the following Options : ", font="Helvetica 12 italic").grid(column=3, row=0, sticky=N, pady=20, padx=40)
        
        ttk.Label(IntroFrameRight, text="    Mixed-mode Runway Only :", font="Helvetica 12 italic").grid(column=3, row=0, sticky=N, pady=20, padx=40)
        
        ttk.Label(SchImportFrame, text="                    Note - Ensure that the dataset includes SIBT/SOBT times             ", font="Helvetica 9 italic").grid(column=1, row=1, sticky=N, pady=10, padx=20, ipadx=45)
        ttk.Button(SchImportFrame, text="Import a Flight Schedule ->", command=sch_load_file).grid(column=1, row=2, sticky=N, padx=10, pady=10, ipadx=5, ipady=5)
        
        SchGenFrame = LabelFrame(IntroFrameLeft, text="   OR   [ B ] - Generate a Flight Schedule   ", font="Helvetica 12 bold")
        SchGenFrame.grid(row=2, column=1, columnspan=7, sticky='N', \
                  padx=40, pady=15, ipadx=5, ipady=5)
        
        ttk.Label(SchGenFrame, text="   ( Generated Schedule will use Wake/SID/Stand Group proportions based on operational data )   ", font="Helvetica 9 italic").grid(column=1, row=1, sticky=N, pady=10)
        
        General = LabelFrame(SchGenFrame, text=" General Settings ")
        General.grid(column=1, row=2, rowspan=1, sticky='N', \
                  padx=10, pady=10, ipadx=40, ipady=10)
        
        perc_arrivals = IntVar(General, value='20')
        total_aircraft = IntVar(General, value='60')
        
        ttk.Label(General, text=" Set the peak number of hourly Aircraft ").grid(column=1, row=1, sticky=W, padx=20, pady=5)
        Gen_entry1 = ttk.Entry(General, width=7, textvariable=total_aircraft)
        Gen_entry1.grid(column=2, row=1, sticky=(W, E))
        
        ttk.Label(General, text=" Set Arrival Percentage in Schedule (%) ").grid(column=1, row=2, sticky=W, padx=20, pady=5)
        Gen_entry2 = ttk.Entry(General, width=7, textvariable=perc_arrivals)
        Gen_entry2.grid(column=2, row=2, sticky=(W, E))
        
        
        ttk.Button(SchGenFrame, text=" Generate and Save  ", command=define_gen_parameters).grid(column=1, row=4, sticky=N, ipadx=5, ipady=5)
        
        SchGenFrame.bind('<Return>', define_gen_parameters)
        
        #Strategy Assessment:
        
        StratInputFrame = LabelFrame(IntroFrameRight, text="   Generate a Spacing Strategy ->  ", font="Helvetica 12 bold")
        StratInputFrame.grid(row=1, column=1, columnspan=7, sticky='N', \
                  padx=10, pady=5, ipadx=5, ipady=5)
        
        ttk.Label(StratInputFrame, text="      Each Scheduled Aircraft is assigned a Strategy (15min intervals) :", font="Helvetica 9 italic").grid(column=1, row=0, sticky=N, pady=10, padx=40)
        
        ttk.Button(StratInputFrame, text=" Assign Strategy  ", command=assign_strat_tool).grid(column=1, row=2, sticky=N, ipadx=5, ipady=5)
        
        raise_frame(f1)
       
    tk.Button(f0_buttons, text='NEXT ->', command=lambda:raise_subframe_f0(f1), activebackground = "pink", font=16, height = 1, overrelief="raised", width = 15).pack(side="right")

import1 = ttk.Button(f0_content_a, text="Select Operational Data", command=generate_new_input).grid(column=1, row=0, sticky=N, padx=10, pady=10, ipadx=5, ipady=5)
f0.bind('<Return>', create_dataframe_operational_data)

f0_content_b = ttk.LabelFrame(f0_content, text=" [ B ] - Load existing INPUT file ")
f0_content_b.grid(row=2, column=1, sticky=E, \
             padx=5, pady=15, ipadx=5, ipady=5)
def load_input_file():
    input_file_excel = filedialog.askopenfilename()
    name_excel_sheet.set(input_file_excel)  
    
    ttk.Label(f0_content_b, text="File Successfully Loaded!").grid(column=1, row=3, sticky=N, pady=10)
    f0_buttons= tk.Frame(f0) 
    f0_buttons.grid(column = 0, row=1, columnspan = 7, sticky=(N, W, E, S))
    tk.Button(f0_buttons, text='NEXT ->', command=lambda:raise_main_frame(core_module), activebackground = "pink", font=16, height = 1, overrelief="raised", width = 15).pack(side="right")
import2 = ttk.Button(f0_content_b, text="Select INPUT File", command=load_input_file).grid(column=1, row=0, sticky=N, padx=10, pady=10, ipadx=5, ipady=5)



#-------------------------------CORE GUI--------------------------------------#



# buttons frame
buttons_core = ttk.Frame(core_module)
buttons_core.grid(row=2, sticky=(N, W, E, S))

#tabControl.add(tab2, text='CORE MODULE')      # Add the tab
#tabControl.pack(expand=1, fill="both")  # Pack to make visible

helpCoreFrame = ttk.LabelFrame(tab2, text=" Quick Help ")
helpCoreFrame.grid(row=2, column=10, sticky=E, \
             padx=5, pady=5, ipadx=5, ipady=5)
ttk.Label(helpCoreFrame, text="(1) Please use the following format: ", font=10).grid(column=10, row=2, sticky=W)
ttk.Label(helpCoreFrame, text="     (SID1,SID2)(SID3,SID4)...", font=10).grid(column=10, row=3, sticky=W)
ttk.Label(helpCoreFrame, text="(2) Please use the following format: ", font=10).grid(column=10, row=4, sticky=W)
ttk.Label(helpCoreFrame, text="      SID queue 1, SID queue 1... | SID queue 2 , ... | ...", font=10).grid(column=10, row=5, sticky=W)
ttk.Label(helpCoreFrame, text="(3) Chose how many times do you want to run  ", font=10).grid(column=10, row=6, sticky=W)
ttk.Label(helpCoreFrame, text="      the program, thus how many outputs ", font=10).grid(column=10, row=7, sticky=W)
ttk.Label(helpCoreFrame, text="      do you want to get.", font=10).grid(column=10, row=8, sticky=W)
ttk.Label(helpCoreFrame, text="(4) To increase the degree of confidence in  ", font=10).grid(column=10, row=9, sticky=W)
ttk.Label(helpCoreFrame, text="      the results, the model will average the values", font=10).grid(column=10, row=10, sticky=W)
ttk.Label(helpCoreFrame, text="      of the total throughputs and will stop when", font=10).grid(column=10, row=11, sticky=W)
ttk.Label(helpCoreFrame, text="      the variation will be less de 2 aircrafts per hour ", font=10).grid(column=10, row=12, sticky=W)
ttk.Label(helpCoreFrame, text="(5) Useful only for TIME-based separation", font=10).grid(column=10, row=13, sticky=W)
ttk.Label(helpCoreFrame, text="(6) By defaylt everything is delivered to Threshold", font=10).grid(column=10, row=14, sticky=W)


stepTwoFirstFrame = ttk.LabelFrame(tab2, text=" Mandatory Fields ")
stepTwoFirstFrame.grid(row=2, column=2, columnspan=5, sticky='W', \
             padx=5, pady=5, ipadx=5, ipady=5)
stepTwoSecondFrame = ttk.LabelFrame(tab2, text=" Enablers (Optional) ")
stepTwoSecondFrame.grid(row=2, column=8, columnspan=2, sticky='W', \
             padx=5, pady=5, ipadx=5, ipady=5)
stepTwoThirdFrame = ttk.LabelFrame(tab2)
stepTwoThirdFrame.grid(row=11, column=4, columnspan=7, sticky='W', \
             padx=5, pady=5, ipadx=5, ipady=5)

stepTwoFourthFrame = ttk.LabelFrame(stepTwoSecondFrame, text=" Arrivals Separation ")
stepTwoFourthFrame.grid(row=8, column=1, columnspan=7, sticky='W', \
             padx=5, pady=5, ipadx=5, ipady=5)

n_input = IntVar(win, value='50')
ADA_x_input = IntVar(win, value = '10')
minDep_altSID_input = IntVar(win, value='60')
minDep_sameSID_input = IntVar(win, value='109')
SIDmax_input = IntVar(win, value = '4')
SIDgroup_separation_input = StringVar(win, value='(2,4)(3,4)')
SID_queue_assign_input = StringVar(win, value = '1 3 | 2 4' )
n_times_input= IntVar(win, value='1')



button_check = StringVar(mainframe, value='0')
n_output = IntVar()
minDep_altSID_output = IntVar()
minDep_sameSID_output = IntVar()
SIDmax_output = IntVar()
SIDgroup_separation_output = StringVar()
SID_queue_assign_output = StringVar()
VTT_output = IntVar()
debug_output = IntVar()
Tower_sep_output = IntVar()
TBS_output = IntVar()
debug3_output = IntVar()
RECat_output = IntVar()
RECAT_PWS_output = IntVar()
queue1_output = IntVar()
queue2_output = IntVar()
queue3_output = IntVar()
queue4_output = IntVar()
n_times_output = IntVar()
averagethr_output = IntVar()
distance_based_output = IntVar()
time_based_output = IntVar()
ADA_x_output = IntVar()

MRS4dme_output= IntVar()
WAKE4dme_output= IntVar()
ADA4dme_output= IntVar()
ADDA4dme_output= IntVar()
MRSthr_output= IntVar()
WAKEthr_output= IntVar()
ADAthr_output= IntVar()
ADDAthr_output= IntVar()

name_excel_sheet = StringVar()

m_input = IntVar(win, value='0')
Throughput_check_output = IntVar()
Delay_check_output = IntVar()
Seq_check_output = IntVar()
op_yes_output = IntVar()
new_set_output = IntVar()
m_output = IntVar()
arr_delay_output = IntVar()
convergence_output = IntVar()


### Frame 1

input_entry1 = tk.Entry(stepTwoFirstFrame, width=7, textvariable=n_input,font=16)
input_entry1.grid(column=2, row=1, sticky=(W, E))
ttk.Label(stepTwoFirstFrame, text="[Arrival lead time] 'n' value (in secs) =", font=16).grid(column=1, row=1, sticky=W)

input_entry2 = tk.Entry(stepTwoFirstFrame, width=7, textvariable=minDep_altSID_input,font=16)
input_entry2.grid(column=2, row=2, sticky=(W, E))
ttk.Label(stepTwoFirstFrame, text="Minimum Separation (secs) alternating SIDs =", font=16).grid(column=1, row=2, sticky=W)

input_entry3 = tk.Entry(stepTwoFirstFrame, width=7, textvariable=minDep_sameSID_input,font=16)
input_entry3.grid(column=2, row=3, sticky=(W, E))
ttk.Label(stepTwoFirstFrame, text="Minimum Separation (secs) same SIDs =", font=16).grid(column=1, row=3, sticky=W)

input_entry4 = tk.Entry(stepTwoFirstFrame, width=7, textvariable=SIDmax_input,font=16)
input_entry4.grid(column=2, row=4, sticky=(W, E))
ttk.Label(stepTwoFirstFrame, text="Maximum number of SID groups", font=16).grid(column=1, row=4, sticky=W)

input_entry5 = tk.Entry(stepTwoFirstFrame, width=14,  textvariable=SIDgroup_separation_input,font=16)
input_entry5.grid(column=2, row=5, sticky=(W, E))
ttk.Label(stepTwoFirstFrame, text="Enter the pairs of SID groups that require minimum separation. (1)", font=16).grid(column=1, row=5, sticky=W)

ttk.Label(stepTwoFirstFrame, text="Select the type of queue", font=16).grid(column=1, row=6, sticky=W)
queue1 = IntVar()
tk.Checkbutton(stepTwoFirstFrame, text="1x8", variable=queue1, font=16).grid(column=2, row=6, sticky=W)
queue2 = IntVar()
tk.Checkbutton(stepTwoFirstFrame, text="2x4", variable=queue2, font=16).grid(column=3, row=6, sticky=W)
queue3 = IntVar()
tk.Checkbutton(stepTwoFirstFrame, text="4x2", variable=queue3, font=16).grid(column=4, row=6, sticky=W)
queue4 = IntVar()
tk.Checkbutton(stepTwoFirstFrame, text="8x1", variable=queue4, font=16).grid(column=5, row=6, sticky=W)

input_entry6 = tk.Entry(stepTwoFirstFrame, width=14,  textvariable=SID_queue_assign_input,font=16)
input_entry6.grid(column=2, row=7, sticky=(W, E))
ttk.Label(stepTwoFirstFrame, text="Assign SID groups to each RWY queue. (2)", font=16).grid(column=1, row=7, sticky=W) 


### Frame 2



var6 = IntVar()
tk.Checkbutton(stepTwoSecondFrame, text="RECAT", variable=var6, font=16).grid(column=1, row=1, sticky=W)

var17 = IntVar()
tk.Checkbutton(stepTwoSecondFrame, text="RECAT-PWS", variable=var17, font=16).grid(column=1, row=2, sticky=W)


var2 = IntVar()
tk.Checkbutton(stepTwoSecondFrame, text="Debug", variable=var2,font=16).grid(column=1, row=3, sticky=W)


var15 = IntVar()
tk.Checkbutton(stepTwoFourthFrame, text="DISTANCE-based Arrivals separation (5) ", variable=var15,font=16).grid(column=1, row=1, sticky=W)

var16 = IntVar()
tk.Checkbutton(stepTwoFourthFrame, text="TIME-based Arrivals separation (6) ", variable=var16,font=16).grid(column=1, row=2, sticky=W)

input_entry15 = tk.Entry(stepTwoFourthFrame, width=7, textvariable=ADA_x_input,font=16)
input_entry15.grid(column=2, row=3, sticky=(W, E))
tk.Label(stepTwoFourthFrame, text="ADA target time X-value (5) = ",font=16).grid(column=1, row=3, sticky=W)


delievery = ttk.LabelFrame(stepTwoFourthFrame, text="Separation Delievery (6)")
delievery.grid(row=4, column=1, columnspan=2, sticky='W', \
             padx=5, pady=5, ipadx=5, ipady=5)

tk.Label(delievery, text="4dme : ",font=16).grid(column=1, row=1, sticky=W)

MRS_4dme = IntVar()
tk.Checkbutton(delievery, text="MRS", variable=MRS_4dme,font=16).grid(column=2, row=1, sticky=W)

WAKE_4dme = IntVar()
tk.Checkbutton(delievery, text="WAKE", variable=WAKE_4dme,font=16).grid(column=3, row=1, sticky=W)

ADA_4dme = IntVar()
tk.Checkbutton(delievery, text="ADA", variable=ADA_4dme,font=16).grid(column=4, row=1, sticky=W)

ADDA_4dme = IntVar()
tk.Checkbutton(delievery, text="ADDA", variable=ADDA_4dme,font=16).grid(column=5, row=1, sticky=W)


tk.Label(delievery, text="Threshold : ",font=16).grid(column=1, row=2, sticky=W)

MRS_thr = IntVar()
tk.Checkbutton(delievery, text="MRS", variable=MRS_thr,font=16).grid(column=2, row=2, sticky=W)

WAKE_thr = IntVar()
tk.Checkbutton(delievery, text="WAKE", variable=WAKE_thr,font=16).grid(column=3, row=2, sticky=W)

ADA_thr = IntVar()
tk.Checkbutton(delievery, text="ADA", variable=ADA_thr,font=16).grid(column=4, row=2, sticky=W)

ADDA_thr = IntVar()
tk.Checkbutton(delievery, text="ADDA", variable=ADDA_thr,font=16).grid(column=5, row=2, sticky=W)


### Frame 3

input_entry7 = tk.Entry(stepTwoThirdFrame, width=7, textvariable=n_times_input,font=16)
input_entry7.grid(column=2, row=1, sticky=(W, E))
ttk.Label(stepTwoThirdFrame, text="Number of runs (3) = ", font=16).grid(column=1, row=1, sticky=W)

var7 = IntVar()
tk.Checkbutton(stepTwoThirdFrame, text="I want to feel confident! (4) ", variable=var7, font=16).grid(column=1, row=2, sticky=W) #average
var14 = IntVar()
tk.Checkbutton(stepTwoThirdFrame, text = 'Print a debug tab', variable=var14, font=16).grid(column=1, row=3, sticky=W)


# Buttons


buttons_frame_core = ttk.Frame(buttons_core)
buttons_frame_core.pack(side="right")
tk.Button(buttons_frame_core, text="VISUALIZATION MODULE ->", command=lambda:raise_main_frame(visual_module), activebackground = "pink", font=16, height = 1, overrelief="raised", width = 25).pack(side="top")
tk.Button(buttons_frame_core, text="Just run the model", command=define_input_parameters, activebackground = "pink", font=16, height = 1, overrelief="raised", width = 25).pack(side="bottom")


#------------------------------VISUAL GUI-------------------------------------#


button_check = StringVar(win, value='0')
m_input = IntVar(win, value='0')
convergence_output = IntVar()
Throughput_check_output = IntVar()
Delay_check_output = IntVar()
Seq_check_output = IntVar()
ADA_buffer_output = IntVar()
op_yes_output = IntVar()
new_set_output = IntVar()
m_output = IntVar()
arr_delay_output = IntVar()
name_excel_sheet = StringVar()


helpVisFrame = ttk.LabelFrame(tab3, text=" Quick Help ")
helpVisFrame.grid(row=1, column=8, sticky=E, \
             padx=5, pady=5, ipadx=5, ipady=5)
ttk.Label(helpVisFrame, text="(1) If this box is checked, another window will pop-up after pressing 'Visualize results'.", font=12).grid(column=8, row=1, sticky=W)
ttk.Label(helpVisFrame, text="      Please make sure that the operational data are in the right format.", font=12).grid(column=8, row=2, sticky=W)

ttk.Label(helpVisFrame, text="(2) If this box is checked, another window will pop-up after pressing 'Visualize results'. ", font=12).grid(column=8, row=3, sticky=W)
ttk.Label(helpVisFrame, text="      Please make sure that the new data are in the same format as the outputs of the model.", font=12).grid(column=8, row=4, sticky=W)



stepThree = ttk.LabelFrame(tab3)
stepThree.grid(row=1, columnspan=7, sticky='N', \
               padx=5, pady=5, ipadx=5, ipady=5)

var0 = IntVar()
tk.Checkbutton(stepThree, text="Convergence", variable=var0, font=16).grid(column=1, row=1, sticky=W)
var8 = IntVar()
tk.Checkbutton(stepThree, text="Throughput", variable=var8, font=16).grid(column=1, row=2, sticky=W)
var9 = IntVar()
tk.Checkbutton(stepThree, text="Departures Delay", variable=var9, font=16).grid(column=1, row=3, sticky=W)
var13 = IntVar()
tk.Checkbutton(stepThree, text="Arrivals Delay", variable=var13, font=16).grid(column=1, row=4, sticky=W)

var10 = IntVar()
tk.Checkbutton(stepThree, text="Sequence", variable=var10, font=16).grid(column=1, row=5, sticky=W)

var18 = IntVar()
tk.Checkbutton(stepThree, text="ADA Buffer", variable=var18, font=16).grid(column=1, row=6, sticky=W)

ttk.Label(stepThree, text="Compare results to operational data (1)", font=16).grid(column=2, row=1, sticky=W)
var11 = IntVar()
tk.Checkbutton(stepThree, variable=var11, font=16).grid(column=3, row=1, sticky=W)

ttk.Label(stepThree, text="I want to compare the results to other set of results. (2)", font=16).grid(column=2, row=2, sticky=W)
var12 = IntVar()
tk.Checkbutton(stepThree, variable=var12, font=16).grid(column=3, row=2, sticky=W)

ttk.Label(stepThree, text="How many? (Up to 5)", font=16).grid(column=2, row=3, sticky=W)

input_entry1 = ttk.Entry(stepThree, width=7, textvariable=m_input)
input_entry1.grid(column=3, row=3, sticky=(W, E))

buttons_visual = ttk.Frame(visual_module)
buttons_visual.grid(row=2, sticky=(N, W, E, S))

tk.Button(buttons_visual, text="RUN", command=define_input_parameters, activebackground = "pink", font=16, height = 1, overrelief="raised", width = 15).pack(side="right")
tk.Button(buttons_visual, text='<- CORE MODULE', command=lambda:raise_main_frame(core_module), activebackground = "pink", font=16, height = 1, overrelief="raised", width = 15).pack(side="right")


raise_frame(f0)
raise_main_frame(input_module)

win.columnconfigure(0, weight=1)
win.rowconfigure(0, weight=1)
win.rowconfigure(1, weight=1)

win.mainloop()      


n_times = n_times_output.get()
m = m_output.get()

# Switch for Convergence
if convergence_output.get() == 1:
    convergenceFLAG = True
else:
    convergenceFLAG = False
# Switch for Throughput
if Throughput_check_output.get() == 1:
    Thr_FLAG = True
else:
    Thr_FLAG = False
# Switch for Departures Delay
if Delay_check_output.get() == 1:
    Delay_FLAG = True
else:
    Delay_FLAG = False
#Switch for Sequence
if Seq_check_output.get() == 1:
    Seq_FLAG = True
else:
    Seq_FLAG = False
#Switch for ADA Buffer 
if ADA_buffer_output.get() == 1:
    ADA_buffer_FLAG = True
else:
    ADA_buffer_FLAG = False   
#Switch for op data    
if op_yes_output.get() == 1:
    OP_FLAG = True
else:
    OP_FLAG = False
#Switch for comparison   
if new_set_output.get() == 1:
    new_set_FLAG = True
else:
    new_set_FLAG = False
# Switch for Arrivals Delay
if arr_delay_output.get() == 1:
    arr_delay_FLAG = True
else:
    arr_delay_FLAG = False
    
#----Operational Data -----#
if OP_FLAG == True:
    win = Tk()
    win.title("Operational Data import")
    mainframe = ttk.Frame(win, padding="10 10 30 40")
    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
    mainframe.columnconfigure(0, weight=1)
    mainframe.rowconfigure(0, weight=1)
    innerframe = ttk.Frame(win, padding="5 5 0 0")
    # Defines expected inputs (i.e. GUI expects integers) and assigns default values
    button_check = StringVar(win, value='0')
    op_data_sheet = StringVar()
    ttk.Label(mainframe, text="Import the Operational Data File : ").grid(column=1, row=1, sticky=W)
    open_op_data = ttk.Button(mainframe, text="Import operational data", command=load_op_data).grid(column=2, row=1, sticky=W)
    inner = Frame(win, bg='pink', width=0, height=0, padx=20, pady=20)
    inner.grid(column=0, row=1)
    inner.columnconfigure(0, weight=1)
    inner.rowconfigure(0, weight=1)
    ttk.Button(inner, text="Visualize results", command=define_input_parameters).grid(column=0, row=0, sticky=W)
    
    win.columnconfigure(0, weight=1)
    win.rowconfigure(0, weight=1)
    win.rowconfigure(1, weight=1)
    
    win.mainloop()
    operational_data = op_data_sheet.get() 
#----New set of data---#
if new_set_FLAG == True:
    if m >= 1 :
        win = Tk()
        win.title("New set of data import")
        
        mainframe = ttk.Frame(win, padding="10 10 30 40")
        mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
        mainframe.columnconfigure(0, weight=1)
        mainframe.rowconfigure(0, weight=1)
        
        innerframe = ttk.Frame(win, padding="5 5 0 0")
        
        # Defines expected inputs (i.e. GUI expects integers) and assigns default values
        button_check = StringVar(win, value='0')
        
        average_check_output = IntVar() 
        new_data_sheet2 = StringVar()        
        ttk.Label(mainframe, text="Import new data set 2: ").grid(column=1, row=1, sticky=W)        
        open_new_data2 = ttk.Button(mainframe, text="Import data 2", command=load_new_data2).grid(column=2, row=1, sticky=W)
        
        if m >=2:
            new_data_sheet3 = StringVar()
            ttk.Label(mainframe, text="Import new data set 3: ").grid(column=1, row=2, sticky=W)            
            open_new_data3 = ttk.Button(mainframe, text="Import data 3", command=load_new_data3).grid(column=2, row=2, sticky=W)
            
            if m >=3:
                new_data_sheet4 = StringVar()
                ttk.Label(mainframe, text="Import new data set 4: ").grid(column=1, row=3, sticky=W)            
                open_new_data4 = ttk.Button(mainframe, text="Import data 4", command=load_new_data4).grid(column=2, row=3, sticky=W)
                if m >=4:
                    new_data_sheet5 = StringVar()
                    ttk.Label(mainframe, text="Import new data set 5: ").grid(column=1, row=4, sticky=W)            
                    open_new_data5 = ttk.Button(mainframe, text="Import data 5", command=load_new_data5).grid(column=2, row=4, sticky=W)
                    if m >=5:
                        new_data_sheet6 = StringVar()
                        ttk.Label(mainframe, text="Import new data set 6: ").grid(column=1, row=5, sticky=W)            
                        open_new_data6 = ttk.Button(mainframe, text="Import data 6", command=load_new_data6).grid(column=2, row=5, sticky=W)
        
        
        inner = Frame(win, bg='pink', width=0, height=0, padx=20, pady=20)
        inner.grid(column=0, row=1)
        inner.columnconfigure(0, weight=1)
        inner.rowconfigure(0, weight=1)
        ttk.Button(inner, text="Visualize results", command=define_input_parameters2).grid(column=0, row=0, sticky=W)
        
        win.columnconfigure(0, weight=1)
        win.rowconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)
           
        win.mainloop()
        if average_check_output.get() == 1:
            average_FLAG = True
        else:
            average_FLAG = False
        
    if m>=1:    
        new_data2 = new_data_sheet2.get() 
        xls2 = pd.ExcelFile(new_data2) 
        df_thr2 = xls2.parse(5)
        df_delay2 = xls2.parse(6)
        df_dep_output2 = xls2.parse(4)
        df_arr_output2 = xls2.parse(3)
        df_rwy_calcs2 = xls2.parse(2)
        
        if m>=2:
            new_data3 = new_data_sheet3.get() 
            xls3 = pd.ExcelFile(new_data3) 
            df_thr3 = xls3.parse(5)
            df_delay3 = xls3.parse(6)
            df_dep_output3 = xls3.parse(4)
            df_arr_output3 = xls3.parse(3)
            df_rwy_calcs3 = xls3.parse(2)
            if m>=3:
                new_data4 = new_data_sheet4.get() 
                xls4 = pd.ExcelFile(new_data4) 
                df_thr4 = xls4.parse(5)
                df_delay4 = xls4.parse(6)
                df_dep_output4 = xls4.parse(4)
                df_arr_output4 = xls4.parse(3)
                df_rwy_calcs4 = xls4.parse(2)
                if m>=4:
                    new_data5 = new_data_sheet5.get() 
                    xls5 = pd.ExcelFile(new_data5) 
                    df_thr5 = xls5.parse(5)
                    df_delay5 = xls5.parse(6)
                    df_dep_output5 = xls5.parse(4)
                    df_arr_output5 = xls5.parse(3)
                    df_rwy_calcs5 = xls5.parse(2)
                    if m>=5:
                        new_data6 = new_data_sheet6.get() 
                        xls6 = pd.ExcelFile(new_data6) 
                        df_thr6 = xls6.parse(5)
                        df_delay6 = xls6.parse(6)
                        df_dep_output6 = xls6.parse(4)
                        df_arr_output6 = xls6.parse(3)
                        df_rwy_calcs6 = xls6.parse(2)



#*****************************************************************************#
#=============================================================================#
#                                                                             #
#                              CORE MODULE                                    #
#                                                                             #
# ============================================================================#
#*****************************************************************************#

n_times = n_times_output.get()
limit_to_change = 15
lower_limit_to_change = 10

big_list = []
averages =[]
difference=[]
iter2 = 0
iter1 = 0 
if averagethrFLAG == True:
    maxIter = 10
else:
    maxIter = n_times
 #How many times you want to do it....
while (iter1 < maxIter):

    n = n_output.get()
    ADA_x = ADA_x_output.get()
    minDep_altSID = minDep_altSID_output.get()
    minDep_sameSID = minDep_sameSID_output.get()
    SIDmax = SIDmax_output.get()
    SIDgroup_separation = SIDgroup_separation_output.get()
    SID_queue_assign = SID_queue_assign_output.get()    
    input_excel_sheet = name_excel_sheet.get()
    
    
    # Switch for including 'Debug' output
    if debug_output.get() == 1:
        debugFLAG = True
    else:
        debugFLAG = False
        
    if debug3_output.get() == 1:
        debugFLAG3 = True
    else:
        debugFLAG3 = False
    # Switch for modelling 'Radar Tower Separation' concept
        
    if TBS_output.get() == 1:
        TBS_Flag = True
    else:
        TBS_Flag = False
    
    if RECat_output.get() == 1:
        RECatFLAG = True
    else:
        RECatFLAG = False
    
    if RECAT_PWS_output.get() == 1:
        RECAT_PWS_FLAG = True
    else:
        RECAT_PWS_FLAG = False
        
        
        
    if averagethr_output.get() ==1:
        averagethrFLAG = True
    else:
        averagethrFLAG = False
        
#    if distance_based_output.get() ==1:
#        distance_based_FLAG = True
#    else:
#        distance_based_FLAG = False
        
    if time_based_output.get() ==1:
        time_based_FLAG = True
        distance_based_FLAG = False
    else:
        time_based_FLAG = False
        
        
    #Delievery:
    if MRS4dme_output.get() ==1:
        MRS_4dme_FLAG = True
    else:
        MRS_4dme_FLAG = False
    
    if WAKE4dme_output.get() ==1:
        WAKE_4dme_FLAG = True
    else:
        WAKE_4dme_FLAG = False
    
    if ADA4dme_output.get() ==1:
        ADA_4dme_FLAG = True
    else:
        ADA_4dme_FLAG = False
    if ADDA4dme_output.get() ==1:
        ADDA_4dme_FLAG = True
    else:
        ADDA_4dme_FLAG = False
    if MRSthr_output.get() ==1:
        MRS_thr_FLAG = True
    else:
        MRS_thr_FLAG = False
    
    if WAKEthr_output.get() ==1:
        WAKE_thr_FLAG = True
    else:
        WAKE_thr_FLAG = False
    
    if ADAthr_output.get() ==1:
        ADA_thr_FLAG = True
    else:
        ADA_thr_FLAG = False
    if ADDAthr_output.get() ==1:
        ADDA_thr_FLAG = True
    else:
        ADDA_thr_FLAG = False
        

    # Queue selection    
    if queue1_output.get() == 1:
        Use_queue = 1
        maxRWYqueue1Length = 8
        previousRWYqueue = 1
    elif queue2_output.get() == 1:
        Use_queue = 2
        maxRWYqueue1Length = 4
        maxRWYqueue2Length = 4
        # Forces RWYqueue1 to go first
        previousRWYqueue = 2
    elif queue3_output.get() == 1:
        Use_queue = 3
        maxRWYqueue1Length = 2
        maxRWYqueue2Length = 2
        maxRWYqueue3Length = 2
        maxRWYqueue4Length = 2
        # Forces RWYqueue1/2 to go first
        previousRWYqueue = 3
    elif queue4_output.get() == 1:
        Use_queue = 4
        maxRWYqueue1Length = 4 # Will use 2x4 methods (accounts for 4 Queues x 1 in length)
        maxRWYqueue2Length = 4 # Will use 2x4 methods (accounts for 4 Queues x 1 in length)
        # Forces RWYqueue1 to go first
        previousRWYqueue = 2
    else:
        print("Please select ONE of the queue options. Exiting..")
        sys.exit(0)
    
    print("Queue [", Use_queue,"] Selected")
    # If GUI is exited, model will exit too
    if button_check.get() == '0':
        print("GUI closed! Exiting model ...")
        sys.exit(0)
    # If no input file selected
    if input_excel_sheet == '':
        print("Please select an input file ...")
        sys.exit(0)
    

    #####################################################################
    #                           START OF MODEL                          #
    #####################################################################
    
    program_runtime_start = time.time() # RUNTIME CALCULATION
    #-------------------------USEFUL FILES ---------------------------#
    
    #----- SCHEDULE -----#
    wb = openpyxl.load_workbook(input_excel_sheet, data_only=True)
    #------AROT/DROT LOOKUP TABLE ------#
    df_distributions = pd.read_csv('utility/AROTDROT_distributions.csv')
    #------ACTUAL SPEED PROFILE -------#
    df_speed_profiles = pd.read_csv('utility/actual_speed_profiles.csv')
    #------- WAKE ---------#
    df_wake = pd.read_csv('utility/wake.csv')
    
    df_wake_WTC=pd.DataFrame()
    df_wake_WTC['ICAO'] = df_wake['ICAO']
    df_wake_WTC['WTC'] = df_wake['WTC']
    df_wake_WTC = df_wake_WTC.set_index('ICAO')
    
    if RECatFLAG==True:
        df_wake_RECAT=pd.DataFrame()
        df_wake_RECAT['ICAO'] = df_wake['ICAO']
        df_wake_RECAT['RECAT-EU'] = df_wake['RECAT-EU']
        df_wake_RECAT = df_wake_RECAT.set_index('ICAO')
    
        
    #-----RECAT-EU separation-------#
    if RECatFLAG==True:
        df_RECAT_EU_separation = pd.read_csv('utility/RECAT_EU_separation.csv')
        df_RECAT_EU_separation = df_RECAT_EU_separation.set_index("LEAD")
    #----WTC separation ------#
    if RECatFLAG==False:
        df_WTC_separation = pd.read_csv('utility/UK_wake_separation.csv')
        df_WTC_separation = df_WTC_separation.set_index("LEAD")
    #----RECAT-PWS separation ------#   
    if RECAT_PWS_FLAG == True:
        df_RECAT_PWS = pd.read_csv('utility/RECAT_PWS.csv')
        df_RECAT_PWS = df_RECAT_PWS.fillna(0)
        df_RECAT_PWS = df_RECAT_PWS.set_index('FOLLOW')
        
        df_RECAT20 = pd.DataFrame()
        df_RECAT20['ICAO'] = df_wake['ICAO']
        df_RECAT20['RECAT20'] = df_wake['RECAT20']
        df_RECAT20 = df_RECAT20.set_index('ICAO')
        
        #----RECAT-EU 20cat separation ------#
        df_RECAT20_separation = pd.read_csv('utility/RECAT20_separation.csv')
        df_RECAT20_separation = df_RECAT20_separation.fillna(0)
        df_RECAT20_separation = df_RECAT20_separation.set_index('LEAD')
        #df_RECAT20_separation = df_RECAT_PWS.set_index('LEAD')
    
    
    
    
    df_speed_profiles = df_speed_profiles.drop(columns=['Unnamed: 0'])
    dict_actual_speed_profiles= {k: v for k, v in df_speed_profiles.groupby('Aircraft_Type')}
    for key in list(dict_actual_speed_profiles.keys()):
        dict_actual_speed_profiles[key] = dict_actual_speed_profiles[key].reset_index()
        dict_actual_speed_profiles[key] = dict_actual_speed_profiles[key].drop(columns = 'index')
    
    #print(A_temp_df['NO PROFILE'].size)
         
    
    
    
    # process distributions AROT/DROT/TAXI-in/TAXI-out
    def process_each_column_in_distributions(dataframe,name):      
        
        dataframe[name] = df_distributions[name]     
        dataframe = dataframe.dropna(subset=[name])
        return dataframe
    
    df_AROT_H =  pd.DataFrame()
    df_AROT_H = process_each_column_in_distributions(df_AROT_H, 'AROT_H')
    df_AROT_M =  pd.DataFrame()
    df_AROT_M = process_each_column_in_distributions(df_AROT_M, 'AROT_M')
    df_AROT_L =  pd.DataFrame()
    df_AROT_L = process_each_column_in_distributions(df_AROT_L, 'AROT_L')
    df_AROT_J =  pd.DataFrame()
    df_AROT_J = process_each_column_in_distributions(df_AROT_J, 'AROT_J')
    df_AROT_UM =  pd.DataFrame()
    df_AROT_UM = process_each_column_in_distributions(df_AROT_UM, 'AROT_UM')
    df_AROT_S =  pd.DataFrame()
    df_AROT_S = process_each_column_in_distributions(df_AROT_S, 'AROT_S')
    
    df_DROT_H =  pd.DataFrame()
    df_DROT_H = process_each_column_in_distributions(df_DROT_H, 'DROT_H')
    df_DROT_M =  pd.DataFrame()
    df_DROT_M = process_each_column_in_distributions(df_DROT_M, 'DROT_M')
    df_DROT_L =  pd.DataFrame()
    df_DROT_L = process_each_column_in_distributions(df_DROT_L, 'DROT_L')
    df_DROT_J =  pd.DataFrame()
    df_DROT_J = process_each_column_in_distributions(df_DROT_J, 'DROT_J')
    df_DROT_UM =  pd.DataFrame()
    df_DROT_UM = process_each_column_in_distributions(df_DROT_UM, 'DROT_UM')
    df_DROT_S =  pd.DataFrame()
    df_DROT_S = process_each_column_in_distributions(df_DROT_S, 'DROT_S')
    
#    df_TAXI_OUT_S1 =  pd.DataFrame()
#    df_TAXI_OUT_S1 = process_each_column_in_distributions(df_TAXI_OUT_S1, 'TAXI_OUT_S1')
#    df_TAXI_OUT_S2 =  pd.DataFrame()
#    df_TAXI_OUT_S2 = process_each_column_in_distributions(df_TAXI_OUT_S2, 'TAXI_OUT_S2')
#    df_TAXI_OUT_S3 =  pd.DataFrame()
#    df_TAXI_OUT_S3 = process_each_column_in_distributions(df_TAXI_OUT_S3, 'TAXI_OUT_S3')
#    df_TAXI_OUT_S4 =  pd.DataFrame()
#    df_TAXI_OUT_S4 = process_each_column_in_distributions(df_TAXI_OUT_S4, 'TAXI_OUT_S4')
#    df_TAXI_OUT_S5 =  pd.DataFrame()
#    df_TAXI_OUT_S5 = process_each_column_in_distributions(df_TAXI_OUT_S5, 'TAXI_OUT_S5')
#    df_TAXI_OUT_S6 =  pd.DataFrame()
#    df_TAXI_OUT_S6 = process_each_column_in_distributions(df_TAXI_OUT_S6, 'TAXI_OUT_S6')
#    df_TAXI_OUT_S7 =  pd.DataFrame()
#    df_TAXI_OUT_S7 = process_each_column_in_distributions(df_TAXI_OUT_S7, 'TAXI_OUT_S7')
#    df_TAXI_OUT_S8 =  pd.DataFrame()
#    df_TAXI_OUT_S8 = process_each_column_in_distributions(df_TAXI_OUT_S8, 'TAXI_OUT_S8')
#    df_TAXI_OUT_S9 =  pd.DataFrame()
#    df_TAXI_OUT_S9 = process_each_column_in_distributions(df_TAXI_OUT_S9, 'TAXI_OUT_S9')
#    df_TAXI_OUT_S10 =  pd.DataFrame()
#    df_TAXI_OUT_S10 = process_each_column_in_distributions(df_TAXI_OUT_S10, 'TAXI_OUT_S10')
#    df_TAXI_OUT_S11 =  pd.DataFrame()
#    df_TAXI_OUT_S11 = process_each_column_in_distributions(df_TAXI_OUT_S11, 'TAXI_OUT_S11')
#    df_TAXI_OUT_S12 =  pd.DataFrame()
#    df_TAXI_OUT_S12 = process_each_column_in_distributions(df_TAXI_OUT_S12, 'TAXI_OUT_S12')
#    df_TAXI_OUT_S13 =  pd.DataFrame()
#    df_TAXI_OUT_S13 = process_each_column_in_distributions(df_TAXI_OUT_S13, 'TAXI_OUT_S13')
#    df_TAXI_OUT_S14 =  pd.DataFrame()
#    df_TAXI_OUT_S14 = process_each_column_in_distributions(df_TAXI_OUT_S14, 'TAXI_OUT_S14')
#    df_TAXI_OUT_S15 =  pd.DataFrame()
#    df_TAXI_OUT_S15 = process_each_column_in_distributions(df_TAXI_OUT_S15, 'TAXI_OUT_S15')
#    
#    df_TAXI_IN_S1 =  pd.DataFrame()
#    df_TAXI_IN_S1 = process_each_column_in_distributions(df_TAXI_IN_S1, 'TAXI_IN_S1')
#    df_TAXI_IN_S2 =  pd.DataFrame()
#    df_TAXI_IN_S2 = process_each_column_in_distributions(df_TAXI_IN_S2, 'TAXI_IN_S2')
#    df_TAXI_IN_S3 =  pd.DataFrame()
#    df_TAXI_IN_S3 = process_each_column_in_distributions(df_TAXI_IN_S3, 'TAXI_IN_S3')
#    df_TAXI_IN_S4 =  pd.DataFrame()
#    df_TAXI_IN_S4 = process_each_column_in_distributions(df_TAXI_IN_S4, 'TAXI_IN_S4')
#    df_TAXI_IN_S5 =  pd.DataFrame()
#    df_TAXI_IN_S5 = process_each_column_in_distributions(df_TAXI_IN_S5, 'TAXI_IN_S5')
#    df_TAXI_IN_S6 =  pd.DataFrame()
#    df_TAXI_IN_S6 = process_each_column_in_distributions(df_TAXI_IN_S6, 'TAXI_IN_S6')
#    df_TAXI_IN_S7 =  pd.DataFrame()
#    df_TAXI_IN_S7 = process_each_column_in_distributions(df_TAXI_IN_S7, 'TAXI_IN_S7')
#    df_TAXI_IN_S8 =  pd.DataFrame()
#    df_TAXI_IN_S8 = process_each_column_in_distributions(df_TAXI_IN_S8, 'TAXI_IN_S8')
#    df_TAXI_IN_S9 =  pd.DataFrame()
#    df_TAXI_IN_S9 = process_each_column_in_distributions(df_TAXI_IN_S9, 'TAXI_IN_S9')
#    df_TAXI_IN_S10 =  pd.DataFrame()
#    df_TAXI_IN_S10 = process_each_column_in_distributions(df_TAXI_IN_S10, 'TAXI_IN_S10')
#    df_TAXI_IN_S11 =  pd.DataFrame()
#    df_TAXI_IN_S11 = process_each_column_in_distributions(df_TAXI_IN_S11, 'TAXI_IN_S11')
#    df_TAXI_IN_S12 =  pd.DataFrame()
#    df_TAXI_IN_S12 = process_each_column_in_distributions(df_TAXI_IN_S12, 'TAXI_IN_S12')
#    df_TAXI_IN_S13 =  pd.DataFrame()
#    df_TAXI_IN_S13 = process_each_column_in_distributions(df_TAXI_IN_S13, 'TAXI_IN_S13')
#    df_TAXI_IN_S14 =  pd.DataFrame()
#    df_TAXI_IN_S14 = process_each_column_in_distributions(df_TAXI_IN_S14, 'TAXI_IN_S14')
#    df_TAXI_IN_S15 =  pd.DataFrame()
#    df_TAXI_IN_S15 = process_each_column_in_distributions(df_TAXI_IN_S15, 'TAXI_IN_S15')
    
    
    
    ###########################################


    
    
    # Data frame
    xls = pd.ExcelFile(input_excel_sheet) 
    df_dep = xls.parse(1)
    df_arr = xls.parse(0)

#---------------------------GLOBAL VARIABLES----------------------------------#


    # SET WAKE RULES for departures
    H_H_d = 90
    H_M_d = 139 # Used for H_ UM/M/S/L
    J_H_d = 139 #120
    J_M_d = 204 #180 #2016 data 
    J_S_d = 204 #180 #2016 data 
    J_L_d = 204 #180 #2016 data 
    M_L_d = 139 #Used for UM_L & M_L
    S_L_d = 139 #120
    
    #(DBS)#ICAO wake rules. obs: L=S and M=UM
    j_h = 6
    j_m = 7
    j_l = 8
    h_h = 4
    h_m = 5
    h_l = 6
    m_l = 5

    #(DBS)#RECAT separation (RECAT-EU edition 1.1 = 15/07/2015) 
    A_A = 3 #same for : B_B , C_C , C_D , F_F
    A_B = 4 #same for : B_C , B_D , C_E , E_F
    A_C = 5 #same for : A_D , B_E , D_F
    A_D = 6 #same for : A_E , C_F
    A_F = 8
    B_F = 7
    
    #(TBS)#RECAT separation(RECAT-EU edition 1.1 = 15/07/2015)
    a_b = 100 #same for : B_C , B_D , C_E , E_F
    a_c = 120 #same for : B_E , C_F , D_F
    a_d = 140 #same for : B_F 
    a_e = 160 
    a_f = 180
    c_d = 80 #same for : F_F

    c_dme = 4
    d_dme = 3
    min_radar_separation_distance = 3 #NM
    STT = 600 # Standard Taxi Time - used for arrivals. Actually, it is from the landing point to the stand
    # Initiate global variables:
    RWY_status = "E"
    # Initialised 'dict' for storing Taxiing-in Arrivals
    ARRIVALqueue = {}
    #Initialisez 'dict' for the Arrivals hold queue
    ArrHOLDqueue = {}
    APPqueue = {}
    # Initialised 'dict' for holding Departures on Stands (Push/Start Delay)
    DepSTANDqueue = {}
    TAXIqueue = {}
    TAXIhold = {}
    # Place Dep A/C with SID group 1 into RWYqueue1 etc - UNLESS there's no A/C of this type available
    RWYqueue1 = {}
    RWYqueue2 = {}
    # Other queues instigated for 4x2 (and future 8x1) arrangements
    RWYqueue3 = {}
    RWYqueue4 = {}
    
    GoAroundCount = {}
    # maxTAXIaircraft = 23 # OLD VALUE - Didn't result in any HOLDdelay!
    maxTAXIaircraft = 15
    
    # If no queued Departures arrive at Alpha_box - counts as 'debug'
    countDEPdebug = 0
    countARRdebug = 0
    
    STANDholdDelay = 0 # Instigate new variable 'STANDholdDelay' ready to store STANDholdDelay values in DepSTANDhold queue
    TAXIholdDelay = 0 # Instigate new variable 'TAXIholdDelay' ready to store TAXIholdDelay values in TAXIqueue
    RWYqueueDelay = 0 # Calc time each A/C spent in RWYqueue
    
    currentGap = 86400
    ArrthroughputRow = 2     # For throughput calcs
    # Count number of times each method called for Debug
    countArr = 0
    countDep = 0
    # Arrival 'go-around' case
    goAroundCase = False
    number_of_goArounds_queued = 0
    #Timing variables
    SOBTtime = 0
    dep2time = 0
    deptime = 0
    currentGapErrorFLAG = False
    #ARRIVALS
    ARRkey = 2
    ArrOutput = 2
    #DEPARTURES
    SOBTrow = 2
    DepOutput = 2
    seqRow = 2
    
    x_buffer = 15
    
    throughput =[]
    ####### Create list of hours from 15 to 15 mins
    list_for_15min =[0]
    entry_in_15min=900
    while entry_in_15min <= 86400:
        entry_in_15min = entry_in_15min + 900
        list_for_15min.append(entry_in_15min)
    start_value_15min = 0
    
    
    
    
    #########################################################################
    #                             RECAT                                     # 
    #########################################################################
    RECAT_categories = {'A': ['A388','A124'],
                        'B': ['A332','A333','A343','A345','A346','A359','B744','B748','B772','B773','B77L','B77W','B788','B789','IL96'],
                        'C' : ['A306','A30B ','A310','B703 ','B752','B753 ','B762','B763','B764','B783','C135','DC10','DC85','IL76','MD11','TU22','TU95'],
                        'D' : ['A318', 'A319', 'A320', 'A321', 'AN12', 'B736', 'B737', 'B738', 'B739', 'C130', 'IL18', 'MD81', 'MD82', 'MD83', 'MD87', 'MD88', 'MD90', 'T204', 'TU16'],
                        'E': ['AT43', 'AT45', 'AT72', 'B712', 'B732', 'B733', 'B734', 'B735', 'CL60', 'CRJ1', 'CRJ2', 'CRJ7', 'CRJ9', 'DH8D', 'E135', 'E145', 'E170', 'E175', 'E190', 'E195', 'F70', 'F100', 'GLF4', 'RJ85', 'RJ1H'],
                        'F' : ['FA10', 'FA20', 'D328', 'E120', 'BE40', 'BE45', 'H25B', 'JS32', 'JS41', 'LJ35', 'LJ60', 'SF34', 'P180', 'C650', 'C525', 'C180', 'C152']}
    AC_types_list = []
    wake_cat_list = []
    for wake_cat in list(RECAT_categories.keys()):
        wake_list_temp = []
        AC_types_list += RECAT_categories[wake_cat]
        a = len(RECAT_categories[wake_cat])
        wake_list_temp = [str(wake_cat)] *a
        wake_cat_list.append(wake_list_temp)
    wake_cat_list = wake_cat_list[0] + wake_cat_list[1] + wake_cat_list[2] + wake_cat_list[3] + wake_cat_list[4] + wake_cat_list[5]
    



#------------------------------INPUT PRE-PROCESS -----------------------------#

    
    #Initialise Arrival input - N.B. must be outside method
    arrivalInput = wb.get_sheet_by_name('Arrivals')
    departureInput = wb.get_sheet_by_name('Departures')
    max_ARRIVAL = arrivalInput.max_row + 1
    max_DEPARTURE = departureInput.max_row + 1
    # Create 3 new excel tabs - Intermediate calculations, Arrival Output, Departure Output
    wb.create_sheet(index=3, title='Runway_calcs')
    runwayCalculations = wb.get_sheet_by_name('Runway_calcs')
    wb.create_sheet(index=4, title='Arrival_Output')
    arrivalOutput = wb.get_sheet_by_name('Arrival_Output')
    wb.create_sheet(index=5, title='Departure_Output')
    departureOutput = wb.get_sheet_by_name('Departure_Output')
    wb.create_sheet(index=6, title='Throughput')
    throughputTab = wb.get_sheet_by_name('Throughput')
    wb.create_sheet(index=7, title='Delay')
    delayTab = wb.get_sheet_by_name('Delay')
    wb.create_sheet(index=7, title='Sequence')
    sequenceTab = wb.get_sheet_by_name('Sequence')
    
    if debugFLAG3 == True:
        wb.create_sheet(index=8,title='Debug')
        debugTab = wb.get_sheet_by_name('Debug')
    # Function to write headers for the output excel sheets
    def set_Output_Excel_headers():
        # runway Calculations (intermediate step) headers
        runwayCalculations['A' + str(1)].value = 'Arrival ID'
        runwayCalculations['B' + str(1)].value = 'TAXI-IN'
        runwayCalculations['C' + str(1)].value = 'AROT'
        runwayCalculations['D' + str(1)].value = 'ADA'
        runwayCalculations['E' + str(1)].value = 'ADDA'
        runwayCalculations['F' + str(1)].value = 'ATCO variability'
        runwayCalculations['G' + str(1)].value = 'WIND1'
        runwayCalculations['H' + str(1)].value = 'SPEED1'
        runwayCalculations['I' + str(1)].value = 'WIND2'
        runwayCalculations['J' + str(1)].value = 'SPEED2'
        runwayCalculations['K' + str(1)].value = 'VTGT'
        runwayCalculations['L' + str(1)].value = 'SAE'
        runwayCalculations['M' + str(1)].value = 'PREDICTED Landing Time'
        runwayCalculations['N' + str(1)].value = 'MAX Constraint'
        runwayCalculations['O' + str(1)].value = 'MAX Constraint Label'
        runwayCalculations['R' + str(1)].value = 'Departure ID' 
        runwayCalculations['S' + str(1)].value = 'TAXI-OUT'
        runwayCalculations['T' + str(1)].value = 'DROT'
        runwayCalculations['U' + str(1)].value = 'ARRIVAL actual WAKE'

        
        
        # Arrival Output tab headers
        arrivalOutput['A' + str(1)].value = 'Arrival ID'
        arrivalOutput['B' + str(1)].value = 'Arrival HOUR'
        arrivalOutput['C' + str(1)].value = 'ACTUAL Landing Time'
        arrivalOutput['D' + str(1)].value = 'Arrival RWY_EXIT'
        arrivalOutput['E' + str(1)].value = 'WAKE'
        arrivalOutput['F' + str(1)].value = 'In Blocks Time'
        arrivalOutput['G' + str(1)].value = 'AROT'
        arrivalOutput['H' + str(1)].value = 'TAXI-IN Duration'
        arrivalOutput['I' + str(1)].value = 'MAX Constraint'
        arrivalOutput['J' + str(1)].value = 'MAX Constraint Label'
        arrivalOutput['K' + str(1)].value = 'len(ArrHOLDqueue)'
        arrivalOutput['L' + str(1)].value = 'Arrival DELAY'
        
        #Arrival Input tab added columns (ACTUAL SPEED PROFILE)
        arrivalInput['AD' + str(1)].value = 'GS_0_1dme'
        arrivalInput['AE' + str(1)].value = 'GS_1_2dme'
        arrivalInput['AF' + str(1)].value = 'GS_2_3dme'
        arrivalInput['AG' + str(1)].value = 'GS_3_4dme'
        arrivalInput['AH' + str(1)].value = 'GS_4_5dme'
        arrivalInput['AI' + str(1)].value = 'GS_5_6dme'
        arrivalInput['AJ' + str(1)].value = 'GS_6_7dme'
        arrivalInput['AK' + str(1)].value = 'GS_7_8dme'
        arrivalInput['AL' + str(1)].value = 'GS_8_9dme'
        arrivalInput['AM' + str(1)].value = 'GS_9_10dme'
        
        arrivalInput['AN' + str(1)].value = 'IAS_0_1dme'
        arrivalInput['AO' + str(1)].value = 'IAS_1_2dme'
        arrivalInput['AP' + str(1)].value = 'IAS_2_3dme'
        arrivalInput['AQ' + str(1)].value = 'IAS_3_4dme'
        arrivalInput['AR' + str(1)].value = 'IAS_4_5dme'
        arrivalInput['AS' + str(1)].value = 'IAS_5_6dme'
        arrivalInput['AT' + str(1)].value = 'IAS_6_7dme'
        arrivalInput['AU' + str(1)].value = 'IAS_7_8dme'
        arrivalInput['AV' + str(1)].value = 'IAS_8_9dme'
        arrivalInput['AW' + str(1)].value = 'IAS_9_10dme'
        
        
        # Departure Output tab headers
        departureOutput['A' + str(1)].value = 'Departure ID'
        departureOutput['B' + str(1)].value = 'Departure HOUR'
        departureOutput['C' + str(1)].value = 'Departure_RWY_ENTRY'
        departureOutput['D' + str(1)].value = 'Departure_RWY_EXIT'
        departureOutput['E' + str(1)].value = 'WAKE'
        departureOutput['F' + str(1)].value = 'SID GROUP'
        departureOutput['G' + str(1)].value = 'DROT'
        departureOutput['H' + str(1)].value = 'TAXI-OUT'
        departureOutput['I' + str(1)].value = 'Dep MIN Separation'
        departureOutput['J' + str(1)].value = 'Dep MIN Separation Label'
        departureOutput['K' + str(1)].value = 'currentGap'
        departureOutput['L' + str(1)].value = 'len(DepSTANDqueue)'
        departureOutput['M' + str(1)].value = 'len(TAXIhold)'
        departureOutput['N' + str(1)].value = 'len(RWYqueue1)'
        departureOutput['O' + str(1)].value = 'len(RWYqueue2)'
        departureOutput['P' + str(1)].value = 'len(RWYqueue3)'
        departureOutput['Q' + str(1)].value = 'len(RWYqueue4)'
        departureOutput['R' + str(1)].value = 'DELAY DepSTANDqueue'
        departureOutput['S' + str(1)].value = 'DELAY TAXIhold'
        departureOutput['T' + str(1)].value = 'DELAY RWYqueue'
        departureOutput['U' + str(1)].value = 'RWY queue USED'

        
        # Throughput tab headers
        throughputTab['A' + str(1)].value = 'Hour'
        throughputTab['B' + str(1)].value = 'Departure Throughput'
        throughputTab['C' + str(1)].value = 'Arrival Throughput'
        throughputTab['D' + str(1)].value = 'Total Throughput'
        throughputTab['E' + str(1)].value = 'Cum. No. of Go-Arounds'
#        
#        # Delay tab headers 
        delayTab['A' + str(1)].value = 'Departure ID'
        delayTab['B' + str(1)].value = 'HOUR'
        delayTab['C' + str(1)].value = 'RWY HOLD Delay'
        delayTab['D' + str(1)].value = 'Push/Start Delay'
        
        delayTab['I' + str(1)].value = 'Arrival ID'
        delayTab['J' + str(1)].value = 'HOUR'
        delayTab['K' + str(1)].value = 'Arrival Delay'
        
        
        # Sequence tab headers
        sequenceTab['A' + str(1)].value = 'Type'
        sequenceTab['B' + str(1)].value = 'ID'
        sequenceTab['C' + str(1)].value = 'RWY ENTRY'
        sequenceTab['D' + str(1)].value = 'RWY EXIT'
        sequenceTab['E' + str(1)].value = 'ROT'
        sequenceTab['F' + str(1)].value = 'Arr ID start ADA pair'
        sequenceTab['G' + str(1)].value = 'ADA Buffer'
        
        #Debug tab headers
        if debugFLAG3 == True:
            debugTab['A' + str(1)].value = 'Time'
            debugTab['B' + str(1)].value = 'Runway status'
            debugTab['C' + str(1)].value = 'Current Gap - D'
            debugTab['D' + str(1)].value = 'Current Gap - A'
            debugTab['E' + str(1)].value = 'Current Gap - E'
            debugTab['L' + str(1)].value = 'Arrival Hold Delay'
        return
    
   ####################### ARRIVALS SEPARATION FUNCTIONS########################
                
    def distance_to_time_assumed_speed_profile_IAS(i, d_dme, distance): #DELIVERED at THR
        #fixed d_dme at 3dme, variable c_dme because max deceleration speed is 20kts/NM
        c_dme = 4
        
        ##### JI - THESE THREE LINES LOOK FISHY!
        deceleration_difference= runwayCalculations['H' + str(i)].value - runwayCalculations['K' + str(i)].value
        if deceleration_difference > 20 :
            c_dme = deceleration_difference / 20                    
        #####
        
        TBS_assumed_speed_profile_value = 0
         #time between d_dme - THR:
        t1 = (d_dme *3600)/(runwayCalculations['K' + str(i)].value)
        #time between c_dme - d_dme:
        t2 = (2*3600*(c_dme-d_dme))/(runwayCalculations['H' + str(i)].value+runwayCalculations['K' + str(i)].value)
        
        if distance >= c_dme:
            TBS_assumed_speed_profile_value = int(t1+t2+((distance-c_dme)*3600/runwayCalculations['H' + str(i)].value))
        elif (distance < c_dme) and (distance>d_dme):
            d1 = distance-d_dme
            speed_at_d1 = (d1*( runwayCalculations['H' + str(i)].value - runwayCalculations['K' + str(i)].value ) /(c_dme-d_dme)) + runwayCalculations['K' + str(i)].value
            TBS_assumed_speed_profile_value = int(d1*3600*2/(speed_at_d1 + runwayCalculations['K' + str(i)].value) + t1)
        elif distance <= d_dme:
            TBS_assumed_speed_profile_value = int(distance*3600/runwayCalculations['K' + str(i)].value)
        #print('TBS - on' )   
        
        return(TBS_assumed_speed_profile_value)
        
        
    def DBS_assumed_speed_profile(i, d_dme, distance): #DELIVERED at THR
        
        c_dme = 4
        deceleration_difference= (runwayCalculations['H' + str(i)].value - runwayCalculations['G' + str(i)].value) - (runwayCalculations['K' + str(i)].value - runwayCalculations['I' + str(i)].value)
        if deceleration_difference > 20 :
            c_dme = deceleration_difference / 20    
            
        DBS_assumed_speed_profile_value = 0
        #time between d_dme - THR:
        t1 = (d_dme *3600)/(runwayCalculations['K' + str(i)].value-runwayCalculations['I' + str(i)].value)
        #time between c_dme - d_dme:
        t2 = (2*3600*(c_dme-d_dme))/((runwayCalculations['H' + str(i)].value-runwayCalculations['G' + str(i)].value)+(runwayCalculations['K' + str(i)].value-runwayCalculations['I' + str(i)].value))
        
        if distance >= c_dme:
            DBS_assumed_speed_profile_value = int(t1+t2+((distance-c_dme)*3600/(runwayCalculations['H' + str(i)].value-runwayCalculations['G' + str(i)].value)))     
        elif (distance < c_dme) and (distance>d_dme):
            d1 = distance-d_dme
            speed_at_d1 = (d1*( (runwayCalculations['H' + str(i)].value-runwayCalculations['G' + str(i)].value) - (runwayCalculations['K' + str(i)].value-runwayCalculations['I' + str(i)].value) ) /(c_dme-d_dme)) + (runwayCalculations['K' + str(i)].value-runwayCalculations['I' + str(i)].value)
            DBS_assumed_speed_profile_value = int(d1*3600*2/(speed_at_d1 + runwayCalculations['K' + str(i)].value - runwayCalculations['I' + str(i)].value) + t1)
        elif distance <= d_dme:
            DBS_assumed_speed_profile_value = int(distance*3600/(runwayCalculations['K' + str(i)].value-runwayCalculations['I' + str(i)].value))
                    
        return(DBS_assumed_speed_profile_value)
    
    def DBS_actual_speed_profile(distance,row): #DELIVERED at THR # use GS
        T=0
        def full_segments(n,row):
            T= 0
            if n >= 1:
                T = 2*3600/(arrivalInput['AD'+str(row)].value+arrivalInput['AE'+str(row)].value)
                if n >=2:
                    T += 2*3600/(arrivalInput['AE'+str(row)].value +arrivalInput['AF'+str(row)].value)
                    if n>=3:
                        T += 2*3600/(arrivalInput['AF'+str(row)].value + arrivalInput['AG'+str(row)].value)
                        if n>=4:
                            T += 2*3600/(arrivalInput['AG'+str(row)].value + arrivalInput['AH'+str(row)].value)
                            if n>=5:
                                T += 2*3600/(arrivalInput['AH'+str(row)].value+arrivalInput['AI'+str(row)].value)
                                if n>=6:
                                    T += 2*3600/(arrivalInput['AI'+str(row)].value+arrivalInput['AJ'+str(row)].value)
                                    if n>=7:
                                        T += 2*3600/(arrivalInput['AJ'+str(row)].value+arrivalInput['AK'+str(row)].value)
                                        if n>=8:
                                            T += 2*3600/(arrivalInput['AK'+str(row)].value+arrivalInput['AL'+str(row)].value)
                                            if n==9:
                                                T += 2*3600/(arrivalInput['AL'+str(row)].value + arrivalInput['AM'+str(row)].value)
                                            elif n>9:
                                                T += (n-9)*3600/arrivalInput['AM'+str(row)].value
                                    
                                                
                                                    
            return T
        def fraction_of_segments(n,f,row):
            T = 0
            if n==1:
                S = f*(arrivalInput['AF'+str(row)].value - arrivalInput['AE'+str(row)].value) + arrivalInput['AE'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AE'+str(row)].value)
            elif n==2:
                S = f*(arrivalInput['AG'+str(row)].value - arrivalInput['AF'+str(row)].value) + arrivalInput['AF'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AF'+str(row)].value)
            elif n==3:
                S = f*(arrivalInput['AH'+str(row)].value - arrivalInput['AG'+str(row)].value) + arrivalInput['AG'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AG'+str(row)].value)
            elif n==4:
                S = f*(arrivalInput['AI'+str(row)].value - arrivalInput['AH'+str(row)].value) + arrivalInput['AH'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AH'+str(row)].value)
            elif n==5:
                S = f*(arrivalInput['AJ'+str(row)].value - arrivalInput['AI'+str(row)].value) + arrivalInput['AI'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AI'+str(row)].value)
            elif n==6:
                S = f*(arrivalInput['AK'+str(row)].value - arrivalInput['AJ'+str(row)].value) + arrivalInput['AJ'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AJ'+str(row)].value)
            elif n==7:
                S = f*(arrivalInput['AL'+str(row)].value - arrivalInput['AK'+str(row)].value) + arrivalInput['AK'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AK'+str(row)].value)
            elif n==8:
                S = f*(arrivalInput['AM'+str(row)].value - arrivalInput['AL'+str(row)].value) + arrivalInput['AL'+str(row)].value
                T = (f*3600)/(arrivalInput['AM'+str(row)].value)
            
            return T
#        if distance > 0:
        X = distance + runwayCalculations['F' + str(row)].value # Actual distance + ATCO var
        D = X - 0.5
        if D <0:
            T = (X*3600)/arrivalInput['AD'+str(row)].value
        elif D > 0:
            n = math.floor(D)  
            f = D - n
            T1 = full_segments(n,row)
            if (f != 0) and (n<=8):
                T2 = fraction_of_segments(n,f,row)
                T = T1 + T2 + (0.5*3600)/arrivalInput['AD'+str(row)].value
            else:
                T = T1 + (0.5*3600)/arrivalInput['AD'+str(row)].value
        return T
    
    def TBS_actual_speed_profile(distance,row): #DELIVERED at THR # use IAS
        def full_segments(n,row):
            if n >= 1:
                T = 2*3600/(arrivalInput['AN'+str(row)].value+arrivalInput['AO'+str(row)].value)
                if n >=2:
                    T += 2*3600/(arrivalInput['AO'+str(row)].value+arrivalInput['AP'+str(row)].value)
                    if n>=3:
                        T += 2*3600/(arrivalInput['AP'+str(row)].value + arrivalInput['AQ'+str(row)].value)
                        if n>=4:
                            T += 2*3600/(arrivalInput['AQ'+str(row)].value+arrivalInput['AR'+str(row)].value)
                            if n>=5:
                                T += 2*3600/(arrivalInput['AR'+str(row)].value + arrivalInput['AS'+str(row)].value)
                                if n>=6:
                                    T += 2*3600/(arrivalInput['AS'+str(row)].value+arrivalInput['AT'+str(row)].value)
                                    if n>=7:
                                        T += 2*3600/(arrivalInput['AT'+str(row)].value+arrivalInput['AU'+str(row)].value)
                                        if n>=8:
                                            T += 2*3600/(arrivalInput['AU'+str(row)].value+arrivalInput['AV'+str(row)].value)
                                            if n==9:
                                                T += 2*3600/(arrivalInput['AV'+str(row)].value + arrivalInput['AW'+str(row)].value)
                                            elif n>9:
                                                T += (n-9)*3600/arrivalInput['AW'+str(row)].value
                                                
                                                
            return T
        def fraction_of_segments(n,f,row):
            if n==1:
                S = f*(arrivalInput['AP'+str(row)].value - arrivalInput['AO'+str(row)].value) + arrivalInput['AO'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AO'+str(row)].value)
            elif n==2:
                S = f*(arrivalInput['AQ'+str(row)].value - arrivalInput['AP'+str(row)].value) + arrivalInput['AP'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AP'+str(row)].value)
            elif n==3:
                S = f*(arrivalInput['AR'+str(row)].value - arrivalInput['AQ'+str(row)].value) + arrivalInput['AQ'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AQ'+str(row)].value)
            elif n==4:
                S = f*(arrivalInput['AS'+str(row)].value - arrivalInput['AR'+str(row)].value) + arrivalInput['AR'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AR'+str(row)].value)
            elif n==5:
                S = f*(arrivalInput['AT'+str(row)].value - arrivalInput['AS'+str(row)].value) + arrivalInput['AS'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AS'+str(row)].value)
            elif n==6:
                S = f*(arrivalInput['AU'+str(row)].value - arrivalInput['AT'+str(row)].value) + arrivalInput['AT'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AT'+str(row)].value)
            elif n==7:
                S = f*(arrivalInput['AV'+str(row)].value - arrivalInput['AU'+str(row)].value) + arrivalInput['AU'+str(row)].value
                T = (f*2*3600)/(S+arrivalInput['AU'+str(row)].value)
            elif n==8: 
                S = f*(arrivalInput['AW'+str(row)].value - arrivalInput['AV'+str(row)].value) + arrivalInput['AV'+str(row)].value
                T = (f*3600)/arrivalInput['AW'+str(row)].value
            return T
#        if distance > 0:
        X = distance + runwayCalculations['F' + str(row)].value # Actual distance + ATCO var
        D = X - 0.5
        if D <0:
            T = (X*3600)/arrivalInput['AN'+str(row)].value
        elif D > 0:
            n = math.floor(D)            
            f = D - n
            T1 = full_segments(n,row)
            if (f != 0) and (n<=8):
                T2 = fraction_of_segments(n,f,row)
                T = T1 + T2 + (0.5*3600)/arrivalInput['AN'+str(row)].value
            else:
                T = T1  
               
        return T
            
    def time_to_distance_assumed_speed_profile_IAS(row, d_dme, T):
        c_dme = 4
        deceleration_difference= (runwayCalculations['H' + str(row)].value - runwayCalculations['K' + str(row)].value)
        if deceleration_difference > 20 :
            c_dme = deceleration_difference / 20    
        t1 = d_dme*3600/runwayCalculations['K' + str(row)].value
        t2 = (c_dme - d_dme)*3600*2/(runwayCalculations['K' + str(row)].value + runwayCalculations['H' + str(row)].value) + t1
        
        
        if T <= t1 :
            D = (runwayCalculations['K' + str(row)].value*T)/3600
        elif (T > t1) and (T < t2):
            t = T- t1
            S = (t*(runwayCalculations['H' + str(row)].value-runwayCalculations['K' + str(row)].value))/t2 + runwayCalculations['K' + str(row)].value
            D = (t*(runwayCalculations['K' + str(row)].value+S))/(2*3600) + d_dme
        elif T >= t2:
            D = c_dme + (T-t2)*runwayCalculations['H' + str(row)].value/3600
        return D
    
    def time_to_distance_assumed_speed_profile_GS(row, d_dme,T):
        c_dme = 4
        deceleration_difference= (runwayCalculations['H' + str(row)].value - runwayCalculations['G' + str(row)].value) - (runwayCalculations['K' + str(row)].value - runwayCalculations['I' + str(row)].value)
        if deceleration_difference > 20 :
            c_dme = deceleration_difference / 20    
        t1 = d_dme*3600/(runwayCalculations['K' + str(row)].value- runwayCalculations['I' + str(row)].value)
        t2 = (c_dme - d_dme)*3600*2/((runwayCalculations['K' + str(row)].value- runwayCalculations['I' + str(row)].value) + (runwayCalculations['H' + str(row)].value- runwayCalculations['G' + str(row)].value)) + t1
        
        
        if T <= t1 :
            D = ((runwayCalculations['K' + str(row)].value- runwayCalculations['I' + str(row)].value)*T)/3600
        elif (T > t1) and (T < t2):
            t = T- t1
            S = (t*((runwayCalculations['H' + str(row)].value- runwayCalculations['G' + str(row)].value)-(runwayCalculations['K' + str(row)].value- runwayCalculations['I' + str(row)].value)))/t2 + (runwayCalculations['K' + str(row)].value- runwayCalculations['I' + str(row)].value)
            D = (t*((runwayCalculations['K' + str(row)].value- runwayCalculations['I' + str(row)].value)+S))/(2*3600) + d_dme
            
            D = (T*((runwayCalculations['K' + str(row)].value- runwayCalculations['I' + str(row)].value)+(runwayCalculations['H' + str(row)].value- runwayCalculations['G' + str(row)].value))/(2*3600)) + d_dme
        elif T >= t2:
            D = c_dme + (T-t2)*(runwayCalculations['H' + str(row)].value- runwayCalculations['G' + str(row)].value)/3600
        return D

            
    # Function to pre-process the Arrival input file and make initial calculations
    def Arrival_Input_pre_process():       
        
        
       
        def write_actual_speed_profile_to_output(row, AC_type):
           
            df_row = random.randint(0,(dict_actual_speed_profiles[AC_type]['Aircraft_Type'].size-1)) 
            arrivalInput['AD' + str(row)].value = dict_actual_speed_profiles[AC_type]['GSPD_0_1DME'][df_row]
            arrivalInput['AE' + str(row)].value = dict_actual_speed_profiles[AC_type]['GSPD_1_2DME'][df_row]
            arrivalInput['AF' + str(row)].value = dict_actual_speed_profiles[AC_type]['GSPD_2_3DME'][df_row]
            arrivalInput['AG' + str(row)].value = dict_actual_speed_profiles[AC_type]['GSPD_3_4DME'][df_row]
            arrivalInput['AH' + str(row)].value = dict_actual_speed_profiles[AC_type]['GSPD_4_5DME'][df_row]
            arrivalInput['AI' + str(row)].value = dict_actual_speed_profiles[AC_type]['GSPD_5_6DME'][df_row]
            arrivalInput['AJ' + str(row)].value = dict_actual_speed_profiles[AC_type]['GSPD_6_7DME'][df_row]
            arrivalInput['AK' + str(row)].value = dict_actual_speed_profiles[AC_type]['GSPD_7_8DME'][df_row]
            arrivalInput['AL' + str(row)].value = dict_actual_speed_profiles[AC_type]['GSPD_8_9DME'][df_row]
            arrivalInput['AM' + str(row)].value = dict_actual_speed_profiles[AC_type]['GSPD_9_10DME'][df_row]
            
            arrivalInput['AN' + str(row)].value = dict_actual_speed_profiles[AC_type]['IAS_0_1DME'][df_row]
            arrivalInput['AO' + str(row)].value = dict_actual_speed_profiles[AC_type]['IAS_1_2DME'][df_row]
            arrivalInput['AP' + str(row)].value = dict_actual_speed_profiles[AC_type]['IAS_2_3DME'][df_row]
            arrivalInput['AQ' + str(row)].value = dict_actual_speed_profiles[AC_type]['IAS_3_4DME'][df_row]
            arrivalInput['AR' + str(row)].value = dict_actual_speed_profiles[AC_type]['IAS_4_5DME'][df_row]
            arrivalInput['AS' + str(row)].value = dict_actual_speed_profiles[AC_type]['IAS_5_6DME'][df_row]
            arrivalInput['AT' + str(row)].value = dict_actual_speed_profiles[AC_type]['IAS_6_7DME'][df_row]
            arrivalInput['AU' + str(row)].value = dict_actual_speed_profiles[AC_type]['IAS_7_8DME'][df_row]
            arrivalInput['AV' + str(row)].value = dict_actual_speed_profiles[AC_type]['IAS_8_9DME'][df_row]
            arrivalInput['AW' + str(row)].value = dict_actual_speed_profiles[AC_type]['IAS_9_10DME'][df_row]
        
        # Read in Arrival data from an excel workbook
        for row in range(2, arrivalInput.max_row + 1):
            
            if arrivalInput['A' + str(row)].value == None: ## Blank space in input
                print("Blank space detected in input file, terminating 'Read Input' here")
                break
            
            #If SIBTs are in time format convert them into seconds.
            if arrivalInput['C' + str(row)].value == None: ## SIBT seconds are not printed:                
                SIBT = arrivalInput['B' + str(row)].value                
                SIBT_sec = (SIBT.hour * 3600) + (SIBT.minute * 60) + SIBT.second                
                arrivalInput['C' + str(row)].value = SIBT_sec # Used from initial schedule
            # Write WTC in arrival Input | it will be used for AROT
            
            AC_type = arrivalInput['D' +str(row)].value
            arrivalInput['E' +str(row)].value = df_wake_WTC.at[AC_type,'WTC']
            
            # Write wake categories in runway calcs | used for wake separation:
            if RECatFLAG == True: 
                AC_type = arrivalInput['D' +str(row)].value                
                runwayCalculations['U' +str(row)].value = df_wake_RECAT.at[AC_type,'RECAT-EU'] #RECT-EU cat
            elif RECAT_PWS_FLAG == True:
                AC_type = arrivalInput['D' +str(row)].value 
                runwayCalculations['U' +str(row)].value = df_RECAT20.at[AC_type,'RECAT20']
            
            else:
                runwayCalculations['U' +str(row)].value = arrivalInput['E' +str(row)].value #WTC cat
            
                
            
            
            ################# ACTUAL _ SPEED _ PROFILE ####################
            
            AC_type = arrivalInput['D' + str(row)].value
            if AC_type in list(dict_actual_speed_profiles.keys()):
                write_actual_speed_profile_to_output(row, AC_type) 
                #print(AC_type)
            else:# search for which wake category the AC belongs to
                #print(AC_type, ' does not have actual speed profiles')
                for category in list(RECAT_categories.keys()):
                    if AC_type in RECAT_categories[category]:          
                        wake_cat = category
                        #print('AC :', AC_type, ' in wake_cat = ',wake_cat)
                        break
                list_AC_in_wake_cat = list(range(0,(len(RECAT_categories[wake_cat]))))
                
                while len(list_AC_in_wake_cat) !=0: #while there are AC in the list
                    index_picked = random.choice(list_AC_in_wake_cat) #pick a random one
                    #print('index_picked= ', index_picked)
                    AC_replacement = RECAT_categories[wake_cat][index_picked]
                    #print('AC_replacement = ',AC_replacement)
                    if AC_replacement in list(dict_actual_speed_profiles.keys()): #check if if has a speed profile
                        #print(AC_replacement, ' has speed profile')
                        write_actual_speed_profile_to_output(row, AC_replacement)
                        break #stop searching for
                    else: #if it doesn't have a speed profile, delete it from the list and check againg
                        #print(AC_replacement, ' does not have a speed profile')
                        list_AC_in_wake_cat.remove(index_picked)
                        
 
                
            ######################## INTERMEDIATE CALCULATIONS ###########################
        
            # Arrival ID
            runwayCalculations['A' + str(row)].value = arrivalInput['A' + str(row)].value
            ##################### TAXI-IN - normal distribution ###############
           
            Arrival_Taxiin_mean = arrivalInput['I' + str(row)].value
            Arrival_Taxiin_SD = arrivalInput['J' + str(row)].value
            #Taxiinlookup = arrivalInput['M' + str(row)].value
#            if taxi_outliers == False:
            tempTaxiIn = random.normalvariate(Arrival_Taxiin_mean, Arrival_Taxiin_SD)
#            else:
#                tempTaxiIn = Taxiinlookup   
            runwayCalculations['B' + str(row)].value = round(tempTaxiIn, 0)

            ####################### AROT - from lookup ########################
            
            if arrivalInput['E'+str(row)].value=="H":
                random_arot = np.random.choice(df_AROT_H['AROT_H'], 1)[0] 
                runwayCalculations['C' + str(row)].value = random_arot
            elif arrivalInput['E'+str(row)].value=="M":
                random_arot = np.random.choice(df_AROT_M['AROT_M'], 1)[0] 
                runwayCalculations['C' + str(row)].value = random_arot
            elif arrivalInput['E'+str(row)].value=="L":
                random_arot = np.random.choice(df_AROT_L['AROT_L'], 1)[0] 
                runwayCalculations['C' + str(row)].value= random_arot   
            elif arrivalInput['E'+str(row)].value=="UM":
                random_arot = np.random.choice(df_AROT_UM['AROT_UM'], 1)[0] 
                runwayCalculations['C' + str(row)].value = random_arot
            elif arrivalInput['E'+str(row)].value=="J":
                random_arot = np.random.choice(df_AROT_J['AROT_J'], 1)[0] 
                runwayCalculations['C' + str(row)].value = random_arot
            elif arrivalInput['E'+str(row)].value=="S":
                random_arot = np.random.choice(df_AROT_S['AROT_S'], 1)[0] 
                runwayCalculations['C' + str(row)].value = random_arot
            
            ##################### ADA - normal distribution ###############
            ADA_mean = arrivalInput['O' + str(row)].value
            ADA_sd = arrivalInput['P' + str(row)].value
            actualADA = random.normalvariate(ADA_mean, ADA_sd)
            runwayCalculations['D' + str(row)].value  = int(actualADA)
            
            ##################### ADDA - normal distribution ###############
            ADDA_mean = arrivalInput['Q' + str(row)].value
            ADDA_sd = arrivalInput['R' + str(row)].value
            actualADDA = random.normalvariate(ADDA_mean, ADDA_sd)
            runwayCalculations['E' + str(row)].value  = int(actualADDA)
            
            ################ ATCO variability - normal distribution ###########
            ATCO_mean = arrivalInput['S' + str(row)].value
            ATCO_sd = arrivalInput['T' + str(row)].value
            actualATCO = random.normalvariate(ATCO_mean, ATCO_sd)
            runwayCalculations['F' + str(row)].value  = int(actualATCO)          
            
            ################## ASSUMED_SPEED_PROFILE ######################
            
            # --- WIND 1  ---#
            WIND1_mean = arrivalInput['V' + str(row)].value
            WIND1_sd = arrivalInput['W' + str(row)].value
            actualWIND1 = random.normalvariate(WIND1_mean, WIND1_sd)
            runwayCalculations['G' + str(row)].value  = actualWIND1
            # --- SPEED 1  ---#
            SPEED1_mean = arrivalInput['X' + str(row)].value
            SPEED1_sd = arrivalInput['Y' + str(row)].value
            actualSPEED1 = random.normalvariate(SPEED1_mean, SPEED1_sd)
            runwayCalculations['H' + str(row)].value  = actualSPEED1
            # --- WIND 2  ---#
            WIND2_mean = arrivalInput['Z' + str(row)].value
            WIND2_sd = arrivalInput['AA' + str(row)].value
            actualWIND2 = random.normalvariate(WIND2_mean, WIND2_sd)
            runwayCalculations['I' + str(row)].value  = actualWIND2
            # --- SPEED 2  ---#
            SPEED2_mean = arrivalInput['AB' + str(row)].value
            SPEED2_sd = arrivalInput['AC' + str(row)].value
            actualSPEED2 = random.normalvariate(SPEED2_mean, SPEED2_sd)
            runwayCalculations['J' + str(row)].value  = actualSPEED2
            # --- VTGT  ---#
            if (actualWIND2 < 5) or (actualWIND2>20):
                wind_adjustment = 5
            else:
                wind_adjustment = actualWIND2*0.5
            V_TGT = actualSPEED2 + wind_adjustment
            runwayCalculations['K' + str(row)].value = V_TGT
            
            #------ SAE -------#
            runwayCalculations['L' + str(row)].value = arrivalInput['C' + str(row)].value - STT - 200 # SAE = SIBT - Standard Taxi Time - App length*
            #---- Predicted Landing Time --------#
            runwayCalculations['M' + str(row)].value = runwayCalculations['L' + str(row)].value + 60 # PLT = SAE + MRS*
            
 
            
            ############################ MAX CONSTRAINT CALCS ##################################

            
            def min_wake_separation_arrs(key_of_nextArrival): # delievered at THR ACTUAL SPEED PROFILE
                minWakeSepArr = 0 # Initialise local variable (reset on each iteration)
                
                if RECAT_PWS_FLAG == True: # analyse by ac type
                    previousArrival = arrivalInput['D' +str(key_of_nextArrival-1)].value
                    currentArrival = arrivalInput['D' +str(key_of_nextArrival)].value
                    previousArrivalWake = runwayCalculations['U' +str(key_of_nextArrival-1)].value #20cat classification
                    currentArrivalWake = runwayCalculations['U' +str(key_of_nextArrival)].value #20cat classification
                    
                    if key_of_nextArrival == 2: #FirstArrival
                       minWakeSepArr = 0
                    else:
                        if (currentArrival in df_RECAT_PWS) and (previousArrival in df_RECAT_PWS):
                            wakeDistance = df_RECAT_PWS.at[currentArrival,previousArrival]
                            if wakeDistance==0:
                                wakeDistance = df_RECAT20_separation.at[previousArrivalWake,currentArrivalWake]
                        else: # if the pair is not in the 96x96 table, search in the 20cat
                            wakeDistance = df_RECAT20_separation.at[previousArrivalWake,currentArrivalWake]
                        
                        if wakeDistance == 0:
                            minWakeSepArr =0
                        else:    
                            if distance_based_FLAG == True:     
                                if  WAKE_4dme_FLAG== True:
                                    Total_time_follow = int(DBS_actual_speed_profile((wakeDistance+4),key_of_nextArrival))
                                    Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(key_of_nextArrival-1)))
                                    minWakeSepArr = Total_time_follow - Time_lead_4dme_to_thr
                                elif WAKE_thr_FLAG == True:
                                    minWakeSepArr = int(DBS_actual_speed_profile(wakeDistance,key_of_nextArrival))  #time
                                else: # the same as the previous one but it's the default condition                                    
                                    minWakeSepArr = int(DBS_actual_speed_profile(wakeDistance,key_of_nextArrival))  #time
                                    
                            elif time_based_FLAG == True:                            
                                time1 = distance_to_time_assumed_speed_profile_IAS(key_of_nextArrival, d_dme, wakeDistance) #time
                                distance = time_to_distance_assumed_speed_profile_GS(key_of_nextArrival, d_dme,int(time1))#distance
                                if WAKE_4dme_FLAG == True:
                                    Total_time_follow = int(DBS_actual_speed_profile((distance+4),key_of_nextArrival))
                                    Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(key_of_nextArrival-1)))
                                    minWakeSepArr = Total_time_follow - Time_lead_4dme_to_thr
                                    
                                elif WAKE_thr_FLAG == True:
                                    minWakeSepArr = int(DBS_actual_speed_profile(distance,key_of_nextArrival)) #time
                                else:
                                    minWakeSepArr = int(DBS_actual_speed_profile(distance,key_of_nextArrival)) #time
                                    
                                    
                                    
                else: #analyze by wake
                    previousArrivalWake = runwayCalculations['U' +str(key_of_nextArrival-1)].value
                    currentArrivalWake = runwayCalculations['U' +str(key_of_nextArrival)].value
                    if key_of_nextArrival == 2: #FirstArrival
                       minWakeSepArr = 0
                    else: #next arrivals
                          
                        if RECatFLAG == True: # delievered to THR
                            wakeDistance = df_RECAT_EU_separation.at[previousArrivalWake,currentArrivalWake] 
                        else: #UK cat *********** should be delievered to 4dme
                            wakeDistance = df_WTC_separation.at[previousArrivalWake,currentArrivalWake] #distance   
                         
                        if wakeDistance == 0:
                            minWakeSepArr =0
                        else:    
                            if distance_based_FLAG == True:                   
                                if WAKE_4dme_FLAG == True:
                                    Total_time_follow = int(DBS_actual_speed_profile((wakeDistance+4),key_of_nextArrival))
                                    Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(key_of_nextArrival-1)))
                                    minWakeSepArr = Total_time_follow - Time_lead_4dme_to_thr
                                elif WAKE_thr_FLAG == True:
                                    minWakeSepArr = int(DBS_actual_speed_profile(wakeDistance,key_of_nextArrival))  #time
                                else: # the same as the previous one but it's the default condition                                    
                                    minWakeSepArr = int(DBS_actual_speed_profile(wakeDistance,key_of_nextArrival))  #time
                                
                            elif time_based_FLAG == True:
                                time1 = distance_to_time_assumed_speed_profile_IAS(key_of_nextArrival, d_dme, wakeDistance) #time
                                distance = time_to_distance_assumed_speed_profile_GS(row, d_dme,int(time1))#distance
                                if WAKE_4dme_FLAG == True:
                                    Total_time_follow = int(DBS_actual_speed_profile((distance+4),key_of_nextArrival))
                                    Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(key_of_nextArrival-1)))
                                    minWakeSepArr = Total_time_follow - Time_lead_4dme_to_thr
                                    
                                elif WAKE_thr_FLAG == True:
                                    minWakeSepArr = int(DBS_actual_speed_profile(distance,key_of_nextArrival)) #time
                                else:
                                    minWakeSepArr = int(DBS_actual_speed_profile(distance,key_of_nextArrival)) #time
                                
                       
                return(minWakeSepArr)   
            
            
            
            
            runwayCalculations['P' + str(1)].value = "WAKE SEPARATION"
            runwayCalculations['P' + str(row)].value = int(min_wake_separation_arrs(row)) #always Distance-based
            
            
            runwayCalculations['Q' + str(1)].value = "MRS" 
            MRSArr = 0
            if (MRS_4dme_FLAG == True) and (row>2):
                Total_time_follow = int(DBS_actual_speed_profile((min_radar_separation_distance+4),row))
                Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(row-1)))
                MRSArr = Total_time_follow - Time_lead_4dme_to_thr
            elif MRS_thr_FLAG == True:
                MRSArr = int(DBS_actual_speed_profile(min_radar_separation_distance,row))  #time
            else: # the same as the previous one but it's the default condition                                    
                MRSArr = int(DBS_actual_speed_profile(min_radar_separation_distance,row))  #time
            runwayCalculations['Q' + str(row)].value = MRSArr
                
            def max_constraint_generator(row):
                wake_constraint = runwayCalculations['P' + str(row)].value
                MRS_constraint = runwayCalculations['Q' + str(row)].value                
                spFLAG = "None"
                max_constraint = 0
                if row == 2 :                   
                    max_constraint = max(wake_constraint,MRS_constraint)
                    spFLAG = "First Arrival"
                    
                else: #not he first arrival
                    AROT_constraint = runwayCalculations['C' + str(row-1)].value + 5
                    
                    if (df_dep.empty == True): #no departures
                        max_constraint = int(max(wake_constraint, MRS_constraint ,AROT_constraint))
                        if max_constraint == wake_constraint:
                            spFLAG = "WAKE"
                        elif max_constraint == MRS_constraint:
                            spFLAG = "MRS"
                        else:
                            spFLAG = "AROT"
                            
                    elif (df_dep.empty == False) and (df_arr.empty == False): #there are both arrivals and departures scheduled
                        if time_based_FLAG == True:
                            
                            max_constraint = int(max(wake_constraint, MRS_constraint, AROT_constraint))
                            if max_constraint == wake_constraint:
                                spFLAG = "WAKE"
                            elif max_constraint == MRS_constraint:
                                spFLAG = "MRS"
                            else:
                                spFLAG = "AROT"
                                
                        elif distance_based_FLAG == True:
                            if (arrivalInput['U' + str(row)].value) == "ADDA" :
                                ADDA_distance = runwayCalculations['E' + str(row)].value
                                if (ADDA_4dme_FLAG == True) and (row>2):
                                    Total_time_follow = int(DBS_actual_speed_profile((ADDA_distance+4),row))
                                    Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(row-1)))
                                    ADDA_separation = Total_time_follow - Time_lead_4dme_to_thr
                                elif ADDA_thr_FLAG == True:
                                    ADDA_separation = int(DBS_actual_speed_profile(ADDA_distance,row))  #time
                                else: # the same as the previous one but it's the default condition                                    
                                    ADDA_separation = int(DBS_actual_speed_profile(ADDA_distance,row))  #time
                                
                                
                                #ADDA_separation = int(DBS_actual_speed_profile(ADDA_distance,row))
                                
                                max_constraint = int(max(wake_constraint, ADDA_separation, MRS_constraint,AROT_constraint))
                                if max_constraint ==wake_constraint:
                                    spFLAG = "WAKE"
                                elif max_constraint == ADDA_separation:
                                    spFLAG = "ADDA"
                                elif max_constraint == MRS_constraint:
                                    spFLAG = "MRS"
                                else:
                                    spFLAG = "AROT"
                            elif (arrivalInput['U' + str(row)].value) == "ADA" :
                                
                                ADA_distance = runwayCalculations['D' + str(row)].value
                                
                                if (ADA_4dme_FLAG == True) and (row>2):
                                    Total_time_follow = int(DBS_actual_speed_profile((ADA_distance+4),row))
                                    Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(row-1)))
                                    ADA_separation = Total_time_follow - Time_lead_4dme_to_thr
                                elif ADA_thr_FLAG == True:
                                    ADA_separation = int(DBS_actual_speed_profile(ADA_distance,row))  #time
                                else: # the same as the previous one but it's the default condition                                    
                                    ADA_separation = int(DBS_actual_speed_profile(ADA_distance,row))  #time
                                
                                                                
                                max_constraint = int(max(wake_constraint, ADA_separation, MRS_constraint,AROT_constraint))
                                if max_constraint == wake_constraint:
                                    spFLAG = "WAKE"
                                elif max_constraint == ADA_separation:
                                    spFLAG = "ADA"
                                elif max_constraint == MRS_constraint:
                                    spFLAG = "MRS"
                                else:
                                    spFLAG = "AROT"
                            else:
                                max_constraint = int(max(wake_constraint, MRS_constraint,AROT_constraint))
                                if max_constraint == wake_constraint:
                                    spFLAG = "WAKE"                                
                                elif max_constraint == MRS_constraint:
                                    spFLAG = "MRS"
                                else:
                                    spFLAG = "AROT"
                                 
                return{'a' : max_constraint,
                       'b' : spFLAG}
            
            ################################### MAX CONSTRAINT PRINT ###################################
            
            runwayCalculations['N' + str(row)].value = max_constraint_generator(row)['a']
            runwayCalculations['O' + str(row)].value = max_constraint_generator(row)['b']
            
            
        return
    
    # Function to pre-process the Departure input file and make initial calculations
    def Departure_Input_pre_process():
        
        #Initialise Departure input
        
        
        # Read in Departure data from an excel workbook
        for row in range(2, departureInput.max_row + 1):
            if departureInput['A' + str(row)].value == None:  # NO (more) DEPARTURES
                print("Blank space detected in input file, terminating 'Read Input' here")
                break
            if departureInput['C' + str(row)].value == None: #SOBT are in time-format
                SOBT = departureInput['B' + str(row)].value
                SOBT_sec = (SOBT.hour * 3600) + (SOBT.minute * 60) + SOBT.second                
                departureInput['C' + str(row)].value = SOBT_sec # Used from initial schedule
            
            ################### INTERMEDIATE CALCULATIONS #####################
            
            #-----Departure WAKE category-----#
            AC_type = departureInput['F' +str(row)].value
            departureInput['H' +str(row)].value = df_wake_WTC.at[AC_type,'WTC']
            
            #----Departure ID -----#
            runwayCalculations['R' + str(row)].value = departureInput['A' + str(row)].value
            
            #------TAXI-OUT------#
           
            Departure_Taxiout_mean = departureInput['K' + str(row)].value
            Departure_Taxiout_SD = departureInput['L' + str(row)].value
            actualTAXIOUT = random.normalvariate(Departure_Taxiout_mean, Departure_Taxiout_SD)
            runwayCalculations['S' + str(row)].value = round(actualTAXIOUT,0)
            
            #------ DROT-------#
            if departureInput['H'+ str(row)].value=="H":
                random_drot = np.random.choice(df_DROT_H['DROT_H'], 1)[0] 
                runwayCalculations['T' + str(row)].value = random_drot
            elif departureInput['H'+str(row)].value=="M":
                random_drot = np.random.choice(df_DROT_M['DROT_M'], 1)[0] 
                runwayCalculations['T' + str(row)].value = random_drot
            elif departureInput['H'+str(row)].value=="L":
                random_drot = np.random.choice(df_DROT_L['DROT_L'], 1)[0] 
                runwayCalculations['T' + str(row)].value = random_drot   
            elif departureInput['H'+str(row)].value=="UM":
                random_drot = np.random.choice(df_DROT_UM['DROT_UM'], 1)[0] 
                runwayCalculations['T' + str(row)].value = random_drot
            elif departureInput['H'+str(row)].value=="J":
                random_drot = np.random.choice(df_DROT_J['DROT_J'], 1)[0] 
                runwayCalculations['T' + str(row)].value = random_drot
            elif departureInput['H'+str(row)].value=="S":
                random_drot = np.random.choice(df_DROT_S['DROT_S'], 1)[0] 
                runwayCalculations['T' + str(row)].value = random_drot
            
        return
    
   
    #####################################################################
    #                EXECUTE PRE-PROCESSING FUNCTIONS                   #
    #####################################################################
    
    set_Output_Excel_headers()
    Arrival_Input_pre_process()
    arrivalInputmaxRow = arrivalInput.max_row
    Departure_Input_pre_process()
    print("Input file successfully read")
    
#===== SID separation ====#
    
    index_key_of_nextDeparture_Q1 = 0
    index_key_of_nextDeparture_Q2 = 0
    pair_SID = 0
    #SIDgroup_separation = "(2,4)(3,4)"
    SIDgroup_separation = SIDgroup_separation.replace("("," ")
    SIDgroup_separation = SIDgroup_separation.replace(")","")
    SIDgroup_separation = SIDgroup_separation.replace(","," ")      
    #print("SIDgroup_separation :  ",SIDgroup_separation)
    list_SIDgroup = SIDgroup_separation.split(" ")
    del list_SIDgroup[0]   
    SIDgroup_list1 = [int(x) for x in list_SIDgroup]
    SIDgroup_list2 = SIDgroup_list1[::-1]
    SIDgroup_list = SIDgroup_list1 + SIDgroup_list2
    chunks2 = [SIDgroup_list[x:x+2] for x in range(0, len(SIDgroup_list), 2)]
    
#===== SID queue ====#
    
    #SID_queue_assign = "1 3 | 2 4"
    Q1_SID = []
    Q2_SID = []
    Q3_SID = []
    Q4_SID = []
    Q5_SID = []
    Q6_SID = []
    Q7_SID = []
    Q8_SID = []
    SID_queue_assign = SID_queue_assign.split(' ')
    index_queue_separator = SID_queue_assign.index('|')
    SID_queue_assign.remove('|')
    chunks = [SID_queue_assign[x:x+index_queue_separator] for x in range(0, len(SID_queue_assign), index_queue_separator)]
    if len(chunks) >= 1 :
        Q1_SID =[int(x) for x in chunks[0]]
        if len(chunks) >= 2 :
            Q2_SID =[int(x) for x in chunks[1]]
            if len(chunks) >= 4 :
                Q3_SID =[int(x) for x in chunks[2]]
                Q4_SID =[int(x) for x in chunks[3]]
                if len(chunks) >= 8 :
                    Q5_SID =[int(x) for x in chunks[4]]
                    Q6_SID =[int(x) for x in chunks[5]]
                    Q7_SID =[int(x) for x in chunks[6]]
                    Q8_SID =[int(x) for x in chunks[7]]
                    
#===TIME limits====#
    
    if arrivalInput['A' + str(2)].value==None: #no arrivals:
        Start_time = departureInput['C' +str(2)].value - 3000
    elif departureInput['A' +str(2)].value == None: #no departures
        Start_time = arrivalInput['C' + str(2)].value - 3000
    else:
        Start_time = min(arrivalInput['C' + str(2)].value,departureInput['C' +str(2)].value) - 3000
    
    if arrivalInput['A' + str(2)].value==None: #no arrivals:
        End_time = departureInput['C' +str(max_DEPARTURE-1)].value + 10000
    elif departureInput['A' +str(2)].value == None: #no departures
        End_time = arrivalInput['C' + str(max_ARRIVAL-1)].value + 10000
    else:
        End_time = min(arrivalInput['C' + str(max_ARRIVAL-1)].value,departureInput['C' +str(max_DEPARTURE-1)].value) + 10000
        
    Current_time = Start_time
#--------------------------- Movement Functions---------------------------------#


 
######################## DEPARTURES #######################################
    
     
    def update_Departure_Delays(Current_time):
        if len(DepSTANDqueue)>0:
            for AC in list(DepSTANDqueue.keys()):
                DepSTANDqueue_Delay = Current_time - DepSTANDqueue[AC][1]
                DepSTANDqueue[AC][4] = DepSTANDqueue_Delay
        if len(TAXIhold)>0:
            for AC in list(TAXIhold.keys()):           
                TAXIhold_Delay = Current_time - TAXIhold[AC][7]
                TAXIhold[AC][7] = TAXIhold_Delay
        if len(RWYqueue1)>0:
            for AC in list(RWYqueue1.keys()):
                RWYqueue1_delay = Current_time - RWYqueue1[AC][9] #Current_time - RWYqueue entry_time
                RWYqueue1[AC][10] = RWYqueue1_delay
        if len(RWYqueue2)>0:
            for AC in list(RWYqueue2.keys()):
                RWYqueue2_delay = Current_time - RWYqueue2[AC][9] #Current_time - RWYqueue entry_time
                RWYqueue2[AC][10] = RWYqueue2_delay

        
    def SOBTlookup(Current_time, SOBTrow):
        if SOBTrow<max_DEPARTURE:
            if Current_time >= departureInput['C' + str(SOBTrow)].value :# Current time = SOBT
                DepSTANDqueue[SOBTrow]=[departureInput['A' + str(SOBTrow)].value,departureInput['C' + str(SOBTrow)].value,runwayCalculations['T' + str(SOBTrow)].value,departureInput['I' + str(SOBTrow)].value,0]  
                SOBTrow += 1
        return(SOBTrow)
        
    def TAXIqueue_update(Current_time):
#        if ((len(TAXIqueue) + len(ARRIVALqueue)+ len(TAXIhold))<15) and len(DepSTANDqueue)> 0: # if there are less than 15 AC moving on the TAXIway and there's something in DepSTANDqueue
        #check who has to go first
        first_in_line_DepSTANDqueue = min(list(DepSTANDqueue.keys()))
    
        DepSTANDqueue[first_in_line_DepSTANDqueue].append(Current_time) #TAXIqueue entry time
        DepSTANDqueue[first_in_line_DepSTANDqueue].append(runwayCalculations['S' + str(first_in_line_DepSTANDqueue)].value) #TAXI-out
        
        
        #ADD first_in_line_DepSTANDqueue to TAXIqueue:
        TAXIqueue[first_in_line_DepSTANDqueue]=DepSTANDqueue[first_in_line_DepSTANDqueue]
        del DepSTANDqueue[first_in_line_DepSTANDqueue] 
    
    def first_in_line_TAXIqueue_func(End_time):
        first_in_line_TAXIqueue = 0
        min_TAXIqueue_out = End_time
        for AC in list(TAXIqueue.keys()):
            TAXIqueue_out = TAXIqueue[AC][5]+TAXIqueue[AC][6]
            if TAXIqueue_out<min_TAXIqueue_out:
                min_TAXIqueue_out = TAXIqueue_out
                first_in_line_TAXIqueue = AC
        return(first_in_line_TAXIqueue)
    
    def TAXIhold_update(Current_time,End_time):
        if len(TAXIqueue)>0:
            first_in_line_TAXIqueue = first_in_line_TAXIqueue_func(End_time)
            if Current_time >= (TAXIqueue[first_in_line_TAXIqueue][5] + TAXIqueue[first_in_line_TAXIqueue][6]): # current_time = TAXIqueue_entry_time + Taxi-out
                TAXIqueue[first_in_line_TAXIqueue].append(Current_time) #TAXIhold entry time
                TAXIqueue[first_in_line_TAXIqueue].append(0) #TAXIhold delay
                
                #ADD first_in_line_TAXIqueue to TAXIhold
                TAXIhold[first_in_line_TAXIqueue] = TAXIqueue[first_in_line_TAXIqueue]
                del TAXIqueue[first_in_line_TAXIqueue]
    
    
    def transfer_to_2x4_RWYqueues(first_in_line_TAXIhold,Current_time):
        if TAXIhold[first_in_line_TAXIhold][3] in Q1_SID: # First check if SID group belongs to RWYqueue1
            if len(RWYqueue1) < maxRWYqueue1Length: # if there is space in RWYqueue1 add A/C to the queue
                TAXIhold[first_in_line_TAXIhold].append(Current_time) #RWYqueue1 entry time
                TAXIhold[first_in_line_TAXIhold].append(0) # RWYqueue1 Delay
                TAXIhold[first_in_line_TAXIhold].append(1) #RWYqueue used 
                
                RWYqueue1[first_in_line_TAXIhold] = TAXIhold[first_in_line_TAXIhold]
                del TAXIhold[first_in_line_TAXIhold]
               
        elif TAXIhold[first_in_line_TAXIhold][3] in Q2_SID: # First check if SID group belongs to RWYqueue2
            if len(RWYqueue2) < maxRWYqueue2Length: # if there is space in RWYqueue1 add A/C to the queue
                TAXIhold[first_in_line_TAXIhold].append(Current_time) #RWYqueue2 entry time
                TAXIhold[first_in_line_TAXIhold].append(0) # RWYqueue2 Delay
                TAXIhold[first_in_line_TAXIhold].append(2) #RWYqueue used 
                 
                RWYqueue2[first_in_line_TAXIhold] = TAXIhold[first_in_line_TAXIhold]  
                del TAXIhold[first_in_line_TAXIhold]

    def RWYqueues_update(Current_time):
        previous_in_line = 0
        first_in_line_TAXIhold = 0
        while (len(TAXIhold)>0 and (len(RWYqueue1)+len(RWYqueue2))<8): # while there is something in TAXIhold and there's space in RWY queues
            previous_in_line = first_in_line_TAXIhold
            first_in_line_TAXIhold = min(list(TAXIhold.keys()))
            
            if (first_in_line_TAXIhold!=0) and (previous_in_line!= first_in_line_TAXIhold):
                transfer_to_2x4_RWYqueues(first_in_line_TAXIhold,Current_time)                
            else:
                break
 
    def first_in_line_RWYqueue_funct(DepOutput,End_time):
        first_in_line_RWYqueue = 0
        currentRWYqueue = 0
        
        #ONLY FOR 2x4
        if DepOutput==2: #for the first departure check in which queue is the first departure:
            min_entry_time1 = End_time
            min_entry_time2 = End_time
            first_in_line_RWYqueue1=0
            first_in_line_RWYqueue2=0
            if len(RWYqueue1)>0:
                for AC in list(RWYqueue1.keys()):
                    if RWYqueue1[AC][9]<min_entry_time1:
                        min_entry_time1 = RWYqueue1[AC][9]
                        first_in_line_RWYqueue1 = AC
            if len(RWYqueue2)>0:
                for AC in list(RWYqueue2.keys()):
                    if RWYqueue2[AC][9]<min_entry_time2:
                        min_entry_time2 = RWYqueue2[AC][9]
                        first_in_line_RWYqueue2 = AC
            if min_entry_time1<min_entry_time2:
                first_in_line_RWYqueue = first_in_line_RWYqueue1
                currentRWYqueue = 1
            else:
                first_in_line_RWYqueue = first_in_line_RWYqueue2
                currentRWYqueue = 2
               
        elif DepOutput!=2 :
            if departureOutput['U' + str(DepOutput-1)].value == 1: #If previous departure started from queue 1
                
                # use RWYqueue2
                if len(RWYqueue2)>0: #There is smth in the queue
                    
                    min_entry_time = End_time
                    for AC in list(RWYqueue2.keys()):
                        if RWYqueue2[AC][9]<min_entry_time:
                            min_entry_time = RWYqueue2[AC][9]
                            first_in_line_RWYqueue = AC
                    
                    currentRWYqueue = 2
                else: #there is nobody in RWYqueue2
                    #use RWYqueue1 again
                    min_entry_time = End_time
                    for AC in list(RWYqueue1.keys()):
                        if RWYqueue1[AC][9]<min_entry_time:
                            min_entry_time = RWYqueue1[AC][9]
                            first_in_line_RWYqueue = AC
                    currentRWYqueue = 1
            elif departureOutput['U' + str(DepOutput-1)].value == 2: #If previous departure started from queue 2
                
                # use RWYqueue1
                if len(RWYqueue1)>0: #There is smth in the queue
                    min_entry_time = End_time
                    for AC in list(RWYqueue1.keys()):
                        if RWYqueue1[AC][9]<min_entry_time:
                            min_entry_time = RWYqueue1[AC][9]
                            first_in_line_RWYqueue = AC
                    currentRWYqueue = 1
                    
                else: #there is nobody in RWYqueue2
                    #use RWYqueue2 again
                    min_entry_time = End_time
                    for AC in list(RWYqueue2.keys()):
                        if RWYqueue2[AC][9]<min_entry_time:
                            min_entry_time = RWYqueue2[AC][9]
                            first_in_line_RWYqueue = AC
                    currentRWYqueue = 2
        return(first_in_line_RWYqueue, currentRWYqueue)
        
        
    def second_in_line_RWYqueues(previousRWYqueue,End_time): #used for target ADDA time
        min_entry_time = End_time
        second_in_line_RWYqueue = 0
        currentRWYqueue = 0
        if previousRWYqueue == 1: #now use queue2
            if len(RWYqueue2)>0:
                currentRWYqueue = 2
                for AC in list(RWYqueue2.keys()):
                    if RWYqueue2[AC][9]<min_entry_time:
                        min_entry_time = RWYqueue2[AC][9]
                        second_in_line_RWYqueue = AC
            else:
                currentRWYqueue = 1
                for AC in list(RWYqueue1.keys()):
                    if RWYqueue1[AC][9]<min_entry_time:
                        min_entry_time = RWYqueue1[AC][9]
                        second_in_line_RWYqueue = AC
                        
        elif previousRWYqueue == 2: #now use 1
            if len(RWYqueue1)>0:
                currentRWYqueue = 1
                for AC in list(RWYqueue1.keys()):
                    if RWYqueue1[AC][9]<min_entry_time:
                        min_entry_time = RWYqueue1[AC][9]
                        second_in_line_RWYqueue = AC
               
            else:
                currentRWYqueue = 2
                for AC in list(RWYqueue2.keys()):
                    if RWYqueue2[AC][9]<min_entry_time:
                        min_entry_time = RWYqueue2[AC][9]
                        second_in_line_RWYqueue = AC
        return(second_in_line_RWYqueue,currentRWYqueue)
        
       
    def dep_Wake_separation(first_in_line_RWYqueue, DepOutput):
        minWakeSep = 0 # Initialise local variable (reset on each iteration)
        if DepOutput == 2: #first departure:
            minWakeSep = 0
        else:
            previousDepartureWake = departureOutput['E' + str(DepOutput-1)].value 
            currentDepartureWake = departureInput['H' + str(first_in_line_RWYqueue)].value
                                  
            if previousDepartureWake == "J":
                if currentDepartureWake == "J":
                    minWakeSep = 0
                elif currentDepartureWake == "H":
                    minWakeSep = J_H_d
                elif (currentDepartureWake == "UM") or (currentDepartureWake == "M"):
                    minWakeSep = J_M_d
                elif (currentDepartureWake == "S") or (currentDepartureWake == "L"):
                    minWakeSep = J_L_d
                else:
                    print("[J-] Wake Category other than normal detected - check Input file")
    
            elif previousDepartureWake == "H":
                if currentDepartureWake == "J":
                    minWakeSep = 0
                elif currentDepartureWake == "H":
                    minWakeSep = H_H_d
                elif (currentDepartureWake == "UM") or (currentDepartureWake == "M"):
                    minWakeSep = H_M_d
                elif (currentDepartureWake == "S") or (currentDepartureWake == "L"):
                    minWakeSep = H_M_d
                else:
                    print("[H-] Wake Category other than normal detected - check Input file")
            
            elif (previousDepartureWake == "UM") or (previousDepartureWake == "M"):
                if currentDepartureWake == "L":
                    minWakeSep = M_L_d
                else:
                    minWakeSep = 0
                                
            elif (previousDepartureWake == "S") or (previousDepartureWake == "S"):
                if currentDepartureWake == "L":
                    minWakeSep = 0
                  
            else:
                minWakeSep = 0
        
        return(minWakeSep)
        
    def dep_SID_separation(first_in_line_RWYqueue, DepOutput):
        minSIDsep = 0 # Initialise local variable (reset on each iteration)
        if DepOutput == 2: #first departure:
            minSIDsep = 0
        else:
            # Compares SID groups between the previous and current A/C - then sets 'minSIDsep' variable as either altSID or sameSID
            previousDepartureSID = departureOutput['F' + str(DepOutput-1)].value
            nextDepartureSID = departureInput['I' + str(first_in_line_RWYqueue)].value
                    
            if nextDepartureSID == previousDepartureSID: #IF the next departure SID is tha same as the previous departure SID => maximum separation
                minSIDsep = minDep_sameSID 
            # If they are not equal, check if the SID group has some more separation rules
            elif nextDepartureSID != previousDepartureSID: 
                minSIDsep = minDep_altSID
                for item in chunks2:
                    if nextDepartureSID == item[0] and previousDepartureSID == item[1]:
                       #if previousDepartureSID == item[1]: # IF the previous departure SID matches the partner, apply maximum separation
                        minSIDsep = minDep_sameSID
        
        
        return (minSIDsep)
    
    

    def departure_separation(first_in_line_RWYqueue,DepOutput):
        minDeptime = 0
        minDepLabel = ""
        #WAKE
        minWakeSep = dep_Wake_separation(first_in_line_RWYqueue, DepOutput)
        #SID
        minSIDsep = dep_SID_separation(first_in_line_RWYqueue, DepOutput)
        #compare the two and take the largest constraint
        if minSIDsep>minWakeSep:
            minDeptime = minSIDsep
            minDepLabel = "SID"
        else:
            minDeptime = minWakeSep
            minDepLabel = "WAKE"
        return(minDeptime,minDepLabel)
        
                

            
    def Dep_TAKE_OFF(Current_time, DepOutput, currentGap,End_time,seqRow): 
        #if (len(RWYqueue1) != 0) or (len(RWYqueue2)!=0): #there is something in the queues:
        #print('Something in the RWYqueues') 
        
        first_in_line_RWYqueue, currentRWYqueue = first_in_line_RWYqueue_funct(DepOutput, End_time) 
        if first_in_line_RWYqueue !=0: # there's someone in line
            minDepTime,minDepLabel = departure_separation(first_in_line_RWYqueue,DepOutput)
            
            if DepOutput == 2: # First departure, no wake/sid constraints
                if (currentGap > n): 
                    #TAKE-OFF
                    departureOutput['B' + str(DepOutput)].value = int(Current_time/3600) # Dep HOUR
                    departureOutput['C' + str(DepOutput)].value = Current_time # Departure RWY Entry
                    
                    if currentRWYqueue == 1:
                        departureOutput['A' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][0] # AC ID                        
                        departureOutput['D' + str(DepOutput)].value = departureOutput['C' + str(DepOutput)].value + RWYqueue1[first_in_line_RWYqueue][2] # Dep RWY EXIT = Dep RWY ENTRY + DROT
                        departureOutput['E' + str(DepOutput)].value = departureInput['H'+ str(first_in_line_RWYqueue)].value #WAKE
                        departureOutput['F' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][3] #SID
                        departureOutput['G' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][2]#DROT
                        departureOutput['H' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][6]#TAXIOUT
                        departureOutput['I' + str(DepOutput)].value = minDepTime#DEP MIN SEPARATION
                        departureOutput['J' + str(DepOutput)].value = minDepLabel#DEP MIN SEPARATION LABEL
                        departureOutput['K' + str(DepOutput)].value = currentGap#currentGap
                        departureOutput['L' + str(DepOutput)].value = len(DepSTANDqueue)
                        departureOutput['M' + str(DepOutput)].value = len(TAXIhold)
                        departureOutput['N' + str(DepOutput)].value = len(RWYqueue1)
                        departureOutput['O' + str(DepOutput)].value = len(RWYqueue2)
#                        departureOutput['P' + str(DepOutput)].value = len(RWYqueue3)
#                        departureOutput['Q' + str(DepOutput)].value = len(RWYqueue4)
                        departureOutput['R' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][4]#DELAY DepSTANDqueue
                        departureOutput['S' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][8]#DELAY TAXIhold
                        departureOutput['T' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][10]#DELAY RWYqueue
                        departureOutput['U' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][11]#RWYqueue USED
                        
                        del RWYqueue1[first_in_line_RWYqueue]
                        
                        
                    elif currentRWYqueue == 2:
                        departureOutput['A' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][0] # AC ID                        
                        departureOutput['D' + str(DepOutput)].value = departureOutput['C' + str(DepOutput)].value + RWYqueue2[first_in_line_RWYqueue][2] # Dep RWY EXIT = Dep RWY ENTRY + DROT
                        departureOutput['E' + str(DepOutput)].value = departureInput['H'+ str(first_in_line_RWYqueue)].value #WAKE
                        departureOutput['F' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][3] #SID
                        departureOutput['G' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][2]#DROT
                        departureOutput['H' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][6]#TAXIOUT
                        departureOutput['I' + str(DepOutput)].value = minDepTime#DEP MIN SEPARATION
                        departureOutput['J' + str(DepOutput)].value = minDepLabel#DEP MIN SEPARATION LABEL
                        departureOutput['K' + str(DepOutput)].value = currentGap#currentGap
                        departureOutput['L' + str(DepOutput)].value = len(DepSTANDqueue)
                        departureOutput['M' + str(DepOutput)].value = len(TAXIhold)
                        departureOutput['N' + str(DepOutput)].value = len(RWYqueue1)
                        departureOutput['O' + str(DepOutput)].value = len(RWYqueue2)
#                        departureOutput['P' + str(DepOutput)].value = len(RWYqueue3)
#                        departureOutput['Q' + str(DepOutput)].value = len(RWYqueue4)
                        departureOutput['R' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][4]#DELAY DepSTANDqueue
                        departureOutput['S' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][8]#DELAY TAXIhold
                        departureOutput['T' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][10]#DELAY RWYqueue
                        departureOutput['U' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][11]#RWYqueue USED
                        del RWYqueue2[first_in_line_RWYqueue]
                        
                    DepOutput += 1
            elif DepOutput != 2:
                if (currentGap > n) and (Current_time>(departureOutput['C' + str(DepOutput-1)].value)+minDepTime) : 
                    #print(first_in_line_RWYqueue,' condition met', DepOutput)
                    #TAKE-OFF
                    departureOutput['B' + str(DepOutput)].value = int(Current_time/3600) # Dep HOUR
                    departureOutput['C' + str(DepOutput)].value = Current_time # Departure RWY Entry
                    
                    if currentRWYqueue == 1:
                        departureOutput['A' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][0] # AC ID                        
                        departureOutput['D' + str(DepOutput)].value = departureOutput['C' + str(DepOutput)].value + RWYqueue1[first_in_line_RWYqueue][2] # Dep RWY EXIT = Dep RWY ENTRY + DROT
                        departureOutput['E' + str(DepOutput)].value = departureInput['H'+ str(first_in_line_RWYqueue)].value #WAKE
                        departureOutput['F' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][3] #SID
                        departureOutput['G' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][2]#DROT
                        departureOutput['H' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][6]#TAXIOUT
                        departureOutput['I' + str(DepOutput)].value = minDepTime#DEP MIN SEPARATION
                        departureOutput['J' + str(DepOutput)].value = minDepLabel#DEP MIN SEPARATION LABEL
                        departureOutput['K' + str(DepOutput)].value = currentGap#currentGap
                        departureOutput['L' + str(DepOutput)].value = len(DepSTANDqueue)
                        departureOutput['M' + str(DepOutput)].value = len(TAXIhold)
                        departureOutput['N' + str(DepOutput)].value = len(RWYqueue1)
                        departureOutput['O' + str(DepOutput)].value = len(RWYqueue2)
#                        departureOutput['P' + str(DepOutput)].value = len(RWYqueue3)
#                        departureOutput['Q' + str(DepOutput)].value = len(RWYqueue4)
                        departureOutput['R' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][4]#DELAY DepSTANDqueue
                        departureOutput['S' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][8]#DELAY TAXIhold
                        departureOutput['T' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][10]#DELAY RWYqueue
                        departureOutput['U' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][11]#RWYqueue USED
                        del RWYqueue1[first_in_line_RWYqueue]
                        
                        
                    elif currentRWYqueue == 2:
                        departureOutput['A' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][0] # AC ID                        
                        departureOutput['D' + str(DepOutput)].value = departureOutput['C' + str(DepOutput)].value + RWYqueue2[first_in_line_RWYqueue][2] # Dep RWY EXIT = Dep RWY ENTRY + DROT
                        departureOutput['E' + str(DepOutput)].value = departureInput['H'+ str(first_in_line_RWYqueue)].value #WAKE
                        departureOutput['F' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][3] #SID
                        departureOutput['G' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][2]#DROT
                        departureOutput['H' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][6]#TAXIOUT
                        departureOutput['I' + str(DepOutput)].value = minDepTime#DEP MIN SEPARATION
                        departureOutput['J' + str(DepOutput)].value = minDepLabel#DEP MIN SEPARATION LABEL
                        departureOutput['K' + str(DepOutput)].value = currentGap#currentGap
                        departureOutput['L' + str(DepOutput)].value = len(DepSTANDqueue)
                        departureOutput['M' + str(DepOutput)].value = len(TAXIhold)
                        departureOutput['N' + str(DepOutput)].value = len(RWYqueue1)
                        departureOutput['O' + str(DepOutput)].value = len(RWYqueue2)
#                        departureOutput['P' + str(DepOutput)].value = len(RWYqueue3)
#                        departureOutput['Q' + str(DepOutput)].value = len(RWYqueue4)
                        departureOutput['R' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][4]#DELAY DepSTANDqueue
                        departureOutput['S' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][8]#DELAY TAXIhold
                        departureOutput['T' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][10]#DELAY RWYqueue
                        departureOutput['U' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][11]#RWYqueue USED
                        del RWYqueue2[first_in_line_RWYqueue]
                        
                    sequenceTab['A' + str(seqRow)].value = 'D'
                    sequenceTab['B' + str(seqRow)].value = departureOutput['A' + str(DepOutput)].value
                    sequenceTab['C' + str(seqRow)].value = departureOutput['C' + str(DepOutput)].value
                    sequenceTab['D' + str(seqRow)].value = departureOutput['D' + str(DepOutput)].value
                    sequenceTab['E' + str(seqRow)].value = departureOutput['G' + str(DepOutput)].value
                    seqRow+=1
                    DepOutput += 1
                    

        return(DepOutput,seqRow) 


    ########################### ARRIVALS #####################################
    
    def update_ArrHOLDqueue_Delay(Current_time):
        for AC in list(ArrHOLDqueue.keys()):
            ArrHOLDqueue_Delay = Current_time - ArrHOLDqueue[AC][1] # Delay = Current_time - SAE 
            ArrHOLDqueue[AC][4] = ArrHOLDqueue_Delay
            
    def SAE_lookup(Current_time, ARRkey):
        if ARRkey != (max_ARRIVAL):
            if Current_time >= runwayCalculations['L' + str(ARRkey)].value : # Current_time = SAE
                ArrHOLDqueue[ARRkey] = [arrivalInput['A' + str(ARRkey)].value, runwayCalculations['L' + str(ARRkey)].value, runwayCalculations['C' + str(ARRkey)].value, runwayCalculations['M' + str(ARRkey)].value, 0]
                ARRkey += 1 
        update_ArrHOLDqueue_Delay(Current_time)
         
        return (ARRkey)
    
    def update_APPqueue(Current_time,DepOutput,End_time,distance_based_FLAG,time_based_FLAG,ArrOutput): # add to APPqueue
        #print(Current_time, ' app queue called')
        if (len(ArrHOLDqueue)>0) and (len(APPqueue)==0): # There is something in the hold but nothing on approach
            first_in_line_ArrHOLDqueue = min(list(ArrHOLDqueue.keys()))

            max_constraint = 0
            arrivalOutput['I' + str(ArrOutput)].value = runwayCalculations['N' + str(first_in_line_ArrHOLDqueue)].value
            arrivalOutput['J' + str(ArrOutput)].value = runwayCalculations['O' + str(first_in_line_ArrHOLDqueue)].value
            #target time, optimised gaps
            if time_based_FLAG == True:
                if (len(RWYqueue1) + len(RWYqueue2))>0: #ther is a departure ready to go
                    if (arrivalInput['U' + str(first_in_line_ArrHOLDqueue)].value) == "ADDA" :    #*********to be changed 
                        AROT = ArrHOLDqueue[first_in_line_ArrHOLDqueue][2]
                        firstDeparture, currentRWYqueue = first_in_line_RWYqueue_funct(DepOutput,End_time)
                        if currentRWYqueue ==1:
                            DROT1 = RWYqueue1[firstDeparture][2]
                        else:
                            DROT1 = RWYqueue2[firstDeparture][2]
#                            secondDeparture, nextRWYqueue = second_in_line_RWYqueues(currentRWYqueue,End_time)
#                            if nextRWYqueue ==1:
#                                DROT2 = RWYqueue1[secondDeparture][2]
#                            else:
#                                DROT2 = RWYqueue2[secondDeparture][2] 
                        ADDA_target_time = AROT + DROT1 + DROT1 + x_buffer# AROT + NextDep DROT + NextDep2 DROT
                        ADDA_target_distance = time_to_distance_assumed_speed_profile_GS(first_in_line_ArrHOLDqueue, d_dme,int(ADDA_target_time))#distance
                        if (ADDA_4dme_FLAG == True) and (ArrOutput>2):
    	                    Total_time_follow = int(DBS_actual_speed_profile((ADDA_target_distance+4),first_in_line_ArrHOLDqueue))
    	                    Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,ArrOutput))
    	                    ADDA_separation = Total_time_follow - Time_lead_4dme_to_thr
                                
                        elif ADDA_thr_FLAG == True:
                            ADDA_separation = int(DBS_actual_speed_profile(ADDA_target_distance,first_in_line_ArrHOLDqueue)) #time
                        else:
                            ADDA_separation = int(DBS_actual_speed_profile(ADDA_target_distance,first_in_line_ArrHOLDqueue)) #time - default
                       
                        
                        if ADDA_separation > arrivalOutput['I' + str(ArrOutput)].value:
                            arrivalOutput['J' + str(ArrOutput)].value= "ADDA"
                            arrivalOutput['I' + str(ArrOutput)].value = ADDA_separation
                        
                    elif (arrivalInput['U' + str(first_in_line_ArrHOLDqueue)].value) == "ADA" :  
                        AROT = ArrHOLDqueue[first_in_line_ArrHOLDqueue][2]
                        firstDeparture, currentRWYqueue = first_in_line_RWYqueue_funct(DepOutput,End_time)
                        if currentRWYqueue ==1:
                            DROT1 = RWYqueue1[firstDeparture][2]
                        else:
                            DROT1 = RWYqueue2[firstDeparture][2]
                        ADA_target_time = AROT + DROT1 + x_buffer# AROT + NextDep DROT + NextDep2 DROT
                        ADA_target_distance = time_to_distance_assumed_speed_profile_GS(first_in_line_ArrHOLDqueue, d_dme, int(ADA_target_time))#distance
                        if (ADA_4dme_FLAG == True) and (ArrOutput>2):
                            Total_time_follow = int(DBS_actual_speed_profile((ADA_target_distance+4),first_in_line_ArrHOLDqueue))
                            Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,ArrOutput-1))
                            ADA_separation = Total_time_follow - Time_lead_4dme_to_thr
                                
                        elif ADA_thr_FLAG == True:
                            ADA_separation = int(DBS_actual_speed_profile(ADA_target_distance,first_in_line_ArrHOLDqueue)) #time
                        else:
                            ADA_separation = int(DBS_actual_speed_profile(ADA_target_distance,first_in_line_ArrHOLDqueue)) #time

                        
                        if ADA_separation > arrivalOutput['I' + str(ArrOutput)].value:
                            arrivalOutput['J' + str(ArrOutput)].value = "ADA"
                            arrivalOutput['I' + str(ArrOutput)].value = ADA_separation
                #else: # no departure ready to go
                    #max_constraint = arrivalOutput['I' + str(ArrOutput)].value
           # elif distance_based_FLAG == True:
            max_constraint = arrivalOutput['I' + str(ArrOutput)].value
                
            #print(Current_time, ArrOutput, ' | max_constraint = ', max_constraint)
            
            #if max_constraint != 0:
                
            ArrHOLDqueue[first_in_line_ArrHOLDqueue].append(Current_time)#APPqueue entry time
            ALT = ArrHOLDqueue[first_in_line_ArrHOLDqueue][5]+ int(max_constraint) #(ALT = APPqueue_entry_time + max_constraint)
            ArrHOLDqueue[first_in_line_ArrHOLDqueue].append(ALT)               
            RWY_EXIT = ArrHOLDqueue[first_in_line_ArrHOLDqueue][6] + ArrHOLDqueue[first_in_line_ArrHOLDqueue][2] # ALT + AROT
            ArrHOLDqueue[first_in_line_ArrHOLDqueue].append(RWY_EXIT)
            
            APPqueue[first_in_line_ArrHOLDqueue]=ArrHOLDqueue[first_in_line_ArrHOLDqueue]
  
            del ArrHOLDqueue[first_in_line_ArrHOLDqueue]
            
            #print to sequence tab
            
           
                          
            
    def Arr_LANDING(Current_time, ArrOutput,first_in_line_APPqueue,seqRow): 
        
        
#        if len(APPqueue)!=0:
#            first_in_line_APPqueue = min(list(APPqueue.keys()))#there is only one AC in the APPqueue  
#            #print('There is something in the APPqueue')
#           
#            #print('NEXT ARRIVAL = ', AC)
#            if Current_time == APPqueue[first_in_line_APPqueue][6]: #it's time to land
        if RWY_status == "D":
            print('*** GO AROUND ***', APPqueue[first_in_line_APPqueue])
            del APPqueue[first_in_line_APPqueue]
            goAroundHour = int(Current_time/3600)
            if goAroundHour in list(GoAroundCount.keys()): #if there was already a goAround at that hour:
                GoAroundCount[goAroundHour].append(1)
            else:
                GoAroundCount[goAroundHour]=[1]
            
        elif RWY_status == "E":
            #print(RWY_status)
            arrivalOutput['A' + str(ArrOutput)].value = arrivalInput['A' + str(first_in_line_APPqueue)].value #ARR ID
            arrivalOutput['B' + str(ArrOutput)].value = int(Current_time/3600) #LANDING HOUR
            arrivalOutput['C' + str(ArrOutput)].value = Current_time #ACTUAL LANDING TIME 
            arrivalOutput['D' + str(ArrOutput)].value = APPqueue[first_in_line_APPqueue][7] # RWY EXIT
            arrivalOutput['E' + str(ArrOutput)].value = runwayCalculations['U' + str(first_in_line_APPqueue)].value #WAKE
            arrivalOutput['F' + str(ArrOutput)].value = arrivalOutput['D' + str(ArrOutput)].value + runwayCalculations['B' + str(first_in_line_APPqueue)].value #In blocks time
            arrivalOutput['G' + str(ArrOutput)].value = runwayCalculations['C' + str(first_in_line_APPqueue)].value#AROT
            arrivalOutput['H' + str(ArrOutput)].value = runwayCalculations['B' + str(first_in_line_APPqueue)].value# Taxi-in duration
            
            
            arrivalOutput['K' + str(ArrOutput)].value = len(ArrHOLDqueue)#length ArrHOLDqueue
            arrivalOutput['L' + str(ArrOutput)].value = APPqueue[first_in_line_APPqueue][4]# ArrHOLDqueue delay   
            AIBT = arrivalOutput['F' + str(ArrOutput)].value
           
            #Add Arrival to ARRIVALqueue
            ARRIVALqueue[first_in_line_APPqueue]=[arrivalOutput['A' + str(ArrOutput)].value, AIBT, ArrOutput]
            #print('ARRIVALqueue = ', list(ARRIVALqueue.keys()))
            del APPqueue[first_in_line_APPqueue]
            sequenceTab['A' + str(seqRow)].value = 'A'
            sequenceTab['B' + str(seqRow)].value = arrivalOutput['A' + str(ArrOutput)].value
            sequenceTab['C' + str(seqRow)].value = arrivalOutput['C' + str(ArrOutput)].value
            sequenceTab['D' + str(seqRow)].value = arrivalOutput['D' + str(ArrOutput)].value
            sequenceTab['E' + str(seqRow)].value = arrivalOutput['G' + str(ArrOutput)].value
            ArrOutput+=1
            seqRow += 1
            
                
                    
        return (ArrOutput,seqRow)
                   
    def first_in_line_ARRIVALqueue_func(End_time):
        min_IBT = End_time
        first_in_line_ARRIVALqueue = 0
        for AC in list(ARRIVALqueue.keys()):
            if ARRIVALqueue[AC][1]<min_IBT:
                min_IBT=ARRIVALqueue[AC][1]
                first_in_line_ARRIVALqueue = AC
        return(first_in_line_ARRIVALqueue)
    
    def update_ARRIVALqueue(Current_time,End_time):
        if len(ARRIVALqueue)>0:
            
            #Check first in line in arrival queue
            first_in_line_ARRIVALqueue = first_in_line_ARRIVALqueue_func(End_time)              
                
            if Current_time > ARRIVALqueue[first_in_line_ARRIVALqueue][1]:
                #print(Current_time, 'ARR { ',AC,' } deleted from ARRIVALqueue ')
                del ARRIVALqueue[first_in_line_ARRIVALqueue]
    
    def update_currentGap(Current_time, End_time):
        if RWY_status == "E" or RWY_status == "D":
            if len(APPqueue)==0: #Nothing in the queue
                currentGap = End_time # Huuuuge currentGap
            else:
                next_Arrival = min(list(APPqueue.keys())) # should be only one key in the list
                currentGap = APPqueue[next_Arrival][6] - Current_time # ALT - Current_time
        elif RWY_status == "A":
            currentGap = 0 
            
        return (currentGap)
    
        

    
   
   
#--------------------------------MODEL RUNS-----------------------------------#


    print('distance_based_FLAG = ',distance_based_FLAG)
    print('time_based_FLAG =',time_based_FLAG)
    while Current_time < End_time:
        #print(Current_time)
        #print(RWY_status)

        if RWY_status == "E":
            if df_dep.empty == False: #there are departures
                
                SOBTrow = SOBTlookup(Current_time, SOBTrow)
                if ((len(TAXIqueue) + len(ARRIVALqueue)+ len(TAXIhold))<15) and len(DepSTANDqueue)> 0: # if there are less than 15 AC moving on the TAXIway and there's something in DepSTANDqueue
                    TAXIqueue_update(Current_time)
                TAXIhold_update(Current_time,End_time)                
                RWYqueues_update(Current_time)
                update_Departure_Delays(Current_time)
                
            if df_arr.empty == False: # there are arrivals
                ARRkey = SAE_lookup(Current_time, ARRkey)
                if len(APPqueue) == 0:
                    update_APPqueue(Current_time,DepOutput,End_time,distance_based_FLAG,time_based_FLAG,ArrOutput)             
                update_ARRIVALqueue(Current_time,End_time)                 
                currentGap = update_currentGap(Current_time, End_time)
                
                
            else:#if there aren't any arrivals
                currentGap = End_time #huuuuuge gap
                
            #DEPARTURES TAKE OFF    
            if df_dep.empty == False:
                if (len(RWYqueue1)+len(RWYqueue2))>0:#there is something waiting to takeoff
                     #print('TAKE OFF called')
                     DepOutput,seqRow = Dep_TAKE_OFF(Current_time,DepOutput,currentGap,End_time,seqRow)
                     #print('dep took off')
                     # Note : DepOurputROW was already increased so (DepOutputROW-1) will reffer to the current departure
                     #if type(departureOutput['C' + str(DepOutput-1)].value) == int:   
                         #print(departureOutput['C' + str(DepOutput-1)].value)
                     if Current_time < departureOutput['D' + str(DepOutput-1)].value : # while the Departure is still on the runway
                         #print(Current_time,' Departure {',(DepOutput-1),'} is about to take-off')
                         RWY_status = "D"
        
            #ARRIVALS LANDING
            if df_arr.empty == False:                        
                if len(APPqueue)!=0:
                    first_in_line_APPqueue = min(list(APPqueue.keys()))#there is only one AC in the APPqueue  
                    if Current_time == APPqueue[first_in_line_APPqueue][6]: #it's time to land
                        #print('Current_time = ', Current_time, '| ALT = ',APPqueue[first_in_line_APPqueue][6])
                        ArrOutput,seqRow = Arr_LANDING(Current_time, ArrOutput,first_in_line_APPqueue,seqRow)
                        if Current_time < arrivalOutput['D' + str(ArrOutput-1)].value : #while Arrival is still on the runway
                            #print(Current_time,' Arrival {',ArrOutput-1,'} is about to land ')
                            RWY_status = "A"
                            
                           
            
            
        elif RWY_status == "D":
            #print(Current_time,' | ', RWY_status)
            SOBTrow = SOBTlookup(Current_time, SOBTrow)
            if ((len(TAXIqueue) + len(ARRIVALqueue)+ len(TAXIhold))<15) and len(DepSTANDqueue)> 0: # if there are less than 15 AC moving on the TAXIway and there's something in DepSTANDqueue
                TAXIqueue_update(Current_time)
            TAXIhold_update(Current_time,End_time)                
            RWYqueues_update(Current_time)
            update_Departure_Delays(Current_time)
            
            if df_arr.empty == False: #there are arrivals
                ARRkey = SAE_lookup(Current_time, ARRkey)
                if len(APPqueue) == 0:
                    update_APPqueue(Current_time,DepOutput,End_time,distance_based_FLAG,time_based_FLAG,ArrOutput)              
                update_ARRIVALqueue(Current_time,End_time)                 
                currentGap = update_currentGap(Current_time,End_time)
            else:#if there aren't any arrivals
                currentGap = End_time #huuuuuge gap
                
            if Current_time == departureOutput['D' + str(DepOutput-1)].value : # when current_time > departure RWY_EXIT the rwy is empty again
                RWY_status = "E"
            
            #ARRIVALS LANDING (GO-AROUND case)
            if df_arr.empty == False:                        
                if len(APPqueue)!=0:
                    first_in_line_APPqueue = min(list(APPqueue.keys()))#there is only one AC in the APPqueue  
                    if Current_time == APPqueue[first_in_line_APPqueue][6]: #it's time to land
                        #print('It is time to land but GOaround')
                        ArrOutput,seqRow = Arr_LANDING(Current_time, ArrOutput,first_in_line_APPqueue,seqRow)
                        
                        #print(ArrOutput,'******GO AROUND************')
            
            
        elif RWY_status == "A":
            #print(Current_time,' | ', RWY_status)
            if df_dep.empty == False: #there are departures
                SOBTrow = SOBTlookup(Current_time, SOBTrow)
                if ((len(TAXIqueue) + len(ARRIVALqueue)+ len(TAXIhold))<15) and len(DepSTANDqueue)> 0: # if there are less than 15 AC moving on the TAXIway and there's something in DepSTANDqueue
                    TAXIqueue_update(Current_time)
                TAXIhold_update(Current_time,End_time)                
                RWYqueues_update(Current_time)
                update_Departure_Delays(Current_time)
                
            ARRkey = SAE_lookup(Current_time, ARRkey)
            if len(APPqueue) == 0:
                update_APPqueue(Current_time,DepOutput,End_time,distance_based_FLAG,time_based_FLAG,ArrOutput)                
            update_ARRIVALqueue(Current_time,End_time)                 
            currentGap = update_currentGap(Current_time,End_time)
            
            if Current_time == arrivalOutput['D' + str(ArrOutput-1)].value : #while Arrival is still on the runway
                RWY_status = "E"
        Current_time += 1


# ============================================================================#
#                       Buffer Calculations                                   #      
# ============================================================================#
    bufferRow = 2
    
    for row in range (2, (sequenceTab.max_row-2)):
       if (sequenceTab['A'+str(row)].value == "A") and (sequenceTab['A'+str(row+1)].value == "D") and  (sequenceTab['A'+str(row+2)].value == "A") :#ADA sequence:
           sequenceTab['G' + str(bufferRow)].value = (sequenceTab['C'+str(row+1)].value - sequenceTab['D'+str(row)].value) + (sequenceTab['C'+str(row+2)].value - sequenceTab['D'+str(row+1)].value)
           sequenceTab['F' + str(bufferRow)].value = sequenceTab['B' + str(row)].value 
           bufferRow+=1

# ============================================================================#
#                   THROUGHPUT AND DELAYS CALCULATIONS                        #
# ============================================================================#
 
    min_thr_HOUR = min(arrivalOutput['B' + str(2)].value,departureOutput['B' + str(2)].value)
    print('arrivalOutput["B" + str(ArrOutput-1)].value = ',arrivalOutput['B' + str(ArrOutput-1)].value)
    print('departureOutput["B" + str(DepOutput-1)].value = ', departureOutput['B' + str(DepOutput-1)].value)
    max_thr_HOUR = min(arrivalOutput['B' + str(ArrOutput-1)].value,departureOutput['B' + str(DepOutput-1)].value)
    diff_thr_HOUR = max_thr_HOUR-min_thr_HOUR
    
    for row in range(2,(diff_thr_HOUR+3)):
        dep_thr_count = 0
        arr_thr_count = 0
        throughputTab['A' + str(row)].value = min_thr_HOUR
        
        for i in range(2, (DepOutput)):
            if departureOutput['B' + str(i)].value == None: #no departures
                break
            if departureOutput['B' + str(i)].value == min_thr_HOUR:
                dep_thr_count +=1       
        throughputTab['B' + str(row)].value = dep_thr_count
        for i in range(2, (ArrOutput)):
            if arrivalOutput['B' + str(i)].value == None: #No arrivals
                break
            if arrivalOutput['B' + str(i)].value == min_thr_HOUR:
                arr_thr_count +=1    
                
        throughputTab['C' + str(row)].value = arr_thr_count        
        throughputTab['D' + str(row)].value = throughputTab['B' + str(row)].value + throughputTab['C' + str(row)].value
        total_thr = throughputTab['D' + str(row)].value
        throughput.append(total_thr)
        
        if min_thr_HOUR in list(GoAroundCount.keys()):#there was at least a goAround at that hour:
            
            throughputTab['E' + str(row)].value = sum(GoAroundCount[min_thr_HOUR])
            
        else:
            throughputTab['E' + str(row)].value = 0
        min_thr_HOUR +=1
  
# ============================================================================#
#                               DELAYS                                        #
# ============================================================================#
    
    for row in range(2, DepOutput):
        delayTab['A' + str(row)].value = departureOutput['A' + str(row)].value
        delayTab['B' + str(row)].value = departureOutput['B' + str(row)].value
        delayTab['C' + str(row)].value = departureOutput['T' + str(row)].value + departureOutput['S' + str(row)].value
        delayTab['D' + str(row)].value = departureOutput['R' + str(row)].value
        
    for row in range(2, ArrOutput):
        delayTab['I' + str(row)].value = arrivalOutput['A' + str(row)].value
        delayTab['J' + str(row)].value = arrivalOutput['B' + str(row)].value
        delayTab['K' + str(row)].value = arrivalOutput['L' + str(row)].value 
        
       
    
    number_of_goArounds_queued = 0
    for i in list(GoAroundCount.keys()):
        number_of_goArounds_queued+=sum(GoAroundCount[i])
    print('End_time = ',End_time)
    print("There are [", str(len(TAXIhold)),"] Departure A/C remaining in the TAXI ,", str(len(RWYqueue1)+len(RWYqueue2)),"Departure A/C remaining in the RWY queues,")
    print("There are [", str(len(ArrHOLDqueue)),"] Arrivals remaining in the Arrival Hold Queue ,", str(len(APPqueue)),"Arrivals remaining in the APPqueue,")
    #print("Out of interest - [",countArr,"] times 'departureLookup' method (for Arrivals) was non-zero + [",countDep,"] times 'departureLookup2' method (for queued A/C) was non-zero")
    #print("DEBUG - countARRdebug =", countARRdebug, "and countDEPdebug =", countDEPdebug)
    print("Final number of go-around Arrival cases (Queued):", number_of_goArounds_queued)
    print("Model took %s seconds to run" % round((time.time() - program_runtime_start),2))
    if (len(DepSTANDqueue)>0 or len(TAXIhold)>0):
        print("ERROR!!!  Check DEPARTURES")
    if (len(ARRIVALqueue)>0) or (len(APPqueue)>0) or (len(ArrHOLDqueue)>0):
        print("ERROR!!!  Check ARRIVALS")
    
    if averagethrFLAG == False:
        output_extension = time.strftime("%H_%M", time.localtime(time.time()))
        throughputTab['F' + str(1)].value = 'Difference in thr averages'
        extra_diff=[0]*(throughputTab.max_row-1)        
        difference.append(extra_diff)
        throughputTab['F' + str(2)].value = str(difference)
        
        wb.save('OUTPUT_RAPID_v3.0_' + str(output_extension) +  '.xlsx') # Choose file name once complete?
        name_output_file = 'OUTPUT_RAPID_v3.0_' + str(output_extension) +  '.xlsx'
        iter1 += 1 
        
    else:            
        big_list.append(throughput)
        average_run = []
        diff2=[]
        diff=0
        
        average_hour = 0
        summ = 0
        if maxIter <2:
            averages.append(throughput)
            maxIter +=1
        elif (maxIter >=2) and (maxIter <10):       # minimum number of runs = 10
            for j in range(0,len(throughput)): 
                for i in range (0, len(big_list)):  #len(biglist)
                    print('element sum = ',big_list[i][j])
                    summ = summ + big_list[i][j]
                
                average_hour = summ/ len(big_list)
                average_run.append(average_hour)
                print('average_run=' ,average_run)
                summ = 0
            averages.append(average_run)
            average_run=[]
            print(averages)
            for j in range(0,len(averages[0])):
                print('______________________ ', j)
                compare = averages[len(averages)-1][j]
                print('last run = ',compare)
                diff = compare - averages[len(averages)-2][j]
                print('diff = ',diff)
                diff2.append(diff)
                print('diff2 = ',diff2)
               
            difference.append(diff2)
            diff2=[]
            print('diff = ',diff2)
            print('difference list', difference)

          
            maxIter +=1    
            summ=0
        else:
            for j in range(0,len(throughput)): 
                for i in range (0, len(big_list)):  #len(biglist)
                    print('element sum = ',big_list[i][j])
                    summ = summ + big_list[i][j]
                
                average_hour = summ/ len(big_list)
                average_run.append(average_hour)
                print('average_run=' ,average_run)
                summ = 0
            averages.append(average_run)
            average_run=[]
            print(averages)
            for j in range(0,len(averages[0])):
                print('______________________ ', j)
                compare = averages[len(averages)-1][j]
                print('last run = ',compare)
                diff = compare - averages[len(averages)-2][j]
                print('diff = ',diff)
                diff2.append(diff)
                print('diff2 = ',diff2)
               
            difference.append(diff2)
            diff2=[]
            print('diff = ',diff2)
            print('difference list', difference)
            for i in range(0, len(difference[0])):
                print("OK so far")
                if (difference[len(difference)-1][i] <= 0.1) and (difference[len(difference)-1][i] >=-0.1) :
                    print('####### difference in averages = ',difference[len(difference)-1][i])
                    iter2 = 0
                else:
                    print('############condition false')
                    iter2 = 1
                    
                    break
            summ=0    
            if iter2 == 0:
                print('n_times 1 =', maxIter)
                throughputTab['F' + str(1)].value = 'Difference in thr averages'
                extra_diff=[0]*(throughputTab.max_row-1)
                
                difference.append(extra_diff)
                throughputTab['F' + str(2)].value = str(difference)
                output_extension = time.strftime("%H_%M", time.localtime(time.time()))
                output_extension2 = iter1+1
                arrivalOutput.delete_cols(13)
                arrivalOutput.delete_cols(13)
                wb.save('OUTPUT_RAPID_v3.0_' + str(output_extension) + '_iteration_' + str(output_extension2) +  '.xlsx') # Choose file name once complete?
                name_output_file = 'OUTPUT_RAPID_v3.0_' + str(output_extension) + '_iteration_' + str(output_extension2) +  '.xlsx'
            else:
                maxIter += 1
                print('n_times 2 =', maxIter)
        iter1 += 1


#*****************************************************************************#
#=============================================================================#
#                                                                             #
#                             VISUAL MODULE                                   #
#                                                                             #
# ============================================================================#
#*****************************************************************************#

print('Output file name = ', name_output_file)  
if (Thr_FLAG == True) or (Delay_FLAG == True) or (arr_delay_FLAG == True) or (Seq_FLAG == True) or (convergenceFLAG == True) or (ADA_buffer_FLAG == True):
    xls = pd.ExcelFile(name_output_file) 
    df_thr = xls.parse(5)
    df_delay = xls.parse(6)
    df_dep_output = xls.parse(4)
    df_arr_output = xls.parse(3)
    df_rwy_calcs = xls.parse(2)
    df_sequence_output =xls.parse(7)
    df_thr['Mean Thr'] = 0

    if OP_FLAG == True:
        opr_xls = pd.ExcelFile(operational_data) 
        op_data = opr_xls.parse(0)
        #Transform date format into seconds
        op_data['Time Bin-H'] = op_data['Time Bin-H'] .astype(str)
        op_data['Time Bin-H'] = pd.DatetimeIndex(op_data['Time Bin-H'])
        op_data['Time Bin-H'] = pd.to_timedelta(op_data['Time Bin-H']) # convert to timedelta to calculate seconds
        op_data['Time Bin-H'] = op_data['Time Bin-H'].dt.seconds
        
        op_data['Time Bin - PS'] = op_data['Time Bin - PS'] .astype(str)
        op_data['Time Bin - PS'] = pd.DatetimeIndex(op_data['Time Bin - PS'])
        op_data['Time Bin - PS'] = pd.to_timedelta(op_data['Time Bin - PS']) # convert to timedelta to calculate seconds
        op_data['Time Bin - PS'] = op_data['Time Bin - PS'].dt.seconds 
        
        
        #Extract lists to plot them
        OD_Thr_Hour = op_data['Hour'].tolist()
        OD_Thr = op_data['Total Throughput'].tolist()
        OP_H_Delay = op_data['Hold-Delay'].tolist()
        OP_Time_H_Delay = op_data['Time Bin-H'].tolist()
        OP_PS_Delay = op_data['PS-Delay'].tolist()
        OP_Time_PS_Delay = op_data['Time Bin - PS'].tolist()
    
    LARGE_FONT= ("Verdana", 12)
    
    Flag_another_run =  False
    class RAPIDvisual(tk.Tk):
    
        def __init__(self, *args, **kwargs):
            
            tk.Tk.__init__(self, *args, **kwargs)
    
            #tk.Tk.iconbitmap(self, text="clienticon.ico")
            tk.Tk.wm_title(self, "RAPID VISUAL")
            
            
            container = tk.Frame(self)
            container.pack(side="top", fill="both", expand = True)
            container.grid_rowconfigure(0, weight=1)
            container.grid_columnconfigure(0, weight=1)
    
            self.frames = {}
            for F in (StartPage, Conv, Thr, DepDelay, DepDelay2, ArrivalDelay, Seq, ADAbuffer):
        
                frame = F(container, self)
    
                self.frames[F] = frame
    
                frame.grid(row=0, column=0, sticky="nsew")
    
            self.show_frame(StartPage)
        def show_frame(self, cont):
    
            frame = self.frames[cont]
            frame.tkraise()
    
    class StartPage(tk.Frame):
    
        def __init__(self, parent, controller):
            tk.Frame.__init__(self,parent)
            label = tk.Label(self, text="Start Page", font=LARGE_FONT)
            label.pack(pady=10,padx=10)
            
            
           
            if convergenceFLAG ==True:
                button1 = ttk.Button(self, text="Convergence Throughput",
                                    command=lambda: controller.show_frame(Conv))
                button1.pack()
            if Thr_FLAG == True:
                button = ttk.Button(self, text="Throughput",
                                    command=lambda: controller.show_frame(Thr))
                button.pack()
            if Delay_FLAG == True:
                button2 = ttk.Button(self, text="RWY Hold Delay",
                                    command=lambda: controller.show_frame(DepDelay))
                button2.pack()
            if Delay_FLAG == True:
                button5 = ttk.Button(self, text="Push/Start Delay",
                                    command=lambda: controller.show_frame(DepDelay2))
                button5.pack()
            if arr_delay_FLAG == True:
                button3 = ttk.Button(self, text="Arrivals Delay",
                                    command=lambda: controller.show_frame(ArrivalDelay))
                button3.pack()
            if Seq_FLAG == True:
                button4 = ttk.Button(self, text="Sequence",
                                    command=lambda: controller.show_frame(Seq))
                button4.pack()
            if ADA_buffer_FLAG == True:
                button5 = ttk.Button(self, text="ADA Buffer",
                                    command=lambda: controller.show_frame(ADAbuffer))
                button5.pack()
       
            label1_1 = ttk.Label(self, text=".")
            label1_1.pack()
            label1_2 = ttk.Label(self, text=".")
            label1_2.pack()
            label1_3 = ttk.Label(self, text=".")
            label1_3.pack()
            label1 = ttk.Label(self, text="I want to compare my results")
            label1.pack()
            label2 = ttk.Label(self, text="_____________________________")
            label2.pack()
            label3 = ttk.Label(self, text="How many new sets?")
            label3.pack()
    
            
            def show_button():
                m2 = int(m2_input.get())
                m2_output.set(m2)
                          
                
                app.destroy()
                m2 = m2_output.get()
                def load_new_data2():
                    new_data2 = filedialog.askopenfilename()
                    print(new_data2)
                    new_data_sheet2.set(new_data2)
                    return()
                        
                def load_new_data3():
                    new_data3 = filedialog.askopenfilename()
                    print(new_data3)
                    new_data_sheet3.set(new_data3) 
                    return()
                
                def load_new_data4():
                    new_data4 = filedialog.askopenfilename()
                    print(new_data4)
                    new_data_sheet4.set(new_data4) 
                    return() 
                
                def load_new_data5():
                    new_data5 = filedialog.askopenfilename()
                    print(new_data5)
                    new_data_sheet5.set(new_data5) 
                    return()    
                    
                def load_new_data6():
                    new_data6 = filedialog.askopenfilename()
                    print(new_data6)
                    new_data_sheet6.set(new_data6) 
                    return() 
                    
                    
                def define_input_parameters3():   
                    convergence = int(var0.get())
                    convergence_output.set(convergence)
                    Throughput_check = int(var8.get())
                    Throughput_check_output.set(Throughput_check)
                    Delay_check = int(var9.get())
                    Delay_check_output.set(Delay_check)
                    Seq_check = int(var10.get())
                    Seq_check_output.set(Seq_check)
                    arr_delay = int(var13.get())
                    arr_delay_output.set(arr_delay) 
                    ADA_buffer = int(var18.get())
                    ADA_buffer_output.set(ADA_buffer)
                    op_yes = int(var11.get())
                    op_yes_output.set(op_yes)
                    new_set = int(var12.get())
                    new_set_output.set(new_set)
                    button_check.set(True)
                    average_check = int(var6.get())
                    average_check_output.set(average_check)        
                    window.destroy()
                if m2 >= 1 :
                    
                    window = Tk()
                    window.title("New set of data import")
                    
                    mainframe = ttk.Frame(window, padding="10 10 30 40")
                    mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
                    mainframe.columnconfigure(0, weight=1)
                    mainframe.rowconfigure(0, weight=1)
                    
                    #innerframe = ttk.Frame(window, padding="5 5 0 0")
                    
                    # Defines expected inputs (i.e. GUI expects integers) and assigns default values
                    button_check = StringVar(window, value='0')
                    
                    average_check_output = IntVar() 
                    new_data_sheet2 = StringVar()        
                    ttk.Label(mainframe, text="Import new data set 2: ").grid(column=1, row=1, sticky=W)        
                    ttk.Button(mainframe, text="Import data 2", command=load_new_data2).grid(column=2, row=1, sticky=W)
                    
                    
                    
                    if m2 >=2:
                        new_data_sheet3 = StringVar()
                        ttk.Label(mainframe, text="Import new data set 3: ").grid(column=1, row=2, sticky=W)            
                        ttk.Button(mainframe, text="Import data 3", command=load_new_data3).grid(column=2, row=2, sticky=W)
                        
                        if m2 >=3:
                            new_data_sheet4 = StringVar()
                            ttk.Label(mainframe, text="Import new data set 4: ").grid(column=1, row=3, sticky=W)            
                            ttk.Button(mainframe, text="Import data 4", command=load_new_data4).grid(column=2, row=3, sticky=W)
                            if m2 >=4:
                                new_data_sheet5 = StringVar()
                                ttk.Label(mainframe, text="Import new data set 5: ").grid(column=1, row=4, sticky=W)            
                                ttk.Button(mainframe, text="Import data 5", command=load_new_data5).grid(column=2, row=4, sticky=W)
                                if m2 >=5:
                                    new_data_sheet6 = StringVar()
                                    ttk.Label(mainframe, text="Import new data set 6: ").grid(column=1, row=5, sticky=W)            
                                    ttk.Button(mainframe, text="Import data 6", command=load_new_data6).grid(column=2, row=5, sticky=W)
                    
                    
                    inner = Frame(window, bg='pink', width=0, height=0, padx=20, pady=20)
                    inner.grid(column=0, row=1)
                    inner.columnconfigure(0, weight=1)
                    inner.rowconfigure(0, weight=1)
                    ttk.Button(inner, text="Visualize results", command=define_input_parameters3).grid(column=0, row=0, sticky=W)
                    
                    
                    for child in mainframe.winfo_children(): child.grid_configure(padx=5, pady=5)        
                    window.bind('<Return>', define_input_parameters)
                    
                    window.columnconfigure(0, weight=1)
                    window.rowconfigure(0, weight=1)
                    window.rowconfigure(1, weight=1)
                    
                    window.mainloop()
    
                    
                if m2>=1:    
                    new_data2 = new_data_sheet2.get() 
                    xls2 = pd.ExcelFile(new_data2) 
                    df_thr2 = xls2.parse(5)
                    df_delay2 = xls2.parse(6)
                    df_dep_output2 = xls2.parse(4)
                    df_arr_output2 = xls2.parse(3)
                    df_rwy_calcs2 = xls2.parse(2)
                    
                    if m2>=2:
                        new_data3 = new_data_sheet3.get() 
                        xls3 = pd.ExcelFile(new_data3) 
                        df_thr3 = xls3.parse(5)
                        df_delay3 = xls3.parse(6)
                        df_dep_output3 = xls3.parse(4)
                        df_arr_output3 = xls3.parse(3)
                        df_rwy_calcs3 = xls3.parse(2)
                        if m2>=3:
                            new_data4 = new_data_sheet4.get() 
                            xls4 = pd.ExcelFile(new_data4) 
                            df_thr4 = xls4.parse(5)
                            df_delay4 = xls4.parse(6)
                            df_dep_output4 = xls4.parse(4)
                            df_arr_output4 = xls4.parse(3)
                            df_rwy_calcs4 = xls4.parse(2)
                            if m2>=4:
                                new_data5 = new_data_sheet5.get() 
                                xls5 = pd.ExcelFile(new_data5) 
                                df_thr5 = xls5.parse(5)
                                df_delay5 = xls5.parse(6)
                                df_dep_output5 = xls5.parse(4)
                                df_arr_output5 = xls5.parse(3)
                                df_rwy_calcs5 = xls5.parse(2)
                                if m2>=5:
                                    new_data6 = new_data_sheet6.get() 
                                    xls6 = pd.ExcelFile(new_data6) 
                                    df_thr6 = xls6.parse(5)
                                    df_delay6 = xls6.parse(6)
                                    df_dep_output6 = xls6.parse(4)
                                    df_arr_output6 = xls6.parse(3)
                                    df_rwy_calcs6 = xls6.parse(2)
                # ============================================================================#
                #                       VISUAL GUI                                            #
                # ============================================================================#
                LARGE_FONT= ("Verdana", 12)
                
                
                class RAPIDvisual2(tk.Tk):
                
                    def __init__(self, *args, **kwargs):
                        
                        tk.Tk.__init__(self, *args, **kwargs)
                
                        #tk.Tk.iconbitmap(self, text="clienticon.ico")
                        tk.Tk.wm_title(self, "RAPID VISUAL")
                        
                        
                        container = tk.Frame(self)
                        container.pack(side="top", fill="both", expand = True)
                        container.grid_rowconfigure(0, weight=1)
                        container.grid_columnconfigure(0, weight=1)
                
                        self.frames = {}
                        for F in (StartPage2, Conv2, Thr2, DepDelay22, DepDelay222, ArrivalDelay2, Seq2, ADAbuffer2):
                    
                            frame = F(container, self)
                
                            self.frames[F] = frame
                
                            frame.grid(row=0, column=0, sticky="nsew")
                
                        self.show_frame(StartPage2)
                
                    def show_frame(self, cont):
                
                        frame = self.frames[cont]
                        frame.tkraise()
                
                class StartPage2(tk.Frame):
                
                    def __init__(self, parent, controller):
                        tk.Frame.__init__(self,parent)
                        label = tk.Label(self, text="Start Page", font=LARGE_FONT)
                        label.pack(pady=10,padx=10)
                        
                        
                       
                        if convergenceFLAG ==True:
                            button1 = ttk.Button(self, text="Convergence Throughput",
                                                command=lambda: controller.show_frame(Conv2))
                            button1.pack()
                        if Thr_FLAG == True:
                            button = ttk.Button(self, text="Throughput",
                                                command=lambda: controller.show_frame(Thr2))
                            button.pack()
                        if Delay_FLAG == True:
                            button2 = ttk.Button(self, text="RWY Hold Delay",
                                                command=lambda: controller.show_frame(DepDelay22))
                            button2.pack()
                        if Delay_FLAG == True:
                            button5 = ttk.Button(self, text="Push/Start Delay",
                                                command=lambda: controller.show_frame(DepDelay222))
                            button5.pack()
                        if arr_delay_FLAG == True:
                            button3 = ttk.Button(self, text="Arrivals Delay",
                                                command=lambda: controller.show_frame(ArrivalDelay2))
                            button3.pack()
                        if Seq_FLAG == True:
                            button4 = ttk.Button(self, text="Sequence",
                                                command=lambda: controller.show_frame(Seq2))
                            button4.pack()
                        
                       
                        
                # ============================================================================#
                #                           CONVERGENCE                                       #
                # ============================================================================#
                #if convergenceFLAG ==True:
                class Conv2(tk.Frame):
                
                    def __init__(self, parent, controller):
                        tk.Frame.__init__(self, parent)
                        label = tk.Label(self, text="THROUGHPUT CONVERGENCE", font=LARGE_FONT)
                        label.pack(pady=10,padx=10)
                
                        button1 = ttk.Button(self, text="Back to Home",
                                            command=lambda: controller.show_frame(StartPage2))
                        button1.pack()
                        
                        
                        if Thr_FLAG == True:
                            button = ttk.Button(self, text="Throughput",
                                                command=lambda: controller.show_frame(Thr2))
                            button.pack()
                            
                        if Delay_FLAG == True:
                            button2 = ttk.Button(self, text="RWY Hold Delay",
                                                command=lambda: controller.show_frame(DepDelay22))
                            button2.pack()
                        if Delay_FLAG == True:
                            button5 = ttk.Button(self, text="Push/Start Delay",
                                                command=lambda: controller.show_frame(DepDelay222))
                            button5.pack()
                        if arr_delay_FLAG == True:
                            button3 = ttk.Button(self, text="Arrivals Delay",
                                                command=lambda: controller.show_frame(ArrivalDelay2))
                            button3.pack()
                        if Seq_FLAG == True:
                            button4 = ttk.Button(self, text="Sequence",
                                                command=lambda: controller.show_frame(Seq2))
                            button4.pack() 
                        
                        if ADA_buffer_FLAG == True:
                            button5 = ttk.Button(self, text="ADA Buffer",
                                                command=lambda: controller.show_frame(ADAbuffer))
                            button5.pack()
                
                       
                        thr_av_diff = pd.DataFrame(df_thr['Difference in thr averages'])
                        thr_av_diff=thr_av_diff.dropna(subset=['Difference in thr averages'])
                       
                        thr_av_diff['Tags'] = thr_av_diff['Difference in thr averages'].apply(lambda x: json.loads(x))
                        thr_av_difference = thr_av_diff['Tags'].tolist()
                        
                        
                        
                        print(len(thr_av_difference[0][0]))
                        print(len(thr_av_difference[0]))
                        thr_abc=[]
                        total_runs =[]
                        runs_Thr = []
                        for i in range (1,(len(thr_av_difference[0])+1)):
                            runs_Thr.append(i) 
                       
                        f = Figure(figsize=(5,5), dpi=100)
                        A = f.add_subplot(111)
                        
                        for j in range (0, len(thr_av_difference[0][0])):
                            for i in range (0, len(thr_av_difference[0])):            
                                total_thr = thr_av_difference[0][i][j]
                                runs_Thrx = runs_Thr[i]
                                total_runs.append(runs_Thrx)
                                thr_abc.append(total_thr) 
                            A.plot(total_runs,thr_abc)
                            thr_abc = []
                            total_runs =[]
                            
                        A.set_xlabel('No. of Runs')
                        A.set_ylabel('Difference in average throughput per hour')
                        A.grid(color='b', linestyle='-', linewidth=0.1) 
                        #A.title('Throughput')   
                        canvas = FigureCanvasTkAgg(f, self)
                        canvas.show()
                        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=False)
                
                        toolbar = NavigationToolbar2TkAgg(canvas, self)
                        toolbar.update()
                        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                ##            
                # ============================================================================#
                #                          THROUGHPUT                                         #
                # ============================================================================#
                #if Thr_FLAG == True:
                class Thr2(tk.Frame):
                
                    def __init__(self, parent, controller):
                        tk.Frame.__init__(self, parent)
                        label = tk.Label(self, text="THROUGHPUT", font=LARGE_FONT)
                        label.pack(pady=10,padx=10)
                
                        button1 = ttk.Button(self, text="Back to Home",
                                            command=lambda: controller.show_frame(StartPage2))
                        button1.pack()
                        
                        
                        if convergenceFLAG ==True:
                            button1 = ttk.Button(self, text="Convergence Throughput",
                                                command=lambda: controller.show_frame(Conv2))
                            button1.pack()
                        if Delay_FLAG == True:
                            button2 = ttk.Button(self, text="RWY Hold Delay",
                                                command=lambda: controller.show_frame(DepDelay22))
                            button2.pack()
                        
                            button5 = ttk.Button(self, text="Push/Start Delay",
                                                command=lambda: controller.show_frame(DepDelay222))
                            button5.pack()
                        if arr_delay_FLAG == True:
                            button3 = ttk.Button(self, text="Arrivals Delay",
                                                command=lambda: controller.show_frame(ArrivalDelay2))
                            button3.pack()
                        if Seq_FLAG == True:
                            button4 = ttk.Button(self, text="Sequence",
                                                command=lambda: controller.show_frame(Seq2))
                            button4.pack()
                        if ADA_buffer_FLAG == True:
                            button5 = ttk.Button(self, text="ADA Buffer",
                                                command=lambda: controller.show_frame(ADAbuffer2))
                            button5.pack()
                        
                        def create_first_df_thr():                   
                            df_thr_to_plot = pd.DataFrame()
                            df_thr_to_plot['Hour'] = df_thr['Hour']
                            x = 'Hour'
                            df_thr_to_plot['RUN 1'] = df_thr['Total Throughput']
                            y = 'RUN 1'
                            
                            df_thr_to_plot2 = df_thr_to_plot[[x,y]].groupby(x).sum()
                            
                            return df_thr_to_plot2
                        
                       
                        
                        def create_multiple_df_thr(df_thr_to_plot2, df_thr_input, name):
                            df_thr_to_plot2[name] = df_thr_input['Total Throughput']
                            df_thr_to_plot_temp = pd.DataFrame()
                            df_thr_to_plot_temp['Hour'] = df_thr_input['Hour']
                            a = 'Hour'
                            df_thr_to_plot_temp[name] = df_thr_input['Total Throughput']
                            b = name
                            df_thr_to_plot_temp = df_thr_to_plot_temp.dropna(subset=['Hour'])    
                            df_thr_to_plot_temp2 = df_thr_to_plot_temp[[a,b]].groupby(a).sum()    
                                    
                            df_thr_to_plot2[name] = df_thr_to_plot_temp2[name]
                            
                            
                            return df_thr_to_plot2 
                            
                        def plot_bar_thr(df_thr_to_plot2, A):                      
                            df_thr_to_plot2.plot(kind='bar', legend=False, ax=A) 
                        
                        def total_thr(df_thr_input):
                            total_thr = df_thr_input['Total Throughput']
                            total_thr = total_thr.reset_index()
                            return(total_thr)
                        def hour_thr(df_thr_input):
                            hour_Thr = df_thr_input['Hour'].tolist()
                            return(hour_Thr)
                            
                        
                        if m2>= 1:
                            f = Figure(figsize=(5,5), dpi=100)
                            A = f.add_subplot(111)
            #                throughput(df_thr, 'k')
            #                throughput(df_thr2, 'g')
                            
                            df_thr_to_plot2 = create_first_df_thr()
                            df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, df_thr2, 'RUN 2')
                            
                            
                            if OP_FLAG == True:
                                df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                                A.legend(('Model','Operational Data', 'Model 2'), loc = 'upper right')
                            else:
                                A.legend(('Model', 'Model 2'), loc = 'upper right')
                            if m2>= 2:
                                df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, df_thr3, 'RUN 3')
                                
                                if OP_FLAG ==True:
                                    df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                                    A.legend(('Model','Operational Data', 'Model 2', 'Model 3'), loc = 'upper right')
                                else:
                                    A.legend(('Model','Model 2', 'Model 3'), loc = 'upper right')
                                if m2>=3:
                                    #throughput(df_thr4, 'm') 
                                    df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, df_thr4, 'RUN 4')
                                    
                                    if OP_FLAG ==True:
                                        df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                                        A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4'), loc = 'upper right')
                                    else:
                                        A.legend(('Model','Model 2', 'Model 3','Model 4'), loc = 'upper right')
                                    if m2>=4:
                                        #throughput(df_thr5, 'y') 
                                        df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, df_thr5, 'RUN 5')
                                        
                                        if OP_FLAG ==True:
                                            df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                                            A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5'), loc = 'upper right')
                                        else:
                                            df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                                            A.legend(('Model','Model 2', 'Model 3','Model 4', 'Model 5'), loc = 'upper right')
                                        if m2>=5:
                                            df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, df_thr6, 'RUN 6')
                                            #throughput(df_thr6, color = 'purple')
                                            
                                            
                                            if OP_FLAG ==True:
                                                df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                                                A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                                            else:
                                                A.legend(('Model','Model 2', 'Model 3','Model 4','Model 5','Model 6'), loc = 'upper right')
                            plot_bar_thr(df_thr_to_plot2, A)
                            if OP_FLAG ==True:
                                df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                                A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                            A.set_xlabel('Hours of the day')
                            A.set_ylabel('No of A/C')
                            A.grid(color='b', linestyle='-', linewidth=0.1) 
                            #A.title('Throughput')   
                            canvas = FigureCanvasTkAgg(f, self)
                            canvas.show()
                            canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                    
                            toolbar = NavigationToolbar2TkAgg(canvas, self)
                            toolbar.update()
                            canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                        
                
                
                # ============================================================================#
                #                         DEPARTURE DELAY - RWY hold delay                    #
                # ============================================================================#
                #if Delay_FLAG == True:
                class DepDelay22(tk.Frame):
                
                    def __init__(self, parent, controller):
                        tk.Frame.__init__(self, parent)
                        label = tk.Label(self, text="RWY HOLD DELAY", font=LARGE_FONT)
                        label.pack(pady=10,padx=10)
                
                        button1 = ttk.Button(self, text="Back to Home",
                                            command=lambda: controller.show_frame(StartPage2))
                        button1.pack() 
                        
                        
                        if convergenceFLAG ==True:
                            button1 = ttk.Button(self, text="Convergence Throughput",
                                                command=lambda: controller.show_frame(Conv2))
                            button1.pack()
                        if Thr_FLAG == True:
                            button = ttk.Button(self, text="Throughput",
                                                command=lambda: controller.show_frame(Thr2))
                            button.pack()
                            
                        if Delay_FLAG == True:
                            button5 = ttk.Button(self, text="Push/Start Delay",
                                                command=lambda: controller.show_frame(DepDelay222))
                            button5.pack()                
                        if arr_delay_FLAG == True:
                            button3 = ttk.Button(self, text="Arrivals Delay",
                                                command=lambda: controller.show_frame(ArrivalDelay2))
                            button3.pack()
                        if Seq_FLAG == True:
                            button4 = ttk.Button(self, text="Sequence",
                                                command=lambda: controller.show_frame(Seq2))
                            button4.pack()
                        if ADA_buffer_FLAG == True:
                            button5 = ttk.Button(self, text="ADA Buffer",
                                                command=lambda: controller.show_frame(ADAbuffer2))
                            button5.pack()
                        if (df_delay.empty == True):
                            print("No departures. Nothing to show")
                        else:   
                            def delay(df_delay,df_dep_output):
                                
                                interval15=[]
                                df_ps_delay = pd.DataFrame()
                                # Add arrival delay values
                                df_ps_delay['RWYhold_Delay'] = df_delay['RWY HOLD Delay']
                                # print when those values occure
                                df_ps_delay['Time1'] = df_dep_output['Departure_RWY_ENTRY']
                                # round each time value to 15 minutes
                                df_ps_delay= df_ps_delay.dropna(subset=['Time1'])
                                time1_list = df_ps_delay['Time1'].tolist()
                                #a=[]
                                for a in time1_list:
                                    b = int(int(a/900)*900)
                                    interval15.append(b)
                                    
                                df_ps_delay['interval15'] = interval15
                                df_ps_delay = df_ps_delay.drop(columns=['Time1'])
                                #Group data by the time interval, if there are multiple values for the same time interval, take the mean.
                                df_ps_delay = df_ps_delay.groupby(['interval15'])['RWYhold_Delay'].mean()
                                #make the rolling average
                                df_ps_delay = df_ps_delay.reset_index()
                                df_ps_delay2 = df_ps_delay.rolling(window=4, on='interval15')['RWYhold_Delay'].mean()
                                RWYhold_Delay = df_ps_delay2.tolist()
                                df_ps_delay['DATE'] = pd.to_datetime(df_ps_delay['interval15'],unit='s')
                                df_ps_delay['DATE'] = df_ps_delay['DATE'].apply(lambda x: x.time())
                                
                                time_interval = df_ps_delay['DATE']
                                return {'a': time_interval,
                                        'b': RWYhold_Delay}
                            def print_Hold_delay(ab, color):
                                
                                H_delay_time = ab['a']
                                H_delay = ab['b']
                                
                                A = f.add_subplot(111)
                                A.plot(H_delay_time, H_delay, color)
                                #A.set_title('RWY HOLD DELAY', loc='left')
                                if OP_FLAG == True: #Plot Operational Data
                                    A.plot(OP_Time_H_Delay,OP_H_Delay, 'b')    
                                    A.legend(('Model','Operational Data'), loc = 'upper right')
                                    
                            
                         
                            
                            if m2 >=1:          
                                f = Figure(figsize=(5,5), dpi=100)
                                A = f.add_subplot(111)
                                hold_delay = delay(df_delay,df_dep_output)
                                print_Hold_delay(hold_delay,'k')
                                hold_delay2 = delay(df_delay2,df_dep_output2)
                                print_Hold_delay(hold_delay2,'g')
                                
                                
                                
                                if OP_FLAG == True:
                                    A.legend(('Model','Operational Data', 'Model 2'), loc = 'upper right')
                                else:
                                    A.legend(('Model', 'Model 2'), loc = 'upper right')
                                if m2>=2:
                                    hold_delay3 = delay(df_delay3,df_dep_output3)
                                    print_Hold_delay(hold_delay3,'c')
                                    
                                    
                                    if OP_FLAG == True:
                                        A.legend(('Model','Operational Data', 'Model 2', 'Model 3'), loc = 'upper right')
                                    else:
                                        A.legend(('Model', 'Model 2', 'Model 3'), loc = 'upper right')
                                    if m2>=3:
                                        hold_delay4 = delay(df_delay4,df_dep_output4)
                                        print_Hold_delay(hold_delay4,'m')
                                        
                                       
                                        
                                        if OP_FLAG == True:
                                            A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4'), loc = 'upper right')
                                        else:
                                            A.legend(('Model', 'Model 2', 'Model 3', 'Model 4'), loc = 'upper right')
                                        if m2>=4:
                                            hold_delay5 = delay(df_delay5,df_dep_output5)
                                            print_Hold_delay(hold_delay5,'y')
                                            
                                            
                                            if OP_FLAG == True:
                                                A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5'), loc = 'upper right')
                                            else:
                                                A.legend(('Model', 'Model 2', 'Model 3', 'Model 4', 'Model 5'), loc = 'upper right')
                                            if m2>=5:
                                                hold_delay6 = delay(df_delay6,df_dep_output6)
                                                print_Hold_delay(hold_delay6,color='purple')
                                                
                                                
                                                if OP_FLAG == True:
                                                    A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                                                else:
                                                    A.legend(('Model', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                                A.set_xlabel('Seconds of the day')
                                A.set_ylabel('Seconds of delay')                
                                A.grid(color='b', linestyle='-', linewidth=0.1)
            
                              
                                canvas = FigureCanvasTkAgg(f, self)
                                canvas.show()
                                canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                        
                                toolbar = NavigationToolbar2TkAgg(canvas, self)
                                toolbar.update()
                                canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                           
                               
                
                # ============================================================================#
                #                   DEPARTURE DELAY - PS delay                                #
                # ============================================================================#
                
                class DepDelay222(tk.Frame):
                
                    def __init__(self, parent, controller):
                        tk.Frame.__init__(self, parent)
                        label = tk.Label(self, text="PUSH/START DELAY", font=LARGE_FONT)
                        label.pack(pady=10,padx=10)
                
                        button1 = ttk.Button(self, text="Back to Home",
                                            command=lambda: controller.show_frame(StartPage2))
                        button1.pack() 
                        
                        
                        if convergenceFLAG ==True:
                            button1 = ttk.Button(self, text="Convergence Throughput",
                                                command=lambda: controller.show_frame(Conv2))
                            button1.pack()   
                        if Thr_FLAG == True:
                            button = ttk.Button(self, text="Throughput",
                                                command=lambda: controller.show_frame(Thr2))
                            button.pack()   
                        if Delay_FLAG == True:
                            button2 = ttk.Button(self, text="RWY HOLD DELAY",
                                                command=lambda: controller.show_frame(DepDelay22))
                            button2.pack()                
                        if arr_delay_FLAG == True:
                            button3 = ttk.Button(self, text="Arrivals Delay",
                                                command=lambda: controller.show_frame(ArrivalDelay2))
                            button3.pack()
                        if Seq_FLAG == True:
                            button4 = ttk.Button(self, text="Sequence",
                                                command=lambda: controller.show_frame(Seq2))
                            button4.pack()
                        if ADA_buffer_FLAG == True:
                            button5 = ttk.Button(self, text="ADA Buffer",
                                                command=lambda: controller.show_frame(ADAbuffer2))
                            button5.pack()
                        if (df_delay.empty == True):
                            print("No departures. Nothing to show")
                        else:   
                            def delay(df_delay, df_dep_output):                
                                interval15=[]
                                df_ps_delay = pd.DataFrame()
                                # Add arrival delay values
                                df_ps_delay['PS_Delay'] = df_delay['Push/Start Delay']
                                # print when those values occure
                                df_ps_delay['Time1'] = df_dep_output['Departure_RWY_ENTRY']
                                # round each time value to 15 minutes
                                df_ps_delay= df_ps_delay.dropna(subset=['Time1'])
                                time1_list = df_ps_delay['Time1'].tolist()
                                #a=[]
                                for a in time1_list:
                                    b = int(int(a/900)*900)
                                    interval15.append(b)
                                    
                                df_ps_delay['interval15'] = interval15
                                df_ps_delay = df_ps_delay.drop(columns=['Time1'])
                                #Group data by the time interval, if there are multiple values for the same time interval, take the mean.
                                df_ps_delay = df_ps_delay.groupby(['interval15'])['PS_Delay'].mean()
                                #make the rolling average
                                df_ps_delay = df_ps_delay.reset_index()
                                df_ps_delay2 = df_ps_delay.rolling(window=4, on='interval15')['PS_Delay'].mean()
                                PS_delay = df_ps_delay2.tolist()
                                df_ps_delay['DATE'] = pd.to_datetime(df_ps_delay['interval15'],unit='s')
                                df_ps_delay['DATE'] = df_ps_delay['DATE'].apply(lambda x: x.time())
                                
                                time_interval = df_ps_delay['DATE']
      
                                return {'c': time_interval,
                                        'd': PS_delay}
                                
                       
                            
                            
                            def print_PS_delay(ab, color):
                                PS_time = ab['c']
                                PS_delay = ab['d']
                                B = f.add_subplot(111)
                                B.plot(PS_time, PS_delay, color)
                                #B.set_title('PUSH/START DELAY', loc = 'right')
                                if OP_FLAG == True:#Plot Operational Data
                                    
                                    B.plot(OP_Time_PS_Delay,OP_PS_Delay, 'b')    
                                    #plt.legend(('Model','Operational Data'), loc = 'upper right')
                            
                                
                            if m2 >=1: 
                                f = Figure(figsize=(5,5), dpi=100)
                                B = f.add_subplot(111)
                                Push_delay = delay(df_delay, df_dep_output)
                                print_PS_delay(Push_delay,'k')
                                Push_delay2 = delay(df_delay2, df_dep_output2)
                                print_PS_delay(Push_delay2,'g')
                                
                                
                                if OP_FLAG == True:
                                    B.legend(('Model','Operational Data', 'Model 2'), loc = 'upper right')
                                else:
                                    B.legend(('Model', 'Model 2'), loc = 'upper right')
                                if m2>=2:
                                    Push_delay3 = delay(df_delay3, df_dep_output3)
                                    print_PS_delay(Push_delay3,'c')
                                    
                                    if OP_FLAG == True:
                                        B.legend(('Model','Operational Data', 'Model 2', 'Model 3'), loc = 'upper right')
                                    else:
                                        B.legend(('Model', 'Model 2', 'Model 3'), loc = 'upper right')
                                    if m2>=3:
                                        Push_delay4 = delay(df_delay4, df_dep_output4)
                                        print_PS_delay(Push_delay4,'m')
                                        
                                        
                                        if OP_FLAG == True:
                                            B.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4'), loc = 'upper right')
                                        else:
                                            B.legend(('Model', 'Model 2', 'Model 3', 'Model 4'), loc = 'upper right')
                                        if m2>=4:
                                            Push_delay5 = delay(df_delay5, df_dep_output5)
                                            print_PS_delay(Push_delay5,'y')
                                            
                                            
                                            if OP_FLAG == True:
                                                B.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5'), loc = 'upper right')
                                            else:
                                                B.legend(('Model', 'Model 2', 'Model 3', 'Model 4', 'Model 5'), loc = 'upper right')
                                            if m2>=5:
                                                Push_delay6 = delay(df_delay6, df_dep_output6)
                                                print_PS_delay(Push_delay6, color='purple')
                                                
                                               
                                                if OP_FLAG == True:
                                                    B.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                                                else:
                                                    B.legend(('Model', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                                
                                B.grid(color='b', linestyle='-', linewidth=0.1)
                                B.set_xlabel('Time')
                                B.set_ylabel('Seconds of delay')    
                                canvas = FigureCanvasTkAgg(f, self)
                                canvas.show()
                                canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                        
                                toolbar = NavigationToolbar2TkAgg(canvas, self)
                                toolbar.update()
                                canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                           
                
                # ============================================================================#
                #                          ARRIVAL DELAY                                      # 
                # ============================================================================#
                #if arr_delay_FLAG == True:
                class ArrivalDelay2(tk.Frame):
                
                    def __init__(self, parent, controller):
                        tk.Frame.__init__(self, parent)
                        label = tk.Label(self, text="ARRIVALS DELAY", font=LARGE_FONT)
                        label.pack(pady=10,padx=10)
                
                        button1 = ttk.Button(self, text="Back to Home",
                                            command=lambda: controller.show_frame(StartPage2))
                        button1.pack()
                        
                        
                        if convergenceFLAG ==True:
                            button1 = ttk.Button(self, text="Convergence Throughput",
                                                command=lambda: controller.show_frame(Conv2))
                            button1.pack()
                        if Thr_FLAG == True:
                            button = ttk.Button(self, text="Throughput",
                                                command=lambda: controller.show_frame(Thr2))
                            button.pack()
                        if Delay_FLAG == True:
                            button2 = ttk.Button(self, text="RWY Hold Delay",
                                                command=lambda: controller.show_frame(DepDelay22))
                            button2.pack()
                        if Delay_FLAG == True:
                            button5 = ttk.Button(self, text="Push/Start Delay",
                                                command=lambda: controller.show_frame(DepDelay222))
                            button5.pack()
                        
                        if Seq_FLAG == True:
                            button4 = ttk.Button(self, text="Sequence",
                                                command=lambda: controller.show_frame(Seq2))
                            button4.pack()
                        if ADA_buffer_FLAG == True:
                            button5 = ttk.Button(self, text="ADA Buffer",
                                                command=lambda: controller.show_frame(ADAbuffer2))
                            button5.pack()
                
                        
                        
                        if (df_delay.empty == True):
                                print("No departures. Nothing to show")
                        else:     
                            def ArrDelay(df_delay, df_arr_output):
                                interval15=[]
                                df_arr_delay = pd.DataFrame()
                                # Add arrival delay values
                                df_arr_delay['ARR_Delay'] = df_delay['Arrival Delay']
                                # print when those values occure
                                df_arr_delay['Time1'] = df_arr_output['ACTUAL Landing Time']
                                # round each time value to 15 minutes
                                df_arr_delay= df_arr_delay.dropna(subset=['Time1'])
                                time1_list = df_arr_delay['Time1'].tolist()
                                #a=[]
                                for a in time1_list:
                                    b = int(int(a/900)*900)
                                    interval15.append(b)
                                    
                                df_arr_delay['interval15'] = interval15
                                df_arr_delay = df_arr_delay.drop(columns=['Time1'])
                                #Group data by the time interval, if there are multiple values for the same time interval, take the mean.
                                df_arr_delay = df_arr_delay.groupby(['interval15'])['ARR_Delay'].mean()
                                #make the rolling average
                                df_arr_delay = df_arr_delay.reset_index()
                                df_arr_delay2 = df_arr_delay.rolling(window=4, on='interval15')['ARR_Delay'].mean()
                                ARR_delay = df_arr_delay2.tolist()
                                df_arr_delay['DATE'] = pd.to_datetime(df_arr_delay['interval15'],unit='s')
                                df_arr_delay['DATE'] = df_arr_delay['DATE'].apply(lambda x: x.time())
                                
                                time_interval = df_arr_delay['DATE']
                                #        return(H_delay_time,H_delay,PS_time,PS_delay)
                                return {'a': time_interval,
                                        'b': ARR_delay}
                           
                            def plotArrDelay(ab, color):            
                                arr_delay_time = ab['a']
                                arr_delay = ab['b'] 
                                A = f.add_subplot(111)                      
                                A.plot(arr_delay_time, arr_delay, color)     
                            
                            
                                 
                            if m2 >=1:          
                                f = Figure(figsize=(5,5), dpi=100)
                                A = f.add_subplot(111)  
                                plotArrDelay(ArrDelay(df_delay, df_arr_output),'k')
                                plotArrDelay(ArrDelay(df_delay2, df_arr_output2),'g')
                                
                                A.legend(('Model', 'Model 2'), loc = 'upper right')
                                if m2>=2:                        
                                    plotArrDelay(ArrDelay(df_delay3, df_arr_output3),'c')
                                    
                                    A.legend(('Model', 'Model 2', 'Model 3'), loc = 'upper right')
                                    if m2>=3:
                                        plotArrDelay(ArrDelay(df_delay4, df_arr_output4),'m')
                                       
                                        A.legend(('Model', 'Model 2', 'Model 3', 'Model 4'), loc = 'upper right')
                                        if m2>=4:
                                            plotArrDelay(ArrDelay(df_delay5, df_arr_output5),'y')
                                            
                                            A.legend(('Model', 'Model 2', 'Model 3', 'Model 4', 'Model 5'), loc = 'upper right')
                                            if m2>=5:
                                                plotArrDelay(ArrDelay(df_delay6, df_arr_output6),'purple')
                                                
                                                A.legend(('Model', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                                A.set_xlabel('Time')
                                A.set_ylabel('Seconds of delay')                
                                A.grid(color='b', linestyle='-', linewidth=0.1)
                                canvas = FigureCanvasTkAgg(f, self)
                                canvas.show()
                                canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                        
                                toolbar = NavigationToolbar2TkAgg(canvas, self)
                                toolbar.update()
                                canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                                    
                           
                        
                # ============================================================================#
                #                           SEQUENCE                                          #
                # ============================================================================#
                #if Seq_FLAG == True:
                class Seq2(tk.Frame):
                
                    def __init__(self, parent, controller):
                        tk.Frame.__init__(self, parent)
                        label = tk.Label(self, text="SEQUENCE", font=LARGE_FONT)
                        label.pack(pady=10,padx=10)
                        
                        button1 = ttk.Button(self, text="Back to Home",
                                            command=lambda: controller.show_frame(StartPage2))
                        button1.pack() 
                        if convergenceFLAG ==True:
                            button1 = ttk.Button(self, text="Convergence Throughput",
                                                command=lambda: controller.show_frame(Conv2))
                            button1.pack()
                        if Thr_FLAG == True:
                            button = ttk.Button(self, text="Throughput",
                                                command=lambda: controller.show_frame(Thr2))
                            button.pack()
                        
                        if Delay_FLAG == True:
                            button2 = ttk.Button(self, text="RWY Hold Delay",
                                                command=lambda: controller.show_frame(DepDelay22))
                            button2.pack()
                        if Delay_FLAG == True:
                            button5 = ttk.Button(self, text="Push/Start Delay",
                                                command=lambda: controller.show_frame(DepDelay222))
                            button5.pack()
                        
                        if arr_delay_FLAG == True:
                            button3 = ttk.Button(self, text="Arrivals Delay",
                                                command=lambda: controller.show_frame(ArrivalDelay2))
                            button3.pack()
                            
                        if ADA_buffer_FLAG == True:
                            button5 = ttk.Button(self, text="ADA Buffer",
                                                command=lambda: controller.show_frame(ADAbuffer2))
                            button5.pack()
                            
                        
                        def sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, number):
                
                            df_sequence = pd.DataFrame()
                            #Arrivals 
                            
                            df_sequence['ARRIVAL'] = df_arr_output['ACTUAL Landing Time'] + (df_arr_output['AROT']/2)
                            df_sequence['ARRIVAL_error'] = df_arr_output['AROT']/2
                            df_sequence['ARRIVAL_spacing'] = df_arr_output['ACTUAL Landing Time']
                            df_temp = pd.DataFrame()
                            df_temp['MAX Constraint'] = df_arr_output['MAX Constraint']
                            df_temp = df_temp.drop([0])
                            df_temp = df_temp.reset_index()
                            df_temp = df_temp.drop(columns=['index'])
                            df_temp['Arrival_ZERO'] = 0
                            df_sequence['ARRIVAL_spacing_error'] = df_temp['MAX Constraint']           
                            df_sequence['Arrival_ZERO'] = df_temp['Arrival_ZERO']
                            
                            #Positions
                            df_sequence['main_position'] = number
                            df_sequence['arr_spacing_position'] = number+0.005
                            
                            #Annotation
                            
                            df_temp = pd.DataFrame()
                            df_temp['ARR_ID'] = df_arr_output['Arrival ID'].astype(str)
                            df_temp['ARR_WAKE'] = df_rwy_calcs['ARRIVAL actual WAKE'].astype(str)
                            df_temp['ARR_DELAY'] = df_arr_output['Arrival DELAY'].astype(str)
                            df_temp['ARRIVAL_LABEL'] = 'ID = ' + df_temp['ARR_ID'] + ' | WAKE = ' + df_temp['ARR_WAKE'] + ' | Delay = ' + df_temp['ARR_DELAY']
                            
                            df_sequence['ARRIVAL_LABEL'] = df_temp['ARRIVAL_LABEL']
                            
                            df_temp = pd.DataFrame()
                            df_temp['reason'] = df_arr_output['MAX Constraint Label']
                            df_temp = df_temp.drop([0])
                            df_temp = df_temp.reset_index()
                            df_temp = df_temp.drop(columns=['index'])
                            df_temp['value'] = df_sequence['ARRIVAL_spacing_error'].astype(str)
                            df_temp['ARRIVAL_spacing_LABEL'] = df_temp['reason'] + ' : ' + df_temp['value']
                            
                            df_sequence['ARRIVAL_spacing_LABEL'] = df_temp['ARRIVAL_spacing_LABEL']
                            
                            # LISTS to plot
                            
                            #----ARRIVALS----#
                            main_arrival = df_sequence['ARRIVAL'].tolist()
                            main_arrival_error = df_sequence['ARRIVAL_error'].tolist()
                            arrival_spacing = df_sequence['ARRIVAL_spacing'].tolist()
                            arrival_spacing_error = df_sequence['ARRIVAL_spacing_error'].tolist()
                            arrival_zero = df_sequence['Arrival_ZERO'].tolist()
                            #-Labels:
                            arrival_label = df_sequence['ARRIVAL_LABEL'].tolist()
                            arrival_spacing_label = df_sequence['ARRIVAL_spacing_LABEL'].tolist()
                            
                            #-----POSITIONS------#
                            main_data_position = df_sequence['main_position'].tolist()
                            arrival_spacing_position = df_sequence['arr_spacing_position'].tolist()
                            #Data prep for tags
                            labels = arrival_label +  arrival_spacing_label  
                            labels_y = main_data_position +  arrival_spacing_position 
                            labels_x = main_arrival +  arrival_spacing 
                        
                        
                            return{'0a': arrival_zero,
                                   '1' : main_arrival, 
                                   '2' : main_arrival_error,
                                   '3' : arrival_spacing,
                                   '4' : arrival_spacing_error,                   
                                   '9' : main_data_position,
                                   '10' : arrival_spacing_position,                   
                                   '12' : labels,
                                   '13' : labels_y,
                                   '14' : labels_x}
                         
                        def sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, number):
                            df_sequence = pd.DataFrame()
                            #Departure
                            
                            df_sequence['DEPARTURES'] = df_dep_output['Departure_RWY_ENTRY'] + (df_dep_output['DROT']/2)
                            df_sequence['DEPARTURES_error'] = df_dep_output['DROT']/2
                            df_sequence['DEPARTURES_spacing'] = df_dep_output['Departure_RWY_ENTRY']
                            df_temp = pd.DataFrame()
                            df_temp['Dep MIN Separation'] = df_dep_output['Dep MIN Separation']
                            df_temp=df_temp.drop([0])
                            df_temp = df_temp.reset_index()
                            df_temp = df_temp.drop(columns=['index'])
                            df_temp['Departure_ZERO'] = 0
                            df_sequence['DEPARTURES_spacing_error'] = df_temp['Dep MIN Separation']            
                            df_sequence['Departure_ZERO'] = df_temp['Departure_ZERO']
                                        
                            #Positions
                            df_sequence['main_position'] = number
                            df_sequence['dep_spacing_position'] = number-0.005
                            
                            df_temp = pd.DataFrame()
                            df_temp['DEP_ID'] = df_dep_output['Departure ID'].astype(str)
                            df_temp['DEP_SID'] = df_dep_output['SID GROUP'].astype(str)
                            df_temp['DEP_WAKE'] = df_dep_output['WAKE'].astype(str)
                            df_temp['DEP_DELAY1'] = df_dep_output['DELAY DepSTANDqueue'] + df_dep_output['DELAY TAXIhold'] + df_dep_output['DELAY RWYqueue']
                            df_temp['DEP_DELAY'] = df_temp['DEP_DELAY1'].astype(str)
                            df_temp['DEPARTURE_LABEL'] = 'ID = ' + df_temp['DEP_ID'] + ' | SID = ' +df_temp['DEP_SID'] +' | WAKE = ' + df_temp['DEP_WAKE'] + ' | Delay = ' + df_temp['DEP_DELAY']
                            
                            df_sequence['DEPARTURE_LABEL'] = df_temp['DEPARTURE_LABEL']
                            
                            df_temp = pd.DataFrame()
                            df_temp['reason'] = df_dep_output['Dep MIN Separation Label']
                            df_temp = df_temp.drop([0])
                            df_temp = df_temp.reset_index()
                            df_temp = df_temp.drop(columns=['index'])
                            df_temp['value'] = df_sequence['DEPARTURES_spacing_error'].astype(str)
                            df_temp['DEPARTURES_spacing_LABEL'] = df_temp['reason'] + ' : ' + df_temp['value']
                            
                            df_sequence['DEPARTURES_spacing_LABEL'] = df_temp['DEPARTURES_spacing_LABEL']
                            
                            # LISTS to plot
                           
                            #----DEPARTURES----#
                            main_departure = df_sequence['DEPARTURES'].tolist()
                            main_departure_error = df_sequence['DEPARTURES_error'].tolist()
                            departure_spacing = df_sequence['DEPARTURES_spacing'].tolist()
                            departure_spacing_error = df_sequence['DEPARTURES_spacing_error'].tolist()
                            departure_zero = df_sequence['Departure_ZERO']
                            #-Labels:
                            departure_label = df_sequence['DEPARTURE_LABEL'].tolist()
                            departure_spacing_label = df_sequence['DEPARTURES_spacing_LABEL'].tolist()
                            
                            #-----POSITIONS------#
                            main_data_position = df_sequence['main_position'].tolist()
                            departure_spacing_position = df_sequence['dep_spacing_position'].tolist()
                            
                            
                            
                            #Data prep for tags
                            labels = departure_label +  departure_spacing_label
                            labels_y = main_data_position + departure_spacing_position
                            labels_x = main_departure +  departure_spacing
                        
                        
                            return{'0b': departure_zero,
                                   '5' : main_departure, 
                                   '6' : main_departure_error,
                                   '7' : departure_spacing,
                                   '8' : departure_spacing_error,
                                   '9' : main_data_position,
                                   '11' : departure_spacing_position,
                                   '12' : labels,
                                   '13' : labels_y,
                                   '14' : labels_x}
                        
                        def sequence(df_arr_output, df_rwy_calcs, df_dep_output, number): ######MIX MODE
                            
                            df_sequence = pd.DataFrame()
                            #Arrivals 
                            
                            df_sequence['ARRIVAL'] = df_arr_output['ACTUAL Landing Time'] + (df_arr_output['AROT']/2)
                            df_sequence['ARRIVAL_error'] = df_arr_output['AROT']/2
                            df_sequence['ARRIVAL_spacing'] = df_arr_output['ACTUAL Landing Time']
                            df_temp = pd.DataFrame()
                            df_temp['MAX Constraint'] = df_arr_output['MAX Constraint']
                            df_temp = df_temp.drop([0])
                            df_temp = df_temp.reset_index()
                            df_temp = df_temp.drop(columns=['index'])
                            df_temp['Arrival_ZERO'] = 0
                            df_sequence['ARRIVAL_spacing_error'] = df_temp['MAX Constraint']           
                            df_sequence['Arrival_ZERO'] = df_temp['Arrival_ZERO']
                            
                            
                            #Departure
                            
                            df_sequence['DEPARTURES'] = df_dep_output['Departure_RWY_ENTRY'] + (df_dep_output['DROT']/2)
                            df_sequence['DEPARTURES_error'] = df_dep_output['DROT']/2
                            df_sequence['DEPARTURES_spacing'] = df_dep_output['Departure_RWY_ENTRY']
                            df_temp = pd.DataFrame()
                            df_temp['Dep MIN Separation'] = df_dep_output['Dep MIN Separation']
                            df_temp=df_temp.drop([0])
                            df_temp = df_temp.reset_index()
                            df_temp = df_temp.drop(columns=['index'])
                            df_temp['Departure_ZERO'] = 0
                            df_sequence['DEPARTURES_spacing_error'] = df_temp['Dep MIN Separation']            
                            df_sequence['Departure_ZERO'] = df_temp['Departure_ZERO']
                                        
                            #Positions
                            df_sequence['main_position'] = number
                            df_sequence['arr_spacing_position'] = number+0.005
                            df_sequence['dep_spacing_position'] = number-0.005
                            
                            
                            #Annotation
                            
                            df_temp = pd.DataFrame()
                            df_temp['ARR_ID'] = df_arr_output['Arrival ID'].astype(str)
                            df_temp['ARR_WAKE'] = df_rwy_calcs['ARRIVAL actual WAKE'].astype(str)
                            df_temp['ARR_DELAY'] = df_arr_output['Arrival DELAY'].astype(str)
                            df_temp['ARRIVAL_LABEL'] = 'ID = ' + df_temp['ARR_ID'] + ' | WAKE = ' + df_temp['ARR_WAKE'] + ' | Delay = ' + df_temp['ARR_DELAY']
                            
                            df_sequence['ARRIVAL_LABEL'] = df_temp['ARRIVAL_LABEL']
                            
                            df_temp = pd.DataFrame()
                            df_temp['DEP_ID'] = df_dep_output['Departure ID'].astype(str)
                            df_temp['DEP_SID'] = df_dep_output['SID GROUP'].astype(str)
                            df_temp['DEP_WAKE'] = df_dep_output['WAKE'].astype(str)
                            df_temp['DEP_DELAY1'] = df_dep_output['DELAY DepSTANDqueue'] + df_dep_output['DELAY TAXIhold'] + df_dep_output['DELAY RWYqueue']
                            df_temp['DEP_DELAY'] = df_temp['DEP_DELAY1'].astype(str)
                            df_temp['DEPARTURE_LABEL'] = 'ID = ' + df_temp['DEP_ID'] + ' | SID = ' +df_temp['DEP_SID'] +' | WAKE = ' + df_temp['DEP_WAKE'] + ' | Delay = ' + df_temp['DEP_DELAY']
                            
                            df_sequence['DEPARTURE_LABEL'] = df_temp['DEPARTURE_LABEL']
                            
                            df_temp = pd.DataFrame()
                            df_temp['reason'] = df_arr_output['MAX Constraint Label']
                            df_temp = df_temp.drop([0])
                            df_temp = df_temp.reset_index()
                            df_temp = df_temp.drop(columns=['index'])
                            df_temp['value'] = df_sequence['ARRIVAL_spacing_error'].astype(str)
                            df_temp['ARRIVAL_spacing_LABEL'] = df_temp['reason'] + ' : ' + df_temp['value']
                            
                            df_sequence['ARRIVAL_spacing_LABEL'] = df_temp['ARRIVAL_spacing_LABEL']
                            
                            df_temp = pd.DataFrame()
                            df_temp['reason'] = df_dep_output['Dep MIN Separation Label']
                            df_temp = df_temp.drop([0])
                            df_temp = df_temp.reset_index()
                            df_temp = df_temp.drop(columns=['index'])
                            df_temp['value'] = df_sequence['DEPARTURES_spacing_error'].astype(str)
                            df_temp['DEPARTURES_spacing_LABEL'] = df_temp['reason'] + ' : ' + df_temp['value']
                            
                            df_sequence['DEPARTURES_spacing_LABEL'] = df_temp['DEPARTURES_spacing_LABEL']
                            
                            # LISTS to plot
                            
                            #----ARRIVALS----#
                            main_arrival = df_sequence['ARRIVAL'].tolist()
                            main_arrival_error = df_sequence['ARRIVAL_error'].tolist()
                            arrival_spacing = df_sequence['ARRIVAL_spacing'].tolist()
                            arrival_spacing_error = df_sequence['ARRIVAL_spacing_error'].tolist()
                            arrival_zero = df_sequence['Arrival_ZERO'].tolist()
                            #-Labels:
                            arrival_label = df_sequence['ARRIVAL_LABEL'].tolist()
                            arrival_spacing_label = df_sequence['ARRIVAL_spacing_LABEL'].tolist()
                            
                            #----DEPARTURES----#
                            main_departure = df_sequence['DEPARTURES'].tolist()
                            main_departure_error = df_sequence['DEPARTURES_error'].tolist()
                            departure_spacing = df_sequence['DEPARTURES_spacing'].tolist()
                            departure_spacing_error = df_sequence['DEPARTURES_spacing_error'].tolist()
                            departure_zero = df_sequence['Departure_ZERO']
                            #-Labels:
                            departure_label = df_sequence['DEPARTURE_LABEL'].tolist()
                            departure_spacing_label = df_sequence['DEPARTURES_spacing_LABEL'].tolist()
                            
                            #-----POSITIONS------#
                            main_data_position = df_sequence['main_position'].tolist()
                            arrival_spacing_position = df_sequence['arr_spacing_position'].tolist()
                            departure_spacing_position = df_sequence['dep_spacing_position'].tolist()
                            
                            
                            
                            #Data prep for tags
                            labels = arrival_label + departure_label + arrival_spacing_label  + departure_spacing_label
                            labels_y = main_data_position + main_data_position + arrival_spacing_position + departure_spacing_position
                            labels_x = main_arrival + main_departure + arrival_spacing + departure_spacing
                        
                        
                            return{'0a': arrival_zero,
                                   '0b': departure_zero,
                                   '1' : main_arrival, 
                                   '2' : main_arrival_error,
                                   '3' : arrival_spacing,
                                   '4' : arrival_spacing_error,
                                   '5' : main_departure, 
                                   '6' : main_departure_error,
                                   '7' : departure_spacing,
                                   '8' : departure_spacing_error,
                                   '9' : main_data_position,
                                   '10' : arrival_spacing_position,
                                   '11' : departure_spacing_position,
                                   '12' : labels,
                                   '13' : labels_y,
                                   '14' : labels_x}
                        
                        #------ ARRIVALS only -------#
                        
                        if df_dep_output.empty ==True:
                              
                            if df_dep_output2.empty ==True:#arr only
                                if m2==1: # two arr only comparison
                                    f = Figure(figsize=(5,5), dpi=100)  
                                    A = f.add_subplot(111)
                                    ax = f.add_subplot(111)
                                    
                                    labels =  sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']  + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                                    tag_text_use = np.array(list(labels))
                                    labels_y = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13']
                                    labels_x = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14']
                                    
                                    tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                    #plt.axhline(y = 10, color='w')                    
                                    #A.axhline(y=0.5, color='w')
                                       
                                    
                                    A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='maroon', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='indigo', fmt='o', markersize=8, capsize=10, )
                                    
                                                    
                                    annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                    bbox=dict(boxstyle="round", fc="w"),
                                                    arrowprops=dict(arrowstyle="->"))
                                    annot.set_visible(False)
                                    
                                
                                    def update_annot(ind):
                                    
                                        pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                        annot.xy = pos
                                        text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                        annot.set_text(text)
                                    
                                
                                                    
                                    def hover(event):
                                        vis = annot.get_visible()
                                        if event.inaxes == ax:
                                            cont, ind = tags_main_data.contains(event)
                                            if cont:
                                                update_annot(ind)
                                                annot.set_visible(True)
                                                canvas.draw_idle()
                                            else:
                                                if vis:
                                                    annot.set_visible(False)
                                                    canvas.draw_idle()
                                    
                                
                                    
                                    A.set_xlabel('Seconds of the day')        
                                    A.axes.get_yaxis().set_visible(False) 
                                    A.grid(color='b', linestyle='-', linewidth=0.1)    
                                    A.legend(("Legend","Arrivals M1","Arrivals Spacing M1","Arrivals M2", "Arrivals Spacing M2"), loc = 'upper right')
                                    #plt.title("Sequence analysis")
                                    canvas = FigureCanvasTkAgg(f, self)
                                    canvas.show()
                                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                    canvas.mpl_connect("motion_notify_event", hover)
                                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                                    toolbar.update()
                                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                                elif m2==2: # 3 ARRonly comparison
                                    f = Figure(figsize=(5,5), dpi=100)  
                                    A = f.add_subplot(111)
                                    ax = f.add_subplot(111)
                                    
                                    labels =  sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']  + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] + sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['12']
                                    tag_text_use = np.array(list(labels))
                                    labels_y = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] + sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['13'] 
                                    labels_x = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] + sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['14']
                                    
                                    tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                    #plt.axhline(y = 10, color='w')                    
                                    #A.axhline(y=0.5, color='w')
                                       
                                    
                                    A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='maroon', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='indigo', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['1'], sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2'], sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2']], color='salmon', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['3'], sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['10'], xerr=[sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0a'], sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['4']], color='orchid', fmt='o', markersize=8, capsize=10, )
                                                    
                                    annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                    bbox=dict(boxstyle="round", fc="w"),
                                                    arrowprops=dict(arrowstyle="->"))
                                    annot.set_visible(False)
                                    
                                
                                    def update_annot(ind):
                                    
                                        pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                        annot.xy = pos
                                        text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                        annot.set_text(text)
                                    
                                
                                                    
                                    def hover(event):
                                        vis = annot.get_visible()
                                        if event.inaxes == ax:
                                            cont, ind = tags_main_data.contains(event)
                                            if cont:
                                                update_annot(ind)
                                                annot.set_visible(True)
                                                canvas.draw_idle()
                                            else:
                                                if vis:
                                                    annot.set_visible(False)
                                                    canvas.draw_idle()
                                    
                                
                                    
                                    A.set_xlabel('Seconds of the day')        
                                    A.axes.get_yaxis().set_visible(False) 
                                    A.grid(color='b', linestyle='-', linewidth=0.1)    
                                    A.legend(("Legend","Arrivals M1","Arrivals Spacing M1","Arrivals M2", "Arrivals Spacing M2","Arrivals M3", "Arrivals Spacing M3"), loc = 'upper right')
                                    #plt.title("Sequence analysis")
                                    canvas = FigureCanvasTkAgg(f, self)
                                    canvas.show()
                                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                    canvas.mpl_connect("motion_notify_event", hover)
                                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                                    toolbar.update()
                                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                                    #f.canvas.figure.savefig('sequence.png')
                            
                            
                            elif df_arr_output2.empty ==True: # ARR only + DEP only                    
                                f = Figure(figsize=(5,5), dpi=100)  
                                A = f.add_subplot(111)
                                ax = f.add_subplot(111)
                                
                                labels =  sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']  + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                                tag_text_use = np.array(list(labels))
                                labels_y = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                                labels_x = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] 
                                                        
                                tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                #A.axhline(y=0.5, color='w')
                                
                                A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['3'], sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['4'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['5'], sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['6']], color='r', fmt='o', markersize=8, capsize=10, )
                                A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['8'], sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['10'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['11'], sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['9']], color='purple', fmt='o', markersize=8, capsize=10)
                                
                                A.errorbar(sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                                A.errorbar(sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], fmt='o', markersize=8, capsize=10)
            
                                annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                bbox=dict(boxstyle="round", fc="w"),
                                                arrowprops=dict(arrowstyle="->"))
                                annot.set_visible(False)
                                
                            
                                def update_annot(ind):
                                
                                    pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                    annot.xy = pos
                                    text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                    annot.set_text(text)
                                
                            
                                                
                                def hover(event):
                                    vis = annot.get_visible()
                                    if event.inaxes == ax:
                                        cont, ind = tags_main_data.contains(event)
                                        if cont:
                                            update_annot(ind)
                                            annot.set_visible(True)
                                            canvas.draw_idle()
                                        else:
                                            if vis:
                                                annot.set_visible(False)
                                                canvas.draw_idle()
                                
                            
                                
                                A.set_xlabel('Seconds of the day')        
                                A.axes.get_yaxis().set_visible(False) 
                                A.grid(color='b', linestyle='-', linewidth=0.1)    
                                A.legend(("Legend","Arrivals M1","Arrivals Spacing M1","Departures M2", "Departures Spacing M2"), loc = 'upper right')
                                #plt.title("Sequence analysis")
                                canvas = FigureCanvasTkAgg(f, self)
                                canvas.show()
                                canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                canvas.mpl_connect("motion_notify_event", hover)
                                toolbar = NavigationToolbar2TkAgg(canvas, self)
                                toolbar.update()
                                canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                                
                            else:#ARR only + MIXED
                                f = Figure(figsize=(5,5), dpi=100)  
                                A = f.add_subplot(111)
                                ax = f.add_subplot(111)
                                
                                
                                labels =  sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']  + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                                tag_text_use = np.array(list(labels))
                                labels_y = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                                labels_x = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] 
                                                        
                                tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                #A.axhline(y=0.5, color='w')
                                                    
                                A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                                A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
            
                                A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                                A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                                A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], fmt='o', markersize=8, capsize=10)
                                A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
            
                                annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                bbox=dict(boxstyle="round", fc="w"),
                                                arrowprops=dict(arrowstyle="->"))
                                annot.set_visible(False)
                                def update_annot2(ind):
                                
                                    pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                    annot.xy = pos
                                    text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                    annot.set_text(text)
                                def hover(event):
                                    vis = annot.get_visible()
                                    if event.inaxes == ax:
                                        cont, ind = tags_main_data.contains(event)
                                        if cont:
                                            update_annot2(ind)
                                            annot.set_visible(True)
                                            canvas.draw_idle()
                                        else:
                                            if vis:
                                                annot.set_visible(False)
                                                canvas.draw_idle()
                                A.set_xlabel('Seconds of the day')        
                                A.axes.get_yaxis().set_visible(False)                     
                                A.grid(color='b', linestyle='-', linewidth=0.1)    
                                A.legend(("Legend","Arrivals M1","Arrivals Spacing M1","Departures M2","Arrivals M2", "Departures Spacing M2","Arrivals Spacing M2"), loc = 'upper right')
                                #plt.title("Sequence analysis")
                                canvas = FigureCanvasTkAgg(f, self)
                                canvas.show()
                                canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                canvas.mpl_connect("motion_notify_event", hover)
                                toolbar = NavigationToolbar2TkAgg(canvas, self)
                                toolbar.update()
                                canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                                    
                           
                        
                        #-------DEPARTURES only --------#
                        
                        elif df_arr_output.empty ==True:
                                        
                                              
                            if df_dep_output2.empty ==True: # DEPonly + ARRonly                    
                                f = Figure(figsize=(5,5), dpi=100)  
                                A = f.add_subplot(111)
                                ax = f.add_subplot(111)
                                
                                labels =  sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']+ sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12']
                                tag_text_use = np.array(list(labels))
                                labels_x = sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14']
                                labels_y = sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13']
                                tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                #A.axhline(y=0.5, color='w')
                                   
                                A.errorbar(sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                                A.errorbar(sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                                
                                A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                                A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                              
                                                
                                annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                bbox=dict(boxstyle="round", fc="w"),
                                                arrowprops=dict(arrowstyle="->"))
                                annot.set_visible(False)
                                
                            
                                def update_annot(ind):
                                
                                    pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                    annot.xy = pos
                                    text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                    annot.set_text(text)
                                
                            
                                                
                                def hover(event):
                                    vis = annot.get_visible()
                                    if event.inaxes == ax:
                                        cont, ind = tags_main_data.contains(event)
                                        if cont:
                                            update_annot(ind)
                                            annot.set_visible(True)
                                            canvas.draw_idle()
                                        else:
                                            if vis:
                                                annot.set_visible(False)
                                                canvas.draw_idle()
                                
                            
                                
                                A.set_xlabel('Seconds of the day')        
                                A.axes.get_yaxis().set_visible(False) 
                                A.grid(color='b', linestyle='-', linewidth=0.1)    
                                A.legend(("Legend","Departures M1","Departures Spacing M1","Arrivals M2", "Arrivals Spacing M2"), loc = 'upper right')
                                #plt.title("Sequence analysis")
                                canvas = FigureCanvasTkAgg(f, self)
                                canvas.show()
                                canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                canvas.mpl_connect("motion_notify_event", hover)
                                toolbar = NavigationToolbar2TkAgg(canvas, self)
                                toolbar.update()
                                canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                            
                            elif df_arr_output2.empty ==True: # DEPonly + DEPonly
                                f = Figure(figsize=(5,5), dpi=100)  
                                A = f.add_subplot(111)
                                ax = f.add_subplot(111)
                                
                                labels =  sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']  + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                                tag_text_use = np.array(list(labels))
                                labels_y = sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                                labels_x = sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] 
                                                        
                                tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                #A.axhline(y=0.5, color='w')
                                
                                A.errorbar(sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                                A.errorbar(sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
            
                                A.errorbar(sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='darkgreen', fmt='o', markersize=8, capsize=10)    
                                A.errorbar(sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], color='indigo',fmt = 'o', markersize=8, capsize=10)
            
            
                                
                                annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                bbox=dict(boxstyle="round", fc="w"),
                                                arrowprops=dict(arrowstyle="->"))
                                annot.set_visible(False)
                                
                            
                                def update_annot(ind):
                                
                                    pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                    annot.xy = pos
                                    text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                    annot.set_text(text)
                                
                            
                                                
                                def hover(event):
                                    vis = annot.get_visible()
                                    if event.inaxes == ax:
                                        cont, ind = tags_main_data.contains(event)
                                        if cont:
                                            update_annot(ind)
                                            annot.set_visible(True)
                                            canvas.draw_idle()
                                        else:
                                            if vis:
                                                annot.set_visible(False)
                                                canvas.draw_idle()
                                
                            
                                
                                A.set_xlabel('Seconds of the day')        
                                A.axes.get_yaxis().set_visible(False) 
                                A.grid(color='b', linestyle='-', linewidth=0.1)    
                                A.legend(("Legend","Departures M1","Departures Spacing M1","Departures M2", "Departures Spacing M2"), loc = 'upper right')
                                #plt.title("Sequence analysis")
                                canvas = FigureCanvasTkAgg(f, self)
                                canvas.show()
                                canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                canvas.mpl_connect("motion_notify_event", hover)
                                toolbar = NavigationToolbar2TkAgg(canvas, self)
                                toolbar.update()
                                canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                                
                            else: # DEPonly + MIXED
                                f = Figure(figsize=(5,5), dpi=100)  
                                A = f.add_subplot(111)
                                ax = f.add_subplot(111)
                                
                                
                                labels =  sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']  + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                                tag_text_use = np.array(list(labels))
                                labels_y = sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                                labels_x = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14']
                                                        
                                tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                #A.axhline(y=0.5, color='w')
                                                           
                                                    
                                A.errorbar(sequenceDepOnly(df_dep_output, 1)['3'], sequenceDepOnly(df_dep_output, 1)['4'], xerr=[sequenceDepOnly(df_dep_output, 1)['5'], sequenceDepOnly(df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)
                                A.errorbar(sequenceDepOnly(df_dep_output, 1)['10'], sequenceDepOnly(df_dep_output, 1)['9'], xerr=[sequenceDepOnly(df_dep_output, 1)['11'], sequenceDepOnly(df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                                
                                A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11']], color='darkgreen', fmt='o', markersize=8, capsize=10)    
                                A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='maroon', fmt='o', markersize=8, capsize=10, )
                                A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['15']], color='navy', fmt='o', markersize=8, capsize=10)
                                A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['18'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['19'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['17'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['16']], color='indigo', fmt='o', markersize=8, capsize=10, )
                            
                                annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                bbox=dict(boxstyle="round", fc="w"),
                                                arrowprops=dict(arrowstyle="->"))
                                annot.set_visible(False)
                                def update_annot2(ind):
                                
                                    pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                    annot.xy = pos
                                    text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                    annot.set_text(text)
                                def hover(event):
                                    vis = annot.get_visible()
                                    if event.inaxes == ax:
                                        cont, ind = tags_main_data.contains(event)
                                        if cont:
                                            update_annot2(ind)
                                            annot.set_visible(True)
                                            canvas.draw_idle()
                                        else:
                                            if vis:
                                                annot.set_visible(False)
                                                canvas.draw_idle()
                                A.set_xlabel('Seconds of the day')        
                                A.axes.get_yaxis().set_visible(False)                     
                                A.grid(color='b', linestyle='-', linewidth=0.1)    
                                A.legend(("Legend","Departures M1","Departures Spacing M1","Departures M2","Arrivals M2", "Departures Spacing M2","Arrivals Spacing M2"), loc = 'upper right')
                                #plt.title("Sequence analysis")
                                canvas = FigureCanvasTkAgg(f, self)
                                canvas.show()
                                canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                canvas.mpl_connect("motion_notify_event", hover)
                                toolbar = NavigationToolbar2TkAgg(canvas, self)
                                toolbar.update()
                                canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                                    
                           
                        
                        #------------------ MIXED MODE ------------------------#
                        
                        else: 
                            
                            if (df_dep_output2.empty ==False) and (df_arr_output2.empty ==False): #mixed both
                                if m2==2: #MIXED +MIXED + MIXED
                                            
                                    f = Figure(figsize=(5,5), dpi=100)  
                                    A = f.add_subplot(111)
                                    ax = f.add_subplot(111)
                                    
                                    
                                    labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] +sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['12']
                                    tag_text_use = np.array(list(labels))
                                    labels_x = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] +sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['14']
                                    labels_y = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] + sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['13']
                                    tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                    #plt.axhline(y = 10, color='w')                    
                                    ##A.axhline(y=0.5, color='w')
                                       
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='darkgreen', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='maroon', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], color='navy',fmt = 'o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='indigo', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['5'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['6'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['6']], color='limegreen', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['1'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2']], color='orangered', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['7'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['11'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0b'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['8']], color='royalblue',fmt = 'o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['3'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['10'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0a'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['4']], color='magenta', fmt='o', markersize=8, capsize=10, )
                                    
                                    
                                    
                                    
                                    annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                    bbox=dict(boxstyle="round", fc="w"),
                                                    arrowprops=dict(arrowstyle="->"))
                                    annot.set_visible(False)
                                    def update_annot2(ind):
                                    
                                        pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                        annot.xy = pos
                                        text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                        annot.set_text(text)
                                    def hover(event):
                                        vis = annot.get_visible()
                                        if event.inaxes == ax:
                                            cont, ind = tags_main_data.contains(event)
                                            if cont:
                                                update_annot2(ind)
                                                annot.set_visible(True)
                                                canvas.draw_idle()
                                            else:
                                                if vis:
                                                    annot.set_visible(False)
                                                    canvas.draw_idle()
                                    A.set_xlabel('Seconds of the day')        
                                    A.axes.get_yaxis().set_visible(False)                     
                                    A.grid(color='b', linestyle='-', linewidth=0.1)    
                                    A.legend(("Legend","Departures M1","Arrivals M1","Departures Spacing M1","Arrivals Spacing M1","Departures M2","Arrivals M2", "Departures Spacing M2","Arrivals Spacing M2", "Departures M3","Arrivals M3","Departures Spacing M3","Arrivals Spacing M3"), loc = 'upper right')
                                    #plt.title("Sequence analysis")
                                    canvas = FigureCanvasTkAgg(f, self)
                                    canvas.show()
                                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                    canvas.mpl_connect("motion_notify_event", hover)
                                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                                    toolbar.update()
                                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                                    #f.canvas.figure.savefig('sequence.png')
                                
                                
                                elif m2==3:#MIXED +MIXED + MIXED + MIXED
                                            
                                    f = Figure(figsize=(5,5), dpi=100)  
                                    A = f.add_subplot(111)
                                    ax = f.add_subplot(111)
                                    
                                    labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] +sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['12'] + sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['12']
                                    tag_text_use = np.array(list(labels))
                                    labels_x = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] +sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['14'] + sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['14']
                                    labels_y = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] + sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['13'] + sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['13']
                                    tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                    #plt.axhline(y = 10, color='w')                    
                                    #A.axhline(y=0.5, color='w')
                                       
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='darkgreen', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='maroon', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], color='navy',fmt = 'o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='indigo', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['5'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['6'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['6']], color='limegreen', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['1'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2']], color='orangered', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['7'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['11'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0b'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['8']], color='royalblue',fmt = 'o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['3'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['10'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0a'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['4']], color='magenta', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['5'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['9'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['6'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['6']], color='lime', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['1'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['9'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['2'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['2']], color='salmon', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['7'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['11'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['0b'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['8']], color='cyan',fmt = 'o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['3'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['10'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['0a'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['4']], color='orchid', fmt='o', markersize=8, capsize=10, )
                                    
                                    
                                    
                                    annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                    bbox=dict(boxstyle="round", fc="w"),
                                                    arrowprops=dict(arrowstyle="->"))
                                    annot.set_visible(False)
                                    def update_annot2(ind):
                                    
                                        pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                        annot.xy = pos
                                        text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                        annot.set_text(text)
                                    def hover(event):
                                        vis = annot.get_visible()
                                        if event.inaxes == ax:
                                            cont, ind = tags_main_data.contains(event)
                                            if cont:
                                                update_annot2(ind)
                                                annot.set_visible(True)
                                                canvas.draw_idle()
                                            else:
                                                if vis:
                                                    annot.set_visible(False)
                                                    canvas.draw_idle()
                                    A.set_xlabel('Seconds of the day')        
                                    A.axes.get_yaxis().set_visible(False)                     
                                    A.grid(color='b', linestyle='-', linewidth=0.1)    
                                    A.legend(("Legend","Departures M1","Arrivals M1","Departures Spacing M1","Arrivals Spacing M1","Departures M2","Arrivals M2", "Departures Spacing M2","Arrivals Spacing M2", "Departures M3","Arrivals M3","Departures Spacing M3","Arrivals Spacing M3", "Departures M4","Arrivals M4","Departures Spacing M4","Arrivals Spacing M4"), loc = 'upper right')
                                    #plt.title("Sequence analysis")
                                    canvas = FigureCanvasTkAgg(f, self)
                                    canvas.show()
                                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                    canvas.mpl_connect("motion_notify_event", hover)
                                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                                    toolbar.update()
                                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                                    #f.canvas.figure.savefig('sequence.png')
                                elif m2==4: #MIXED +MIXED + MIXED +MIXED + MIXED
                                                                    
                                    f = Figure(figsize=(5,5), dpi=100)  
                                    A = f.add_subplot(111)
                                    ax = f.add_subplot(111)
                                    
                                    
                                    labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] +sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['12'] + sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['12'] + sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['12']
                                    tag_text_use = np.array(list(labels))
                                    labels_x = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] +sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['14'] + sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['14'] + sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['14']
                                    labels_y = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] + sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['13'] + sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['13'] + sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['13']
                                    tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                    #plt.axhline(y = 10, color='w')                    
                                    #A.axhline(y=0.5, color='w')
                                       
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='darkgreen', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='maroon', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], color='navy',fmt = 'o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='indigo', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['5'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['6'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['6']], color='limegreen', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['1'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2']], color='orangered', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['7'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['11'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0b'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['8']], color='royalblue',fmt = 'o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['3'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['10'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0a'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['4']], color='magenta', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['5'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['9'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['6'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['6']], color='lime', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['1'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['9'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['2'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['2']], color='salmon', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['7'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['11'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['0b'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['8']], color='cyan',fmt = 'o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['3'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['10'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['0a'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['4']], color='orchid', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['5'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['9'], xerr=[sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['6'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['6']], color='olive', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['1'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['9'], xerr=[sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['2'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['2']], color='firebrick', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['7'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['11'], xerr=[sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['0b'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['8']], color='mediumblue',fmt = 'o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['3'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['10'], xerr=[sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['0a'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['4']], color='deeppink', fmt='o', markersize=8, capsize=10, )
                                    
                                
                                    
                                    
                                    annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                    bbox=dict(boxstyle="round", fc="w"),
                                                    arrowprops=dict(arrowstyle="->"))
                                    annot.set_visible(False)
                                    def update_annot2(ind):
                                    
                                        pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                        annot.xy = pos
                                        text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                        annot.set_text(text)
                                    def hover(event):
                                        vis = annot.get_visible()
                                        if event.inaxes == ax:
                                            cont, ind = tags_main_data.contains(event)
                                            if cont:
                                                update_annot2(ind)
                                                annot.set_visible(True)
                                                canvas.draw_idle()
                                            else:
                                                if vis:
                                                    annot.set_visible(False)
                                                    canvas.draw_idle()
                                    A.set_xlabel('Seconds of the day')        
                                    A.axes.get_yaxis().set_visible(False)                    
                                    A.grid(color='b', linestyle='-', linewidth=0.1)    
                                    A.legend(("Legend","Departures M1","Arrivals M1","Departures Spacing M1","Arrivals Spacing M1","Departures M2","Arrivals M2", "Departures Spacing M2","Arrivals Spacing M2", "Departures M3","Arrivals M3","Departures Spacing M3","Arrivals Spacing M3", "Departures M4","Arrivals M4","Departures Spacing M4","Arrivals Spacing M4"), loc = 'upper right')
                                    #plt.title("Sequence analysis")
                                    canvas = FigureCanvasTkAgg(f, self)
                                    canvas.show()
                                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                    canvas.mpl_connect("motion_notify_event", hover)
                                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                                    toolbar.update()
                                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                                    #f.canvas.figure.savefig('sequence.png')
                                
                                else: # only 2 to compare
                                            
                                    f = Figure(figsize=(5,5), dpi=100)  
                                    A = f.add_subplot(111)
                                    ax = f.add_subplot(111)
                                    
                                    
                                    labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                                    tag_text_use = np.array(list(labels))
                                    labels_x = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] 
                                    labels_y = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                                    tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                    #plt.axhline(y = 10, color='w')                    
                                    ##A.axhline(y=0.5, color='w')
                                       
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                                    
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='darkgreen', fmt='o', markersize=8, capsize=10)    
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='maroon', fmt='o', markersize=8, capsize=10, )
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], color='navy',fmt = 'o', markersize=8, capsize=10)
                                    A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='indigo', fmt='o', markersize=8, capsize=10, )
                                    
                                    annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                    bbox=dict(boxstyle="round", fc="w"),
                                                    arrowprops=dict(arrowstyle="->"))
                                    annot.set_visible(False)
                                    def update_annot2(ind):
                                    
                                        pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                        annot.xy = pos
                                        text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                        annot.set_text(text)
                                    def hover(event):
                                        vis = annot.get_visible()
                                        if event.inaxes == ax:
                                            cont, ind = tags_main_data.contains(event)
                                            if cont:
                                                update_annot2(ind)
                                                annot.set_visible(True)
                                                canvas.draw_idle()
                                            else:
                                                if vis:
                                                    annot.set_visible(False)
                                                    canvas.draw_idle()
                                    A.set_xlabel('Seconds of the day')        
                                    A.axes.get_yaxis().set_visible(False)                     
                                    A.grid(color='b', linestyle='-', linewidth=0.1)    
                                    A.legend(("Legend","Departures M1","Arrivals M1","Departures Spacing M1","Arrivals Spacing M1","Departures M2","Arrivals M2", "Departures Spacing M2","Arrivals Spacing M2"), loc = 'upper right')
                                    #plt.title("Sequence analysis")
                                    canvas = FigureCanvasTkAgg(f, self)
                                    canvas.show()
                                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                    canvas.mpl_connect("motion_notify_event", hover)
                                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                                    toolbar.update()
                                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                                
                                
                                
                            elif df_dep_output2.empty ==True:#arr only
                                f = Figure(figsize=(5,5), dpi=100)  
                                A = f.add_subplot(111)
                                ax = f.add_subplot(111)
                                
                                labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                                tag_text_use = np.array(list(labels))
                                labels_x = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] 
                                labels_y = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                                tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                #plt.axhline(y = 10, color='w')                    
                                #A.axhline(y=0.5, color='w')
                                   
                                A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                                A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                                A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                                A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                                
                                A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                                A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
            
                            
                                annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                bbox=dict(boxstyle="round", fc="w"),
                                                arrowprops=dict(arrowstyle="->"))
                                annot.set_visible(False)
                                def update_annot2(ind):
                                
                                    pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                    annot.xy = pos
                                    text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                    annot.set_text(text)
                                def hover(event):
                                    vis = annot.get_visible()
                                    if event.inaxes == ax:
                                        cont, ind = tags_main_data.contains(event)
                                        if cont:
                                            update_annot2(ind)
                                            annot.set_visible(True)
                                            canvas.draw_idle()
                                        else:
                                            if vis:
                                                annot.set_visible(False)
                                                canvas.draw_idle()
                                A.set_xlabel('Seconds of the day')        
                                A.axes.get_yaxis().set_visible(False)                     
                                A.grid(color='b', linestyle='-', linewidth=0.1)    
                                A.legend(("Legend","Departures M1","Arrivals M1","Departures Spacing M1","Arrivals Spacing M1","Arrivals M2", "Arrivals Spacing M2"), loc = 'upper right')
                                #plt.title("Sequence analysis")
                                canvas = FigureCanvasTkAgg(f, self)
                                canvas.show()
                                canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                canvas.mpl_connect("motion_notify_event", hover)
                                toolbar = NavigationToolbar2TkAgg(canvas, self)
                                toolbar.update()
                                canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                            
                            elif df_arr_output2.empty ==True: # MIXED + ARRonly
                                
                                f = Figure(figsize=(5,5), dpi=100)  
                                A = f.add_subplot(111)
                                ax = f.add_subplot(111)
                                
                                #plt.axhline(y = 10, color='w')                    
                                
                                labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                                tag_text_use = np.array(list(labels))
                                labels_x = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] 
                                labels_y = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                                tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                                #plt.axhline(y = 10, color='w')                    
                                #A.axhline(y=0.5, color='w')
                                   
                                A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                                A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                                A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                                A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                                
                                
                                A.errorbar(sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                                A.errorbar(sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], fmt='o', markersize=8, capsize=10)
            
                                annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                bbox=dict(boxstyle="round", fc="w"),
                                                arrowprops=dict(arrowstyle="->"))
                                annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                                bbox=dict(boxstyle="round", fc="w"),
                                                arrowprops=dict(arrowstyle="->"))
                                annot.set_visible(False)
                                def update_annot2(ind):
                                
                                    pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                    annot.xy = pos
                                    text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                    annot.set_text(text)
                                def hover(event):
                                    vis = annot.get_visible()
                                    if event.inaxes == ax:
                                        cont, ind = tags_main_data.contains(event)
                                        if cont:
                                            update_annot2(ind)
                                            annot.set_visible(True)
                                            canvas.draw_idle()
                                        else:
                                            if vis:
                                                annot.set_visible(False)
                                                canvas.draw_idle()
                                A.set_xlabel('Seconds of the day')        
                                A.axes.get_yaxis().set_visible(False)                     
                                A.grid(color='b', linestyle='-', linewidth=0.1)    
                                A.legend(("Legend","Departures M1","Arrivals M1","Departures Spacing M1","Arrivals Spacing M1","Departures M2", "Departures Spacing M2"), loc = 'upper right')
                                #plt.title("Sequence analysis")
                                canvas = FigureCanvasTkAgg(f, self)
                                canvas.show()
                                canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                                canvas.mpl_connect("motion_notify_event", hover)
                                toolbar = NavigationToolbar2TkAgg(canvas, self)
                                toolbar.update()
                                canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                                
                                
                          
                                
                # ============================================================================#
                #                           ADA buffer                                          #
                # ============================================================================#
                class ADAbuffer2(tk.Frame):
                
                    def __init__(self, parent, controller):
                        tk.Frame.__init__(self, parent)
                        label = tk.Label(self, text="SEQUENCE", font=LARGE_FONT)
                        label.pack(pady=10,padx=10)
                        
                        button1 = ttk.Button(self, text="Back to Home",
                                            command=lambda: controller.show_frame(StartPage2))
                        button1.pack() 
                        if convergenceFLAG ==True:
                            button1 = ttk.Button(self, text="Convergence Throughput",
                                                command=lambda: controller.show_frame(Conv2))
                            button1.pack()
                        if Thr_FLAG == True:
                            button = ttk.Button(self, text="Throughput",
                                                command=lambda: controller.show_frame(Thr2))
                            button.pack()
                        
                        if Delay_FLAG == True:
                            button2 = ttk.Button(self, text="RWY Hold Delay",
                                                command=lambda: controller.show_frame(DepDelay22))
                            button2.pack()
                        if Delay_FLAG == True:
                            button5 = ttk.Button(self, text="Push/Start Delay",
                                                command=lambda: controller.show_frame(DepDelay222))
                            button5.pack()
                        
                        if arr_delay_FLAG == True:
                            button3 = ttk.Button(self, text="Arrivals Delay",
                                                command=lambda: controller.show_frame(ArrivalDelay2))
                            button3.pack()
                            
                        
                        
                        df_Buffer = pd.DataFrame()
                        df_Buffer['ADA_Buffer'] = df_sequence_output['ADA Buffer']        
                        df_Buffer = df_Buffer.dropna(subset = ['ADA_Buffer'])
                        #df_Buffer = df_Buffer.drop([0])
                        ADA_buffer = df_Buffer['ADA_Buffer'].tolist()
                        a = int(min(ADA_buffer))
                        b = int(max(ADA_buffer))
                        number_bins = b-a
                        h = [0]+sorted(ADA_buffer)
                        
                        coord = [[0,0], [15,0], [15,0.013], [0,0.013]]
                        coord.append(coord[0]) #repeat the first point to create a 'closed loop'
                        
                        xs, ys = zip(*coord) #create lists of x and y values
                
                
                        fit = stats.norm.pdf(h, np.mean(h), np.std(h))  #this is a fitting indeed
                        
                        
                        f = Figure(figsize=(5,5), dpi=100)
                        A = f.add_subplot(111)
                        A.plot(h,fit,'-o')
                        A.hist(h,normed=True,bins=number_bins)
                        
                        A.plot(xs,ys,"r") 
                        
                        A.set_xlabel('SECONDS')
                        A.set_ylabel('%')
                        A.grid(color='b', linestyle='-', linewidth=0.1)     
                        #A.title('Throughput')   
                        canvas = FigureCanvasTkAgg(f, self)
                        canvas.show()
                        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                
                        toolbar = NavigationToolbar2TkAgg(canvas, self)
                        toolbar.update()
                        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    
    
    
        
                    
                app2 = RAPIDvisual2()
                
                app2.columnconfigure(0, weight=1)
                app2.rowconfigure(0, weight=1)
                app2.rowconfigure(1, weight=1)
                
                app2.mainloop()            
            
                           
            m2_input = IntVar(self, value='0')
            m2_output = IntVar()
            entry4 = ttk.Entry(self, width=7, textvariable=m2_input)
            entry4.pack()
            
            
            button6 = tk.Button(self, text="SHOW", command=show_button)
            button6.pack()
            
          
        
            
            
    # ============================================================================#
    #                           CONVERGENCE                                       #
    # ============================================================================#
    #if convergenceFLAG ==True:
    class Conv(tk.Frame):
    
        def __init__(self, parent, controller):
            tk.Frame.__init__(self, parent)
            label = tk.Label(self, text="THROUGHPUT CONVERGENCE", font=LARGE_FONT)
            label.pack(pady=10,padx=10)
    
            button1 = ttk.Button(self, text="Back to Home",
                                command=lambda: controller.show_frame(StartPage))
            button1.pack()
            
            
            if Thr_FLAG == True:
                button = ttk.Button(self, text="Throughput",
                                    command=lambda: controller.show_frame(Thr))
                button.pack()
                
            if Delay_FLAG == True:
                button2 = ttk.Button(self, text="RWY Hold Delay",
                                    command=lambda: controller.show_frame(DepDelay))
                button2.pack()
            if Delay_FLAG == True:
                button5 = ttk.Button(self, text="Push/Start Delay",
                                    command=lambda: controller.show_frame(DepDelay2))
                button5.pack()
            if arr_delay_FLAG == True:
                button3 = ttk.Button(self, text="Arrivals Delay",
                                    command=lambda: controller.show_frame(ArrivalDelay))
                button3.pack()
            if Seq_FLAG == True:
                button4 = ttk.Button(self, text="Sequence",
                                    command=lambda: controller.show_frame(Seq))
                button4.pack() 
            
            if ADA_buffer_FLAG == True:
                button5 = ttk.Button(self, text="ADA Buffer",
                                    command=lambda: controller.show_frame(ADAbuffer))
                button5.pack()
    
           
            thr_av_diff = pd.DataFrame(df_thr['Difference in thr averages'])
            thr_av_diff=thr_av_diff.dropna(subset=['Difference in thr averages'])
           
            thr_av_diff['Tags'] = thr_av_diff['Difference in thr averages'].apply(lambda x: json.loads(x))
            thr_av_difference = thr_av_diff['Tags'].tolist()
            
            
            
            print(len(thr_av_difference[0][0]))
            print(len(thr_av_difference[0]))
            thr_abc=[]
            total_runs =[]
            runs_Thr = []
            for i in range (1,(len(thr_av_difference[0])+1)):
                runs_Thr.append(i) 
           
            f = Figure(figsize=(5,5), dpi=100)
            A = f.add_subplot(111)
            
            for j in range (0, len(thr_av_difference[0][0])):
                for i in range (0, len(thr_av_difference[0])):            
                    total_thr = thr_av_difference[0][i][j]
                    runs_Thrx = runs_Thr[i]
                    total_runs.append(runs_Thrx)
                    thr_abc.append(total_thr) 
                A.plot(total_runs,thr_abc)
                thr_abc = []
                total_runs =[]
                
            A.set_xlabel('No. of Runs')
            A.set_ylabel('Difference in average throughput per hour')
            A.grid(color='b', linestyle='-', linewidth=0.1) 
            #A.title('Throughput')   
            canvas = FigureCanvasTkAgg(f, self)
            canvas.show()
            canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=False)
    
            toolbar = NavigationToolbar2TkAgg(canvas, self)
            toolbar.update()
            canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                
    # ============================================================================#
    #                          THROUGHPUT                                         #
    # ============================================================================#
    #if Thr_FLAG == True:
    class Thr(tk.Frame):
    
        def __init__(self, parent, controller):
            tk.Frame.__init__(self, parent)
            label = tk.Label(self, text="THROUGHPUT", font=LARGE_FONT)
            label.pack(pady=10,padx=10)
    
            button1 = ttk.Button(self, text="Back to Home",
                                command=lambda: controller.show_frame(StartPage))
            button1.pack()
            
            
            if convergenceFLAG ==True:
                button1 = ttk.Button(self, text="Convergence Throughput",
                                    command=lambda: controller.show_frame(Conv))
                button1.pack()
            if Delay_FLAG == True:
                button2 = ttk.Button(self, text="RWY Hold Delay",
                                    command=lambda: controller.show_frame(DepDelay))
                button2.pack()
            
                button5 = ttk.Button(self, text="Push/Start Delay",
                                    command=lambda: controller.show_frame(DepDelay2))
                button5.pack()
            if arr_delay_FLAG == True:
                button3 = ttk.Button(self, text="Arrivals Delay",
                                    command=lambda: controller.show_frame(ArrivalDelay))
                button3.pack()
            if Seq_FLAG == True:
                button4 = ttk.Button(self, text="Sequence",
                                    command=lambda: controller.show_frame(Seq))
                button4.pack()
            if ADA_buffer_FLAG == True:
                button5 = ttk.Button(self, text="ADA Buffer",
                                    command=lambda: controller.show_frame(ADAbuffer))
                button5.pack()
            
            def create_first_df_thr():                   
                df_thr_to_plot = pd.DataFrame()
                df_thr_to_plot['Hour'] = df_thr['Hour']
                x = 'Hour'
                df_thr_to_plot['RUN 1'] = df_thr['Total Throughput']
                y = 'RUN 1'
                
                df_thr_to_plot2 = df_thr_to_plot[[x,y]].groupby(x).sum()
                
                return df_thr_to_plot2
            
           
            
            def create_multiple_df_thr(df_thr_to_plot2, df_thr_input, name):
                df_thr_to_plot2[name] = df_thr_input['Total Throughput']
                df_thr_to_plot_temp = pd.DataFrame()
                df_thr_to_plot_temp['Hour'] = df_thr_input['Hour']
                a = 'Hour'
                df_thr_to_plot_temp[name] = df_thr_input['Total Throughput']
                b = name
                df_thr_to_plot_temp = df_thr_to_plot_temp.dropna(subset=['Hour'])    
                df_thr_to_plot_temp2 = df_thr_to_plot_temp[[a,b]].groupby(a).sum()    
                        
                df_thr_to_plot2[name] = df_thr_to_plot_temp2[name]
                
                
                return df_thr_to_plot2 
            
    #        def add_operational_data ():
    #            #df_thr_to_plot2['Operational Data'] = op_data['Total Throughput']
    #            df_thr_to_plot_temp = pd.DataFrame()
    #            df_thr_to_plot_temp['Hour'] = op_data['Hour']
    #            a = 'Hour'
    #            df_thr_to_plot_temp['Operational Data'] = op_data['Total Throughput']
    #            b = 'Operational Data'
    #            df_thr_to_plot_temp = df_thr_to_plot_temp.dropna(subset=['Hour'])
    #                
    #            df_thr_to_plot_temp2 = df_thr_to_plot_temp[[a,b]].groupby(a).sum()    
    #                    
    #            df_thr_to_plot2['Operational Data'] = df_thr_to_plot_temp2['Operational Data']
                
            def plot_bar_thr(df_thr_to_plot2, A):                      
                df_thr_to_plot2.plot(kind='bar', legend=False, ax=A) 
            
            def total_thr(df_thr_input):
                total_thr = df_thr_input['Total Throughput']
                total_thr = total_thr.reset_index()
                return(total_thr)
            def hour_thr(df_thr_input):
                hour_Thr = df_thr_input['Hour'].tolist()
                return(hour_Thr)
                
            
            if new_set_FLAG == True:
                if m>= 1:
                    f = Figure(figsize=(5,5), dpi=100)
                    A = f.add_subplot(111)
    #                throughput(df_thr, 'k')
    #                throughput(df_thr2, 'g')
                    
                    df_thr_to_plot2 = create_first_df_thr()
                    df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, df_thr2, 'RUN 2')
                    
                    if (m == 1) and (average_FLAG == True):
                        a=total_thr(df_thr)            
                        b=total_thr(df_thr2)            
                        average = pd.DataFrame()
                        average = (a+b)/2
                        average_thr = average['Total Throughput'].tolist()
                        A.plot(hour_thr(df_thr),average_thr, linewidth=3.0, color = 'red', linestyle='dashed', alpha = 0.5)             
                    
                    if OP_FLAG == True:
                        df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                        A.legend(('Model','Operational Data', 'Model 2'), loc = 'upper right')
                    else:
                        A.legend(('Model', 'Model 2'), loc = 'upper right')
                    if m>= 2:
                        df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, df_thr3, 'RUN 3')
                        #throughput(df_thr3, 'c') 
                        if (m==2) and (average_FLAG ==True):
                            a=total_thr(df_thr)            
                            b=total_thr(df_thr2)
                            c=total_thr(df_thr3)
                            average = pd.DataFrame()
                            average = (a+b+c)/3
                            average_thr = average['Total Throughput'].tolist()
                            A.plot(hour_thr(df_thr),average_thr, linewidth=3.0, color = 'red', linestyle='dashed', alpha = 0.5)
                        
                        if OP_FLAG ==True:
                            df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                            A.legend(('Model','Operational Data', 'Model 2', 'Model 3'), loc = 'upper right')
                        else:
                            A.legend(('Model','Model 2', 'Model 3'), loc = 'upper right')
                        if m>=3:
                            #throughput(df_thr4, 'm') 
                            df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, df_thr4, 'RUN 4')
                            if (m==3) and (average_FLAG ==True):
                                a=total_thr(df_thr)            
                                b=total_thr(df_thr2)
                                c=total_thr(df_thr3)
                                d=total_thr(df_thr4)
                                average = pd.DataFrame()
                                average = (a+b+c+d)/4
                                average_thr = average['Total Throughput'].tolist()
                                A.plot(hour_thr(df_thr),average_thr, linewidth=3.0, color = 'red',linestyle='dashed', alpha = 0.5)
                            
                            if OP_FLAG ==True:
                                df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                                A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4'), loc = 'upper right')
                            else:
                                A.legend(('Model','Model 2', 'Model 3','Model 4'), loc = 'upper right')
                            if m>=4:
                                #throughput(df_thr5, 'y') 
                                df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, df_thr5, 'RUN 5')
                                if (m==4) and (average_FLAG ==True):
                                    a=total_thr(df_thr)            
                                    b=total_thr(df_thr2)
                                    c=total_thr(df_thr3)
                                    d=total_thr(df_thr4)
                                    e=total_thr(df_thr5)
                                    average = pd.DataFrame()
                                    average = (a+b+c+d+e)/5
                                    average_thr = average['Total Throughput'].tolist()
                                    A.plot(hour_thr(df_thr),average_thr, linewidth=3.0, color = 'red', linestyle='dashed', alpha = 0.5)
                                
                                if OP_FLAG ==True:
                                    df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                                    A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5'), loc = 'upper right')
                                else:
                                    df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                                    A.legend(('Model','Model 2', 'Model 3','Model 4', 'Model 5'), loc = 'upper right')
                                if m>=5:
                                    df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, df_thr6, 'RUN 6')
                                    #throughput(df_thr6, color = 'purple')
                                    if (m==5) and (average_FLAG ==True):
                                        a=total_thr(df_thr)            
                                        b=total_thr(df_thr2)
                                        c=total_thr(df_thr3)
                                        d=total_thr(df_thr4)
                                        e=total_thr(df_thr5)
                                        f=total_thr(df_thr6)
                                        average = pd.DataFrame()
                                        average = (a+b+c+d+e+f)/6
                                        average_thr = average['Total Throughput'].tolist()
                                        A.plot(hour_thr(df_thr),average_thr, linewidth=3.0, color = 'red',linestyle='dashed',  alpha = 0.5)
                                    
                                    if OP_FLAG ==True:
                                        df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                                        A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                                    else:
                                        A.legend(('Model','Model 2', 'Model 3','Model 4','Model 5','Model 6'), loc = 'upper right')
                    plot_bar_thr(df_thr_to_plot2, A)
                    if OP_FLAG ==True:
                        df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')
                        A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                    A.set_xlabel('Hours of the day')
                    A.set_ylabel('No of A/C')
                    A.grid(color='b', linestyle='-', linewidth=0.1) 
                    #A.title('Throughput')   
                    canvas = FigureCanvasTkAgg(f, self)
                    canvas.show()
                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
            
                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                    toolbar.update()
                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
            else:
                f = Figure(figsize=(5,5), dpi=100)
                A = f.add_subplot(111)    
                #df_thr_to_plot2 = throughput2(df_thr) 
                df_thr_to_plot2 = create_first_df_thr()
                    
                #df_thr_to_plot2.plot(kind='bar', legend=False, ax=A)          
                
                if OP_FLAG == True:
                    df_thr_to_plot2 = create_multiple_df_thr(df_thr_to_plot2, op_data, 'OPERATIONA_DATA')            
                
                plot_bar_thr(df_thr_to_plot2, A) 
                A.set_xlabel('Hours of the day')
                A.set_ylabel('No of A/C')
                A.grid(color='b', linestyle='-', linewidth=0.1)     
                #A.title('Throughput')   
                canvas = FigureCanvasTkAgg(f, self)
                canvas.show()
                canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
        
                toolbar = NavigationToolbar2TkAgg(canvas, self)
                toolbar.update()
                canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
            
    
    
    # ============================================================================#
    #                         DEPARTURE DELAY - RWY hold delay                    #
    # ============================================================================#
    #if Delay_FLAG == True:
    class DepDelay(tk.Frame):
    
        def __init__(self, parent, controller):
            tk.Frame.__init__(self, parent)
            label = tk.Label(self, text="RWY HOLD DELAY", font=LARGE_FONT)
            label.pack(pady=10,padx=10)
    
            button1 = ttk.Button(self, text="Back to Home",
                                command=lambda: controller.show_frame(StartPage))
            button1.pack() 
            
            
            if convergenceFLAG ==True:
                button1 = ttk.Button(self, text="Convergence Throughput",
                                    command=lambda: controller.show_frame(Conv))
                button1.pack()
            if Thr_FLAG == True:
                button = ttk.Button(self, text="Throughput",
                                    command=lambda: controller.show_frame(Thr))
                button.pack()
                
            if Delay_FLAG == True:
                button5 = ttk.Button(self, text="Push/Start Delay",
                                    command=lambda: controller.show_frame(DepDelay2))
                button5.pack()                
            if arr_delay_FLAG == True:
                button3 = ttk.Button(self, text="Arrivals Delay",
                                    command=lambda: controller.show_frame(ArrivalDelay))
                button3.pack()
            if Seq_FLAG == True:
                button4 = ttk.Button(self, text="Sequence",
                                    command=lambda: controller.show_frame(Seq))
                button4.pack()
            if ADA_buffer_FLAG == True:
                button5 = ttk.Button(self, text="ADA Buffer",
                                    command=lambda: controller.show_frame(ADAbuffer))
                button5.pack()
            if (df_delay.empty == True):
                print("No departures. Nothing to show")
            else:   
                def delay(df_delay,df_dep_output):
                    
                    interval15=[]
                    df_ps_delay = pd.DataFrame()
                    # Add arrival delay values
                    df_ps_delay['RWYhold_Delay'] = df_delay['RWY HOLD Delay']
                    # print when those values occure
                    df_ps_delay['Time1'] = df_dep_output['Departure_RWY_ENTRY']
                    # round each time value to 15 minutes
                    df_ps_delay= df_ps_delay.dropna(subset=['Time1'])
                    time1_list = df_ps_delay['Time1'].tolist()
                    #a=[]
                    for a in time1_list:
                        b = int(int(a/900)*900)
                        interval15.append(b)
                        
                    df_ps_delay['interval15'] = interval15
                    df_ps_delay = df_ps_delay.drop(columns=['Time1'])
                    #Group data by the time interval, if there are multiple values for the same time interval, take the mean.
                    df_ps_delay = df_ps_delay.groupby(['interval15'])['RWYhold_Delay'].mean()
                    #make the rolling average
                    df_ps_delay = df_ps_delay.reset_index()
                    df_ps_delay2 = df_ps_delay.rolling(window=4, on='interval15')['RWYhold_Delay'].mean()
                    RWYhold_Delay = df_ps_delay2.tolist()
                    df_ps_delay['DATE'] = pd.to_datetime(df_ps_delay['interval15'],unit='s')
                    df_ps_delay['DATE'] = df_ps_delay['DATE'].apply(lambda x: x.time())
                    
                    time_interval = df_ps_delay['DATE']
    #            
    #                df_delay_input['Time interval'] = pd.DatetimeIndex(df_delay_input['Time interval'])
    #                #For RWY_Hold Delay
    #                df_rwy = df_delay_input
    #                df_rwy = df_rwy.groupby(['Time interval'])['RWY_Hold Delay'].mean()
    #                df_rwy = df_rwy.reset_index()
    #                df_rwy_avg = df_rwy.rolling(window='3600s', on='Time interval')['RWY_Hold Delay'].mean()
    #                df_rwy_avg = df_rwy_avg.reset_index()
    #                df_rwy = df_rwy.reset_index()
    #                df_rwy = df_rwy.drop(columns=['RWY_Hold Delay'])
    #                
    #                
    #                df_final_rwy = pd.merge(df_rwy, df_rwy_avg,  how='left', on=['index'], copy=True)
    #                df_final_rwy = df_final_rwy.drop(columns=['index'])
    #                df_final_rwy['Time interval'] = pd.to_timedelta(df_final_rwy['Time interval']) # convert to timedelta to calculate seconds
    #                df_final_rwy['Time interval'] = df_final_rwy['Time interval'].dt.seconds
    #                
    #                #For Push/Start Delay
    #                df_ps = df_delay_input
    #                df_ps = df_ps.groupby(['Time interval'])['Push/Start Delay'].mean()
    #                df_ps = df_ps.reset_index() 
    #                dh_ps_avg = df_ps.rolling(window='3600s', on='Time interval')['Push/Start Delay'].mean()
    #                dh_ps_avg = dh_ps_avg.reset_index()
    #                df_ps = df_ps.reset_index()
    #                df_ps = df_ps.drop(columns=['Push/Start Delay'])
    #                #df_ps['Time interval'] = df_ps['Time interval'].apply(lambda x: x.time())
    #                #df_rwy_arr['Arr Time Interval'] = df_rwy_arr['Arr Time Interval'].apply(lambda x: x.time())
    #                
    #               
    #                df_final_ps = pd.merge(df_ps, dh_ps_avg,  how='left', on=['index'], copy=True)
    #                df_final_ps = df_final_ps.drop(columns=['index'])
    #                df_final_ps['Time interval'] = pd.to_timedelta(df_final_ps['Time interval']) # convert to timedelta to calculate seconds
    #                df_final_ps['Time interval'] = df_final_ps['Time interval'].dt.seconds
    #                
    #                df_rwy['Time interval'] = df_rwy['Time interval'].apply(lambda x: x.time())
    #                #Extract lists to plot
    #                H_delay_time = df_rwy['Time interval'].tolist()
    #                H_delay = df_final_rwy['RWY_Hold Delay'].tolist()
    #                
    #                    
    #        #        return(H_delay_time,H_delay,PS_time,PS_delay)
                    return {'a': time_interval,
                            'b': RWYhold_Delay}
    #                
    #        #    def print_Hold_delay(H_delay_time,H_delay,color):
                def print_Hold_delay(ab, color):
                    
                    H_delay_time = ab['a']
                    H_delay = ab['b']
                    
                    A = f.add_subplot(111)
                    A.plot(H_delay_time, H_delay, color)
                    #A.set_title('RWY HOLD DELAY', loc='left')
                    if OP_FLAG == True: #Plot Operational Data
                        A.plot(OP_Time_H_Delay,OP_H_Delay, 'b')    
                        A.legend(('Model','Operational Data'), loc = 'upper right')
                        
                
             
                
                if new_set_FLAG == True:
                    if m >=1:          
                        f = Figure(figsize=(5,5), dpi=100)
                        A = f.add_subplot(111)
                        hold_delay = delay(df_delay,df_dep_output)
                        print_Hold_delay(hold_delay,'k')
                        hold_delay2 = delay(df_delay2,df_dep_output2)
                        print_Hold_delay(hold_delay2,'g')
                        
                        
                        
                        if OP_FLAG == True:
                            A.legend(('Model','Operational Data', 'Model 2'), loc = 'upper right')
                        else:
                            A.legend(('Model', 'Model 2'), loc = 'upper right')
                        if m>=2:
                            hold_delay3 = delay(df_delay3,df_dep_output3)
                            print_Hold_delay(hold_delay3,'c')
                            
                            
                            if OP_FLAG == True:
                                A.legend(('Model','Operational Data', 'Model 2', 'Model 3'), loc = 'upper right')
                            else:
                                A.legend(('Model', 'Model 2', 'Model 3'), loc = 'upper right')
                            if m>=3:
                                hold_delay4 = delay(df_delay4,df_dep_output4)
                                print_Hold_delay(hold_delay4,'m')
                                
                               
                                
                                if OP_FLAG == True:
                                    A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4'), loc = 'upper right')
                                else:
                                    A.legend(('Model', 'Model 2', 'Model 3', 'Model 4'), loc = 'upper right')
                                if m>=4:
                                    hold_delay5 = delay(df_delay5,df_dep_output5)
                                    print_Hold_delay(hold_delay5,'y')
                                    
                                    
                                    if OP_FLAG == True:
                                        A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5'), loc = 'upper right')
                                    else:
                                        A.legend(('Model', 'Model 2', 'Model 3', 'Model 4', 'Model 5'), loc = 'upper right')
                                    if m>=5:
                                        hold_delay6 = delay(df_delay6,df_dep_output6)
                                        print_Hold_delay(hold_delay6,color='purple')
                                        
                                        
                                        if OP_FLAG == True:
                                            A.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                                        else:
                                            A.legend(('Model', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                        A.set_xlabel('Seconds of the day')
                        A.set_ylabel('Seconds of delay')                
                        A.grid(color='b', linestyle='-', linewidth=0.1)
    
                      
                        canvas = FigureCanvasTkAgg(f, self)
                        canvas.show()
                        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                
                        toolbar = NavigationToolbar2TkAgg(canvas, self)
                        toolbar.update()
                        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                else:
                    f = Figure(figsize=(5,5), dpi=100)
                    A = f.add_subplot(111)
                    hold_delay = delay(df_delay,df_dep_output)
                    print_Hold_delay(hold_delay,'r')             
                    
                    A.set_xlabel('Time')
                    A.set_ylabel('Seconds of delay')
                    A.grid(color='b', linestyle='-', linewidth=0.1)
                    
                    canvas = FigureCanvasTkAgg(f, self)
                    canvas.show()
                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)        
                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                    toolbar.update()
                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        #                
                   
    
    # ============================================================================#
    #                   DEPARTURE DELAY - PS delay                                #
    # ============================================================================#
    
    class DepDelay2(tk.Frame):
    
        def __init__(self, parent, controller):
            tk.Frame.__init__(self, parent)
            label = tk.Label(self, text="PUSH/START DELAY", font=LARGE_FONT)
            label.pack(pady=10,padx=10)
    
            button1 = ttk.Button(self, text="Back to Home",
                                command=lambda: controller.show_frame(StartPage))
            button1.pack() 
            
            
            if convergenceFLAG ==True:
                button1 = ttk.Button(self, text="Convergence Throughput",
                                    command=lambda: controller.show_frame(Conv))
                button1.pack()   
            if Thr_FLAG == True:
                button = ttk.Button(self, text="Throughput",
                                    command=lambda: controller.show_frame(Thr))
                button.pack()   
            if Delay_FLAG == True:
                button2 = ttk.Button(self, text="RWY HOLD DELAY",
                                    command=lambda: controller.show_frame(DepDelay))
                button2.pack()                
            if arr_delay_FLAG == True:
                button3 = ttk.Button(self, text="Arrivals Delay",
                                    command=lambda: controller.show_frame(ArrivalDelay))
                button3.pack()
            if Seq_FLAG == True:
                button4 = ttk.Button(self, text="Sequence",
                                    command=lambda: controller.show_frame(Seq))
                button4.pack()
            if ADA_buffer_FLAG == True:
                button5 = ttk.Button(self, text="ADA Buffer",
                                    command=lambda: controller.show_frame(ADAbuffer))
                button5.pack()
            if (df_delay.empty == True):
                print("No departures. Nothing to show")
            else:   
                def delay(df_delay, df_dep_output):                
                    interval15=[]
                    df_ps_delay = pd.DataFrame()
                    # Add arrival delay values
                    df_ps_delay['PS_Delay'] = df_delay['Push/Start Delay']
                    # print when those values occure
                    df_ps_delay['Time1'] = df_dep_output['Departure_RWY_ENTRY']
                    # round each time value to 15 minutes
                    df_ps_delay= df_ps_delay.dropna(subset=['Time1'])
                    time1_list = df_ps_delay['Time1'].tolist()
                    #a=[]
                    for a in time1_list:
                        b = int(int(a/900)*900)
                        interval15.append(b)
                        
                    df_ps_delay['interval15'] = interval15
                    df_ps_delay = df_ps_delay.drop(columns=['Time1'])
                    #Group data by the time interval, if there are multiple values for the same time interval, take the mean.
                    df_ps_delay = df_ps_delay.groupby(['interval15'])['PS_Delay'].mean()
                    #make the rolling average
                    df_ps_delay = df_ps_delay.reset_index()
                    df_ps_delay2 = df_ps_delay.rolling(window=4, on='interval15')['PS_Delay'].mean()
                    PS_delay = df_ps_delay2.tolist()
                    df_ps_delay['DATE'] = pd.to_datetime(df_ps_delay['interval15'],unit='s')
                    df_ps_delay['DATE'] = df_ps_delay['DATE'].apply(lambda x: x.time())
                    
                    time_interval = df_ps_delay['DATE']
    #            
    #                df_delay_input['Time interval'] = pd.DatetimeIndex(df_delay_input['Time interval'])
    #                #For RWY_Hold Delay
    #                df_rwy = df_delay_input
    #                df_rwy = df_rwy.groupby(['Time interval'])['RWY_Hold Delay'].mean()
    #                df_rwy = df_rwy.reset_index()
    #                df_rwy_avg = df_rwy.rolling(window='3600s', on='Time interval')['RWY_Hold Delay'].mean()
    #                df_rwy_avg = df_rwy_avg.reset_index()
    #                df_rwy = df_rwy.reset_index()
    #                df_rwy = df_rwy.drop(columns=['RWY_Hold Delay'])
    #                
    #                
    #                df_final_rwy = pd.merge(df_rwy, df_rwy_avg,  how='left', on=['index'], copy=True)
    #                df_final_rwy = df_final_rwy.drop(columns=['index'])
    #                df_final_rwy['Time interval'] = pd.to_timedelta(df_final_rwy['Time interval']) # convert to timedelta to calculate seconds
    #                df_final_rwy['Time interval'] = df_final_rwy['Time interval'].dt.seconds
    #                
    #                #For Push/Start Delay
    #                df_ps = df_delay_input
    #                df_ps = df_ps.groupby(['Time interval'])['Push/Start Delay'].mean()
    #                df_ps = df_ps.reset_index() 
    #                dh_ps_avg = df_ps.rolling(window='3600s', on='Time interval')['Push/Start Delay'].mean()
    #                dh_ps_avg = dh_ps_avg.reset_index()
    #                df_ps = df_ps.reset_index()
    #                df_ps = df_ps.drop(columns=['Push/Start Delay'])
    #                #df_ps['Time interval'] = df_ps['Time interval'].apply(lambda x: x.time())
    #                #df_rwy_arr['Arr Time Interval'] = df_rwy_arr['Arr Time Interval'].apply(lambda x: x.time())
    #                
    #               
    #                df_final_ps = pd.merge(df_ps, dh_ps_avg,  how='left', on=['index'], copy=True)
    #                df_final_ps = df_final_ps.drop(columns=['index'])
    #                df_final_ps['Time interval'] = pd.to_timedelta(df_final_ps['Time interval']) # convert to timedelta to calculate seconds
    #                df_final_ps['Time interval'] = df_final_ps['Time interval'].dt.seconds
    #                
    #                df_rwy['Time interval'] = df_rwy['Time interval'].apply(lambda x: x.time())
    #                #Extract lists to plot
    ##                H_delay_time = df_rwy['Time interval'].tolist()
    ##                H_delay = df_final_rwy['RWY_Hold Delay'].tolist()
    #                PS_time = df_rwy['Time interval'].tolist()
    #                PS_delay = df_final_ps['Push/Start Delay'].tolist()
    #                    
    
                    return {'c': time_interval,
                            'd': PS_delay}
                    
           
                
                
                def print_PS_delay(ab, color):
                    PS_time = ab['c']
                    PS_delay = ab['d']
                    B = f.add_subplot(111)
                    B.plot(PS_time, PS_delay, color)
                    #B.set_title('PUSH/START DELAY', loc = 'right')
                    if OP_FLAG == True:#Plot Operational Data
                        
                        B.plot(OP_Time_PS_Delay,OP_PS_Delay, 'b')    
                        #plt.legend(('Model','Operational Data'), loc = 'upper right')
                
                if new_set_FLAG == True:
                    
                    if m >=1: 
                        f = Figure(figsize=(5,5), dpi=100)
                        B = f.add_subplot(111)
                        Push_delay = delay(df_delay, df_dep_output)
                        print_PS_delay(Push_delay,'k')
                        Push_delay2 = delay(df_delay2, df_dep_output2)
                        print_PS_delay(Push_delay2,'g')
                        
                        
                        if OP_FLAG == True:
                            B.legend(('Model','Operational Data', 'Model 2'), loc = 'upper right')
                        else:
                            B.legend(('Model', 'Model 2'), loc = 'upper right')
                        if m>=2:
                            Push_delay3 = delay(df_delay3, df_dep_output3)
                            print_PS_delay(Push_delay3,'c')
                            
                            if OP_FLAG == True:
                                B.legend(('Model','Operational Data', 'Model 2', 'Model 3'), loc = 'upper right')
                            else:
                                B.legend(('Model', 'Model 2', 'Model 3'), loc = 'upper right')
                            if m>=3:
                                Push_delay4 = delay(df_delay4, df_dep_output4)
                                print_PS_delay(Push_delay4,'m')
                                
                                
                                if OP_FLAG == True:
                                    B.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4'), loc = 'upper right')
                                else:
                                    B.legend(('Model', 'Model 2', 'Model 3', 'Model 4'), loc = 'upper right')
                                if m>=4:
                                    Push_delay5 = delay(df_delay5, df_dep_output5)
                                    print_PS_delay(Push_delay5,'y')
                                    
                                    
                                    if OP_FLAG == True:
                                        B.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5'), loc = 'upper right')
                                    else:
                                        B.legend(('Model', 'Model 2', 'Model 3', 'Model 4', 'Model 5'), loc = 'upper right')
                                    if m>=5:
                                        Push_delay6 = delay(df_delay6, df_dep_output6)
                                        print_PS_delay(Push_delay6, color='purple')
                                        
                                       
                                        if OP_FLAG == True:
                                            B.legend(('Model','Operational Data', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                                        else:
                                            B.legend(('Model', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                        
                        B.grid(color='b', linestyle='-', linewidth=0.1)
                        B.set_xlabel('Time')
                        B.set_ylabel('Seconds of delay')    
                        canvas = FigureCanvasTkAgg(f, self)
                        canvas.show()
                        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                
                        toolbar = NavigationToolbar2TkAgg(canvas, self)
                        toolbar.update()
                        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                else:
                    f = Figure(figsize=(5,5), dpi=100)
                    
                    B = f.add_subplot(111)
                    Push_delay = delay(df_delay, df_dep_output)
                    print_PS_delay(Push_delay,'r')
                    B.set_xlabel('Time')
                    B.set_ylabel('Seconds of delay')
                    B.grid(color='b', linestyle='-', linewidth=0.1)
                    canvas = FigureCanvasTkAgg(f, self)
                    canvas.show()
                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)        
                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                    toolbar.update()
                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    
    # ============================================================================#
    #                          ARRIVAL DELAY                                      # 
    # ============================================================================#
    #if arr_delay_FLAG == True:
    class ArrivalDelay(tk.Frame):
    
        def __init__(self, parent, controller):
            tk.Frame.__init__(self, parent)
            label = tk.Label(self, text="ARRIVALS DELAY", font=LARGE_FONT)
            label.pack(pady=10,padx=10)
    
            button1 = ttk.Button(self, text="Back to Home",
                                command=lambda: controller.show_frame(StartPage))
            button1.pack()
            
            
            if convergenceFLAG ==True:
                button1 = ttk.Button(self, text="Convergence Throughput",
                                    command=lambda: controller.show_frame(Conv))
                button1.pack()
            if Thr_FLAG == True:
                button = ttk.Button(self, text="Throughput",
                                    command=lambda: controller.show_frame(Thr))
                button.pack()
            if Delay_FLAG == True:
                button2 = ttk.Button(self, text="RWY Hold Delay",
                                    command=lambda: controller.show_frame(DepDelay))
                button2.pack()
            if Delay_FLAG == True:
                button5 = ttk.Button(self, text="Push/Start Delay",
                                    command=lambda: controller.show_frame(DepDelay2))
                button5.pack()
            
            if Seq_FLAG == True:
                button4 = ttk.Button(self, text="Sequence",
                                    command=lambda: controller.show_frame(Seq))
                button4.pack()
            if ADA_buffer_FLAG == True:
                button5 = ttk.Button(self, text="ADA Buffer",
                                    command=lambda: controller.show_frame(ADAbuffer))
                button5.pack()
    
            
            
            if (df_delay.empty == True):
                    print("No departures. Nothing to show")
            else:     
                def ArrDelay(df_delay, df_arr_output):
                    interval15=[]
                    df_arr_delay = pd.DataFrame()
                    # Add arrival delay values
                    df_arr_delay['ARR_Delay'] = df_delay['Arrival Delay']
                    # print when those values occure
                    df_arr_delay['Time1'] = df_arr_output['ACTUAL Landing Time']
                    # round each time value to 15 minutes
                    df_arr_delay= df_arr_delay.dropna(subset=['Time1'])
                    time1_list = df_arr_delay['Time1'].tolist()
                    #a=[]
                    for a in time1_list:
                        b = int(int(a/900)*900)
                        interval15.append(b)
                        
                    df_arr_delay['interval15'] = interval15
                    df_arr_delay = df_arr_delay.drop(columns=['Time1'])
                    #Group data by the time interval, if there are multiple values for the same time interval, take the mean.
                    df_arr_delay = df_arr_delay.groupby(['interval15'])['ARR_Delay'].mean()
                    #make the rolling average
                    df_arr_delay = df_arr_delay.reset_index()
                    df_arr_delay2 = df_arr_delay.rolling(window=4, on='interval15')['ARR_Delay'].mean()
                    ARR_delay = df_arr_delay2.tolist()
                    df_arr_delay['DATE'] = pd.to_datetime(df_arr_delay['interval15'],unit='s')
                    df_arr_delay['DATE'] = df_arr_delay['DATE'].apply(lambda x: x.time())
                    
                    time_interval = df_arr_delay['DATE']
    #                print(df_arr_delay['Time1'].value())
    #                df_delay_input['Arr Time Interval'] = pd.DatetimeIndex(df_delay_input['Arr Time Interval'])
    #                #For RWY_Hold Delay
    #                df_rwy_arr = df_delay_input
    #                df_rwy_arr = df_rwy_arr.groupby(['Arr Time Interval'])['Arr Delay'].mean()
    #                df_rwy_arr = df_rwy_arr.reset_index()
    #                df_rwy_arr_avg = df_rwy_arr.rolling(window='3600s', on='Arr Time Interval')['Arr Delay'].mean()
    #                df_rwy_arr_avg = df_rwy_arr_avg.reset_index()
    #                df_rwy_arr = df_rwy_arr.reset_index()
    #                df_rwy_arr = df_rwy_arr.drop(columns=['Arr Delay'])
    #                
    #                df_final_arr_rwy = pd.merge(df_rwy_arr, df_rwy_arr_avg,  how='left', on=['index'], copy=True)
    #                df_final_arr_rwy = df_final_arr_rwy.drop(columns=['index'])
    #                df_final_arr_rwy['Arr Time Interval'] = pd.to_timedelta(df_final_arr_rwy['Arr Time Interval']) # convert to timedelta to calculate seconds
    #                df_final_arr_rwy['Arr Time Interval'] = df_final_arr_rwy['Arr Time Interval'].dt.seconds
    #                
    #                df_rwy_arr['Arr Time Interval'] = df_rwy_arr['Arr Time Interval'].apply(lambda x: x.time())           
    #                
    #                #Extract lists to plot
    #                arr_delay_time = df_rwy_arr['Arr Time Interval'].tolist()
    #                arr_delay = df_final_arr_rwy['Arr Delay'].tolist()
    #                
                        
                    #        return(H_delay_time,H_delay,PS_time,PS_delay)
                    return {'a': time_interval,
                            'b': ARR_delay}
               
                def plotArrDelay(ab, color):            
                    arr_delay_time = ab['a']
                    arr_delay = ab['b'] 
                    #f = Figure(figsize=(5,5), dpi=100)
                    A = f.add_subplot(111)                      
                    A.plot(arr_delay_time, arr_delay, color)     
                
                
                if new_set_FLAG == True:
                     
                    if m >=1:          
                        f = Figure(figsize=(5,5), dpi=100)
                        A = f.add_subplot(111)  
                        plotArrDelay(ArrDelay(df_delay, df_arr_output),'k')
                        plotArrDelay(ArrDelay(df_delay2, df_arr_output2),'g')
                        
                        A.legend(('Model', 'Model 2'), loc = 'upper right')
                        if m>=2:                        
                            plotArrDelay(ArrDelay(df_delay3, df_arr_output3),'c')
                            
                            A.legend(('Model', 'Model 2', 'Model 3'), loc = 'upper right')
                            if m>=3:
                                plotArrDelay(ArrDelay(df_delay4, df_arr_output4),'m')
                               
                                A.legend(('Model', 'Model 2', 'Model 3', 'Model 4'), loc = 'upper right')
                                if m>=4:
                                    #A.plot(ArrDelay(df_delay5)['a'],ArrDelay(df_delay2)['b'],'y')
                                    plotArrDelay(ArrDelay(df_delay5, df_arr_output5),'y')
                                    
                                    A.legend(('Model', 'Model 2', 'Model 3', 'Model 4', 'Model 5'), loc = 'upper right')
                                    if m>=5:
                                        #A.plot(ArrDelay(df_delay6)['a'],ArrDelay(df_delay2)['b'],'purple')
                                        plotArrDelay(ArrDelay(df_delay6, df_arr_output6),'purple')
                                        
                                        A.legend(('Model', 'Model 2', 'Model 3', 'Model 4', 'Model 5', 'Model 6'), loc = 'upper right')
                        A.set_xlabel('Time')
                        A.set_ylabel('Seconds of delay')                
                        A.grid(color='b', linestyle='-', linewidth=0.1)
                        canvas = FigureCanvasTkAgg(f, self)
                        canvas.show()
                        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                
                        toolbar = NavigationToolbar2TkAgg(canvas, self)
                        toolbar.update()
                        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                        
                else:
                    f = Figure(figsize=(5,5), dpi=100)
                    A = f.add_subplot(111)                                   
                    A.plot(ArrDelay(df_delay, df_arr_output)['a'],ArrDelay(df_delay, df_arr_output)['b'],'k')
                    A.set_xlabel('Time')
                    A.set_ylabel('Seconds of delay')
                    A.grid(color='b', linestyle='-', linewidth=0.1)
                    
                    canvas = FigureCanvasTkAgg(f, self)
                    canvas.show()
                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)        
                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                    toolbar.update()
                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    
            
    # ============================================================================#
    #                           SEQUENCE                                          #
    # ============================================================================#
    #if Seq_FLAG == True:
    class Seq(tk.Frame):
    
        def __init__(self, parent, controller):
            tk.Frame.__init__(self, parent)
            label = tk.Label(self, text="SEQUENCE", font=LARGE_FONT)
            label.pack(pady=10,padx=10)
            
            button1 = ttk.Button(self, text="Back to Home",
                                command=lambda: controller.show_frame(StartPage))
            button1.pack() 
            if convergenceFLAG ==True:
                button1 = ttk.Button(self, text="Convergence Throughput",
                                    command=lambda: controller.show_frame(Conv))
                button1.pack()
            if Thr_FLAG == True:
                button = ttk.Button(self, text="Throughput",
                                    command=lambda: controller.show_frame(Thr))
                button.pack()
            
            if Delay_FLAG == True:
                button2 = ttk.Button(self, text="RWY Hold Delay",
                                    command=lambda: controller.show_frame(DepDelay))
                button2.pack()
            if Delay_FLAG == True:
                button5 = ttk.Button(self, text="Push/Start Delay",
                                    command=lambda: controller.show_frame(DepDelay2))
                button5.pack()
            
            if arr_delay_FLAG == True:
                button3 = ttk.Button(self, text="Arrivals Delay",
                                    command=lambda: controller.show_frame(ArrivalDelay))
                button3.pack()
                
            if ADA_buffer_FLAG == True:
                button5 = ttk.Button(self, text="ADA Buffer",
                                    command=lambda: controller.show_frame(ADAbuffer))
                button5.pack()
                
            
                
           
                
            
                
            
            def sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, number):
    
                df_sequence = pd.DataFrame()
                #Arrivals 
                
                df_sequence['ARRIVAL'] = df_arr_output['ACTUAL Landing Time'] + (df_arr_output['AROT']/2)
                df_sequence['ARRIVAL_error'] = df_arr_output['AROT']/2
                df_sequence['ARRIVAL_spacing'] = df_arr_output['ACTUAL Landing Time']
                df_temp = pd.DataFrame()
                df_temp['MAX Constraint'] = df_arr_output['MAX Constraint']
                df_temp = df_temp.drop([0])
                df_temp = df_temp.reset_index()
                df_temp = df_temp.drop(columns=['index'])
                df_temp['Arrival_ZERO'] = 0
                df_sequence['ARRIVAL_spacing_error'] = df_temp['MAX Constraint']           
                df_sequence['Arrival_ZERO'] = df_temp['Arrival_ZERO']
                
                #Positions
                df_sequence['main_position'] = number
                df_sequence['arr_spacing_position'] = number+0.005
                
                #Annotation
                
                df_temp = pd.DataFrame()
                df_temp['ARR_ID'] = df_arr_output['Arrival ID'].astype(str)
                df_temp['ARR_WAKE'] = df_rwy_calcs['ARRIVAL actual WAKE'].astype(str)
                df_temp['ARR_DELAY'] = df_arr_output['Arrival DELAY'].astype(str)
                df_temp['ARRIVAL_LABEL'] = 'ID = ' + df_temp['ARR_ID'] + ' | WAKE = ' + df_temp['ARR_WAKE'] + ' | Delay = ' + df_temp['ARR_DELAY']
                
                df_sequence['ARRIVAL_LABEL'] = df_temp['ARRIVAL_LABEL']
                
                df_temp = pd.DataFrame()
                df_temp['reason'] = df_arr_output['MAX Constraint Label']
                df_temp = df_temp.drop([0])
                df_temp = df_temp.reset_index()
                df_temp = df_temp.drop(columns=['index'])
                df_temp['value'] = df_sequence['ARRIVAL_spacing_error'].astype(str)
                df_temp['ARRIVAL_spacing_LABEL'] = df_temp['reason'] + ' : ' + df_temp['value']
                
                df_sequence['ARRIVAL_spacing_LABEL'] = df_temp['ARRIVAL_spacing_LABEL']
                
                # LISTS to plot
                
                #----ARRIVALS----#
                main_arrival = df_sequence['ARRIVAL'].tolist()
                main_arrival_error = df_sequence['ARRIVAL_error'].tolist()
                arrival_spacing = df_sequence['ARRIVAL_spacing'].tolist()
                arrival_spacing_error = df_sequence['ARRIVAL_spacing_error'].tolist()
                arrival_zero = df_sequence['Arrival_ZERO'].tolist()
                #-Labels:
                arrival_label = df_sequence['ARRIVAL_LABEL'].tolist()
                arrival_spacing_label = df_sequence['ARRIVAL_spacing_LABEL'].tolist()
                
                #-----POSITIONS------#
                main_data_position = df_sequence['main_position'].tolist()
                arrival_spacing_position = df_sequence['arr_spacing_position'].tolist()
                #Data prep for tags
                labels = arrival_label +  arrival_spacing_label  
                labels_y = main_data_position +  arrival_spacing_position 
                labels_x = main_arrival +  arrival_spacing 
            
            
                return{'0a': arrival_zero,
                       '1' : main_arrival, 
                       '2' : main_arrival_error,
                       '3' : arrival_spacing,
                       '4' : arrival_spacing_error,                   
                       '9' : main_data_position,
                       '10' : arrival_spacing_position,                   
                       '12' : labels,
                       '13' : labels_y,
                       '14' : labels_x}
             
            def sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, number):
                df_sequence = pd.DataFrame()
                #Departure
                
                df_sequence['DEPARTURES'] = df_dep_output['Departure_RWY_ENTRY'] + (df_dep_output['DROT']/2)
                df_sequence['DEPARTURES_error'] = df_dep_output['DROT']/2
                df_sequence['DEPARTURES_spacing'] = df_dep_output['Departure_RWY_ENTRY']
                df_temp = pd.DataFrame()
                df_temp['Dep MIN Separation'] = df_dep_output['Dep MIN Separation']
                df_temp=df_temp.drop([0])
                df_temp = df_temp.reset_index()
                df_temp = df_temp.drop(columns=['index'])
                df_temp['Departure_ZERO'] = 0
                df_sequence['DEPARTURES_spacing_error'] = df_temp['Dep MIN Separation']            
                df_sequence['Departure_ZERO'] = df_temp['Departure_ZERO']
                            
                #Positions
                df_sequence['main_position'] = number
                df_sequence['dep_spacing_position'] = number-0.005
                
                df_temp = pd.DataFrame()
                df_temp['DEP_ID'] = df_dep_output['Departure ID'].astype(str)
                df_temp['DEP_SID'] = df_dep_output['SID GROUP'].astype(str)
                df_temp['DEP_WAKE'] = df_dep_output['WAKE'].astype(str)
                df_temp['DEP_DELAY1'] = df_dep_output['DELAY DepSTANDqueue'] + df_dep_output['DELAY TAXIhold'] + df_dep_output['DELAY RWYqueue']
                df_temp['DEP_DELAY'] = df_temp['DEP_DELAY1'].astype(str)
                df_temp['DEPARTURE_LABEL'] = 'ID = ' + df_temp['DEP_ID'] + ' | SID = ' +df_temp['DEP_SID'] +' | WAKE = ' + df_temp['DEP_WAKE'] + ' | Delay = ' + df_temp['DEP_DELAY']
                
                df_sequence['DEPARTURE_LABEL'] = df_temp['DEPARTURE_LABEL']
                
                df_temp = pd.DataFrame()
                df_temp['reason'] = df_dep_output['Dep MIN Separation Label']
                df_temp = df_temp.drop([0])
                df_temp = df_temp.reset_index()
                df_temp = df_temp.drop(columns=['index'])
                df_temp['value'] = df_sequence['DEPARTURES_spacing_error'].astype(str)
                df_temp['DEPARTURES_spacing_LABEL'] = df_temp['reason'] + ' : ' + df_temp['value']
                
                df_sequence['DEPARTURES_spacing_LABEL'] = df_temp['DEPARTURES_spacing_LABEL']
                
                # LISTS to plot
               
                #----DEPARTURES----#
                main_departure = df_sequence['DEPARTURES'].tolist()
                main_departure_error = df_sequence['DEPARTURES_error'].tolist()
                departure_spacing = df_sequence['DEPARTURES_spacing'].tolist()
                departure_spacing_error = df_sequence['DEPARTURES_spacing_error'].tolist()
                departure_zero = df_sequence['Departure_ZERO']
                #-Labels:
                departure_label = df_sequence['DEPARTURE_LABEL'].tolist()
                departure_spacing_label = df_sequence['DEPARTURES_spacing_LABEL'].tolist()
                
                #-----POSITIONS------#
                main_data_position = df_sequence['main_position'].tolist()
                departure_spacing_position = df_sequence['dep_spacing_position'].tolist()
                
                
                
                #Data prep for tags
                labels = departure_label +  departure_spacing_label
                labels_y = main_data_position + departure_spacing_position
                labels_x = main_departure +  departure_spacing
            
            
                return{'0b': departure_zero,
                       '5' : main_departure, 
                       '6' : main_departure_error,
                       '7' : departure_spacing,
                       '8' : departure_spacing_error,
                       '9' : main_data_position,
                       '11' : departure_spacing_position,
                       '12' : labels,
                       '13' : labels_y,
                       '14' : labels_x}
            
            def sequence(df_arr_output_input, df_rwy_calcs_input, df_dep_output_input, number): ######MIX MODE
                
                df_sequence = pd.DataFrame()
                #Arrivals 
                
                df_sequence['ARRIVAL'] = df_arr_output_input['ACTUAL Landing Time'] + (df_arr_output_input['AROT']/2)
                df_sequence['ARRIVAL_error'] = df_arr_output_input['AROT']/2
                df_sequence['ARRIVAL_spacing'] = df_arr_output_input['ACTUAL Landing Time']
                df_temp = pd.DataFrame()
                df_temp['MAX Constraint'] = df_arr_output_input['MAX Constraint']
                df_temp = df_temp.drop([0])
                df_temp = df_temp.reset_index()
                df_temp = df_temp.drop(columns=['index'])
                df_temp['Arrival_ZERO'] = 0
                df_sequence['ARRIVAL_spacing_error'] = df_temp['MAX Constraint']           
                df_sequence['Arrival_ZERO'] = df_temp['Arrival_ZERO']
                
                
                #Departure
                
                df_sequence['DEPARTURES'] = df_dep_output_input['Departure_RWY_ENTRY'] + (df_dep_output_input['DROT']/2)
                df_sequence['DEPARTURES_error'] = df_dep_output_input['DROT']/2
                df_sequence['DEPARTURES_spacing'] = df_dep_output_input['Departure_RWY_ENTRY']
                df_temp = pd.DataFrame()
                df_temp['Dep MIN Separation'] = df_dep_output_input['Dep MIN Separation']
                df_temp=df_temp.drop([0])
                df_temp = df_temp.reset_index()
                df_temp = df_temp.drop(columns=['index'])
                df_temp['Departure_ZERO'] = 0
                df_sequence['DEPARTURES_spacing_error'] = df_temp['Dep MIN Separation']            
                df_sequence['Departure_ZERO'] = df_temp['Departure_ZERO']
                            
                #Positions
                df_sequence['main_position'] = number
                df_sequence['arr_spacing_position'] = number+0.005
                df_sequence['dep_spacing_position'] = number-0.005
                
                
                #Annotation
                
                df_temp = pd.DataFrame()
                df_temp['ARR_ID'] = df_arr_output_input['Arrival ID'].astype(str)
                df_temp['ARR_WAKE'] = df_rwy_calcs_input['ARRIVAL actual WAKE'].astype(str)
                df_temp['ARR_DELAY'] = df_arr_output_input['Arrival DELAY'].astype(str)
                df_temp['ARRIVAL_LABEL'] = 'ID = ' + df_temp['ARR_ID'] + ' | WAKE = ' + df_temp['ARR_WAKE'] + ' | Delay = ' + df_temp['ARR_DELAY']
                
                df_sequence['ARRIVAL_LABEL'] = df_temp['ARRIVAL_LABEL']
                
                df_temp = pd.DataFrame()
                df_temp['DEP_ID'] = df_dep_output_input['Departure ID'].astype(str)
                df_temp['DEP_SID'] = df_dep_output_input['SID GROUP'].astype(str)
                df_temp['DEP_WAKE'] = df_dep_output_input['WAKE'].astype(str)
                df_temp['DEP_DELAY1'] = df_dep_output_input['DELAY DepSTANDqueue'] + df_dep_output_input['DELAY TAXIhold'] + df_dep_output_input['DELAY RWYqueue']
                df_temp['DEP_DELAY'] = df_temp['DEP_DELAY1'].astype(str)
                df_temp['DEPARTURE_LABEL'] = 'ID = ' + df_temp['DEP_ID'] + ' | SID = ' +df_temp['DEP_SID'] +' | WAKE = ' + df_temp['DEP_WAKE'] + ' | Delay = ' + df_temp['DEP_DELAY']
                
                df_sequence['DEPARTURE_LABEL'] = df_temp['DEPARTURE_LABEL']
                
                df_temp = pd.DataFrame()
                df_temp['reason'] = df_arr_output_input['MAX Constraint Label']
                df_temp = df_temp.drop([0])
                df_temp = df_temp.reset_index()
                df_temp = df_temp.drop(columns=['index'])
                df_temp['value'] = df_sequence['ARRIVAL_spacing_error'].astype(str)
                df_temp['ARRIVAL_spacing_LABEL'] = df_temp['reason'] + ' : ' + df_temp['value']
                
                df_sequence['ARRIVAL_spacing_LABEL'] = df_temp['ARRIVAL_spacing_LABEL']
                
                df_temp = pd.DataFrame()
                df_temp['reason'] = df_dep_output_input['Dep MIN Separation Label']
                df_temp = df_temp.drop([0])
                df_temp = df_temp.reset_index()
                df_temp = df_temp.drop(columns=['index'])
                df_temp['value'] = df_sequence['DEPARTURES_spacing_error'].astype(str)
                df_temp['DEPARTURES_spacing_LABEL'] = df_temp['reason'] + ' : ' + df_temp['value']
                
                df_sequence['DEPARTURES_spacing_LABEL'] = df_temp['DEPARTURES_spacing_LABEL']
                
                # LISTS to plot
                
                #----ARRIVALS----#
                main_arrival = df_sequence['ARRIVAL'].tolist()
                main_arrival_error = df_sequence['ARRIVAL_error'].tolist()
                arrival_spacing = df_sequence['ARRIVAL_spacing'].tolist()
                arrival_spacing_error = df_sequence['ARRIVAL_spacing_error'].tolist()
                arrival_zero = df_sequence['Arrival_ZERO'].tolist()
                #-Labels:
                arrival_label = df_sequence['ARRIVAL_LABEL'].tolist()
                arrival_spacing_label = df_sequence['ARRIVAL_spacing_LABEL'].tolist()
                
                #----DEPARTURES----#
                main_departure = df_sequence['DEPARTURES'].tolist()
                main_departure_error = df_sequence['DEPARTURES_error'].tolist()
                departure_spacing = df_sequence['DEPARTURES_spacing'].tolist()
                departure_spacing_error = df_sequence['DEPARTURES_spacing_error'].tolist()
                departure_zero = df_sequence['Departure_ZERO']
                #-Labels:
                departure_label = df_sequence['DEPARTURE_LABEL'].tolist()
                departure_spacing_label = df_sequence['DEPARTURES_spacing_LABEL'].tolist()
                
                #-----POSITIONS------#
                main_data_position = df_sequence['main_position'].tolist()
                arrival_spacing_position = df_sequence['arr_spacing_position'].tolist()
                departure_spacing_position = df_sequence['dep_spacing_position'].tolist()
                
                
                
                #Data prep for tags
                labels = arrival_label + departure_label + arrival_spacing_label  + departure_spacing_label
                labels_y = main_data_position + main_data_position + arrival_spacing_position + departure_spacing_position
                labels_x = main_arrival + main_departure + arrival_spacing + departure_spacing
            
            
                return{'0a': arrival_zero,
                       '0b': departure_zero,
                       '1' : main_arrival, 
                       '2' : main_arrival_error,
                       '3' : arrival_spacing,
                       '4' : arrival_spacing_error,
                       '5' : main_departure, 
                       '6' : main_departure_error,
                       '7' : departure_spacing,
                       '8' : departure_spacing_error,
                       '9' : main_data_position,
                       '10' : arrival_spacing_position,
                       '11' : departure_spacing_position,
                       '12' : labels,
                       '13' : labels_y,
                       '14' : labels_x}
            
            #------ ARRIVALS only -------#
            
            if df_dep_output.empty ==True:
                  
                if new_set_FLAG == True: 
                    if df_dep_output2.empty ==True:#arr only
                        if m==1: # two arr only comparison
                            f = Figure(figsize=(5,5), dpi=100)  
                            A = f.add_subplot(111)
                            ax = f.add_subplot(111)
                            
                            labels =  sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']  + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                            tag_text_use = np.array(list(labels))
                            labels_y = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13']
                            labels_x = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14']
                            
                            tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                            #plt.axhline(y = 10, color='w')                    
                            #A.axhline(y=0.5, color='w')
                               
                            
                            A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='maroon', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='indigo', fmt='o', markersize=8, capsize=10, )
                            
                                            
                            annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                            bbox=dict(boxstyle="round", fc="w"),
                                            arrowprops=dict(arrowstyle="->"))
                            annot.set_visible(False)
                            
                        
                            def update_annot(ind):
                            
                                pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                annot.xy = pos
                                text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                annot.set_text(text)
                            
                        
                                            
                            def hover(event):
                                vis = annot.get_visible()
                                if event.inaxes == ax:
                                    cont, ind = tags_main_data.contains(event)
                                    if cont:
                                        update_annot(ind)
                                        annot.set_visible(True)
                                        canvas.draw_idle()
                                    else:
                                        if vis:
                                            annot.set_visible(False)
                                            canvas.draw_idle()
                            
                        
                            
                            A.set_xlabel('Seconds of the day')        
                            A.axes.get_yaxis().set_visible(False) 
                            A.grid(color='b', linestyle='-', linewidth=0.1)    
                            A.legend(("Legend","Arrivals M1","Arrivals Spacing M1","Arrivals M2", "Arrivals Spacing M2"), loc = 'upper right')
                            #plt.title("Sequence analysis")
                            canvas = FigureCanvasTkAgg(f, self)
                            canvas.show()
                            canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                            canvas.mpl_connect("motion_notify_event", hover)
                            toolbar = NavigationToolbar2TkAgg(canvas, self)
                            toolbar.update()
                            canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                        elif m==2: # 3 ARRonly comparison
                            f = Figure(figsize=(5,5), dpi=100)  
                            A = f.add_subplot(111)
                            ax = f.add_subplot(111)
                            
                            labels =  sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']  + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] + sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['12']
                            tag_text_use = np.array(list(labels))
                            labels_y = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] + sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['13'] 
                            labels_x = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] + sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['14']
                            
                            tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                            #plt.axhline(y = 10, color='w')                    
                            #A.axhline(y=0.5, color='w')
                               
                            
                            A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='maroon', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='indigo', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['1'], sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2'], sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2']], color='salmon', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['3'], sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['10'], xerr=[sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0a'], sequenceArrOnly(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['4']], color='orchid', fmt='o', markersize=8, capsize=10, )
                                            
                            annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                            bbox=dict(boxstyle="round", fc="w"),
                                            arrowprops=dict(arrowstyle="->"))
                            annot.set_visible(False)
                            
                        
                            def update_annot(ind):
                            
                                pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                annot.xy = pos
                                text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                annot.set_text(text)
                            
                        
                                            
                            def hover(event):
                                vis = annot.get_visible()
                                if event.inaxes == ax:
                                    cont, ind = tags_main_data.contains(event)
                                    if cont:
                                        update_annot(ind)
                                        annot.set_visible(True)
                                        canvas.draw_idle()
                                    else:
                                        if vis:
                                            annot.set_visible(False)
                                            canvas.draw_idle()
                            
                        
                            
                            A.set_xlabel('Seconds of the day')        
                            A.axes.get_yaxis().set_visible(False) 
                            A.grid(color='b', linestyle='-', linewidth=0.1)    
                            A.legend(("Legend","Arrivals M1","Arrivals Spacing M1","Arrivals M2", "Arrivals Spacing M2","Arrivals M3", "Arrivals Spacing M3"), loc = 'upper right')
                            #plt.title("Sequence analysis")
                            canvas = FigureCanvasTkAgg(f, self)
                            canvas.show()
                            canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                            canvas.mpl_connect("motion_notify_event", hover)
                            toolbar = NavigationToolbar2TkAgg(canvas, self)
                            toolbar.update()
                            canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                            #f.canvas.figure.savefig('sequence.png')
                    
                    
                    elif df_arr_output2.empty ==True: # ARR only + DEP only                    
                        f = Figure(figsize=(5,5), dpi=100)  
                        A = f.add_subplot(111)
                        ax = f.add_subplot(111)
                        
                        labels =  sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']  + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                        tag_text_use = np.array(list(labels))
                        labels_y = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                        labels_x = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] 
                                                
                        tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                        #A.axhline(y=0.5, color='w')
                        
                        A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['3'], sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['4'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['5'], sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['6']], color='r', fmt='o', markersize=8, capsize=10, )
                        A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['8'], sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['10'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['11'], sequenceArrOnly(df_arr_output, df_rwy_calcs, 1)['9']], color='purple', fmt='o', markersize=8, capsize=10)
                        
                        A.errorbar(sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                        A.errorbar(sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], fmt='o', markersize=8, capsize=10)
    
                        annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->"))
                        annot.set_visible(False)
                        
                    
                        def update_annot(ind):
                        
                            pos = tags_main_data.get_offsets()[ind["ind"][0]]
                            annot.xy = pos
                            text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                            annot.set_text(text)
                        
                    
                                        
                        def hover(event):
                            vis = annot.get_visible()
                            if event.inaxes == ax:
                                cont, ind = tags_main_data.contains(event)
                                if cont:
                                    update_annot(ind)
                                    annot.set_visible(True)
                                    canvas.draw_idle()
                                else:
                                    if vis:
                                        annot.set_visible(False)
                                        canvas.draw_idle()
                        
                    
                        
                        A.set_xlabel('Seconds of the day')        
                        A.axes.get_yaxis().set_visible(False) 
                        A.grid(color='b', linestyle='-', linewidth=0.1)    
                        A.legend(("Legend","Arrivals M1","Arrivals Spacing M1","Departures M2", "Departures Spacing M2"), loc = 'upper right')
                        #plt.title("Sequence analysis")
                        canvas = FigureCanvasTkAgg(f, self)
                        canvas.show()
                        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                        canvas.mpl_connect("motion_notify_event", hover)
                        toolbar = NavigationToolbar2TkAgg(canvas, self)
                        toolbar.update()
                        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                        
                    else:#ARR only + MIXED
                        f = Figure(figsize=(5,5), dpi=100)  
                        A = f.add_subplot(111)
                        ax = f.add_subplot(111)
                        
                        
                        labels =  sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']  + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                        tag_text_use = np.array(list(labels))
                        labels_y = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                        labels_x = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] 
                                                
                        tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                        #A.axhline(y=0.5, color='w')
                                            
                        A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                        A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
    
                        A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                        A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                        A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], fmt='o', markersize=8, capsize=10)
                        A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
    
                        annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->"))
                        annot.set_visible(False)
                        def update_annot2(ind):
                        
                            pos = tags_main_data.get_offsets()[ind["ind"][0]]
                            annot.xy = pos
                            text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                            annot.set_text(text)
                        def hover(event):
                            vis = annot.get_visible()
                            if event.inaxes == ax:
                                cont, ind = tags_main_data.contains(event)
                                if cont:
                                    update_annot2(ind)
                                    annot.set_visible(True)
                                    canvas.draw_idle()
                                else:
                                    if vis:
                                        annot.set_visible(False)
                                        canvas.draw_idle()
                        A.set_xlabel('Seconds of the day')        
                        A.axes.get_yaxis().set_visible(False)                     
                        A.grid(color='b', linestyle='-', linewidth=0.1)    
                        A.legend(("Legend","Arrivals M1","Arrivals Spacing M1","Departures M2","Arrivals M2", "Departures Spacing M2","Arrivals Spacing M2"), loc = 'upper right')
                        #plt.title("Sequence analysis")
                        canvas = FigureCanvasTkAgg(f, self)
                        canvas.show()
                        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                        canvas.mpl_connect("motion_notify_event", hover)
                        toolbar = NavigationToolbar2TkAgg(canvas, self)
                        toolbar.update()
                        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                        
                        
                else:  #only one ARRonly              
                    f = Figure(figsize=(5,5), dpi=100)  
                    A = f.add_subplot(111)
                    ax = f.add_subplot(111)
                  
                    labels =  sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] 
                    tag_text_use = np.array(list(labels))
                    
                    tags_main_data = A.scatter(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'],sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] , c='w', s=100)      
                    #plt.axhline(y = 10, color='w')                    
                    #A.axhline(y=0.5, color='w')
                       
                    
                    A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                    A.errorbar(sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                    
                    
                    
                    annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                    bbox=dict(boxstyle="round", fc="w"),
                                    arrowprops=dict(arrowstyle="->"))
                    annot.set_visible(False)
                    
                
                    def update_annot(ind):
                    
                        pos = tags_main_data.get_offsets()[ind["ind"][0]]
                        annot.xy = pos
                        text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                        annot.set_text(text)
                    
                
                                    
                    def hover(event):
                        vis = annot.get_visible()
                        if event.inaxes == ax:
                            cont, ind = tags_main_data.contains(event)
                            if cont:
                                update_annot(ind)
                                annot.set_visible(True)
                                canvas.draw_idle()
                            else:
                                if vis:
                                    annot.set_visible(False)
                                    canvas.draw_idle()
                    
                
                    
                    A.set_xlabel('Seconds of the day')        
                    A.axes.get_yaxis().set_visible(False) 
                    A.grid(color='b', linestyle='-', linewidth=0.1)    
                    A.legend(("Legend","Arrivals","Arrivals Spacing"), loc = 'upper right')
                    #plt.title("Sequence analysis")
                    canvas = FigureCanvasTkAgg(f, self)
                    canvas.show()
                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                    canvas.mpl_connect("motion_notify_event", hover)
                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                    toolbar.update()
                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        
            
            
            #-------DEPARTURES only --------#
            
            elif df_arr_output.empty ==True:
                            
                if new_set_FLAG == True:                
                                  
                    if df_dep_output2.empty ==True: # DEPonly + ARRonly                    
                        f = Figure(figsize=(5,5), dpi=100)  
                        A = f.add_subplot(111)
                        ax = f.add_subplot(111)
                        
                        labels =  sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']+ sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12']
                        tag_text_use = np.array(list(labels))
                        labels_x = sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14']
                        labels_y = sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13']
                        tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                        #A.axhline(y=0.5, color='w')
                           
                        A.errorbar(sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                        A.errorbar(sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                        
                        A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                        A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                      
                                        
                        annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->"))
                        annot.set_visible(False)
                        
                    
                        def update_annot(ind):
                        
                            pos = tags_main_data.get_offsets()[ind["ind"][0]]
                            annot.xy = pos
                            text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                            annot.set_text(text)
                        
                    
                                        
                        def hover(event):
                            vis = annot.get_visible()
                            if event.inaxes == ax:
                                cont, ind = tags_main_data.contains(event)
                                if cont:
                                    update_annot(ind)
                                    annot.set_visible(True)
                                    canvas.draw_idle()
                                else:
                                    if vis:
                                        annot.set_visible(False)
                                        canvas.draw_idle()
                        
                    
                        
                        A.set_xlabel('Seconds of the day')        
                        A.axes.get_yaxis().set_visible(False) 
                        A.grid(color='b', linestyle='-', linewidth=0.1)    
                        A.legend(("Legend","Departures M1","Departures Spacing M1","Arrivals M2", "Arrivals Spacing M2"), loc = 'upper right')
                        #plt.title("Sequence analysis")
                        canvas = FigureCanvasTkAgg(f, self)
                        canvas.show()
                        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                        canvas.mpl_connect("motion_notify_event", hover)
                        toolbar = NavigationToolbar2TkAgg(canvas, self)
                        toolbar.update()
                        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                    
                    elif df_arr_output2.empty ==True: # DEPonly + DEPonly
                        f = Figure(figsize=(5,5), dpi=100)  
                        A = f.add_subplot(111)
                        ax = f.add_subplot(111)
                        
                        labels =  sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']  + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                        tag_text_use = np.array(list(labels))
                        labels_y = sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                        labels_x = sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] 
                                                
                        tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                        #A.axhline(y=0.5, color='w')
                        
                        A.errorbar(sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                        A.errorbar(sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
    
                        A.errorbar(sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='darkgreen', fmt='o', markersize=8, capsize=10)    
                        A.errorbar(sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], color='indigo', fmt='o', markersize=8, capsize=10)
    
    
                        
                        annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->"))
                        annot.set_visible(False)
                        
                    
                        def update_annot(ind):
                        
                            pos = tags_main_data.get_offsets()[ind["ind"][0]]
                            annot.xy = pos
                            text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                            annot.set_text(text)
                        
                    
                                        
                        def hover(event):
                            vis = annot.get_visible()
                            if event.inaxes == ax:
                                cont, ind = tags_main_data.contains(event)
                                if cont:
                                    update_annot(ind)
                                    annot.set_visible(True)
                                    canvas.draw_idle()
                                else:
                                    if vis:
                                        annot.set_visible(False)
                                        canvas.draw_idle()
                        
                    
                        
                        A.set_xlabel('Seconds of the day')        
                        A.axes.get_yaxis().set_visible(False) 
                        A.grid(color='b', linestyle='-', linewidth=0.1)    
                        A.legend(("Legend","Departures M1","Departures Spacing M1","Departures M2", "Departures Spacing M2"), loc = 'upper right')
                        #plt.title("Sequence analysis")
                        canvas = FigureCanvasTkAgg(f, self)
                        canvas.show()
                        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                        canvas.mpl_connect("motion_notify_event", hover)
                        toolbar = NavigationToolbar2TkAgg(canvas, self)
                        toolbar.update()
                        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                        
                    else: # DEPonly + MIXED
                        f = Figure(figsize=(5,5), dpi=100)  
                        A = f.add_subplot(111)
                        ax = f.add_subplot(111)
                        
                        
                        labels =  sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12']  + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                        tag_text_use = np.array(list(labels))
                        labels_y = sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                        labels_x = sequenceArrOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14']
                                                
                        tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                        #A.axhline(y=0.5, color='w')
                                                   
                                            
                        A.errorbar(sequenceDepOnly(df_dep_output, 1)['3'], sequenceDepOnly(df_dep_output, 1)['4'], xerr=[sequenceDepOnly(df_dep_output, 1)['5'], sequenceDepOnly(df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)
                        A.errorbar(sequenceDepOnly(df_dep_output, 1)['10'], sequenceDepOnly(df_dep_output, 1)['9'], xerr=[sequenceDepOnly(df_dep_output, 1)['11'], sequenceDepOnly(df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                        
                        A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11']], color='darkgreen', fmt='o', markersize=8, capsize=10)    
                        A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='maroon', fmt='o', markersize=8, capsize=10, )
                        A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['15']], color='navy', fmt='o', markersize=8, capsize=10)
                        A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['18'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['19'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['17'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['16']], color='indigo', fmt='o', markersize=8, capsize=10, )
                    
                        annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->"))
                        annot.set_visible(False)
                        def update_annot2(ind):
                        
                            pos = tags_main_data.get_offsets()[ind["ind"][0]]
                            annot.xy = pos
                            text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                            annot.set_text(text)
                        def hover(event):
                            vis = annot.get_visible()
                            if event.inaxes == ax:
                                cont, ind = tags_main_data.contains(event)
                                if cont:
                                    update_annot2(ind)
                                    annot.set_visible(True)
                                    canvas.draw_idle()
                                else:
                                    if vis:
                                        annot.set_visible(False)
                                        canvas.draw_idle()
                        A.set_xlabel('Seconds of the day')        
                        A.axes.get_yaxis().set_visible(False)                     
                        A.grid(color='b', linestyle='-', linewidth=0.1)    
                        A.legend(("Legend","Departures M1","Departures Spacing M1","Departures M2","Arrivals M2", "Departures Spacing M2","Arrivals Spacing M2"), loc = 'upper right')
                        #plt.title("Sequence analysis")
                        canvas = FigureCanvasTkAgg(f, self)
                        canvas.show()
                        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                        canvas.mpl_connect("motion_notify_event", hover)
                        toolbar = NavigationToolbar2TkAgg(canvas, self)
                        toolbar.update()
                        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                        
                else: # DEPonly
                    f = Figure(figsize=(5,5), dpi=100)  
                    A = f.add_subplot(111)
                    ax = f.add_subplot(111)
                                        
                    labels =  sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] 
                    tag_text_use = np.array(list(labels))
                    labels_x = sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] 
                    labels_y = sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13']
                    tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                    #A.axhline(y=0.5, color='w')
                       
                    A.errorbar(sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                    A.errorbar(sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequenceDepOnly(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                    
                    
     
                    
                    annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                    bbox=dict(boxstyle="round", fc="w"),
                                    arrowprops=dict(arrowstyle="->"))
                    annot.set_visible(False)
                    
                
                    def update_annot(ind):
                    
                        pos = tags_main_data.get_offsets()[ind["ind"][0]]
                        annot.xy = pos
                        text = "{}".format(" ".join([sequenceDepOnly(df_dep_output, 1)['7'][n] for n in ind["ind"]]))
                        annot.set_text(text)
                    
                
                                    
                    def hover(event):
                        vis = annot.get_visible()
                        if event.inaxes == ax:
                            cont, ind = tags_main_data.contains(event)
                            if cont:
                                update_annot(ind)
                                annot.set_visible(True)
                                canvas.draw_idle()
                            else:
                                if vis:
                                    annot.set_visible(False)
                                    canvas.draw_idle()
                    
                
                    
                    A.set_xlabel('Seconds of the day')        
                    A.axes.get_yaxis().set_visible(False) 
                    A.grid(color='b', linestyle='-', linewidth=0.1)    
                    A.legend(("Legend","Departures","Departures Spacing"), loc = 'upper right')
                    #plt.title("Sequence analysis")
                    canvas = FigureCanvasTkAgg(f, self)
                    canvas.show()
                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                    canvas.mpl_connect("motion_notify_event", hover)
                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                    toolbar.update()
                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                    
            
            #------------------ MIXED MODE ------------------------#
            
            else: 
                
                if new_set_FLAG == True:
                    if (df_dep_output2.empty ==False) and (df_arr_output2.empty ==False): #mixed both
                        if m==2: #MIXED +MIXED + MIXED
                                    
                            f = Figure(figsize=(5,5), dpi=100)  
                            A = f.add_subplot(111)
                            ax = f.add_subplot(111)
                            
                            
                            labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] +sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['12']
                            tag_text_use = np.array(list(labels))
                            labels_x = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] +sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['14']
                            labels_y = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] + sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['13']
                            tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                            #plt.axhline(y = 10, color='w')                    
                            ##A.axhline(y=0.5, color='w')
                               
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='darkgreen', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='maroon', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], color='navy',fmt = 'o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='indigo', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['5'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['6'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['6']], color='limegreen', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['1'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2']], color='orangered', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['7'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['11'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0b'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['8']], color='royalblue',fmt = 'o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['3'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['10'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0a'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['4']], color='magenta', fmt='o', markersize=8, capsize=10, )
                            
                            
                            
                            
                            annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                            bbox=dict(boxstyle="round", fc="w"),
                                            arrowprops=dict(arrowstyle="->"))
                            annot.set_visible(False)
                            def update_annot2(ind):
                            
                                pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                annot.xy = pos
                                text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                annot.set_text(text)
                            def hover(event):
                                vis = annot.get_visible()
                                if event.inaxes == ax:
                                    cont, ind = tags_main_data.contains(event)
                                    if cont:
                                        update_annot2(ind)
                                        annot.set_visible(True)
                                        canvas.draw_idle()
                                    else:
                                        if vis:
                                            annot.set_visible(False)
                                            canvas.draw_idle()
                            A.set_xlabel('Seconds of the day')        
                            A.axes.get_yaxis().set_visible(False)                     
                            A.grid(color='b', linestyle='-', linewidth=0.1)    
                            A.legend(("Legend","Departures M1","Arrivals M1","Departures Spacing M1","Arrivals Spacing M1","Departures M2","Arrivals M2", "Departures Spacing M2","Arrivals Spacing M2", "Departures M3","Arrivals M3","Departures Spacing M3","Arrivals Spacing M3"), loc = 'upper right')
                            #plt.title("Sequence analysis")
                            canvas = FigureCanvasTkAgg(f, self)
                            canvas.show()
                            canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                            canvas.mpl_connect("motion_notify_event", hover)
                            toolbar = NavigationToolbar2TkAgg(canvas, self)
                            toolbar.update()
                            canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                            #f.canvas.figure.savefig('sequence.png')
                        
                        
                        elif m==3:#MIXED +MIXED + MIXED + MIXED
                                    
                            f = Figure(figsize=(5,5), dpi=100)  
                            A = f.add_subplot(111)
                            ax = f.add_subplot(111)
                            
                            labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] +sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['12'] + sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['12']
                            tag_text_use = np.array(list(labels))
                            labels_x = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] +sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['14'] + sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['14']
                            labels_y = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] + sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['13'] + sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['13']
                            tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                            #plt.axhline(y = 10, color='w')                    
                            #A.axhline(y=0.5, color='w')
                               
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='darkgreen', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='maroon', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], color='navy',fmt = 'o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='indigo', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['5'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['6'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['6']], color='limegreen', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['1'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2']], color='orangered', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['7'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['11'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0b'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['8']], color='royalblue',fmt = 'o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['3'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['10'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0a'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['4']], color='magenta', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['5'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['9'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['6'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['6']], color='lime', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['1'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['9'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['2'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['2']], color='salmon', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['7'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['11'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['0b'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['8']], color='cyan',fmt = 'o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['3'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['10'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['0a'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['4']], color='orchid', fmt='o', markersize=8, capsize=10, )
                            
                            
                            
                            annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                            bbox=dict(boxstyle="round", fc="w"),
                                            arrowprops=dict(arrowstyle="->"))
                            annot.set_visible(False)
                            def update_annot2(ind):
                            
                                pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                annot.xy = pos
                                text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                annot.set_text(text)
                            def hover(event):
                                vis = annot.get_visible()
                                if event.inaxes == ax:
                                    cont, ind = tags_main_data.contains(event)
                                    if cont:
                                        update_annot2(ind)
                                        annot.set_visible(True)
                                        canvas.draw_idle()
                                    else:
                                        if vis:
                                            annot.set_visible(False)
                                            canvas.draw_idle()
                            A.set_xlabel('Seconds of the day')        
                            A.axes.get_yaxis().set_visible(False)                     
                            A.grid(color='b', linestyle='-', linewidth=0.1)    
                            A.legend(("Legend","Departures M1","Arrivals M1","Departures Spacing M1","Arrivals Spacing M1","Departures M2","Arrivals M2", "Departures Spacing M2","Arrivals Spacing M2", "Departures M3","Arrivals M3","Departures Spacing M3","Arrivals Spacing M3", "Departures M4","Arrivals M4","Departures Spacing M4","Arrivals Spacing M4"), loc = 'upper right')
                            #plt.title("Sequence analysis")
                            canvas = FigureCanvasTkAgg(f, self)
                            canvas.show()
                            canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                            canvas.mpl_connect("motion_notify_event", hover)
                            toolbar = NavigationToolbar2TkAgg(canvas, self)
                            toolbar.update()
                            canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                            #f.canvas.figure.savefig('sequence.png')
                        elif m==4: #MIXED +MIXED + MIXED +MIXED + MIXED
                                                            
                            f = Figure(figsize=(5,5), dpi=100)  
                            A = f.add_subplot(111)
                            ax = f.add_subplot(111)
                            
                            
                            labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] +sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['12'] + sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['12'] + sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['12']
                            tag_text_use = np.array(list(labels))
                            labels_x = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] +sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['14'] + sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['14'] + sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['14']
                            labels_y = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] + sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['13'] + sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['13'] + sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['13']
                            tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                            #plt.axhline(y = 10, color='w')                    
                            #A.axhline(y=0.5, color='w')
                               
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='darkgreen', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='maroon', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], color='navy',fmt = 'o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='indigo', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['5'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['6'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['6']], color='limegreen', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['1'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['9'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['2']], color='orangered', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['7'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['11'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0b'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['8']], color='royalblue',fmt = 'o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['3'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['10'], xerr=[sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['0a'], sequence(df_arr_output3, df_rwy_calcs3, df_dep_output3, 0.96)['4']], color='magenta', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['5'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['9'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['6'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['6']], color='lime', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['1'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['9'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['2'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['2']], color='salmon', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['7'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['11'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['0b'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['8']], color='cyan',fmt = 'o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['3'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['10'], xerr=[sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['0a'], sequence(df_arr_output4, df_rwy_calcs4, df_dep_output4, 0.94)['4']], color='orchid', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['5'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['9'], xerr=[sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['6'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['6']], color='olive', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['1'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['9'], xerr=[sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['2'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['2']], color='firebrick', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['7'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['11'], xerr=[sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['0b'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['8']], color='mediumblue',fmt = 'o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['3'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['10'], xerr=[sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['0a'], sequence(df_arr_output5, df_rwy_calcs5, df_dep_output5, 0.92)['4']], color='deeppink', fmt='o', markersize=8, capsize=10, )
                            
                        
                            
                            
                            annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                            bbox=dict(boxstyle="round", fc="w"),
                                            arrowprops=dict(arrowstyle="->"))
                            annot.set_visible(False)
                            def update_annot2(ind):
                            
                                pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                annot.xy = pos
                                text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                annot.set_text(text)
                            def hover(event):
                                vis = annot.get_visible()
                                if event.inaxes == ax:
                                    cont, ind = tags_main_data.contains(event)
                                    if cont:
                                        update_annot2(ind)
                                        annot.set_visible(True)
                                        canvas.draw_idle()
                                    else:
                                        if vis:
                                            annot.set_visible(False)
                                            canvas.draw_idle()
                            A.set_xlabel('Seconds of the day')        
                            A.axes.get_yaxis().set_visible(False)                    
                            A.grid(color='b', linestyle='-', linewidth=0.1)    
                            A.legend(("Legend","Departures M1","Arrivals M1","Departures Spacing M1","Arrivals Spacing M1","Departures M2","Arrivals M2", "Departures Spacing M2","Arrivals Spacing M2", "Departures M3","Arrivals M3","Departures Spacing M3","Arrivals Spacing M3", "Departures M4","Arrivals M4","Departures Spacing M4","Arrivals Spacing M4"), loc = 'upper right')
                            #plt.title("Sequence analysis")
                            canvas = FigureCanvasTkAgg(f, self)
                            canvas.show()
                            canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                            canvas.mpl_connect("motion_notify_event", hover)
                            toolbar = NavigationToolbar2TkAgg(canvas, self)
                            toolbar.update()
                            canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                            #f.canvas.figure.savefig('sequence.png')
                        
                        else: # only 2 to compare
                                    
                            f = Figure(figsize=(5,5), dpi=100)  
                            A = f.add_subplot(111)
                            ax = f.add_subplot(111)
                            
                            
                            labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                            tag_text_use = np.array(list(labels))
                            labels_x = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] 
                            labels_y = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                            tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                            #plt.axhline(y = 10, color='w')                    
                            ##A.axhline(y=0.5, color='w')
                               
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                            
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='darkgreen', fmt='o', markersize=8, capsize=10)    
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='maroon', fmt='o', markersize=8, capsize=10, )
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], color='navy',fmt = 'o', markersize=8, capsize=10)
                            A.errorbar(sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequence(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='indigo', fmt='o', markersize=8, capsize=10, )
                            
                            annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                            bbox=dict(boxstyle="round", fc="w"),
                                            arrowprops=dict(arrowstyle="->"))
                            annot.set_visible(False)
                            def update_annot2(ind):
                            
                                pos = tags_main_data.get_offsets()[ind["ind"][0]]
                                annot.xy = pos
                                text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                                annot.set_text(text)
                            def hover(event):
                                vis = annot.get_visible()
                                if event.inaxes == ax:
                                    cont, ind = tags_main_data.contains(event)
                                    if cont:
                                        update_annot2(ind)
                                        annot.set_visible(True)
                                        canvas.draw_idle()
                                    else:
                                        if vis:
                                            annot.set_visible(False)
                                            canvas.draw_idle()
                            A.set_xlabel('Seconds of the day')        
                            A.axes.get_yaxis().set_visible(False)                     
                            A.grid(color='b', linestyle='-', linewidth=0.1)    
                            A.legend(("Legend","Departures M1","Arrivals M1","Departures Spacing M1","Arrivals Spacing M1","Departures M2","Arrivals M2", "Departures Spacing M2","Arrivals Spacing M2"), loc = 'upper right')
                            #plt.title("Sequence analysis")
                            canvas = FigureCanvasTkAgg(f, self)
                            canvas.show()
                            canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                            canvas.mpl_connect("motion_notify_event", hover)
                            toolbar = NavigationToolbar2TkAgg(canvas, self)
                            toolbar.update()
                            canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                        
                        
                        
                    elif df_dep_output2.empty ==True:#arr only
                        f = Figure(figsize=(5,5), dpi=100)  
                        A = f.add_subplot(111)
                        ax = f.add_subplot(111)
                        
                        labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                        tag_text_use = np.array(list(labels))
                        labels_x = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] 
                        labels_y = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                        tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                        #plt.axhline(y = 10, color='w')                    
                        #A.axhline(y=0.5, color='w')
                           
                        A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                        A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                        A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                        A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                        
                        A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['1'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                        A.errorbar(sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['3'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['10'], xerr=[sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0a'], sequenceArrOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
    
                    
                        annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->"))
                        annot.set_visible(False)
                        def update_annot2(ind):
                        
                            pos = tags_main_data.get_offsets()[ind["ind"][0]]
                            annot.xy = pos
                            text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                            annot.set_text(text)
                        def hover(event):
                            vis = annot.get_visible()
                            if event.inaxes == ax:
                                cont, ind = tags_main_data.contains(event)
                                if cont:
                                    update_annot2(ind)
                                    annot.set_visible(True)
                                    canvas.draw_idle()
                                else:
                                    if vis:
                                        annot.set_visible(False)
                                        canvas.draw_idle()
                        A.set_xlabel('Seconds of the day')        
                        A.axes.get_yaxis().set_visible(False)                     
                        A.grid(color='b', linestyle='-', linewidth=0.1)    
                        A.legend(("Legend","Departures M1","Arrivals M1","Departures Spacing M1","Arrivals Spacing M1","Arrivals M2", "Arrivals Spacing M2"), loc = 'upper right')
                        #plt.title("Sequence analysis")
                        canvas = FigureCanvasTkAgg(f, self)
                        canvas.show()
                        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                        canvas.mpl_connect("motion_notify_event", hover)
                        toolbar = NavigationToolbar2TkAgg(canvas, self)
                        toolbar.update()
                        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                    
                    elif df_arr_output2.empty ==True: # MIXED + ARRonly
                        
                        f = Figure(figsize=(5,5), dpi=100)  
                        A = f.add_subplot(111)
                        ax = f.add_subplot(111)
                        
                        #plt.axhline(y = 10, color='w')                    
                        
                        labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['12'] 
                        tag_text_use = np.array(list(labels))
                        labels_x = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['14'] 
                        labels_y = sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] + sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['13'] 
                        tags_main_data = A.scatter(labels_x,labels_y , c='w', s=100)      
                        #plt.axhline(y = 10, color='w')                    
                        #A.axhline(y=0.5, color='w')
                           
                        A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                        A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                        A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                        A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                        
                        
                        A.errorbar(sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['5'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['9'], xerr=[sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                        A.errorbar(sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['7'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['11'], xerr=[sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['0b'], sequenceDepOnly(df_arr_output2, df_rwy_calcs2, df_dep_output2, 0.98)['8']], fmt='o', markersize=8, capsize=10)
    
                        annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->"))
                        annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->"))
                        annot.set_visible(False)
                        def update_annot2(ind):
                        
                            pos = tags_main_data.get_offsets()[ind["ind"][0]]
                            annot.xy = pos
                            text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                            annot.set_text(text)
                        def hover(event):
                            vis = annot.get_visible()
                            if event.inaxes == ax:
                                cont, ind = tags_main_data.contains(event)
                                if cont:
                                    update_annot2(ind)
                                    annot.set_visible(True)
                                    canvas.draw_idle()
                                else:
                                    if vis:
                                        annot.set_visible(False)
                                        canvas.draw_idle()
                        A.set_xlabel('Seconds of the day')        
                        A.axes.get_yaxis().set_visible(False)                     
                        A.grid(color='b', linestyle='-', linewidth=0.1)    
                        A.legend(("Legend","Departures M1","Arrivals M1","Departures Spacing M1","Arrivals Spacing M1","Departures M2", "Departures Spacing M2"), loc = 'upper right')
                        #plt.title("Sequence analysis")
                        canvas = FigureCanvasTkAgg(f, self)
                        canvas.show()
                        canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                        canvas.mpl_connect("motion_notify_event", hover)
                        toolbar = NavigationToolbar2TkAgg(canvas, self)
                        toolbar.update()
                        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                        
                    
                        
                else: # only one MIXED
                    f = Figure(figsize=(5,5), dpi=100)  
                    A = f.add_subplot(111)
                    ax = f.add_subplot(111)                
                    
                    
                    
                    labels =  sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['12'] 
                    tag_text_use = np.array(list(labels))
                    
                    tags_main_data = A.scatter(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['14'],sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['13'] , c='w', s=100)      
                    #plt.axhline(y = 10, color='w')                    
                    #A.axhline(y=10, color='w')
                       
                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['5'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['6']], color='g', fmt='o', markersize=8, capsize=10)    
                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['1'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['9'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['2']], color='r', fmt='o', markersize=8, capsize=10, )
                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['7'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['11'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0b'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['8']], fmt='o', markersize=8, capsize=10)
                    A.errorbar(sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['3'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['10'], xerr=[sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['0a'], sequence(df_arr_output, df_rwy_calcs, df_dep_output, 1)['4']], color='purple', fmt='o', markersize=8, capsize=10, )
                    
                    
                    annot = ax.annotate("", xy=(0,0), xytext=(20,20),textcoords="offset points",
                                    bbox=dict(boxstyle="round", fc="w"),
                                    arrowprops=dict(arrowstyle="->"))
                    annot.set_visible(False)
                    def update_annot2(ind):
                    
                        pos = tags_main_data.get_offsets()[ind["ind"][0]]
                        annot.xy = pos
                        text = "{}".format(" ".join([tag_text_use[n] for n in ind["ind"]]))
                        annot.set_text(text)
                    def hover(event):
                        vis = annot.get_visible()
                        if event.inaxes == ax:
                            cont, ind = tags_main_data.contains(event)
                            if cont:
                                update_annot2(ind)
                                annot.set_visible(True)
                                canvas.draw_idle()
                            else:
                                if vis:
                                    annot.set_visible(False)
                                    canvas.draw_idle()
                    
                    A.set_xlabel('Seconds of the day')        
                    A.axes.get_yaxis().set_visible(False)    
                    A.grid(color='b', linestyle='-', linewidth=0.1)    
                    A.legend(("Legend","Departure","Arrivals ","Departures Spacing ","Arrivals Spacing "), loc = 'upper right')
                    #plt.title("Sequence analysis")
                    canvas = FigureCanvasTkAgg(f, self)
                    canvas.show()
                    canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
                    canvas.mpl_connect("motion_notify_event", hover)
                    toolbar = NavigationToolbar2TkAgg(canvas, self)
                    toolbar.update()
                    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
                    
                    
                    
    # ============================================================================#
    #                           ADA buffer                                          #
    # ============================================================================#
    #if Seq_FLAG == True:
    class ADAbuffer(tk.Frame):
    
        def __init__(self, parent, controller):
            tk.Frame.__init__(self, parent)
            label = tk.Label(self, text="SEQUENCE", font=LARGE_FONT)
            label.pack(pady=10,padx=10)
            
            button1 = ttk.Button(self, text="Back to Home",
                                command=lambda: controller.show_frame(StartPage))
            button1.pack() 
            if convergenceFLAG ==True:
                button1 = ttk.Button(self, text="Convergence Throughput",
                                    command=lambda: controller.show_frame(Conv))
                button1.pack()
            if Thr_FLAG == True:
                button = ttk.Button(self, text="Throughput",
                                    command=lambda: controller.show_frame(Thr))
                button.pack()
            
            if Delay_FLAG == True:
                button2 = ttk.Button(self, text="RWY Hold Delay",
                                    command=lambda: controller.show_frame(DepDelay))
                button2.pack()
            if Delay_FLAG == True:
                button5 = ttk.Button(self, text="Push/Start Delay",
                                    command=lambda: controller.show_frame(DepDelay2))
                button5.pack()
            
            if arr_delay_FLAG == True:
                button3 = ttk.Button(self, text="Arrivals Delay",
                                    command=lambda: controller.show_frame(ArrivalDelay))
                button3.pack()
                
            
            
            df_Buffer = pd.DataFrame()
            df_Buffer['ADA_Buffer'] = df_sequence_output['ADA Buffer']        
            df_Buffer = df_Buffer.dropna(subset = ['ADA_Buffer'])
            #df_Buffer = df_Buffer.drop([0])
            ADA_buffer = df_Buffer['ADA_Buffer'].tolist()
            a = int(min(ADA_buffer))
            b = int(max(ADA_buffer))
            number_bins = b-a
            h = [0]+sorted(ADA_buffer)
    #        ba = [15,15]
    #        bb = [0,0.013]
    #        da = [0,15]
    #        db = [0.013,0.013]
            
            coord = [[0,0], [15,0], [15,0.013], [0,0.013]]
            coord.append(coord[0]) #repeat the first point to create a 'closed loop'
            
            xs, ys = zip(*coord) #create lists of x and y values
    
    
            fit = stats.norm.pdf(h, np.mean(h), np.std(h))  #this is a fitting indeed
            
            
            f = Figure(figsize=(5,5), dpi=100)
            A = f.add_subplot(111)
            A.plot(h,fit,'-o')
            A.hist(h,normed=True,bins=number_bins)
            #A.axvline(x=15, color='red', linestyle='--')
            #A.plot(da,db, color='red',linestyle='--')
            
            A.plot(xs,ys,"r") 
            
            A.set_xlabel('SECONDS')
            A.set_ylabel('%')
            A.grid(color='b', linestyle='-', linewidth=0.1)     
            #A.title('Throughput')   
            canvas = FigureCanvasTkAgg(f, self)
            canvas.show()
            canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
    
            toolbar = NavigationToolbar2TkAgg(canvas, self)
            toolbar.update()
            canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    
    
    
    
    
    
    app = RAPIDvisual()
    
    app.columnconfigure(0, weight=1)
    app.rowconfigure(0, weight=1)
    app.rowconfigure(1, weight=1)
    
    app.mainloop()







