import openpyxl
import random
import time
import math
import pandas as pd
import re

# Definitions, creating sheets and columns, input pre-processing
class outputWorkbook():

    def __init__(self, input_filename, debugFlag):

        # --------------------------------- Constants -------------------------------- #

        self.STT = 600 # Standard Taxi Time - used for arrivals. Actually, it is from the landing point to the stand
        self.MIN_RADAR_SEP_DIST = 3 # NM
        self.C_DME = 4
        self.D_DME = 3

        self.RECAT_categories = {
            'A': ['A388','A124'],
            'B': ['A332','A333','A343','A345','A346','A359','B744','B748','B772','B773','B77L','B77W','B788','B789','IL96'],
            'C': ['A306','A30B ','A310','B703 ','B752','B753 ','B762','B763','B764','B783','C135','DC10','DC85','IL76','MD11','TU22','TU95'],
            'D': ['A318','A319','A320','A321','AN12','B736','B737','B738','B739','C130','IL18','MD81','MD82','MD83','MD87','MD88','MD90','T204','TU16'],
            'E': ['AT43','AT45','AT72','B712','B732','B733','B734','B735','CL60','CRJ1','CRJ2','CRJ7','CRJ9','DH8D','E135','E145','E170','E175','E190','E195','F70','F100','GLF4','RJ85','RJ1H'],
            'F': ['FA10','FA20','D328','E120','BE40','BE45','H25B','JS32','JS41','LJ35','LJ60','SF34','P180','C650','C525','C180','C152']
        }

        # Get data from Input file and utilities
        self.importUtility(input_filename = input_filename, utility_dir = 'utility/')

        # Creates new sheets and columns in Workbook
        self.writeRow('Arrivals', [
            'GS_0_1dme', 'GS_1_2dme', 'GS_2_3dme', 'GS_3_4dme', 'GS_4_5dme',
            'GS_5_6dme', 'GS_6_7dme', 'GS_7_8dme', 'GS_8_9dme', 'GS_9_10dme',
            'IAS_0_1dme', 'IAS_1_2dme', 'IAS_2_3dme', 'IAS_3_4dme', 'IAS_4_5dme',
            'IAS_5_6dme', 'IAS_6_7dme', 'IAS_7_8dme', 'IAS_8_9dme', 'IAS_9_10dme'
        ], start_col = 30)

        self.writeRow('Runway_calcs', [
            'Arrival ID', 'TAXI-IN', 'AROT', 'ADA', 'ADDA',
            'ATCO variability', 'WIND1', 'SPEED1', 'WIND2', 'SPEED2',
            'VTGT', 'SAE', 'PREDICTED Landing Time', 'MAX Constraint', 'MAX Constraint Label',
            'WAKE SEPARATION', 'MRS', 'Departure ID', 'TAXI-OUT', 'DROT',
            'ARRIVAL actual WAKE'
        ])

        self.writeRow('Arrival_Output', [
            'Arrival ID', 'Arrival HOUR', 'ACTUAL Landing Time',
            'Arrival RWY_EXIT', 'WAKE', 'In Blocks Time',
            'AROT', 'TAXI-IN Duration', 'MAX Constraint',
            'MAX Constraint Label', 'len(ArrHOLDqueue)', 'Arrival DELAY'
        ])

        self.writeRow('Departure_Output', [
            'Departure ID', 'Departure HOUR', 'Departure_RWY_ENTRY',
            'Departure_RWY_EXIT', 'WAKE', 'SID GROUP',
            'DROT', 'TAXI-OUT', 'Dep MIN Separation',
            'Dep MIN Separation Label', 'currentGap', 'len(DepSTANDqueue)',
            'len(TAXIhold)', 'len(RWYqueue1)', 'len(RWYqueue2)',
            'len(RWYqueue3)', 'len(RWYqueue4)', 'DELAY DepSTANDqueue',
            'DELAY TAXIhold', 'DELAY RWYqueue', 'RWY queue USED'
        ])

        self.writeRow('Throughput', [
            'Hour', 'Departure Throughput', 'Arrival Throughput',
            'Total Throughput', 'Cum. No. of Go-Arounds'
        ])

        self.writeRow('Delay', [
            'Departure ID', 'HOUR', 'RWY HOLD Delay',
            'Push/Start Delay', 'Arrival ID', 'HOUR', 'Arrival Delay'
        ])

        self.writeRow('Sequence', [
            'Type', 'ID', 'RWY ENTRY',
            'RWY EXIT', 'ROT', 'Arr ID start ADA pair', 'ADA Buffer'
        ])

        if debugFlag:
            self.writeRow('Debug', [
                'Time', 'Runway status', 'Current Gap - D',
                'Current Gap - A', 'Current Gap - E', 'Arrival Hold Delay'
            ])
            self.debugTab = self.wb['Debug']

        # Create variables for easy reference
        self.arrivalInput = self.wb['Arrivals']
        self.departureInput = self.wb['Departures']
        self.runwayCalculations = self.wb['Runway_calcs']
        self.arrivalOutput = self.wb['Arrival_Output']
        self.departureOutput = self.wb['Departure_Output']
        self.throughputTab = self.wb['Throughput']
        self.delayTab = self.wb['Delay']
        self.sequenceTab = self.wb['Sequence']


    # Loads all required input/utility data
    def importUtility(self, input_filename, utility_dir):

        # ----------------------------- Input_*.xlsx file ---------------------------- #

        self.wb = openpyxl.load_workbook(input_filename)

        # --------------------------- Actual Speed Profiles -------------------------- #

        self.speed_profile = pd.read_csv(utility_dir + 'actual_speed_profiles.csv')
        self.speed_profile = self.speed_profile.drop(columns=['Unnamed: 0'])
        self.speed_profile_unique_types = list(self.speed_profile['Aircraft_Type'].unique())

        # -------------------------- AROT/DROT Lookup Table -------------------------- #

        self.df_distributions = pd.read_csv(utility_dir + 'AROTDROT_distributions.csv')

        # ----------------------------------- Wake ----------------------------------- #

        self.df_wake = pd.read_csv(utility_dir + 'wake.csv')

        self.df_wake_WTC = pd.DataFrame()
        self.df_wake_WTC['ICAO'] = self.df_wake['ICAO']
        self.df_wake_WTC['WTC'] = self.df_wake['WTC']
        self.df_wake_WTC = self.df_wake_WTC.set_index('ICAO')

        # RECAT-EU separation

        self.df_wake_RECAT = pd.DataFrame()
        self.df_wake_RECAT['ICAO'] = self.df_wake['ICAO']
        self.df_wake_RECAT['RECAT-EU'] = self.df_wake['RECAT-EU']
        self.df_wake_RECAT = self.df_wake_RECAT.set_index('ICAO')

        self.df_RECAT_EU_separation = pd.read_csv(utility_dir + 'RECAT_EU_separation.csv')
        self.df_RECAT_EU_separation = self.df_RECAT_EU_separation.set_index("LEAD")

        # WTC separation

        self.df_WTC_separation = pd.read_csv(utility_dir + 'UK_wake_separation.csv')
        self.df_WTC_separation = self.df_WTC_separation.set_index("LEAD")

        # RECAT-PWS and RECAT-EU 20cat separation

        self.df_RECAT_PWS = pd.read_csv(utility_dir + 'RECAT_PWS.csv')
        self.df_RECAT_PWS = self.df_RECAT_PWS.fillna(0)
        self.df_RECAT_PWS = self.df_RECAT_PWS.set_index('FOLLOW')

        self.df_RECAT20 = pd.DataFrame()
        self.df_RECAT20['ICAO'] = self.df_wake['ICAO']
        self.df_RECAT20['RECAT20'] = self.df_wake['RECAT20']
        self.df_RECAT20 = self.df_RECAT20.set_index('ICAO')

        self.df_RECAT20_separation = pd.read_csv(utility_dir + 'RECAT20_separation.csv')
        self.df_RECAT20_separation = self.df_RECAT20_separation.fillna(0)
        self.df_RECAT20_separation = self.df_RECAT20_separation.set_index('LEAD')
        # self.df_RECAT20_separation = self.df_RECAT_PWS.set_index('LEAD')


    # Writes a list to a worksheet row (creates worksheet if not exist)
    def writeRow(self, ws, value_list, start_col=1, row=1):
        if ws not in self.wb.sheetnames:
            self.wb.create_sheet(ws)
        for i in range(len(value_list)):
            self.wb[ws][openpyxl.utils.get_column_letter(start_col + i) + str(row)].value = value_list[i]


# ---------------------------------------------------------------------------- #
#                                   Functions                                  #
# ---------------------------------------------------------------------------- #

# ---------------------------------------------------------------------------- #
#                                   Run Model                                  #
# ---------------------------------------------------------------------------- #

def runModel(parentFrame):

    v = {
        'filename': parentFrame.filename,
        'RECAT': bool(int(parentFrame.opt['var6'].get())), # Switch for modelling 'Radar Tower Separation' concept
        'RECAT_PWS': bool(int(parentFrame.opt['var17'].get())),
        'avgThr': bool(int(parentFrame.run['var7'].get())),
        'distanceBased': not bool(int(parentFrame.opt['separation_type'].get())),
        'timeBased': bool(int(parentFrame.opt['separation_type'].get())),
        'MRS_4DME': bool(int(parentFrame.opt['MRS_4dme'].get())),
        'WAKE_4DME': bool(int(parentFrame.opt['WAKE_4dme'].get())),
        'ADA_4DME': bool(int(parentFrame.opt['ADA_4dme'].get())),
        'ADDA_4DME': bool(int(parentFrame.opt['ADDA_4dme'].get())),
        'MRS_THR': bool(int(parentFrame.opt['MRS_thr'].get())),
        'WAKE_THR': bool(int(parentFrame.opt['WAKE_thr'].get())),
        'ADA_THR': bool(int(parentFrame.opt['ADA_thr'].get())),
        'ADDA_THR': bool(int(parentFrame.opt['ADDA_thr'].get())),
        'debugTab': bool(int(parentFrame.run['var14'].get())),
        'queueType': str(parentFrame.req['queue_type'].get()),
        'maxRuns': int(parentFrame.run['n_times_input'].get()),
        'n': int(parentFrame.req['n_input'].get()),
        'ADA_x': int(parentFrame.opt['ADA_x_input'].get()), # UNUSED
        'minDep_altSID': int(parentFrame.req['minDep_altSID_input'].get()),
        'minDep_sameSID': int(parentFrame.req['minDep_sameSID_input'].get()),
        'SIDmax': int(parentFrame.req['SIDmax_input'].get()), #UNUSED
        'SIDgroup_separation': str(parentFrame.req['SIDgroup_separation_input'].get()),
        'SID_queue_assign': str(parentFrame.req['SID_queue_assign_input'].get())
    }

    big_list = []
    averages = []
    difference = []
    iter2 = 0
    iter1 = 0

    if v['avgThr']:
        maxIter = 10
    else:
        maxIter = v['maxRuns']

    # SET WAKE RULES for departures
    H_H_d = 90
    H_M_d = 139 # Used for H_ UM/M/S/L
    J_H_d = 139 # 120
    J_M_d = 204 # 180 (2016 data)
    J_S_d = 204 # 180 (2016 data)
    J_L_d = 204 # 180 (2016 data)
    M_L_d = 139 # Used for UM_L & M_L
    S_L_d = 139 # 120

    # ---------------------------------------------------------------------------- #
    #                             Start of Model Run(s)                            #
    # ---------------------------------------------------------------------------- #

    while (iter1 < maxIter):

        program_runtime_start = time.time() # RUNTIME CALCULATION

        wb = outputWorkbook(v['filename'], v['debugTab'])

        RWY_status = "E"
        ARRIVALqueue = {} # Initialised 'dict' for storing Taxiing-in Arrivals
        ArrHOLDqueue = {} # Initialised 'dict' for the Arrivals hold queue
        APPqueue = {}
        DepSTANDqueue = {} # Initialised 'dict' for holding Departures on Stands (Push/Start Delay)
        TAXIqueue = {}
        TAXIhold = {}

        # Place Dep A/C with SID group 1 into RWYqueue1 etc - UNLESS there's no A/C of this type available
        RWYqueue1 = {}
        RWYqueue2 = {}
        # Other queues instigated for 4x2 (and future 8x1) arrangements
        RWYqueue3 = {}
        RWYqueue4 = {}

        GoAroundCount = {}

        currentGap = 86400

        # Arrival 'go-around' case
        number_of_goArounds_queued = 0

        # ARRIVALS
        ARRkey = 2
        ArrOutput = 2

        # DEPARTURES
        SOBTrow = 2
        DepOutput = 2
        seqRow = 2

        x_buffer = 15

        throughput = []

        # SID group separation
        SIDgroups = re.findall(r'\d+', v['SIDgroup_separation'])
        SIDgroups = [[int(SIDgroups[x]), int(SIDgroups[y])] for (x, y) in [(0, 1), (2, 3), (3, 2), (1, 0)]]

        # SID queue
        SIDqueues = re.findall(r'\d+', v['SID_queue_assign']) # Input example: 1 2 | 3 4
        SIDqueues = [[int(SIDqueues[x]), int(SIDqueues[y])] for (x, y) in [(0, 1), (2, 3)]]

        # Time limits
        if wb.arrivalInput['A' + str(2)].value==None: # No arrivals:
            Start_time = wb.departureInput['C' +str(2)].value - 3000
        elif wb.departureInput['A' +str(2)].value == None: # No departures
            Start_time = wb.arrivalInput['C' + str(2)].value - 3000
        else:
            Start_time = min(wb.arrivalInput['C' + str(2)].value, wb.departureInput['C' + str(2)].value) - 3000

        if wb.arrivalInput['A' + str(2)].value==None: # No arrivals:
            End_time = wb.departureInput['C' + str(wb.departureInput.max_row)].value + 10000
        elif wb.departureInput['A' + str(2)].value == None: # No departures
            End_time = wb.arrivalInput['C' + str(wb.arrivalInput.max_row)].value + 10000
        else:
            End_time = min(wb.arrivalInput['C' + str(wb.arrivalInput.max_row)].value, wb.departureInput['C' + str(wb.departureInput.max_row)].value) + 10000

        Current_time = Start_time

        # ----------------------- Arrivals Separation Functions ---------------------- #

        def distance_to_time_assumed_speed_profile_IAS(i, c_dme, d_dme, distance): #DELIVERED at THR
            #fixed d_dme at 3dme, variable c_dme because max deceleration speed is 20kts/NM

            ##### JI - THESE THREE LINES LOOK FISHY!
            deceleration_difference= wb.runwayCalculations['H' + str(i)].value - wb.runwayCalculations['K' + str(i)].value
            if deceleration_difference > 20 :
                c_dme = deceleration_difference / 20
            #####

            TBS_assumed_speed_profile_value = 0
            #time between d_dme - THR:
            t1 = (d_dme *3600)/(wb.runwayCalculations['K' + str(i)].value)
            #time between c_dme - d_dme:
            t2 = (2*3600*(c_dme-d_dme))/(wb.runwayCalculations['H' + str(i)].value+wb.runwayCalculations['K' + str(i)].value)

            if distance >= c_dme:
                TBS_assumed_speed_profile_value = int(t1+t2+((distance-c_dme)*3600/wb.runwayCalculations['H' + str(i)].value))
            elif (distance < c_dme) and (distance>d_dme):
                d1 = distance-d_dme
                speed_at_d1 = (d1*( wb.runwayCalculations['H' + str(i)].value - wb.runwayCalculations['K' + str(i)].value ) /(c_dme-d_dme)) + wb.runwayCalculations['K' + str(i)].value
                TBS_assumed_speed_profile_value = int(d1*3600*2/(speed_at_d1 + wb.runwayCalculations['K' + str(i)].value) + t1)
            elif distance <= d_dme:
                TBS_assumed_speed_profile_value = int(distance*3600/wb.runwayCalculations['K' + str(i)].value)
            #print('TBS - on' )

            return TBS_assumed_speed_profile_value


        def DBS_assumed_speed_profile(i, c_dme, d_dme, distance): #DELIVERED at THR
            deceleration_difference= (wb.runwayCalculations['H' + str(i)].value - wb.runwayCalculations['G' + str(i)].value) - (wb.runwayCalculations['K' + str(i)].value - wb.runwayCalculations['I' + str(i)].value)
            if deceleration_difference > 20 :
                c_dme = deceleration_difference / 20

            DBS_assumed_speed_profile_value = 0
            #time between d_dme - THR:
            t1 = (d_dme *3600)/(wb.runwayCalculations['K' + str(i)].value-wb.runwayCalculations['I' + str(i)].value)
            #time between c_dme - d_dme:
            t2 = (2*3600*(c_dme-d_dme))/((wb.runwayCalculations['H' + str(i)].value-wb.runwayCalculations['G' + str(i)].value)+(wb.runwayCalculations['K' + str(i)].value-wb.runwayCalculations['I' + str(i)].value))

            if distance >= c_dme:
                DBS_assumed_speed_profile_value = int(t1+t2+((distance-c_dme)*3600/(wb.runwayCalculations['H' + str(i)].value-wb.runwayCalculations['G' + str(i)].value)))
            elif (distance < c_dme) and (distance>d_dme):
                d1 = distance-d_dme
                speed_at_d1 = (d1*( (wb.runwayCalculations['H' + str(i)].value-wb.runwayCalculations['G' + str(i)].value) - (wb.runwayCalculations['K' + str(i)].value-wb.runwayCalculations['I' + str(i)].value) ) /(c_dme-d_dme)) + (wb.runwayCalculations['K' + str(i)].value-wb.runwayCalculations['I' + str(i)].value)
                DBS_assumed_speed_profile_value = int(d1*3600*2/(speed_at_d1 + wb.runwayCalculations['K' + str(i)].value - wb.runwayCalculations['I' + str(i)].value) + t1)
            elif distance <= d_dme:
                DBS_assumed_speed_profile_value = int(distance*3600/(wb.runwayCalculations['K' + str(i)].value-wb.runwayCalculations['I' + str(i)].value))

            return DBS_assumed_speed_profile_value


        def DBS_actual_speed_profile(distance, row): #DELIVERED at THR # use GS

            T=0

            def full_segments(n,row):
                T= 0
                if n >= 1:
                    T = 2 * 3600 / (wb.arrivalInput['AD'+str(row)].value+wb.arrivalInput['AE'+str(row)].value)
                    if n >=2:
                        T += 2 * 3600 / (wb.arrivalInput['AE'+str(row)].value +wb.arrivalInput['AF'+str(row)].value)
                        if n>=3:
                            T += 2 * 3600 / (wb.arrivalInput['AF'+str(row)].value + wb.arrivalInput['AG'+str(row)].value)
                            if n>=4:
                                T += 2 * 3600 / (wb.arrivalInput['AG'+str(row)].value + wb.arrivalInput['AH'+str(row)].value)
                                if n>=5:
                                    T += 2 * 3600 / (wb.arrivalInput['AH'+str(row)].value+wb.arrivalInput['AI'+str(row)].value)
                                    if n>=6:
                                        T += 2 * 3600 / (wb.arrivalInput['AI'+str(row)].value+wb.arrivalInput['AJ'+str(row)].value)
                                        if n>=7:
                                            T += 2 * 3600 / (wb.arrivalInput['AJ'+str(row)].value+wb.arrivalInput['AK'+str(row)].value)
                                            if n>=8:
                                                T += 2 * 3600 / (wb.arrivalInput['AK'+str(row)].value+wb.arrivalInput['AL'+str(row)].value)
                                                if n==9:
                                                    T += 2 * 3600 / (wb.arrivalInput['AL'+str(row)].value + wb.arrivalInput['AM'+str(row)].value)
                                                elif n>9:
                                                    T += (n-9)*3600/wb.arrivalInput['AM'+str(row)].value
                return T


            def fraction_of_segments(n,f,row):
                T = 0
                if n== 1:
                    S = f*(wb.arrivalInput['AF'+str(row)].value - wb.arrivalInput['AE'+str(row)].value) + wb.arrivalInput['AE'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AE'+str(row)].value)
                elif n==2:
                    S = f*(wb.arrivalInput['AG'+str(row)].value - wb.arrivalInput['AF'+str(row)].value) + wb.arrivalInput['AF'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AF'+str(row)].value)
                elif n==3:
                    S = f*(wb.arrivalInput['AH'+str(row)].value - wb.arrivalInput['AG'+str(row)].value) + wb.arrivalInput['AG'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AG'+str(row)].value)
                elif n==4:
                    S = f*(wb.arrivalInput['AI'+str(row)].value - wb.arrivalInput['AH'+str(row)].value) + wb.arrivalInput['AH'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AH'+str(row)].value)
                elif n==5:
                    S = f*(wb.arrivalInput['AJ'+str(row)].value - wb.arrivalInput['AI'+str(row)].value) + wb.arrivalInput['AI'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AI'+str(row)].value)
                elif n==6:
                    S = f*(wb.arrivalInput['AK'+str(row)].value - wb.arrivalInput['AJ'+str(row)].value) + wb.arrivalInput['AJ'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AJ'+str(row)].value)
                elif n==7:
                    S = f*(wb.arrivalInput['AL'+str(row)].value - wb.arrivalInput['AK'+str(row)].value) + wb.arrivalInput['AK'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AK'+str(row)].value)
                elif n==8:
                    S = f*(wb.arrivalInput['AM'+str(row)].value - wb.arrivalInput['AL'+str(row)].value) + wb.arrivalInput['AL'+str(row)].value
                    T = (f*3600)/(wb.arrivalInput['AM'+str(row)].value)
                return T

            # if distance > 0:
            X = distance + wb.runwayCalculations['F' + str(row)].value # Actual distance + ATCO var
            D = X - 0.5
            if D <0:
                T = (X*3600)/wb.arrivalInput['AD'+str(row)].value
            elif D > 0:
                v['n'] = math.floor(D)
                f = D - v['n']
                T1 = full_segments(v['n'],row)
                if (f != 0) and (v['n']<=8):
                    T2 = fraction_of_segments(v['n'],f,row)
                    T = T1 + T2 + (0.5*3600)/wb.arrivalInput['AD'+str(row)].value
                else:
                    T = T1 + (0.5*3600)/wb.arrivalInput['AD'+str(row)].value
            return T


        def TBS_actual_speed_profile(distance, row): #DELIVERED at THR # use IAS
            def full_segments(n,row):
                if n >= 1:
                    T = 2 * 3600 / (wb.arrivalInput['AN'+str(row)].value+wb.arrivalInput['AO'+str(row)].value)
                    if n >=2:
                        T += 2 * 3600 / (wb.arrivalInput['AO'+str(row)].value+wb.arrivalInput['AP'+str(row)].value)
                        if n>=3:
                            T += 2 * 3600 / (wb.arrivalInput['AP'+str(row)].value + wb.arrivalInput['AQ'+str(row)].value)
                            if n>=4:
                                T += 2 * 3600 / (wb.arrivalInput['AQ'+str(row)].value+wb.arrivalInput['AR'+str(row)].value)
                                if n>=5:
                                    T += 2 * 3600 / (wb.arrivalInput['AR'+str(row)].value + wb.arrivalInput['AS'+str(row)].value)
                                    if n>=6:
                                        T += 2 * 3600 / (wb.arrivalInput['AS'+str(row)].value+wb.arrivalInput['AT'+str(row)].value)
                                        if n>=7:
                                            T += 2 * 3600 / (wb.arrivalInput['AT'+str(row)].value+wb.arrivalInput['AU'+str(row)].value)
                                            if n>=8:
                                                T += 2 * 3600 / (wb.arrivalInput['AU'+str(row)].value+wb.arrivalInput['AV'+str(row)].value)
                                                if n==9:
                                                    T += 2 * 3600 / (wb.arrivalInput['AV'+str(row)].value + wb.arrivalInput['AW'+str(row)].value)
                                                elif n>9:
                                                    T += (n-9)*3600/wb.arrivalInput['AW'+str(row)].value
                return T


            def fraction_of_segments(n,f,row):
                if n== 1:
                    S = f*(wb.arrivalInput['AP'+str(row)].value - wb.arrivalInput['AO'+str(row)].value) + wb.arrivalInput['AO'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AO'+str(row)].value)
                elif n==2:
                    S = f*(wb.arrivalInput['AQ'+str(row)].value - wb.arrivalInput['AP'+str(row)].value) + wb.arrivalInput['AP'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AP'+str(row)].value)
                elif n==3:
                    S = f*(wb.arrivalInput['AR'+str(row)].value - wb.arrivalInput['AQ'+str(row)].value) + wb.arrivalInput['AQ'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AQ'+str(row)].value)
                elif n==4:
                    S = f*(wb.arrivalInput['AS'+str(row)].value - wb.arrivalInput['AR'+str(row)].value) + wb.arrivalInput['AR'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AR'+str(row)].value)
                elif n==5:
                    S = f*(wb.arrivalInput['AT'+str(row)].value - wb.arrivalInput['AS'+str(row)].value) + wb.arrivalInput['AS'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AS'+str(row)].value)
                elif n==6:
                    S = f*(wb.arrivalInput['AU'+str(row)].value - wb.arrivalInput['AT'+str(row)].value) + wb.arrivalInput['AT'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AT'+str(row)].value)
                elif n==7:
                    S = f*(wb.arrivalInput['AV'+str(row)].value - wb.arrivalInput['AU'+str(row)].value) + wb.arrivalInput['AU'+str(row)].value
                    T = (f*2*3600)/(S+wb.arrivalInput['AU'+str(row)].value)
                elif n==8:
                    S = f*(wb.arrivalInput['AW'+str(row)].value - wb.arrivalInput['AV'+str(row)].value) + wb.arrivalInput['AV'+str(row)].value
                    T = (f*3600)/wb.arrivalInput['AW'+str(row)].value
                return T

            # if distance > 0:
            X = distance + wb.runwayCalculations['F' + str(row)].value # Actual distance + ATCO var
            D = X - 0.5
            if D <0:
                T = (X*3600)/wb.arrivalInput['AN'+str(row)].value
            elif D > 0:
                v['n'] = math.floor(D)
                f = D - v['n']
                T1 = full_segments(v['n'],row)
                if (f != 0) and (v['n']<=8):
                    T2 = fraction_of_segments(v['n'],f,row)
                    T = T1 + T2 + (0.5*3600)/wb.arrivalInput['AN'+str(row)].value
                else:
                    T = T1
            return T


        def time_to_distance_assumed_speed_profile_IAS(row, c_dme, d_dme, T):
            deceleration_difference= (wb.runwayCalculations['H' + str(row)].value - wb.runwayCalculations['K' + str(row)].value)
            if deceleration_difference > 20 :
                c_dme = deceleration_difference / 20
            t1 = d_dme*3600/wb.runwayCalculations['K' + str(row)].value
            t2 = (c_dme - d_dme)*3600*2/(wb.runwayCalculations['K' + str(row)].value + wb.runwayCalculations['H' + str(row)].value) + t1

            if T <= t1 :
                D = (wb.runwayCalculations['K' + str(row)].value*T)/3600
            elif (T > t1) and (T < t2):
                t = T- t1
                S = (t*(wb.runwayCalculations['H' + str(row)].value-wb.runwayCalculations['K' + str(row)].value))/t2 + wb.runwayCalculations['K' + str(row)].value
                D = (t*(wb.runwayCalculations['K' + str(row)].value+S))/(2*3600) + d_dme
            elif T >= t2:
                D = c_dme + (T-t2)*wb.runwayCalculations['H' + str(row)].value/3600
            return D


        def time_to_distance_assumed_speed_profile_GS(row, c_dme, d_dme, T):
            deceleration_difference= (wb.runwayCalculations['H' + str(row)].value - wb.runwayCalculations['G' + str(row)].value) - (wb.runwayCalculations['K' + str(row)].value - wb.runwayCalculations['I' + str(row)].value)
            if deceleration_difference > 20 :
                c_dme = deceleration_difference / 20
            t1 = d_dme*3600/(wb.runwayCalculations['K' + str(row)].value- wb.runwayCalculations['I' + str(row)].value)
            t2 = (c_dme - d_dme)*3600*2/((wb.runwayCalculations['K' + str(row)].value- wb.runwayCalculations['I' + str(row)].value) + (wb.runwayCalculations['H' + str(row)].value- wb.runwayCalculations['G' + str(row)].value)) + t1

            if T <= t1 :
                D = ((wb.runwayCalculations['K' + str(row)].value- wb.runwayCalculations['I' + str(row)].value)*T)/3600
            elif (T > t1) and (T < t2):
                t = T- t1
                S = (t*((wb.runwayCalculations['H' + str(row)].value- wb.runwayCalculations['G' + str(row)].value)-(wb.runwayCalculations['K' + str(row)].value- wb.runwayCalculations['I' + str(row)].value)))/t2 + (wb.runwayCalculations['K' + str(row)].value- wb.runwayCalculations['I' + str(row)].value)
                D = (t*((wb.runwayCalculations['K' + str(row)].value- wb.runwayCalculations['I' + str(row)].value)+S))/(2*3600) + d_dme

                D = (T*((wb.runwayCalculations['K' + str(row)].value- wb.runwayCalculations['I' + str(row)].value)+(wb.runwayCalculations['H' + str(row)].value- wb.runwayCalculations['G' + str(row)].value))/(2*3600)) + d_dme
            elif T >= t2:
                D = c_dme + (T-t2)*(wb.runwayCalculations['H' + str(row)].value- wb.runwayCalculations['G' + str(row)].value)/3600
            return D


        # --------------------------- Pre-Process Functions -------------------------- #

        # Used in preprocessArrivals (Delivered at THR ACTUAL SPEED PROFILE)
        def min_wake_separation_arrs(key_of_nextArrival):
            minWakeSepArr = 0 # Initialise local variable (reset on each iteration)

            if v['RECAT_PWS']: # analyse by ac type
                previousArrival = wb.arrivalInput['D' + str(key_of_nextArrival - 1)].value
                currentArrival = wb.arrivalInput['D' + str(key_of_nextArrival)].value
                previousArrivalWake = wb.runwayCalculations['U' + str(key_of_nextArrival - 1)].value # 20cat classification
                currentArrivalWake = wb.runwayCalculations['U' + str(key_of_nextArrival)].value # 20cat classification

                if key_of_nextArrival == 2: #FirstArrival
                    minWakeSepArr = 0
                else:
                    if (currentArrival in wb.df_RECAT_PWS) and (previousArrival in wb.df_RECAT_PWS):
                        wakeDistance = wb.df_RECAT_PWS.at[currentArrival, previousArrival]
                        if wakeDistance == 0:
                            wakeDistance = wb.df_RECAT20_separation.at[previousArrivalWake, currentArrivalWake]
                    else: # if the pair is not in the 96x96 table, search in the 20cat
                        wakeDistance = wb.df_RECAT20_separation.at[previousArrivalWake, currentArrivalWake]

                    if wakeDistance == 0:
                        minWakeSepArr = 0
                    else:
                        if v['distanceBased']:
                            if v['WAKE_4DME']:
                                Total_time_follow = int(DBS_actual_speed_profile((wakeDistance+4),key_of_nextArrival))
                                Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(key_of_nextArrival-1)))
                                minWakeSepArr = Total_time_follow - Time_lead_4dme_to_thr
                            elif v['WAKE_THR']:
                                minWakeSepArr = int(DBS_actual_speed_profile(wakeDistance,key_of_nextArrival))  #time
                            else: # the same as the previous one but it's the default condition
                                minWakeSepArr = int(DBS_actual_speed_profile(wakeDistance,key_of_nextArrival))  #time

                        elif v['timeBased']:
                            time1 = distance_to_time_assumed_speed_profile_IAS(key_of_nextArrival, wb.C_DME, wb.D_DME, wakeDistance) #time
                            distance = time_to_distance_assumed_speed_profile_GS(key_of_nextArrival, wb.C_DME, wb.D_DME, int(time1))#distance
                            if v['WAKE_4DME']:
                                Total_time_follow = int(DBS_actual_speed_profile((distance+4),key_of_nextArrival))
                                Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(key_of_nextArrival-1)))
                                minWakeSepArr = Total_time_follow - Time_lead_4dme_to_thr

                            elif v['WAKE_THR']:
                                minWakeSepArr = int(DBS_actual_speed_profile(distance,key_of_nextArrival)) #time
                            else:
                                minWakeSepArr = int(DBS_actual_speed_profile(distance,key_of_nextArrival)) #time

            else: #analyze by wake
                previousArrivalWake = wb.runwayCalculations['U' +str(key_of_nextArrival-1)].value
                currentArrivalWake = wb.runwayCalculations['U' +str(key_of_nextArrival)].value
                if key_of_nextArrival == 2: #FirstArrival
                    minWakeSepArr = 0
                else: #next arrivals

                    if v['RECAT']: # delievered to THR
                        wakeDistance = wb.df_RECAT_EU_separation.at[previousArrivalWake,currentArrivalWake]
                    else: #UK cat *********** should be delievered to 4dme
                        wakeDistance = wb.df_WTC_separation.at[previousArrivalWake,currentArrivalWake] #distance

                    if wakeDistance == 0:
                        minWakeSepArr =0
                    else:
                        if v['distanceBased']:
                            if v['WAKE_4DME']:
                                Total_time_follow = int(DBS_actual_speed_profile((wakeDistance+4),key_of_nextArrival))
                                Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(key_of_nextArrival-1)))
                                minWakeSepArr = Total_time_follow - Time_lead_4dme_to_thr
                            elif v['WAKE_THR']:
                                minWakeSepArr = int(DBS_actual_speed_profile(wakeDistance,key_of_nextArrival))  #time
                            else: # the same as the previous one but it's the default condition
                                minWakeSepArr = int(DBS_actual_speed_profile(wakeDistance,key_of_nextArrival))  #time

                        elif v['timeBased']:
                            time1 = distance_to_time_assumed_speed_profile_IAS(key_of_nextArrival, wb.D_DME, wakeDistance) #time
                            distance = time_to_distance_assumed_speed_profile_GS(key_of_nextArrival, wb.D_DME, int(time1))#distance
                            if v['WAKE_4DME']:
                                Total_time_follow = int(DBS_actual_speed_profile((distance+4),key_of_nextArrival))
                                Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(key_of_nextArrival-1)))
                                minWakeSepArr = Total_time_follow - Time_lead_4dme_to_thr

                            elif v['WAKE_THR']:
                                minWakeSepArr = int(DBS_actual_speed_profile(distance,key_of_nextArrival)) #time
                            else:
                                minWakeSepArr = int(DBS_actual_speed_profile(distance,key_of_nextArrival)) #time

            return minWakeSepArr


        # Used in preprocessArrivals
        def max_constraint_generator(row):
            wake_constraint = wb.runwayCalculations['P' + str(row)].value
            MRS_constraint = wb.runwayCalculations['Q' + str(row)].value
            spFLAG = "None"
            max_constraint = 0

            if row == 2 :
                max_constraint = max(wake_constraint,MRS_constraint)
                spFLAG = "First Arrival"
            else: #not he first arrival
                AROT_constraint = wb.runwayCalculations['C' + str(row-1)].value + 5

                if (wb.departureInput.max_row == 0): #no departures
                    max_constraint = int(max(wake_constraint, MRS_constraint ,AROT_constraint))
                    if max_constraint == wake_constraint:
                        spFLAG = "WAKE"
                    elif max_constraint == MRS_constraint:
                        spFLAG = "MRS"
                    else:
                        spFLAG = "AROT"
                elif (wb.departureInput.max_row > 0) and (wb.arrivalInput.max_row > 0): #there are both arrivals and departures scheduled
                    if v['timeBased']:

                        max_constraint = int(max(wake_constraint, MRS_constraint, AROT_constraint))
                        if max_constraint == wake_constraint:
                            spFLAG = "WAKE"
                        elif max_constraint == MRS_constraint:
                            spFLAG = "MRS"
                        else:
                            spFLAG = "AROT"
                    elif v['distanceBased']:
                        if (wb.arrivalInput['U' + str(row)].value) == "ADDA" :
                            ADDA_distance = wb.runwayCalculations['E' + str(row)].value
                            if (v['ADDA_4DME']) and (row>2):
                                Total_time_follow = int(DBS_actual_speed_profile((ADDA_distance+4),row))
                                Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(row-1)))
                                ADDA_separation = Total_time_follow - Time_lead_4dme_to_thr
                            elif v['ADDA_THR']:
                                ADDA_separation = int(DBS_actual_speed_profile(ADDA_distance,row))  #time
                            else: # the same as the previous one but it's the default condition
                                ADDA_separation = int(DBS_actual_speed_profile(ADDA_distance,row))  #time
                            # ADDA_separation = int(DBS_actual_speed_profile(ADDA_distance,row))

                            max_constraint = int(max(wake_constraint, ADDA_separation, MRS_constraint,AROT_constraint))
                            if max_constraint ==wake_constraint:
                                spFLAG = "WAKE"
                            elif max_constraint == ADDA_separation:
                                spFLAG = "ADDA"
                            elif max_constraint == MRS_constraint:
                                spFLAG = "MRS"
                            else:
                                spFLAG = "AROT"
                        elif (wb.arrivalInput['U' + str(row)].value) == "ADA" :

                            ADA_distance = wb.runwayCalculations['D' + str(row)].value

                            if (v['ADA_4DME']) and (row>2):
                                Total_time_follow = int(DBS_actual_speed_profile((ADA_distance+4),row))
                                Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,(row-1)))
                                ADA_separation = Total_time_follow - Time_lead_4dme_to_thr
                            elif v['ADA_THR']:
                                ADA_separation = int(DBS_actual_speed_profile(ADA_distance,row))  #time
                            else: # the same as the previous one but it's the default condition
                                ADA_separation = int(DBS_actual_speed_profile(ADA_distance,row))  #time

                            max_constraint = int(max(wake_constraint, ADA_separation, MRS_constraint,AROT_constraint))
                            if max_constraint == wake_constraint:
                                spFLAG = "WAKE"
                            elif max_constraint == ADA_separation:
                                spFLAG = "ADA"
                            elif max_constraint == MRS_constraint:
                                spFLAG = "MRS"
                            else:
                                spFLAG = "AROT"
                        else:
                            max_constraint = int(max(wake_constraint, MRS_constraint,AROT_constraint))
                            if max_constraint == wake_constraint:
                                spFLAG = "WAKE"
                            elif max_constraint == MRS_constraint:
                                spFLAG = "MRS"
                            else:
                                spFLAG = "AROT"

            return {'a' : max_constraint, 'b' : spFLAG}


        # Writes data to Arrivals and Runway_calcs tabs
        def preprocessArrivals():

            # Reading row-wise from Arrivals tab
            for row in range(2, wb.arrivalInput.max_row + 1):

                AC_type = wb.arrivalInput['D' + str(row)].value

                # Catch missing value in first column
                if wb.arrivalInput['A' + str(row)].value == None:
                    raise ValueError(f'Missing value on row {row} in first column of {v.filename}')

                # Convert SIBT (column B) for SIBT(sec) (column C)
                if wb.arrivalInput['C' + str(row)].value == None:
                    SIBT = wb.arrivalInput['B' + str(row)].value
                    SIBT_sec = (SIBT.hour * 3600) + (SIBT.minute * 60) + SIBT.second
                    wb.arrivalInput['C' + str(row)].value = SIBT_sec # Used from initial schedule

                # Write WTC in arrival Input | it will be used for AROT
                wb.arrivalInput['E' + str(row)].value = wb.df_wake_WTC.at[AC_type,'WTC']

                # Write wake categories in runway calcs | used for wake separation:
                if v['RECAT']:
                    wb.runwayCalculations['U' + str(row)].value = wb.df_wake_RECAT.at[AC_type,'RECAT-EU']
                elif v['RECAT_PWS']:
                    wb.runwayCalculations['U' + str(row)].value = wb.df_RECAT20.at[AC_type,'RECAT20']
                else:
                    wb.runwayCalculations['U' + str(row)].value = wb.arrivalInput['E' + str(row)].value

                # --------------------------- Actual Speed Profile --------------------------- #

                # If AC_type is not found in actual_speed_profiles.csv,
                # find a random available type of same wake (using RECAT_categories),
                # otherwise pick a random available type (no speed profiles at all for that wake)
                if AC_type not in wb.speed_profile_unique_types:
                    for key, value in wb.RECAT_categories.items():
                        if AC_type in value:
                            Available_AC_types = [x for x in wb.RECAT_categories[key] if x in wb.speed_profile_unique_types]
                            if len(Available_AC_types) > 0:
                                AC_type = random.choice(Available_AC_types)
                            else:
                                AC_type = random.choice(wb.speed_profile_unique_types)
                            break

                # Get subset of aircraft type
                sp_type = wb.speed_profile[wb.speed_profile['Aircraft_Type'] == AC_type]

                # Appends the GSPD and IAS of a random aircraft of given type to an Arrivals tab row
                # Assumes columns after GSPD_0_1DME are in order by DME, then followed by same for IAS
                # e.g.: GSPD_0_1DME, GSPD_1_2DME,... GSPD_9_10DME, IAS_0_1DME, IAS_1_2DME,... IAS_9_10DME
                rand_row = random.randint(0, len(sp_type.index) - 1)
                sp_start_col = sp_type.columns.get_loc('GSPD_0_1DME')
                sp_end_col = sp_start_col + 21
                wb.writeRow(
                    'Arrivals',
                    sp_type.iloc[rand_row, sp_start_col:sp_end_col].tolist(),
                    start_col=30,
                    row=row
                )

                # ------------------------- Intermediate Calculations ------------------------ #

                # Arrival ID
                wb.runwayCalculations['A' + str(row)].value = wb.arrivalInput['A' + str(row)].value

                # ----------------------- Taxi-In - Normal Distribution ---------------------- #

                TaxiInLookup = wb.arrivalInput['M' + str(row)].value
                if TaxiInLookup is None:
                    Arrival_Taxiin_mean = wb.arrivalInput['I' + str(row)].value
                    Arrival_Taxiin_SD = wb.arrivalInput['J' + str(row)].value
                    TaxiInLookup = random.normalvariate(Arrival_Taxiin_mean, Arrival_Taxiin_SD)
                wb.runwayCalculations['B' + str(row)].value = round(TaxiInLookup, 0)

                # ---------------------------- AROT - from lookup ---------------------------- #

                if wb.arrivalInput['E' + str(row)].value == 'J':
                    wb.runwayCalculations['C' + str(row)].value = random.choice(wb.df_distributions['AROT_J'].dropna())
                elif wb.arrivalInput['E' + str(row)].value == 'H':
                    wb.runwayCalculations['C' + str(row)].value = random.choice(wb.df_distributions['AROT_H'].dropna())
                elif wb.arrivalInput['E' + str(row)].value == 'UM':
                    wb.runwayCalculations['C' + str(row)].value = random.choice(wb.df_distributions['AROT_UM'].dropna())
                elif wb.arrivalInput['E' + str(row)].value == 'M':
                    wb.runwayCalculations['C' + str(row)].value = random.choice(wb.df_distributions['AROT_M'].dropna())
                elif wb.arrivalInput['E' + str(row)].value == 'S':
                    wb.runwayCalculations['C' + str(row)].value = random.choice(wb.df_distributions['AROT_S'].dropna())
                elif wb.arrivalInput['E' + str(row)].value == 'L':
                    wb.runwayCalculations['C' + str(row)].value = random.choice(wb.df_distributions['AROT_L'].dropna())

                # ------------------------- ADA - Normal Distribution ------------------------ #

                ADA_mean = wb.arrivalInput['O' + str(row)].value
                ADA_sd = wb.arrivalInput['P' + str(row)].value
                actualADA = random.normalvariate(ADA_mean, ADA_sd)
                wb.runwayCalculations['D' + str(row)].value = int(actualADA)

                # ------------------------ ADDA - Normal Distribution ------------------------ #

                ADDA_mean = wb.arrivalInput['Q' + str(row)].value
                ADDA_sd = wb.arrivalInput['R' + str(row)].value
                actualADDA = random.normalvariate(ADDA_mean, ADDA_sd)
                wb.runwayCalculations['E' + str(row)].value = int(actualADDA)

                # ------------------ ATCO Variability - Normal Distribution ------------------ #

                ATCO_mean = wb.arrivalInput['S' + str(row)].value
                ATCO_sd = wb.arrivalInput['T' + str(row)].value
                actualATCO = random.normalvariate(ATCO_mean, ATCO_sd)
                wb.runwayCalculations['F' + str(row)].value = int(actualATCO)

                # ---------------------- Assumed Speed Profile - Wind 1 ---------------------- #

                WIND1_mean = wb.arrivalInput['V' + str(row)].value
                WIND1_sd = wb.arrivalInput['W' + str(row)].value
                actualWIND1 = random.normalvariate(WIND1_mean, WIND1_sd)
                wb.runwayCalculations['G' + str(row)].value = actualWIND1

                # ---------------------- Assumed Speed Profile - Speed 1 --------------------- #

                SPEED1_mean = wb.arrivalInput['X' + str(row)].value
                SPEED1_sd = wb.arrivalInput['Y' + str(row)].value
                actualSPEED1 = random.normalvariate(SPEED1_mean, SPEED1_sd)
                wb.runwayCalculations['H' + str(row)].value = actualSPEED1

                # ---------------------- Assumed Speed Profile - Wind 2 ---------------------- #

                WIND2_mean = wb.arrivalInput['Z' + str(row)].value
                WIND2_sd = wb.arrivalInput['AA' + str(row)].value
                actualWIND2 = random.normalvariate(WIND2_mean, WIND2_sd)
                wb.runwayCalculations['I' + str(row)].value = actualWIND2

                # ---------------------- Assumed Speed Profile - Speed 2 --------------------- #

                SPEED2_mean = wb.arrivalInput['AB' + str(row)].value
                SPEED2_sd = wb.arrivalInput['AC' + str(row)].value
                actualSPEED2 = random.normalvariate(SPEED2_mean, SPEED2_sd)
                wb.runwayCalculations['J' + str(row)].value = actualSPEED2

                # ----------------------- Assumed Speed Profile - Vtgt ----------------------- #

                if (actualWIND2 < 5) or (actualWIND2 > 20):
                    wind_adjustment = 5
                else:
                    wind_adjustment = actualWIND2 * 0.5
                Vtgt = actualSPEED2 + wind_adjustment
                wb.runwayCalculations['K' + str(row)].value = Vtgt

                # ------------------------ Assumed Speed Profile - SAE ----------------------- #

                # SAE = SIBT - Standard Taxi Time - App length*
                wb.runwayCalculations['L' + str(row)].value = wb.arrivalInput['C' + str(row)].value - wb.STT - 200

                # -------------- Assumed Speed Profile - Predicted Landing Time -------------- #

                # PLT = SAE + MRS*
                wb.runwayCalculations['M' + str(row)].value = wb.runwayCalculations['L' + str(row)].value + 60

                # ------------------------ Max Constraint Calculations ----------------------- #

                wb.runwayCalculations['P' + str(row)].value = int(min_wake_separation_arrs(row)) # Always Distance-based

                MRSArr = 0
                if (v['MRS_4DME']) and (row > 2):
                    Total_time_follow = int(DBS_actual_speed_profile((wb.MIN_RADAR_SEP_DIST + 4), row))
                    Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4, row - 1))
                    MRSArr = Total_time_follow - Time_lead_4dme_to_thr
                elif v['MRS_THR']:
                    MRSArr = int(DBS_actual_speed_profile(wb.MIN_RADAR_SEP_DIST, row))  #time
                else: # the same as the previous one but it's the default condition
                    MRSArr = int(DBS_actual_speed_profile(wb.MIN_RADAR_SEP_DIST, row))  #time

                wb.runwayCalculations['Q' + str(row)].value = MRSArr
                wb.runwayCalculations['N' + str(row)].value = max_constraint_generator(row)['a']
                wb.runwayCalculations['O' + str(row)].value = max_constraint_generator(row)['b']


        # Writes data to Departures and Runway_calcs tabs
        def preprocessDepartures():

            # Initialise Departure input
            # Read in Departure data from an excel workbook
            for row in range(2, wb.departureInput.max_row + 1):
                if wb.departureInput['A' + str(row)].value == None:  # NO (more) DEPARTURES
                    print("Blank space detected in input file, terminating 'Read Input' here")
                    break
                if wb.departureInput['C' + str(row)].value == None: #SOBT are in time-format
                    SOBT = wb.departureInput['B' + str(row)].value
                    SOBT_sec = (SOBT.hour * 3600) + (SOBT.minute * 60) + SOBT.second
                    wb.departureInput['C' + str(row)].value = SOBT_sec # Used from initial schedule

                # ------------------------- Intermediate Calculations ------------------------ #

                # -------------------------- Departure Wake Category ------------------------- #

                AC_type = wb.departureInput['F' +str(row)].value
                wb.departureInput['H' +str(row)].value = wb.df_wake_WTC.at[AC_type,'WTC']

                # ------------------------------- Departure ID ------------------------------- #

                wb.runwayCalculations['R' + str(row)].value = wb.departureInput['A' + str(row)].value

                # --------------------------------- Taxi-Out --------------------------------- #

                Departure_Taxiout_mean = wb.departureInput['K' + str(row)].value
                Departure_Taxiout_SD = wb.departureInput['L' + str(row)].value
                actualTAXIOUT = random.normalvariate(Departure_Taxiout_mean, Departure_Taxiout_SD)
                wb.runwayCalculations['S' + str(row)].value = round(actualTAXIOUT,0)

                # ----------------------------------- DROT ----------------------------------- #

                if wb.departureInput['H' + str(row)].value == 'J':
                    wb.runwayCalculations['T' + str(row)].value = random.choice(wb.df_distributions['DROT_J'].dropna())
                elif wb.departureInput['H' + str(row)].value == 'H':
                    wb.runwayCalculations['T' + str(row)].value = random.choice(wb.df_distributions['DROT_H'].dropna())
                elif wb.departureInput['H' + str(row)].value == 'UM':
                    wb.runwayCalculations['T' + str(row)].value = random.choice(wb.df_distributions['DROT_UM'].dropna())
                elif wb.departureInput['H' + str(row)].value == 'M':
                    wb.runwayCalculations['T' + str(row)].value = random.choice(wb.df_distributions['DROT_M'].dropna())
                elif wb.departureInput['H' + str(row)].value == 'S':
                    wb.runwayCalculations['T' + str(row)].value = random.choice(wb.df_distributions['DROT_S'].dropna())
                elif wb.departureInput['H' + str(row)].value == 'L':
                    wb.runwayCalculations['T' + str(row)].value = random.choice(wb.df_distributions['DROT_L'].dropna())


        # ----------------------- Departure Movement Functions ----------------------- #

        def update_Departure_Delays(Current_time):
            if len(DepSTANDqueue)>0:
                for AC in list(DepSTANDqueue.keys()):
                    DepSTANDqueue_Delay = Current_time - DepSTANDqueue[AC][1]
                    DepSTANDqueue[AC][4] = DepSTANDqueue_Delay
            if len(TAXIhold)>0:
                for AC in list(TAXIhold.keys()):
                    TAXIhold_Delay = Current_time - TAXIhold[AC][7]
                    TAXIhold[AC][7] = TAXIhold_Delay
            if len(RWYqueue1)>0:
                for AC in list(RWYqueue1.keys()):
                    RWYqueue1_delay = Current_time - RWYqueue1[AC][9] #Current_time - RWYqueue entry_time
                    RWYqueue1[AC][10] = RWYqueue1_delay
            if len(RWYqueue2)>0:
                for AC in list(RWYqueue2.keys()):
                    RWYqueue2_delay = Current_time - RWYqueue2[AC][9] #Current_time - RWYqueue entry_time
                    RWYqueue2[AC][10] = RWYqueue2_delay


        def SOBTlookup(Current_time, SOBTrow):
            if SOBTrow < wb.departureInput.max_row + 1:
                if Current_time >= wb.departureInput['C' + str(SOBTrow)].value :# Current time = SOBT
                    DepSTANDqueue[SOBTrow]=[wb.departureInput['A' + str(SOBTrow)].value,wb.departureInput['C' + str(SOBTrow)].value,wb.runwayCalculations['T' + str(SOBTrow)].value,wb.departureInput['I' + str(SOBTrow)].value,0]
                    SOBTrow += 1
            return SOBTrow


        def TAXIqueue_update(Current_time):
            #   if ((len(TAXIqueue) + len(ARRIVALqueue)+ len(TAXIhold))<15) and len(DepSTANDqueue)> 0: # if there are less than 15 AC moving on the TAXIway and there's something in DepSTANDqueue
            # check who has to go first
            first_in_line_DepSTANDqueue = min(list(DepSTANDqueue.keys()))

            DepSTANDqueue[first_in_line_DepSTANDqueue].append(Current_time) #TAXIqueue entry time
            DepSTANDqueue[first_in_line_DepSTANDqueue].append(wb.runwayCalculations['S' + str(first_in_line_DepSTANDqueue)].value) #TAXI-out


            #ADD first_in_line_DepSTANDqueue to TAXIqueue:
            TAXIqueue[first_in_line_DepSTANDqueue]=DepSTANDqueue[first_in_line_DepSTANDqueue]
            del DepSTANDqueue[first_in_line_DepSTANDqueue]


        def first_in_line_TAXIqueue_func(End_time):
            first_in_line_TAXIqueue = 0
            min_TAXIqueue_out = End_time
            for AC in list(TAXIqueue.keys()):
                TAXIqueue_out = TAXIqueue[AC][5]+TAXIqueue[AC][6]
                if TAXIqueue_out<min_TAXIqueue_out:
                    min_TAXIqueue_out = TAXIqueue_out
                    first_in_line_TAXIqueue = AC
            return first_in_line_TAXIqueue


        def TAXIhold_update(Current_time, End_time):
            if len(TAXIqueue)>0:
                first_in_line_TAXIqueue = first_in_line_TAXIqueue_func(End_time)
                if Current_time >= (TAXIqueue[first_in_line_TAXIqueue][5] + TAXIqueue[first_in_line_TAXIqueue][6]): # current_time = TAXIqueue_entry_time + Taxi-out
                    TAXIqueue[first_in_line_TAXIqueue].append(Current_time) #TAXIhold entry time
                    TAXIqueue[first_in_line_TAXIqueue].append(0) #TAXIhold delay

                    #ADD first_in_line_TAXIqueue to TAXIhold
                    TAXIhold[first_in_line_TAXIqueue] = TAXIqueue[first_in_line_TAXIqueue]
                    del TAXIqueue[first_in_line_TAXIqueue]


        def transfer_to_2x4_RWYqueues(first_in_line_TAXIhold, Current_time, queueType=v['queueType']):

            # Queue selection
            if queueType == '1x8':
                maxRWYqueue1Length = 8
                previousRWYqueue = 1
            elif queueType == '2x4':
                maxRWYqueue1Length = 4
                maxRWYqueue2Length = 4
                previousRWYqueue = 2 # Forces RWYqueue1 to go first
            elif queueType == '4x2':
                maxRWYqueue1Length = 2
                maxRWYqueue2Length = 2
                maxRWYqueue3Length = 2
                maxRWYqueue4Length = 2
                previousRWYqueue = 3 # Forces RWYqueue1/2 to go first
            elif queueType == '8x1':
                maxRWYqueue1Length = 4 # Will use 2x4 methods (accounts for 4 Queues x 1 in length)
                maxRWYqueue2Length = 4 # Will use 2x4 methods (accounts for 4 Queues x 1 in length)
                previousRWYqueue = 2 # Forces RWYqueue1 to go first

            if TAXIhold[first_in_line_TAXIhold][3] in SIDqueues[0]: # First check if SID group belongs to RWYqueue1
                if len(RWYqueue1) < maxRWYqueue1Length: # if there is space in RWYqueue1 add A/C to the queue
                    TAXIhold[first_in_line_TAXIhold].append(Current_time) #RWYqueue1 entry time
                    TAXIhold[first_in_line_TAXIhold].append(0) # RWYqueue1 Delay
                    TAXIhold[first_in_line_TAXIhold].append(1) #RWYqueue used

                    RWYqueue1[first_in_line_TAXIhold] = TAXIhold[first_in_line_TAXIhold]
                    del TAXIhold[first_in_line_TAXIhold]

            elif TAXIhold[first_in_line_TAXIhold][3] in SIDqueues[1]: # First check if SID group belongs to RWYqueue2
                if len(RWYqueue2) < maxRWYqueue2Length: # if there is space in RWYqueue1 add A/C to the queue
                    TAXIhold[first_in_line_TAXIhold].append(Current_time) #RWYqueue2 entry time
                    TAXIhold[first_in_line_TAXIhold].append(0) # RWYqueue2 Delay
                    TAXIhold[first_in_line_TAXIhold].append(2) #RWYqueue used

                    RWYqueue2[first_in_line_TAXIhold] = TAXIhold[first_in_line_TAXIhold]
                    del TAXIhold[first_in_line_TAXIhold]


        def RWYqueues_update(Current_time):
            previous_in_line = 0
            first_in_line_TAXIhold = 0
            while (len(TAXIhold)>0 and (len(RWYqueue1)+len(RWYqueue2))<8): # while there is something in TAXIhold and there's space in RWY queues
                previous_in_line = first_in_line_TAXIhold
                first_in_line_TAXIhold = min(list(TAXIhold.keys()))

                if (first_in_line_TAXIhold!=0) and (previous_in_line!= first_in_line_TAXIhold):
                    transfer_to_2x4_RWYqueues(first_in_line_TAXIhold,Current_time)
                else:
                    break


        def first_in_line_RWYqueue_funct(DepOutput, End_time):
            first_in_line_RWYqueue = 0
            currentRWYqueue = 0

            #ONLY FOR 2x4
            if DepOutput==2: #for the first departure check in which queue is the first departure:
                min_entry_time1 = End_time
                min_entry_time2 = End_time
                first_in_line_RWYqueue1=0
                first_in_line_RWYqueue2=0
                if len(RWYqueue1)>0:
                    for AC in list(RWYqueue1.keys()):
                        if RWYqueue1[AC][9]<min_entry_time1:
                            min_entry_time1 = RWYqueue1[AC][9]
                            first_in_line_RWYqueue1 = AC
                if len(RWYqueue2)>0:
                    for AC in list(RWYqueue2.keys()):
                        if RWYqueue2[AC][9]<min_entry_time2:
                            min_entry_time2 = RWYqueue2[AC][9]
                            first_in_line_RWYqueue2 = AC
                if min_entry_time1<min_entry_time2:
                    first_in_line_RWYqueue = first_in_line_RWYqueue1
                    currentRWYqueue = 1
                else:
                    first_in_line_RWYqueue = first_in_line_RWYqueue2
                    currentRWYqueue = 2

            elif DepOutput!=2 :
                if wb.departureOutput['U' + str(DepOutput-1)].value == 1: #If previous departure started from queue 1

                    # use RWYqueue2
                    if len(RWYqueue2)>0: #There is smth in the queue

                        min_entry_time = End_time
                        for AC in list(RWYqueue2.keys()):
                            if RWYqueue2[AC][9]<min_entry_time:
                                min_entry_time = RWYqueue2[AC][9]
                                first_in_line_RWYqueue = AC

                        currentRWYqueue = 2
                    else: #there is nobody in RWYqueue2
                        #use RWYqueue1 again
                        min_entry_time = End_time
                        for AC in list(RWYqueue1.keys()):
                            if RWYqueue1[AC][9]<min_entry_time:
                                min_entry_time = RWYqueue1[AC][9]
                                first_in_line_RWYqueue = AC
                        currentRWYqueue = 1
                elif wb.departureOutput['U' + str(DepOutput-1)].value == 2: #If previous departure started from queue 2

                    # use RWYqueue1
                    if len(RWYqueue1)>0: #There is smth in the queue
                        min_entry_time = End_time
                        for AC in list(RWYqueue1.keys()):
                            if RWYqueue1[AC][9]<min_entry_time:
                                min_entry_time = RWYqueue1[AC][9]
                                first_in_line_RWYqueue = AC
                        currentRWYqueue = 1

                    else: #there is nobody in RWYqueue2
                        #use RWYqueue2 again
                        min_entry_time = End_time
                        for AC in list(RWYqueue2.keys()):
                            if RWYqueue2[AC][9]<min_entry_time:
                                min_entry_time = RWYqueue2[AC][9]
                                first_in_line_RWYqueue = AC
                        currentRWYqueue = 2
            return first_in_line_RWYqueue, currentRWYqueue


        def second_in_line_RWYqueues(previousRWYqueue, End_time): #used for target ADDA time
            min_entry_time = End_time
            second_in_line_RWYqueue = 0
            currentRWYqueue = 0
            if previousRWYqueue == 1: #now use queue2
                if len(RWYqueue2)>0:
                    currentRWYqueue = 2
                    for AC in list(RWYqueue2.keys()):
                        if RWYqueue2[AC][9]<min_entry_time:
                            min_entry_time = RWYqueue2[AC][9]
                            second_in_line_RWYqueue = AC
                else:
                    currentRWYqueue = 1
                    for AC in list(RWYqueue1.keys()):
                        if RWYqueue1[AC][9]<min_entry_time:
                            min_entry_time = RWYqueue1[AC][9]
                            second_in_line_RWYqueue = AC

            elif previousRWYqueue == 2: #now use 1
                if len(RWYqueue1)>0:
                    currentRWYqueue = 1
                    for AC in list(RWYqueue1.keys()):
                        if RWYqueue1[AC][9]<min_entry_time:
                            min_entry_time = RWYqueue1[AC][9]
                            second_in_line_RWYqueue = AC

                else:
                    currentRWYqueue = 2
                    for AC in list(RWYqueue2.keys()):
                        if RWYqueue2[AC][9]<min_entry_time:
                            min_entry_time = RWYqueue2[AC][9]
                            second_in_line_RWYqueue = AC
            return second_in_line_RWYqueue,currentRWYqueue


        def dep_Wake_separation(first_in_line_RWYqueue, DepOutput):
            minWakeSep = 0 # Initialise local variable (reset on each iteration)
            if DepOutput == 2: #first departure:
                minWakeSep = 0
            else:
                previousDepartureWake = wb.departureOutput['E' + str(DepOutput-1)].value
                currentDepartureWake = wb.departureInput['H' + str(first_in_line_RWYqueue)].value

                if previousDepartureWake == "J":
                    if currentDepartureWake == "J":
                        minWakeSep = 0
                    elif currentDepartureWake == "H":
                        minWakeSep = J_H_d
                    elif (currentDepartureWake == "UM") or (currentDepartureWake == "M"):
                        minWakeSep = J_M_d
                    elif (currentDepartureWake == "S") or (currentDepartureWake == "L"):
                        minWakeSep = J_L_d
                    else:
                        print("[J-] Wake Category other than normal detected - check Input file")

                elif previousDepartureWake == "H":
                    if currentDepartureWake == "J":
                        minWakeSep = 0
                    elif currentDepartureWake == "H":
                        minWakeSep = H_H_d
                    elif (currentDepartureWake == "UM") or (currentDepartureWake == "M"):
                        minWakeSep = H_M_d
                    elif (currentDepartureWake == "S") or (currentDepartureWake == "L"):
                        minWakeSep = H_M_d
                    else:
                        print("[H-] Wake Category other than normal detected - check Input file")

                elif (previousDepartureWake == "UM") or (previousDepartureWake == "M"):
                    if currentDepartureWake == "L":
                        minWakeSep = M_L_d
                    else:
                        minWakeSep = 0

                elif (previousDepartureWake == "S") or (previousDepartureWake == "S"):
                    if currentDepartureWake == "L":
                        minWakeSep = 0

                else:
                    minWakeSep = 0

            return minWakeSep


        def dep_SID_separation(first_in_line_RWYqueue, DepOutput):
            minSIDsep = 0 # Initialise local variable (reset on each iteration)
            if DepOutput == 2: #first departure:
                minSIDsep = 0
            else:
                # Compares SID groups between the previous and current A/C - then sets 'minSIDsep' variable as either altSID or sameSID
                previousDepartureSID = wb.departureOutput['F' + str(DepOutput-1)].value
                nextDepartureSID = wb.departureInput['I' + str(first_in_line_RWYqueue)].value

                if nextDepartureSID == previousDepartureSID: #IF the next departure SID is tha same as the previous departure SID => maximum separation
                    minSIDsep = v['minDep_sameSID']
                # If they are not equal, check if the SID group has some more separation rules
                elif nextDepartureSID != previousDepartureSID:
                    minSIDsep = v['minDep_altSID']
                    for item in SIDgroups:
                        if nextDepartureSID == item[0] and previousDepartureSID == item[1]:
                        #if previousDepartureSID == item[1]: # IF the previous departure SID matches the partner, apply maximum separation
                            minSIDsep = v['minDep_sameSID']
            return minSIDsep


        def departure_separation(first_in_line_RWYqueue, DepOutput):
            minDeptime = 0
            minDepLabel = ""
            #WAKE
            minWakeSep = dep_Wake_separation(first_in_line_RWYqueue, DepOutput)
            #SID
            minSIDsep = dep_SID_separation(first_in_line_RWYqueue, DepOutput)
            #compare the two and take the largest constraint
            if minSIDsep>minWakeSep:
                minDeptime = minSIDsep
                minDepLabel = "SID"
            else:
                minDeptime = minWakeSep
                minDepLabel = "WAKE"
            return minDeptime,minDepLabel


        def Dep_TAKE_OFF(Current_time, DepOutput, currentGap, End_time, seqRow):
            #if (len(RWYqueue1) != 0) or (len(RWYqueue2)!=0): #there is something in the queues:
            #print('Something in the RWYqueues')

            first_in_line_RWYqueue, currentRWYqueue = first_in_line_RWYqueue_funct(DepOutput, End_time)
            if first_in_line_RWYqueue !=0: # there's someone in line
                minDepTime,minDepLabel = departure_separation(first_in_line_RWYqueue,DepOutput)

                if DepOutput == 2: # First departure, no wake/sid constraints
                    if (currentGap > v['n']):
                        #TAKE-OFF
                        wb.departureOutput['B' + str(DepOutput)].value = int(Current_time/3600) # Dep HOUR
                        wb.departureOutput['C' + str(DepOutput)].value = Current_time # Departure RWY Entry

                        if currentRWYqueue == 1:
                            wb.departureOutput['A' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][0] # AC ID
                            wb.departureOutput['D' + str(DepOutput)].value = wb.departureOutput['C' + str(DepOutput)].value + RWYqueue1[first_in_line_RWYqueue][2] # Dep RWY EXIT = Dep RWY ENTRY + DROT
                            wb.departureOutput['E' + str(DepOutput)].value = wb.departureInput['H'+ str(first_in_line_RWYqueue)].value #WAKE
                            wb.departureOutput['F' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][3] #SID
                            wb.departureOutput['G' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][2]#DROT
                            wb.departureOutput['H' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][6]#TAXIOUT
                            wb.departureOutput['I' + str(DepOutput)].value = minDepTime#DEP MIN SEPARATION
                            wb.departureOutput['J' + str(DepOutput)].value = minDepLabel#DEP MIN SEPARATION LABEL
                            wb.departureOutput['K' + str(DepOutput)].value = currentGap#currentGap
                            wb.departureOutput['L' + str(DepOutput)].value = len(DepSTANDqueue)
                            wb.departureOutput['M' + str(DepOutput)].value = len(TAXIhold)
                            wb.departureOutput['N' + str(DepOutput)].value = len(RWYqueue1)
                            wb.departureOutput['O' + str(DepOutput)].value = len(RWYqueue2)
                            # wb.departureOutput['P' + str(DepOutput)].value = len(RWYqueue3)
                            # wb.departureOutput['Q' + str(DepOutput)].value = len(RWYqueue4)
                            wb.departureOutput['R' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][4]#DELAY DepSTANDqueue
                            wb.departureOutput['S' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][8]#DELAY TAXIhold
                            wb.departureOutput['T' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][10]#DELAY RWYqueue
                            wb.departureOutput['U' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][11]#RWYqueue USED

                            del RWYqueue1[first_in_line_RWYqueue]

                        elif currentRWYqueue == 2:
                            wb.departureOutput['A' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][0] # AC ID
                            wb.departureOutput['D' + str(DepOutput)].value = wb.departureOutput['C' + str(DepOutput)].value + RWYqueue2[first_in_line_RWYqueue][2] # Dep RWY EXIT = Dep RWY ENTRY + DROT
                            wb.departureOutput['E' + str(DepOutput)].value = wb.departureInput['H'+ str(first_in_line_RWYqueue)].value #WAKE
                            wb.departureOutput['F' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][3] #SID
                            wb.departureOutput['G' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][2]#DROT
                            wb.departureOutput['H' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][6]#TAXIOUT
                            wb.departureOutput['I' + str(DepOutput)].value = minDepTime#DEP MIN SEPARATION
                            wb.departureOutput['J' + str(DepOutput)].value = minDepLabel#DEP MIN SEPARATION LABEL
                            wb.departureOutput['K' + str(DepOutput)].value = currentGap#currentGap
                            wb.departureOutput['L' + str(DepOutput)].value = len(DepSTANDqueue)
                            wb.departureOutput['M' + str(DepOutput)].value = len(TAXIhold)
                            wb.departureOutput['N' + str(DepOutput)].value = len(RWYqueue1)
                            wb.departureOutput['O' + str(DepOutput)].value = len(RWYqueue2)
                            # wb.departureOutput['P' + str(DepOutput)].value = len(RWYqueue3)
                            # wb.departureOutput['Q' + str(DepOutput)].value = len(RWYqueue4)
                            wb.departureOutput['R' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][4]#DELAY DepSTANDqueue
                            wb.departureOutput['S' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][8]#DELAY TAXIhold
                            wb.departureOutput['T' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][10]#DELAY RWYqueue
                            wb.departureOutput['U' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][11]#RWYqueue USED
                            del RWYqueue2[first_in_line_RWYqueue]

                        DepOutput += 1
                elif DepOutput != 2:
                    if (currentGap > v['n']) and (Current_time>(wb.departureOutput['C' + str(DepOutput-1)].value)+minDepTime) :
                        #print(first_in_line_RWYqueue,' condition met', DepOutput)
                        #TAKE-OFF
                        wb.departureOutput['B' + str(DepOutput)].value = int(Current_time/3600) # Dep HOUR
                        wb.departureOutput['C' + str(DepOutput)].value = Current_time # Departure RWY Entry

                        if currentRWYqueue == 1:
                            wb.departureOutput['A' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][0] # AC ID
                            wb.departureOutput['D' + str(DepOutput)].value = wb.departureOutput['C' + str(DepOutput)].value + RWYqueue1[first_in_line_RWYqueue][2] # Dep RWY EXIT = Dep RWY ENTRY + DROT
                            wb.departureOutput['E' + str(DepOutput)].value = wb.departureInput['H'+ str(first_in_line_RWYqueue)].value #WAKE
                            wb.departureOutput['F' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][3] #SID
                            wb.departureOutput['G' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][2]#DROT
                            wb.departureOutput['H' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][6]#TAXIOUT
                            wb.departureOutput['I' + str(DepOutput)].value = minDepTime#DEP MIN SEPARATION
                            wb.departureOutput['J' + str(DepOutput)].value = minDepLabel#DEP MIN SEPARATION LABEL
                            wb.departureOutput['K' + str(DepOutput)].value = currentGap#currentGap
                            wb.departureOutput['L' + str(DepOutput)].value = len(DepSTANDqueue)
                            wb.departureOutput['M' + str(DepOutput)].value = len(TAXIhold)
                            wb.departureOutput['N' + str(DepOutput)].value = len(RWYqueue1)
                            wb.departureOutput['O' + str(DepOutput)].value = len(RWYqueue2)
                            # wb.departureOutput['P' + str(DepOutput)].value = len(RWYqueue3)
                            # wb.departureOutput['Q' + str(DepOutput)].value = len(RWYqueue4)
                            wb.departureOutput['R' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][4]#DELAY DepSTANDqueue
                            wb.departureOutput['S' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][8]#DELAY TAXIhold
                            wb.departureOutput['T' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][10]#DELAY RWYqueue
                            wb.departureOutput['U' + str(DepOutput)].value = RWYqueue1[first_in_line_RWYqueue][11]#RWYqueue USED
                            del RWYqueue1[first_in_line_RWYqueue]

                        elif currentRWYqueue == 2:
                            wb.departureOutput['A' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][0] # AC ID
                            wb.departureOutput['D' + str(DepOutput)].value = wb.departureOutput['C' + str(DepOutput)].value + RWYqueue2[first_in_line_RWYqueue][2] # Dep RWY EXIT = Dep RWY ENTRY + DROT
                            wb.departureOutput['E' + str(DepOutput)].value = wb.departureInput['H'+ str(first_in_line_RWYqueue)].value #WAKE
                            wb.departureOutput['F' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][3] #SID
                            wb.departureOutput['G' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][2]#DROT
                            wb.departureOutput['H' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][6]#TAXIOUT
                            wb.departureOutput['I' + str(DepOutput)].value = minDepTime#DEP MIN SEPARATION
                            wb.departureOutput['J' + str(DepOutput)].value = minDepLabel#DEP MIN SEPARATION LABEL
                            wb.departureOutput['K' + str(DepOutput)].value = currentGap#currentGap
                            wb.departureOutput['L' + str(DepOutput)].value = len(DepSTANDqueue)
                            wb.departureOutput['M' + str(DepOutput)].value = len(TAXIhold)
                            wb.departureOutput['N' + str(DepOutput)].value = len(RWYqueue1)
                            wb.departureOutput['O' + str(DepOutput)].value = len(RWYqueue2)
                            # wb.departureOutput['P' + str(DepOutput)].value = len(RWYqueue3)
                            # wb.departureOutput['Q' + str(DepOutput)].value = len(RWYqueue4)
                            wb.departureOutput['R' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][4]#DELAY DepSTANDqueue
                            wb.departureOutput['S' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][8]#DELAY TAXIhold
                            wb.departureOutput['T' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][10]#DELAY RWYqueue
                            wb.departureOutput['U' + str(DepOutput)].value = RWYqueue2[first_in_line_RWYqueue][11]#RWYqueue USED
                            del RWYqueue2[first_in_line_RWYqueue]

                        wb.sequenceTab['A' + str(seqRow)].value = 'D'
                        wb.sequenceTab['B' + str(seqRow)].value = wb.departureOutput['A' + str(DepOutput)].value
                        wb.sequenceTab['C' + str(seqRow)].value = wb.departureOutput['C' + str(DepOutput)].value
                        wb.sequenceTab['D' + str(seqRow)].value = wb.departureOutput['D' + str(DepOutput)].value
                        wb.sequenceTab['E' + str(seqRow)].value = wb.departureOutput['G' + str(DepOutput)].value
                        seqRow+=1
                        DepOutput += 1

            return DepOutput,seqRow


        # ------------------------ Arrival Movement Functions ------------------------ #

        def update_ArrHOLDqueue_Delay(Current_time):
            for AC in list(ArrHOLDqueue.keys()):
                ArrHOLDqueue_Delay = Current_time - ArrHOLDqueue[AC][1] # Delay = Current_time - SAE
                ArrHOLDqueue[AC][4] = ArrHOLDqueue_Delay


        def SAE_lookup(Current_time, ARRkey):
            if ARRkey != (wb.arrivalInput.max_row + 1):
                if Current_time >= wb.runwayCalculations['L' + str(ARRkey)].value : # Current_time = SAE
                    ArrHOLDqueue[ARRkey] = [wb.arrivalInput['A' + str(ARRkey)].value, wb.runwayCalculations['L' + str(ARRkey)].value, wb.runwayCalculations['C' + str(ARRkey)].value, wb.runwayCalculations['M' + str(ARRkey)].value, 0]
                    ARRkey += 1
            update_ArrHOLDqueue_Delay(Current_time)
            return ARRkey


        def update_APPqueue(Current_time, DepOutput, End_time, ArrOutput): # add to APPqueue
            #print(Current_time, ' app queue called')
            if (len(ArrHOLDqueue)>0) and (len(APPqueue)==0): # There is something in the hold but nothing on approach
                first_in_line_ArrHOLDqueue = min(list(ArrHOLDqueue.keys()))

                max_constraint = 0
                wb.arrivalOutput['I' + str(ArrOutput)].value = wb.runwayCalculations['N' + str(first_in_line_ArrHOLDqueue)].value
                wb.arrivalOutput['J' + str(ArrOutput)].value = wb.runwayCalculations['O' + str(first_in_line_ArrHOLDqueue)].value
                #target time, optimised gaps
                if v['timeBased']:
                    if (len(RWYqueue1) + len(RWYqueue2))>0: #ther is a departure ready to go
                        if (wb.arrivalInput['U' + str(first_in_line_ArrHOLDqueue)].value) == "ADDA" :    #*********to be changed
                            AROT = ArrHOLDqueue[first_in_line_ArrHOLDqueue][2]
                            firstDeparture, currentRWYqueue = first_in_line_RWYqueue_funct(DepOutput,End_time)
                            if currentRWYqueue == 1:
                                DROT1 = RWYqueue1[firstDeparture][2]
                            else:
                                DROT1 = RWYqueue2[firstDeparture][2]
                            # secondDeparture, nextRWYqueue = second_in_line_RWYqueues(currentRWYqueue,End_time)
                            # if nextRWYqueue == 1:
                            #     DROT2 = RWYqueue1[secondDeparture][2]
                            # else:
                            #     DROT2 = RWYqueue2[secondDeparture][2]
                            ADDA_target_time = AROT + DROT1 + DROT1 + x_buffer# AROT + NextDep DROT + NextDep2 DROT
                            ADDA_target_distance = time_to_distance_assumed_speed_profile_GS(first_in_line_ArrHOLDqueue, wb.D_DME, int(ADDA_target_time))#distance
                            if (v['ADDA_4DME']) and (ArrOutput>2):
                                Total_time_follow = int(DBS_actual_speed_profile((ADDA_target_distance+4),first_in_line_ArrHOLDqueue))
                                Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,ArrOutput))
                                ADDA_separation = Total_time_follow - Time_lead_4dme_to_thr

                            elif v['ADDA_THR']:
                                ADDA_separation = int(DBS_actual_speed_profile(ADDA_target_distance,first_in_line_ArrHOLDqueue)) #time
                            else:
                                ADDA_separation = int(DBS_actual_speed_profile(ADDA_target_distance,first_in_line_ArrHOLDqueue)) #time - default

                            if ADDA_separation > wb.arrivalOutput['I' + str(ArrOutput)].value:
                                wb.arrivalOutput['J' + str(ArrOutput)].value= "ADDA"
                                wb.arrivalOutput['I' + str(ArrOutput)].value = ADDA_separation

                        elif (wb.arrivalInput['U' + str(first_in_line_ArrHOLDqueue)].value) == "ADA" :
                            AROT = ArrHOLDqueue[first_in_line_ArrHOLDqueue][2]
                            firstDeparture, currentRWYqueue = first_in_line_RWYqueue_funct(DepOutput,End_time)
                            if currentRWYqueue == 1:
                                DROT1 = RWYqueue1[firstDeparture][2]
                            else:
                                DROT1 = RWYqueue2[firstDeparture][2]
                            ADA_target_time = AROT + DROT1 + x_buffer# AROT + NextDep DROT + NextDep2 DROT
                            ADA_target_distance = time_to_distance_assumed_speed_profile_GS(first_in_line_ArrHOLDqueue, wb.D_DME, int(ADA_target_time))#distance
                            if (v['ADA_4DME']) and (ArrOutput>2):
                                Total_time_follow = int(DBS_actual_speed_profile((ADA_target_distance+4),first_in_line_ArrHOLDqueue))
                                Time_lead_4dme_to_thr = int(DBS_actual_speed_profile(4,ArrOutput-1))
                                ADA_separation = Total_time_follow - Time_lead_4dme_to_thr

                            elif v['ADA_THR']:
                                ADA_separation = int(DBS_actual_speed_profile(ADA_target_distance,first_in_line_ArrHOLDqueue)) #time
                            else:
                                ADA_separation = int(DBS_actual_speed_profile(ADA_target_distance,first_in_line_ArrHOLDqueue)) #time

                            if ADA_separation > wb.arrivalOutput['I' + str(ArrOutput)].value:
                                wb.arrivalOutput['J' + str(ArrOutput)].value = "ADA"
                                wb.arrivalOutput['I' + str(ArrOutput)].value = ADA_separation
                    # else: # no departure ready to go
                        # max_constraint = wb.arrivalOutput['I' + str(ArrOutput)].value
                # elif v['distanceBased']:
                max_constraint = wb.arrivalOutput['I' + str(ArrOutput)].value

                # print(Current_time, ArrOutput, ' | max_constraint = ', max_constraint)

                # if max_constraint != 0:

                ArrHOLDqueue[first_in_line_ArrHOLDqueue].append(Current_time)#APPqueue entry time
                ALT = ArrHOLDqueue[first_in_line_ArrHOLDqueue][5]+ int(max_constraint) #(ALT = APPqueue_entry_time + max_constraint)
                ArrHOLDqueue[first_in_line_ArrHOLDqueue].append(ALT)
                RWY_EXIT = ArrHOLDqueue[first_in_line_ArrHOLDqueue][6] + ArrHOLDqueue[first_in_line_ArrHOLDqueue][2] # ALT + AROT
                ArrHOLDqueue[first_in_line_ArrHOLDqueue].append(RWY_EXIT)

                APPqueue[first_in_line_ArrHOLDqueue]=ArrHOLDqueue[first_in_line_ArrHOLDqueue]

                del ArrHOLDqueue[first_in_line_ArrHOLDqueue]
                #print to sequence tab


        def Arr_LANDING(Current_time, ArrOutput, first_in_line_APPqueue, seqRow):
            # if len(APPqueue)!=0:
            #     first_in_line_APPqueue = min(list(APPqueue.keys()))#there is only one AC in the APPqueue
            #     #print('There is something in the APPqueue')

            #     #print('NEXT ARRIVAL = ', AC)
            #     if Current_time == APPqueue[first_in_line_APPqueue][6]: #it's time to land
            if RWY_status == "D":
                print('*** GO AROUND ***', APPqueue[first_in_line_APPqueue])
                del APPqueue[first_in_line_APPqueue]
                goAroundHour = int(Current_time/3600)
                if goAroundHour in list(GoAroundCount.keys()): #if there was already a goAround at that hour:
                    GoAroundCount[goAroundHour].append(1)
                else:
                    GoAroundCount[goAroundHour]=[1]

            elif RWY_status == "E":
                #print(RWY_status)
                wb.arrivalOutput['A' + str(ArrOutput)].value = wb.arrivalInput['A' + str(first_in_line_APPqueue)].value #ARR ID
                wb.arrivalOutput['B' + str(ArrOutput)].value = int(Current_time/3600) #LANDING HOUR
                wb.arrivalOutput['C' + str(ArrOutput)].value = Current_time #ACTUAL LANDING TIME
                wb.arrivalOutput['D' + str(ArrOutput)].value = APPqueue[first_in_line_APPqueue][7] # RWY EXIT
                wb.arrivalOutput['E' + str(ArrOutput)].value = wb.runwayCalculations['U' + str(first_in_line_APPqueue)].value #WAKE
                wb.arrivalOutput['F' + str(ArrOutput)].value = wb.arrivalOutput['D' + str(ArrOutput)].value + wb.runwayCalculations['B' + str(first_in_line_APPqueue)].value #In blocks time
                wb.arrivalOutput['G' + str(ArrOutput)].value = wb.runwayCalculations['C' + str(first_in_line_APPqueue)].value#AROT
                wb.arrivalOutput['H' + str(ArrOutput)].value = wb.runwayCalculations['B' + str(first_in_line_APPqueue)].value# Taxi-in duration


                wb.arrivalOutput['K' + str(ArrOutput)].value = len(ArrHOLDqueue)#length ArrHOLDqueue
                wb.arrivalOutput['L' + str(ArrOutput)].value = APPqueue[first_in_line_APPqueue][4]# ArrHOLDqueue delay
                AIBT = wb.arrivalOutput['F' + str(ArrOutput)].value

                #Add Arrival to ARRIVALqueue
                ARRIVALqueue[first_in_line_APPqueue]=[wb.arrivalOutput['A' + str(ArrOutput)].value, AIBT, ArrOutput]
                #print('ARRIVALqueue = ', list(ARRIVALqueue.keys()))
                del APPqueue[first_in_line_APPqueue]
                wb.sequenceTab['A' + str(seqRow)].value = 'A'
                wb.sequenceTab['B' + str(seqRow)].value = wb.arrivalOutput['A' + str(ArrOutput)].value
                wb.sequenceTab['C' + str(seqRow)].value = wb.arrivalOutput['C' + str(ArrOutput)].value
                wb.sequenceTab['D' + str(seqRow)].value = wb.arrivalOutput['D' + str(ArrOutput)].value
                wb.sequenceTab['E' + str(seqRow)].value = wb.arrivalOutput['G' + str(ArrOutput)].value
                ArrOutput+=1
                seqRow += 1

            return ArrOutput, seqRow


        def first_in_line_ARRIVALqueue_func(End_time):
            min_IBT = End_time
            first_in_line_ARRIVALqueue = 0
            for AC in list(ARRIVALqueue.keys()):
                if ARRIVALqueue[AC][1]<min_IBT:
                    min_IBT=ARRIVALqueue[AC][1]
                    first_in_line_ARRIVALqueue = AC
            return first_in_line_ARRIVALqueue


        def update_ARRIVALqueue(Current_time, End_time):
            if len(ARRIVALqueue)>0:

                #Check first in line in arrival queue
                first_in_line_ARRIVALqueue = first_in_line_ARRIVALqueue_func(End_time)

                if Current_time > ARRIVALqueue[first_in_line_ARRIVALqueue][1]:
                    #print(Current_time, 'ARR { ',AC,' } deleted from ARRIVALqueue ')
                    del ARRIVALqueue[first_in_line_ARRIVALqueue]


        def update_currentGap(Current_time, End_time):
            if RWY_status == "E" or RWY_status == "D":
                if len(APPqueue)==0: #Nothing in the queue
                    currentGap = End_time # Huuuuge currentGap
                else:
                    next_Arrival = min(list(APPqueue.keys())) # should be only one key in the list
                    currentGap = APPqueue[next_Arrival][6] - Current_time # ALT - Current_time
            elif RWY_status == "A":
                currentGap = 0
            return currentGap


        # --------------------------------- Model Run -------------------------------- #

        # Input pre-processing
        preprocessArrivals()
        preprocessDepartures()

        print('distanceBased = ',v['distanceBased'])
        print('timeBased =',v['timeBased'])

        while Current_time < End_time:
            if RWY_status == "E":
                if wb.departureInput.max_row > 0: #there are departures
                    SOBTrow = SOBTlookup(Current_time, SOBTrow)
                    if ((len(TAXIqueue) + len(ARRIVALqueue)+ len(TAXIhold))<15) and len(DepSTANDqueue)> 0: # if there are less than 15 AC moving on the TAXIway and there's something in DepSTANDqueue
                        TAXIqueue_update(Current_time)
                    TAXIhold_update(Current_time,End_time)
                    RWYqueues_update(Current_time)
                    update_Departure_Delays(Current_time)
                if wb.arrivalInput.max_row > 0: # there are arrivals
                    ARRkey = SAE_lookup(Current_time, ARRkey)
                    if len(APPqueue) == 0:
                        update_APPqueue(Current_time,DepOutput,End_time,ArrOutput)
                    update_ARRIVALqueue(Current_time,End_time)
                    currentGap = update_currentGap(Current_time, End_time)
                else:#if there aren't any arrivals
                    currentGap = End_time #huuuuuge gap
                #DEPARTURES TAKE OFF
                if wb.departureInput.max_row > 0:
                    if (len(RWYqueue1)+len(RWYqueue2))>0:#there is something waiting to takeoff
                        #print('TAKE OFF called')
                        DepOutput,seqRow = Dep_TAKE_OFF(Current_time,DepOutput,currentGap,End_time,seqRow)
                        #print('dep took off')
                        # Note : DepOurputROW was already increased so (DepOutputROW-1) will reffer to the current departure
                        #if type(wb.departureOutput['C' + str(DepOutput-1)].value) == int:
                            #print(wb.departureOutput['C' + str(DepOutput-1)].value)
                        if Current_time < wb.departureOutput['D' + str(DepOutput-1)].value : # while the Departure is still on the runway
                            #print(Current_time,' Departure {',(DepOutput-1),'} is about to take-off')
                            RWY_status = "D"
                #ARRIVALS LANDING
                if wb.arrivalInput.max_row > 0:
                    if len(APPqueue)!=0:
                        first_in_line_APPqueue = min(list(APPqueue.keys()))#there is only one AC in the APPqueue
                        if Current_time == APPqueue[first_in_line_APPqueue][6]: #it's time to land
                            #print('Current_time = ', Current_time, '| ALT = ',APPqueue[first_in_line_APPqueue][6])
                            ArrOutput,seqRow = Arr_LANDING(Current_time, ArrOutput,first_in_line_APPqueue,seqRow)
                            if Current_time < wb.arrivalOutput['D' + str(ArrOutput-1)].value : #while Arrival is still on the runway
                                #print(Current_time,' Arrival {',ArrOutput-1,'} is about to land ')
                                RWY_status = "A"
            elif RWY_status == "D":
                #print(Current_time,' | ', RWY_status)
                SOBTrow = SOBTlookup(Current_time, SOBTrow)
                if ((len(TAXIqueue) + len(ARRIVALqueue)+ len(TAXIhold))<15) and len(DepSTANDqueue)> 0: # if there are less than 15 AC moving on the TAXIway and there's something in DepSTANDqueue
                    TAXIqueue_update(Current_time)
                TAXIhold_update(Current_time,End_time)
                RWYqueues_update(Current_time)
                update_Departure_Delays(Current_time)
                if wb.arrivalInput.max_row > 0: #there are arrivals
                    ARRkey = SAE_lookup(Current_time, ARRkey)
                    if len(APPqueue) == 0:
                        update_APPqueue(Current_time,DepOutput,End_time,ArrOutput)
                    update_ARRIVALqueue(Current_time,End_time)
                    currentGap = update_currentGap(Current_time,End_time)
                else:#if there aren't any arrivals
                    currentGap = End_time #huuuuuge gap
                if Current_time == wb.departureOutput['D' + str(DepOutput-1)].value : # when current_time > departure RWY_EXIT the rwy is empty again
                    RWY_status = "E"
                #ARRIVALS LANDING (GO-AROUND case)
                if wb.arrivalInput.max_row > 0:
                    if len(APPqueue)!=0:
                        first_in_line_APPqueue = min(list(APPqueue.keys()))#there is only one AC in the APPqueue
                        if Current_time == APPqueue[first_in_line_APPqueue][6]: #it's time to land
                            #print('It is time to land but GOaround')
                            ArrOutput,seqRow = Arr_LANDING(Current_time, ArrOutput,first_in_line_APPqueue,seqRow)
                            #print(ArrOutput,'******GO AROUND************')
            elif RWY_status == "A":
                #print(Current_time,' | ', RWY_status)
                if wb.departureInput.max_row > 0: #there are departures
                    SOBTrow = SOBTlookup(Current_time, SOBTrow)
                    if ((len(TAXIqueue) + len(ARRIVALqueue)+ len(TAXIhold))<15) and len(DepSTANDqueue)> 0: # if there are less than 15 AC moving on the TAXIway and there's something in DepSTANDqueue
                        TAXIqueue_update(Current_time)
                    TAXIhold_update(Current_time,End_time)
                    RWYqueues_update(Current_time)
                    update_Departure_Delays(Current_time)

                ARRkey = SAE_lookup(Current_time, ARRkey)
                if len(APPqueue) == 0:
                    update_APPqueue(Current_time,DepOutput,End_time,ArrOutput)
                update_ARRIVALqueue(Current_time,End_time)
                currentGap = update_currentGap(Current_time,End_time)

                if Current_time == wb.arrivalOutput['D' + str(ArrOutput-1)].value : #while Arrival is still on the runway
                    RWY_status = "E"
            Current_time += 1

        # ---------------------------- Buffer Calculations --------------------------- #

        bufferRow = 2

        for row in range (2, (wb.sequenceTab.max_row-2)):
            if (wb.sequenceTab['A'+str(row)].value == "A") and (wb.sequenceTab['A'+str(row+1)].value == "D") and  (wb.sequenceTab['A'+str(row+2)].value == "A") :#ADA sequence:
                wb.sequenceTab['G' + str(bufferRow)].value = (wb.sequenceTab['C'+str(row+1)].value - wb.sequenceTab['D'+str(row)].value) + (wb.sequenceTab['C'+str(row+2)].value - wb.sequenceTab['D'+str(row+1)].value)
                wb.sequenceTab['F' + str(bufferRow)].value = wb.sequenceTab['B' + str(row)].value
                bufferRow+=1

        # -------------------- Throughput and Delays Calculations -------------------- #

        min_thr_HOUR = min(wb.arrivalOutput['B' + str(2)].value,wb.departureOutput['B' + str(2)].value)
        print('wb.arrivalOutput["B" + str(ArrOutput-1)].value = ',wb.arrivalOutput['B' + str(ArrOutput-1)].value)
        print('wb.departureOutput["B" + str(DepOutput-1)].value = ', wb.departureOutput['B' + str(DepOutput-1)].value)
        max_thr_HOUR = min(wb.arrivalOutput['B' + str(ArrOutput-1)].value,wb.departureOutput['B' + str(DepOutput-1)].value)
        diff_thr_HOUR = max_thr_HOUR-min_thr_HOUR

        for row in range(2,(diff_thr_HOUR+3)):
            dep_thr_count = 0
            arr_thr_count = 0
            wb.throughputTab['A' + str(row)].value = min_thr_HOUR
            for i in range(2, (DepOutput)):
                if wb.departureOutput['B' + str(i)].value == None: #no departures
                    break
                if wb.departureOutput['B' + str(i)].value == min_thr_HOUR:
                    dep_thr_count +=1
            wb.throughputTab['B' + str(row)].value = dep_thr_count
            for i in range(2, (ArrOutput)):
                if wb.arrivalOutput['B' + str(i)].value == None: #No arrivals
                    break
                if wb.arrivalOutput['B' + str(i)].value == min_thr_HOUR:
                    arr_thr_count +=1
            wb.throughputTab['C' + str(row)].value = arr_thr_count
            wb.throughputTab['D' + str(row)].value = wb.throughputTab['B' + str(row)].value + wb.throughputTab['C' + str(row)].value
            total_thr = wb.throughputTab['D' + str(row)].value
            throughput.append(total_thr)
            if min_thr_HOUR in list(GoAroundCount.keys()):#there was at least a goAround at that hour:
                wb.throughputTab['E' + str(row)].value = sum(GoAroundCount[min_thr_HOUR])
            else:
                wb.throughputTab['E' + str(row)].value = 0
            min_thr_HOUR +=1

        # ---------------------------------- Delays ---------------------------------- #

        for row in range(2, DepOutput):
            wb.delayTab['A' + str(row)].value = wb.departureOutput['A' + str(row)].value
            wb.delayTab['B' + str(row)].value = wb.departureOutput['B' + str(row)].value
            wb.delayTab['C' + str(row)].value = wb.departureOutput['T' + str(row)].value + wb.departureOutput['S' + str(row)].value
            wb.delayTab['D' + str(row)].value = wb.departureOutput['R' + str(row)].value

        for row in range(2, ArrOutput):
            wb.delayTab['I' + str(row)].value = wb.arrivalOutput['A' + str(row)].value
            wb.delayTab['J' + str(row)].value = wb.arrivalOutput['B' + str(row)].value
            wb.delayTab['K' + str(row)].value = wb.arrivalOutput['L' + str(row)].value

        number_of_goArounds_queued = 0
        for i in list(GoAroundCount.keys()):
            number_of_goArounds_queued+=sum(GoAroundCount[i])

        print('End_time = ',End_time)
        print("There are [", str(len(TAXIhold)),"] Departure A/C remaining in the TAXI ,", str(len(RWYqueue1)+len(RWYqueue2)),"Departure A/C remaining in the RWY queues,")
        print("There are [", str(len(ArrHOLDqueue)),"] Arrivals remaining in the Arrival Hold Queue ,", str(len(APPqueue)),"Arrivals remaining in the APPqueue,")
        print("Final number of go-around Arrival cases (Queued):", number_of_goArounds_queued)
        print("Model took %s seconds to run" % round((time.time() - program_runtime_start),2))

        if (len(DepSTANDqueue)>0 or len(TAXIhold)>0):
            print("ERROR!!!  Check DEPARTURES")

        if (len(ARRIVALqueue)>0) or (len(APPqueue)>0) or (len(ArrHOLDqueue)>0):
            print("ERROR!!!  Check ARRIVALS")

        if not v['avgThr']:
            output_extension = time.strftime("%H_%M", time.localtime(time.time()))
            wb.throughputTab['F' + str(1)].value = 'Difference in thr averages'
            extra_diff=[0]*(wb.throughputTab.max_row-1)
            difference.append(extra_diff)
            wb.throughputTab['F' + str(2)].value = str(difference)
            parentFrame.name_output_file = 'OUTPUT_RAPID_v3.0_' + str(output_extension) +  '.xlsx'
            wb.wb.save(parentFrame.name_output_file) # Choose file name once complete?
            iter1 += 1
        else:
            big_list.append(throughput)
            average_run = []
            diff2=[]
            diff=0

            average_hour = 0
            summ = 0
            if maxIter <2:
                averages.append(throughput)
                maxIter +=1
            elif (maxIter >=2) and (maxIter <10):       # minimum number of runs = 10
                for j in range(0,len(throughput)):
                    for i in range (0, len(big_list)):  #len(biglist)
                        print('element sum = ',big_list[i][j])
                        summ = summ + big_list[i][j]

                    average_hour = summ/ len(big_list)
                    average_run.append(average_hour)
                    print('average_run=' ,average_run)
                    summ = 0
                averages.append(average_run)
                average_run=[]
                print(averages)
                for j in range(0,len(averages[0])):
                    print('______________________ ', j)
                    compare = averages[len(averages)-1][j]
                    print('last run = ',compare)
                    diff = compare - averages[len(averages)-2][j]
                    print('diff = ',diff)
                    diff2.append(diff)
                    print('diff2 = ',diff2)

                difference.append(diff2)
                diff2=[]
                print('diff = ',diff2)
                print('difference list', difference)


                maxIter +=1
                summ=0
            else:
                for j in range(0,len(throughput)):
                    for i in range (0, len(big_list)):  #len(biglist)
                        print('element sum = ',big_list[i][j])
                        summ = summ + big_list[i][j]

                    average_hour = summ/ len(big_list)
                    average_run.append(average_hour)
                    print('average_run=' ,average_run)
                    summ = 0
                averages.append(average_run)
                average_run=[]
                print(averages)
                for j in range(0,len(averages[0])):
                    print('______________________ ', j)
                    compare = averages[len(averages)-1][j]
                    print('last run = ',compare)
                    diff = compare - averages[len(averages)-2][j]
                    print('diff = ',diff)
                    diff2.append(diff)
                    print('diff2 = ',diff2)

                difference.append(diff2)
                diff2=[]
                print('diff = ',diff2)
                print('difference list', difference)
                for i in range(0, len(difference[0])):
                    print("OK so far")
                    if (difference[len(difference)-1][i] <= 0.1) and (difference[len(difference)-1][i] >=-0.1) :
                        print('####### difference in averages = ',difference[len(difference)-1][i])
                        iter2 = 0
                    else:
                        print('############condition false')
                        iter2 = 1

                        break
                summ=0
                if iter2 == 0:
                    print('maxRuns 1 =', maxIter)
                    wb.throughputTab['F' + str(1)].value = 'Difference in thr averages'
                    extra_diff=[0]*(wb.throughputTab.max_row-1)

                    difference.append(extra_diff)
                    wb.throughputTab['F' + str(2)].value = str(difference)
                    output_extension = time.strftime("%H_%M", time.localtime(time.time()))
                    output_extension2 = iter1+1
                    wb.arrivalOutput.delete_cols(13)
                    wb.arrivalOutput.delete_cols(13)
                    parentFrame.name_output_file = 'OUTPUT_RAPID_v3.0_' + str(output_extension) + '_iteration_' + str(output_extension2) +  '.xlsx'
                    wb.wb.save(parentFrame.name_output_file) # Choose file name once complete?
                else:
                    maxIter += 1
                    print('maxRuns 2 =', maxIter)
            iter1 += 1
